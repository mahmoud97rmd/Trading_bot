"""
Gold Scalper Bot — v5.2 (Triple-Engine Comparison Backtest)
Strategy : Crypto World TEMA - Adaptive DEMA/HullMA
"""

import asyncio
import aiohttp
import json
import os
import pandas as pd
import numpy as np
from datetime import datetime, timedelta, timezone
from metaapi_cloud_sdk import MetaApi
from aiohttp import web

# ─────────────────────────────────────────────────────────────
# CONFIGURATION
# ─────────────────────────────────────────────────────────────
METAAPI_TOKEN = os.environ.get('METAAPI_TOKEN', 'YOUR_METAAPI_TOKEN')
ACCOUNT_ID    = os.environ.get('ACCOUNT_ID',    'YOUR_ACCOUNT_ID')
TG_TOKEN      = os.environ.get('TG_TOKEN',      '8647261254:AAHICfEZh9o2a2nFAEKvOrM1LqkMBU4UOYg')

# OANDA REST API
OANDA_ACCOUNT  = os.environ.get('OANDA_ACCOUNT', '101-004-28533521-003')
OANDA_TOKEN    = os.environ.get('OANDA_TOKEN',   '0e282d5a3e65ad6fdd809e2c195bb1cd-9e2158e12fa13840e030ee3081b36fab')
OANDA_SYMBOL   = 'XAU_USD'
OANDA_BASE_URL = 'https://api-fxpractice.oanda.com/v3'  

_TFS         = ['1m', '2m', '3m', '5m', '10m', '15m', '30m', '1h', '2h', '4h']
_HTF_OPTIONS = [
    'None',
    '2m','3m','4m','5m','6m','8m','10m','12m','15m',
    '20m','24m','30m','45m','48m',
    '1h','90m','2h','3h','4h','6h','8h','12h','1d',
]

_DEFAULT_HTF = {
    '1m': '15m', '2m': '15m', '3m': '15m',  '5m': '15m',
    '10m': '30m','15m': '30m','30m': '4h',  '1h': '4h',
    '2h': '4h',  '4h': '1d',
}

DD_LIMIT_PCT = 0.03

# ─────────────────────────────────────────────────────────────
# GLOBAL HTTP SESSION
# ─────────────────────────────────────────────────────────────
_http: aiohttp.ClientSession | None = None

def get_http() -> aiohttp.ClientSession:
    global _http
    if _http is None or _http.closed:
        connector = aiohttp.TCPConnector(limit=20, ttl_dns_cache=300)
        timeout   = aiohttp.ClientTimeout(total=30, connect=10)
        _http     = aiohttp.ClientSession(connector=connector, timeout=timeout)
    return _http

def c_log(msg: str) -> None:
    dam = (datetime.now(timezone.utc) + timedelta(hours=3)).strftime('%H:%M:%S')
    print(f"[{dam} DAM] {msg}", flush=True)

# ─────────────────────────────────────────────────────────────
# GLOBAL STATE
# ─────────────────────────────────────────────────────────────
bot_state: dict = {
    'status':           'RUNNING',
    'symbol':           'XAUUSDm',
    'live_connected':   False,
    'chat_id':          None,
    'last_update_id':   0,
    'is_backtesting':   False,
    'connection_obj':   None,
    'account_obj':      None,
    'timeframes':       _TFS,
    'active_tfs': {
        '1m': True, '2m': True, '3m': True, '5m': True,
        '10m': True, '15m': True, '30m': False,
        '1h': False,  '2h': False, '4h': False,
    },
    'htf_per_tf': dict(_DEFAULT_HTF),

    # ── TEMA Strategy ──
    'use_mtf':      True,
    'use_ma':       True,
    'basis_type':   'DEMA',
    'primary_basis_type': 'DEMA',      
    'basis_len':    30,          
    'consecutive_losses': 0,     
    'tracked_positions': {},     
    'ma_len_per_tf': {
        '1m': 30, '2m': 30, '3m': 30, '5m': 30,
        '10m': 20,'15m': 15,'30m': 10,
        '1h': 10, '2h': 10, '4h': 10,
    },
    'use_trailing': False,
    'trail_points': 200,
    'trail_offset': 400,

    # Risk
    'lot_size':         0.05,
    'pip_value':        0.1,
    'spread_pips':      2.2,
    'use_max_spread':   True,
    'max_spread_pips':  3.0,
    'tp_pips': {
        '1m': 50,  '2m': 60,  '3m': 60,   '5m': 60,
        '10m': 70, '15m': 70, '30m': 120,
        '1h': 150, '2h': 200, '4h': 280,
    },
    'sl_pips': {
        '1m': 100, '2m': 100, '3m': 110,  '5m': 110,
        '10m': 110,'15m': 110,'30m': 200,
        '1h': 250, '2h': 300, '4h': 400,
    },
    'use_be': False,

    'menu_button_map': {},
    'sod_balance':  None,
    'sod_date':     None,
    'dd_triggered': False,
    'daily_target_enabled':    False,
    'daily_target_usd':        100.0,
    'profit_target_triggered': False,
    'daily_loss_enabled':      False,
    'daily_loss_usd':          100.0,
    'loss_limit_triggered':    False,
    'use_danger_filter': True,
    'market_data':      {tf: 'Waiting...' for tf in _TFS},
    'last_signal_time': {tf: None for tf in _TFS},
    'last_poll_ok': 0.0,
}

# ─────────────────────────────────────────────────────────────
# TIME FILTER  (Damascus = UTC+3)
# ─────────────────────────────────────────────────────────────
_BLOCKED_DAMASCUS_HOURS = {13, 18, 21, 22}

def is_blocked_time(dt_utc: datetime) -> bool:
    return (dt_utc.hour + 3) % 24 in _BLOCKED_DAMASCUS_HOURS

def blocked_time_label(dt_utc: datetime) -> str:
    h = (dt_utc.hour + 3) % 24
    return f'Danger Zone ({h}:xx Damascus)' if h in _BLOCKED_DAMASCUS_HOURS else ''

# ─────────────────────────────────────────────────────────────
# DAILY DRAWDOWN PROTECTOR
# ─────────────────────────────────────────────────────────────
async def _capture_sod_balance() -> None:
    if not (bot_state['live_connected'] and bot_state['connection_obj']): return
    try:
        info  = await bot_state['connection_obj'].get_account_information()
        bal   = float(info.get('balance', 0))
        today = datetime.now(timezone.utc).strftime('%Y-%m-%d')
        
        # --- نظام التصفير اليومي للمؤشر ---
        if bot_state.get('sod_date') and bot_state['sod_date'] != today:
            p_ma = bot_state.get('primary_basis_type', 'DEMA')
            if bot_state.get('basis_type') != p_ma:
                bot_state['basis_type'] = p_ma
                bot_state['consecutive_losses'] = 0
                for tf in _TFS: bot_state[f'1m_base_{tf}'] = []
                await send_tg_msg(f"🌅 <b>يوم تداول جديد!</b>\nتم تصفير عداد الخسائر والعودة إلى المؤشر الأساسي <b>{p_ma}</b>")
        # ----------------------------------
        
        bot_state['sod_balance']           = bal
        bot_state['sod_date']              = today
        bot_state['dd_triggered']          = False
        bot_state['profit_target_triggered'] = False
        bot_state['loss_limit_triggered']    = False
        target_line = f'\n🎯 Daily Target: <b>${bot_state["daily_target_usd"]:.2f}</b>\nBot pauses at equity: <b>${bal + bot_state["daily_target_usd"]:.2f}</b>' if bot_state['daily_target_enabled'] else ''
        loss_line = f'\n🛑 Daily Loss Limit: <b>${bot_state["daily_loss_usd"]:.2f}</b>\nBot pauses at equity: <b>${bal - bot_state["daily_loss_usd"]:.2f}</b>' if bot_state['daily_loss_enabled'] else ''
        await send_tg_msg(f'📅 <b>New Day — SOD Snapshot</b>\nOpening Balance: <b>${bal:.2f}</b>\nDaily Loss Limit (3%): <b>${bal * DD_LIMIT_PCT:.2f}</b>\nBot pauses at equity: <b>${bal * (1 - DD_LIMIT_PCT):.2f}</b>{target_line}{loss_line}')
    except Exception as e: c_log(f'DD capture SOD error: {e}')

async def daily_drawdown_monitor() -> None:
    while True:
        await asyncio.sleep(30)
        try:
            if not (bot_state['live_connected'] and bot_state['connection_obj']): continue
            today = datetime.now(timezone.utc).strftime('%Y-%m-%d')
            if bot_state['sod_date'] != today or bot_state['sod_balance'] is None:
                await _capture_sod_balance(); continue
            if bot_state['dd_triggered']: continue
            info     = await bot_state['connection_obj'].get_account_information()
            equity   = float(info.get('equity', 0)); sod = bot_state['sod_balance']
            limit    = sod * (1 - DD_LIMIT_PCT); used_pct = round((sod - equity) / sod * 100, 2) if sod else 0
            if equity <= limit:
                bot_state['status'] = 'PAUSED'; bot_state['dd_triggered'] = True; closed = 0
                try:
                    positions = await bot_state['connection_obj'].get_positions()
                    for p in positions: await bot_state['connection_obj'].close_position(p['id']); closed += 1
                except Exception as ce: c_log(f'DD close error: {ce}')
                await send_tg_msg(f'🚨 <b>BOT AUTO-PAUSED — DD LIMIT REACHED</b> 🚨\n\n💰 SOD:    <b>${sod:.2f}</b>\n📉 Equity: <b>${equity:.2f}</b>\n📊 Loss:   <b>${sod-equity:.2f} ({used_pct}%)</b>\n\n{"Closed " + str(closed) + " position(s)." if closed else "No open positions."}\nPress <b>RUN ▶</b> to resume tomorrow.', get_main_keyboard())
        except Exception as e: c_log(f'DD monitor error: {e}')

# ─────────────────────────────────────────────────────────────
# DAILY PROFIT TARGET
# ─────────────────────────────────────────────────────────────
async def _close_all_positions() -> int:
    closed = 0
    try:
        positions = await bot_state['connection_obj'].get_positions()
        for p in positions: await bot_state['connection_obj'].close_position(p['id']); closed += 1
    except Exception as ce: c_log(f'close_all_positions error: {ce}')
    return closed

async def daily_profit_target_monitor() -> None:
    while True:
        await asyncio.sleep(30)
        try:
            if not (bot_state['daily_target_enabled'] or bot_state['daily_loss_enabled']): continue
            if not (bot_state['live_connected'] and bot_state['connection_obj']): continue
            today = datetime.now(timezone.utc).strftime('%Y-%m-%d')
            if bot_state['sod_date'] != today or bot_state['sod_balance'] is None: continue

            info   = await bot_state['connection_obj'].get_account_information()
            equity = float(info.get('equity', 0)); sod = bot_state['sod_balance']; pnl = equity - sod

            if bot_state['daily_target_enabled'] and not bot_state['profit_target_triggered']:
                target = bot_state['daily_target_usd']
                if pnl >= target:
                    bot_state['status'] = 'PAUSED'; bot_state['profit_target_triggered'] = True; closed = await _close_all_positions()
                    await send_tg_msg(f'🎯🎯🎯 <b>DAILY TARGET REACHED!</b> 🎯🎯🎯\n\n💰 SOD:    <b>${sod:.2f}</b>\n📈 Equity: <b>${equity:.2f}</b>\n✅ Profit: <b>${pnl:.2f}</b>  (target: ${target:.2f})\n\n{"Closed " + str(closed) + " position(s) — profit locked in." if closed else "No open positions."}\n\nBot is now <b>PAUSED</b>.\nPress <b>RUN ▶</b> to resume (target resets tomorrow).', get_main_keyboard())
                    continue

            if bot_state['daily_loss_enabled'] and not bot_state['loss_limit_triggered']:
                limit = bot_state['daily_loss_usd']
                if pnl <= -limit:
                    bot_state['status'] = 'PAUSED'; bot_state['loss_limit_triggered'] = True; closed = await _close_all_positions()
                    await send_tg_msg(f'🛑🛑🛑 <b>DAILY LOSS LIMIT HIT!</b> 🛑🛑🛑\n\n💰 SOD:    <b>${sod:.2f}</b>\n📉 Equity: <b>${equity:.2f}</b>\n❌ Loss:   <b>${pnl:.2f}</b>  (limit: -${limit:.2f})\n\n{"Closed " + str(closed) + " position(s) — losses stopped." if closed else "No open positions."}\n\nBot is now <b>PAUSED</b>.\nPress <b>RUN ▶</b> to resume (limit resets tomorrow).', get_main_keyboard())
        except Exception as e: c_log(f'Profit/Loss target monitor error: {e}')

# ─────────────────────────────────────────────────────────────
# MA / TEMA ENGINE
# ─────────────────────────────────────────────────────────────
def _ema(s: pd.Series, n: int) -> pd.Series: return s.ewm(span=n, adjust=False).mean()
def _sma(s: pd.Series, n: int) -> pd.Series: return s.rolling(n).mean()
def _wma(s: pd.Series, n: int) -> pd.Series: w = np.arange(1, n + 1, dtype=float); return s.rolling(n).apply(lambda x: np.dot(x, w) / w.sum(), raw=True)
def _dema(s: pd.Series, n: int) -> pd.Series: e = _ema(s, n); return 2 * e - _ema(e, n)
def _tema(s: pd.Series, n: int) -> pd.Series: e1 = _ema(s, n); e2 = _ema(e1, n); e3 = _ema(e2, n); return 3 * (e1 - e2) + e3
def _hull(s: pd.Series, n: int) -> pd.Series: half = max(1, n // 2); sq = max(1, round(np.sqrt(n))); return _wma(2 * _wma(s, half) - _wma(s, n), sq)
def _lsma(s: pd.Series, n: int) -> pd.Series: return s.rolling(n).apply(lambda x: np.polyval(np.polyfit(range(n), x, 1), n - 1), raw=True)
def _smma(s: pd.Series, n: int) -> pd.Series:
    out = s.copy() * np.nan
    out.iloc[n - 1] = s.iloc[:n].mean()
    for i in range(n, len(s)): out.iloc[i] = (out.iloc[i-1] * (n-1) + s.iloc[i]) / n
    return out

def variant(ma_type: str, s: pd.Series, n: int) -> pd.Series:
    t = ma_type.upper()
    if t == 'EMA': return _ema(s, n)
    if t == 'DEMA': return _dema(s, n)
    if t == 'TEMA': return _tema(s, n)
    if t == 'WMA': return _wma(s, n)
    if t == 'SMMA': return _smma(s, n)
    if t == 'HULLMA': return _hull(s, n)
    if t == 'LSMA': return _lsma(s, n)
    return _sma(s, n)

_HTF_RULES = {
    '1m':'1min', '2m':'2min', '3m':'3min', '4m':'4min', '5m':'5min', '6m':'6min', '8m':'8min', '10m':'10min',
    '12m':'12min','15m':'15min','20m':'20min','24m':'24min', '30m':'30min','45m':'45min','48m':'48min',
    '1h':'1h', '90m':'90min','2h':'2h', '3h':'3h', '4h':'4h', '6h':'6h', '8h':'8h', '12h':'12h', '1d':'1D',
}

def _to_utc(x) -> pd.Timestamp:
    if isinstance(x, pd.Timestamp):
        if x.tzinfo is not None: return x.tz_convert('UTC')
        return x.tz_localize('UTC')
    if isinstance(x, datetime):
        if x.tzinfo is not None:
            utc_dt = x.astimezone(timezone.utc)
            return pd.Timestamp(utc_dt.year, utc_dt.month, utc_dt.day, utc_dt.hour, utc_dt.minute, utc_dt.second, utc_dt.microsecond, tz='UTC')
        return pd.Timestamp(x).tz_localize('UTC')
    import numpy as np
    if isinstance(x, np.datetime64): return pd.Timestamp(x).tz_localize('UTC')
    if isinstance(x, (int, float)): return pd.Timestamp(int(x), unit='s').tz_localize('UTC')
    ts = pd.Timestamp(str(x))
    if ts.tzinfo is not None: return ts.tz_convert('UTC')
    return ts.tz_localize('UTC')

_to_ts = _to_utc
DAM_OFF = timedelta(hours=3)

def _now_dam() -> datetime: return datetime.now(timezone.utc) + DAM_OFF
def _utc_to_dam(dt) -> datetime:
    if isinstance(dt, pd.Timestamp): dt = dt.to_pydatetime()
    if dt.tzinfo is None: dt = dt.replace(tzinfo=timezone.utc)
    return dt + DAM_OFF
def _dam_to_utc(dt) -> datetime:
    if isinstance(dt, str): dt = datetime.strptime(dt, '%Y-%m-%d %H:%M')
    if dt.tzinfo is not None: return dt - DAM_OFF
    return (dt - DAM_OFF).replace(tzinfo=timezone.utc)
def _fmt_dam(t) -> str:
    if t is None: return '-'
    dam = _utc_to_dam(t) if not isinstance(t, str) else t
    if hasattr(dam, 'strftime'): return dam.strftime('%Y-%m-%d %H:%M')
    return str(dam)
def _fmt_utc(t) -> str:
    if hasattr(t, 'strftime'): return t.strftime('%Y-%m-%d %H:%M')
    return str(t)

def calculate_tema_signals(
    candles: list, exec_tf: str = '1m', htf_res: str = '1h',
    use_mtf: bool = True, use_ma: bool = True, ma_type: str = 'TEMA',
    ma_len: int = 10, lookahead: bool = False,
) -> pd.DataFrame:
    if not candles: return pd.DataFrame()
    df1 = pd.DataFrame(candles)
    df1['time'] = df1['time'].apply(_to_utc)
    df1 = df1.drop_duplicates('time').sort_values('time').set_index('time')
    df1 = df1[['open', 'high', 'low', 'close']].astype(float)

    if use_mtf and htf_res != 'None':
        rule = _HTF_RULES.get(htf_res, '1h')
        htf  = df1.resample(rule, closed='right', label='right').agg(
            open=('open', 'first'), high=('high', 'max'), low=('low', 'min'), close=('close', 'last'),
        ).dropna()
        if len(htf) >= ma_len * 3:
            if use_ma:
                htf['tema_close'] = variant(ma_type, htf['close'], ma_len)
                htf['tema_open']  = variant(ma_type, htf['open'],  ma_len)
            else:
                htf['tema_close'] = htf['close']; htf['tema_open']  = htf['open']
            if lookahead:
                htf_open_bar = df1.resample(rule, closed='left', label='left').agg(
                    open=('open', 'first'), high=('high', 'max'), low=('low', 'min'), close=('close', 'last'),
                ).dropna()
                if len(htf_open_bar) >= ma_len * 3:
                    if use_ma:
                        htf_open_bar['tema_close'] = variant(ma_type, htf_open_bar['close'], ma_len)
                        htf_open_bar['tema_open']  = variant(ma_type, htf_open_bar['open'],  ma_len)
                    else:
                        htf_open_bar['tema_close'] = htf_open_bar['close']; htf_open_bar['tema_open']  = htf_open_bar['open']
                    df1['tema_close'] = htf_open_bar['tema_close'].reindex(df1.index, method='ffill')
                    df1['tema_open']  = htf_open_bar['tema_open'].reindex(df1.index,  method='ffill')
                else:
                    df1['tema_close'] = htf['tema_close'].reindex(df1.index, method='ffill')
                    df1['tema_open']  = htf['tema_open'].reindex(df1.index,  method='ffill')
            else:
                df1['tema_close'] = htf['tema_close'].reindex(df1.index, method='ffill')
                df1['tema_open']  = htf['tema_open'].reindex(df1.index,  method='ffill')
        else: use_mtf = False

    if not use_mtf or htf_res == 'None':
        if use_ma:
            df1['tema_close'] = variant(ma_type, df1['close'], ma_len)
            df1['tema_open']  = variant(ma_type, df1['open'],  ma_len)
        else:
            df1['tema_close'] = df1['close']; df1['tema_open']  = df1['open']

    exec_rule = _HTF_RULES.get(exec_tf, '5min')
    if exec_tf == '1m': df_exec = df1.copy()
    else:
        ohlc = df1.resample(exec_rule, closed='right', label='right').agg(
            open=('open', 'first'), high=('high', 'max'), low=('low', 'min'), close=('close', 'last'),
        ).dropna()
        ohlc['tema_close'] = df1['tema_close'].resample(exec_rule, closed='right', label='right').last()
        ohlc['tema_open']  = df1['tema_open'].resample(exec_rule, closed='right', label='right').last()
        df_exec = ohlc

    tc = df_exec['tema_close']; to_s = df_exec['tema_open']
    prev_tc = tc.shift(1); prev_to = to_s.shift(1)
    valid   = tc.notna() & to_s.notna() & prev_tc.notna() & prev_to.notna()

    df_exec['trend'] = 0
    df_exec.loc[tc > to_s, 'trend'] =  1
    df_exec.loc[tc < to_s, 'trend'] = -1
    df_exec['trend'] = df_exec['trend'].replace(0, np.nan).ffill().fillna(0).astype(int)
    df_exec['long_signal']  = valid & (prev_tc <= prev_to) & (tc > to_s)
    df_exec['short_signal'] = valid & (prev_tc >= prev_to) & (tc < to_s)
    return df_exec

def _get_ma_len(tf: str) -> int: return bot_state['ma_len_per_tf'].get(tf, bot_state['basis_len'])

# ─────────────────────────────────────────────────────────────
# DERIV WEBSOCKET FETCHER (OANDA)
# ─────────────────────────────────────────────────────────────
_OANDA_GRAN = {
    '1m': 'M1',  '2m': 'M2',  '3m': 'M3',  '4m': 'M4', '5m': 'M5',  '6m': 'M6',  '8m': 'M8',  '10m': 'M10',
    '12m': 'M12','15m': 'M15','20m': 'M20','30m': 'M30', '45m': 'M45','48m': 'M30',  
    '1h': 'H1',  '2h': 'H2',  '3h': 'H3',  '4h': 'H4', '6h': 'H6',  '8h': 'H8',  '12h': 'H12','1d': 'D', '90m': 'H1',   
}
_oanda_sem: asyncio.Semaphore | None = None
def _get_oanda_sem() -> asyncio.Semaphore:
    global _oanda_sem
    if _oanda_sem is None: _oanda_sem = asyncio.Semaphore(3)
    return _oanda_sem

async def fetch_oanda_candles(granularity_str: str, count: int = 5000, end_time: datetime = None) -> list:
    gran_str = _OANDA_GRAN.get(granularity_str, 'M1'); fetch_count = min(count, 120000)  
    collected = []; remaining = fetch_count
    headers = {'Authorization': f'Bearer {OANDA_TOKEN}', 'Content-Type':  'application/json'}
    url = f'{OANDA_BASE_URL}/instruments/{OANDA_SYMBOL}/candles'
    current_end = end_time if end_time else datetime.now(timezone.utc)

    sem = _get_oanda_sem()
    async with sem:
        while remaining > 0:
            chunk = min(remaining, 5000)
            params = {'granularity': gran_str, 'count': chunk, 'to': current_end.strftime('%Y-%m-%dT%H:%M:%S.000000000Z'), 'price': 'M'}
            candles = []
            for attempt in range(3):
                try:
                    async with get_http().get(url, headers=headers, params=params, timeout=aiohttp.ClientTimeout(total=20)) as resp:
                        if resp.status != 200: break
                        data = await resp.json(); candles = data.get('candles', []); break
                except Exception: await asyncio.sleep(1)

            if not candles: break
            complete = [c for c in candles if c.get('complete', True)]
            if not complete: break

            formatted = [{'time': pd.Timestamp(c['time']).tz_convert('UTC'), 'open': float(c['mid']['o']), 'high': float(c['mid']['h']), 'low': float(c['mid']['l']), 'close': float(c['mid']['c'])} for c in complete]
            collected = formatted + collected; remaining -= len(complete)
            earliest = pd.Timestamp(complete[0]['time']).tz_convert('UTC')
            current_end = earliest.to_pydatetime() - timedelta(seconds=1)
            if len(complete) < chunk: break
            await asyncio.sleep(0.2)
    return collected

fetch_candles = fetch_oanda_candles

# ─────────────────────────────────────────────────────────────
# BACKTEST PROGRESS TRACKER
# ─────────────────────────────────────────────────────────────
class BtProgress:
    BAR_LEN = 14; HEARTBEAT = 15
    def __init__(self, label: str, active_tfs: list):
        self.label = label; self.active_tfs = active_tfs; self.cancelled = False; self.phase = 'Initialising...'
        self.tf_done = 0; self.tf_total = len(active_tfs); self.current_tf = ''
        self.bars_done = 0; self.bars_total = 0; self.win = 0; self.loss = 0; self.be = 0; self.profit = 0.0
        self.chat_id = None; self.msg_id = None; self._last_edit = 0.0; self._lock = asyncio.Lock(); self._hb_task = None; self._start_ts = 0.0

    def _bar(self, done: int, total: int) -> str:
        if total == 0: return chr(9617) * self.BAR_LEN
        filled = round(done / total * self.BAR_LEN)
        return chr(9608) * filled + chr(9617) * (self.BAR_LEN - filled)

    def _elapsed(self) -> str:
        secs = int(datetime.now(timezone.utc).timestamp() - self._start_ts); m, s = divmod(secs, 60); return f'{m}m {s:02d}s'

    @property
    def overall_progress(self) -> float:
        overall = (self.tf_done + self.bars_done / self.bars_total) / max(self.tf_total, 1) if self.bars_total else self.tf_done / max(self.tf_total, 1)
        return overall * 100.0

    def _build_text(self) -> str:
        total = self.win + self.loss + self.be; wr = f'{round(self.win / total * 100)}%' if total else '-'
        pnl = f'+${round(self.profit,2)}' if self.profit >= 0 else f'-${abs(round(self.profit,2))}'; icon = '▲' if self.profit >= 0 else '▼'
        overall = (self.tf_done + self.bars_done / self.bars_total) / max(self.tf_total, 1) if self.bars_total else self.tf_done / max(self.tf_total, 1)
        ov_bar = self._bar(round(overall * 100), 100); ov_pct = f'{round(overall * 100)}%'
        tf_bar = self._bar(self.bars_done, self.bars_total) if self.bars_total else chr(9617) * self.BAR_LEN
        tf_pct = f'{round(self.bars_done / self.bars_total * 100)}%' if self.bars_total else '-'
        lines = [f'Backtest — <b>{self.label}</b>', f'<b>Phase:</b> {self.phase}', '', f'<b>Overall</b>  {ov_pct}', f'<code>[{ov_bar}]</code>']
        if self.current_tf: lines += ['', f'<b>TF:</b> {self.current_tf}  ({self.tf_done}/{self.tf_total})', f'<code>[{tf_bar}] {tf_pct}</code>', f'Bars: {self.bars_done}/{self.bars_total}']
        lines += ['', f'W:{self.win}  L:{self.loss}  BE:{self.be}', f'{icon} {pnl}  WR:{wr}', '', f'Elapsed: {self._elapsed()}']
        if self.cancelled: lines.append('<b>CANCELLED</b>')
        return '\n'.join(lines)

    def _cancel_kbd(self) -> dict: return {'inline_keyboard': [[{'text': '⏹ Cancel', 'callback_data': 'cancel_bt'}]]}

    async def start(self, chat_id: int) -> None:
        self.chat_id = chat_id; self._start_ts = datetime.now(timezone.utc).timestamp(); self._last_edit = self._start_ts
        payload = {'chat_id': chat_id, 'text': self._build_text(), 'parse_mode': 'HTML', 'reply_markup': self._cancel_kbd()}
        try:
            async with aiohttp.ClientSession(connector=aiohttp.TCPConnector(force_close=True), timeout=aiohttp.ClientTimeout(total=12, connect=5)) as sess:
                async with sess.post(f'https://api.telegram.org/bot{TG_TOKEN}/sendMessage', json=payload) as resp:
                    if resp.status == 200: self.msg_id = (await resp.json())['result']['message_id']
        except Exception: pass
        self._hb_task = asyncio.create_task(self._heartbeat())

    async def _heartbeat(self) -> None:
        while not self.cancelled: await asyncio.sleep(self.HEARTBEAT); await self._edit(force=True)

    async def _edit(self, force: bool = False) -> None:
        now = datetime.now(timezone.utc).timestamp()
        if not force and (now - self._last_edit) < 3: return
        if not self.msg_id or not self.chat_id: return
        async with self._lock:
            self._last_edit = now; payload = {'chat_id': self.chat_id, 'message_id': self.msg_id, 'text': self._build_text(), 'parse_mode': 'HTML'}
            if not self.cancelled: payload['reply_markup'] = self._cancel_kbd()
            try: await _tg_post(f'https://api.telegram.org/bot{TG_TOKEN}/editMessageText', json=payload)
            except Exception: pass

    async def set_phase(self, phase: str) -> None: self.phase = phase; await self._edit()
    async def set_tf(self, tf: str, bars_total: int) -> None: self.current_tf = tf; self.bars_done = 0; self.bars_total = bars_total; self.phase = f'Scanning [{tf}]'; await self._edit(force=True)
    async def tick(self, bar_n: int, win: int, loss: int, be: int, profit: float) -> None: self.bars_done = bar_n; self.win = win; self.loss = loss; self.be = be; self.profit = profit; await self._edit()
    async def finish_tf(self) -> None: self.tf_done += 1; self.bars_done = self.bars_total; await self._edit(force=True)
    async def done(self, final_text: str) -> None:
        if self._hb_task: self._hb_task.cancel()
        if not self.msg_id or not self.chat_id: return
        try: await _tg_post(f'https://api.telegram.org/bot{TG_TOKEN}/editMessageText', json={'chat_id': self.chat_id, 'message_id': self.msg_id, 'text': final_text, 'parse_mode': 'HTML'})
        except Exception: pass
    async def cancel(self) -> None:
        self.cancelled = True; self.phase = 'Cancelling...'
        if self._hb_task: self._hb_task.cancel()
        await self._edit(force=True)

_bt_progress: BtProgress | None = None

# ─────────────────────────────────────────────────────────────
# TELEGRAM HELPERS
# ─────────────────────────────────────────────────────────────
async def _tg_post(url: str, **kwargs) -> bool:
    try:
        async with aiohttp.ClientSession(connector=aiohttp.TCPConnector(force_close=True), timeout=aiohttp.ClientTimeout(total=12, connect=5)) as sess:
            async with sess.post(url, **kwargs) as resp: return resp.status == 200
    except Exception as e: return False

def _to_reply_kbd(inline_kbd: dict):
    rows = []; bmap = {}
    for row in inline_kbd.get('inline_keyboard', []):
        new_row = []
        for btn in row:
            text = btn['text']; cb = btn.get('callback_data', 'noop')
            new_row.append({'text': text}); bmap[text] = cb
        rows.append(new_row)
    return {'keyboard': rows, 'resize_keyboard': True, 'is_persistent': True, 'input_field_placeholder': 'اختر من القائمة...'}, bmap

async def send_tg_msg(text: str, reply_markup: dict = None) -> None:
    if not bot_state['chat_id']: return
    if reply_markup and 'inline_keyboard' in reply_markup:
        reply_markup, bmap = _to_reply_kbd(reply_markup); bot_state['menu_button_map'] = bmap
    payload = {'chat_id': bot_state['chat_id'], 'text': text, 'parse_mode': 'HTML'}
    if reply_markup: payload['reply_markup'] = reply_markup
    await _tg_post(f'https://api.telegram.org/bot{TG_TOKEN}/sendMessage', json=payload)

async def edit_tg_msg(chat_id, message_id, text, reply_markup=None) -> None:
    payload = {'chat_id': chat_id, 'message_id': message_id, 'text': text, 'parse_mode': 'HTML'}
    if reply_markup: payload['reply_markup'] = reply_markup
    await _tg_post(f'https://api.telegram.org/bot{TG_TOKEN}/editMessageText', json=payload)

async def _show(chat_id, msg_id, text: str, reply_markup: dict = None) -> None:
    if msg_id: await edit_tg_msg(chat_id, msg_id, text, reply_markup)
    else: await send_tg_msg(text, reply_markup)

async def answer_callback(cbq_id: str, text: str = None) -> None:
    payload = {'callback_query_id': cbq_id}
    if text: payload['text'] = text
    await _tg_post(f'https://api.telegram.org/bot{TG_TOKEN}/answerCallbackQuery', json=payload)

async def send_tg_document(file_path: str, caption: str) -> None:
    if not bot_state['chat_id']: return
    try:
        with open(file_path, 'rb') as f:
            data = aiohttp.FormData()
            data.add_field('chat_id',  str(bot_state['chat_id']))
            data.add_field('document', f, filename=os.path.basename(file_path))
            data.add_field('caption',  caption)
            await _tg_post(f'https://api.telegram.org/bot{TG_TOKEN}/sendDocument', data=data)
    except Exception: pass

# ─────────────────────────────────────────────────────────────
# KEYBOARDS
# ─────────────────────────────────────────────────────────────
def _dd_status_line() -> str:
    if not bot_state['live_connected']: return 'DD: offline'
    sod = bot_state['sod_balance']
    if sod is None: return 'DD: monitoring...'
    lim = sod * (1 - DD_LIMIT_PCT)
    return (f'🔴 DD TRIGGERED ${lim:.0f}' if bot_state['dd_triggered'] else f'🟢 DD OK  ${lim:.0f}')

def get_main_keyboard() -> dict:
    live = '🟢 connected' if bot_state['live_connected'] else '🔴 disconnected'
    st   = '▶ RUNNING'   if bot_state['status'] == 'RUNNING' else '⏸ PAUSED'
    bt   = '⏳ BT Running' if bot_state['is_backtesting'] else '📊 Backtest'
    return {'inline_keyboard': [
        [{'text': f'Server: {live}',    'callback_data': 'toggle_live_conn'}],
        [{'text': f'Bot: {st}',         'callback_data': 'toggle_status'}, {'text': '❌ Close All',       'callback_data': 'close_all'}],
        [{'text': '⚙️ Strategy',        'callback_data': 'menu_strategy'}, {'text': '💰 Risk',            'callback_data': 'menu_risk'}],
        [{'text': '⏱ Timeframes & HTF', 'callback_data': 'menu_tfs'}, {'text': '🔍 Scanner',         'callback_data': 'menu_scanner'}],
        [{'text': '📈 Market Report',   'callback_data': 'report'}, {'text': '💼 Account',         'callback_data': 'account'}],
        [{'text': bt,                   'callback_data': 'menu_backtest'}, {'text': '❌ إخفاء اللوحة',      'callback_data': 'hide_keyboard'}],
        [{'text': _dd_status_line(),    'callback_data': 'dd_status'}],
    ]}

def get_strategy_keyboard() -> dict:
    mtf_i = '✅' if bot_state['use_mtf'] else '⬜'; ma_i = '✅' if bot_state['use_ma'] else '⬜'
    ts_i  = '✅' if bot_state['use_trailing'] else '⬜'; dng_i = '✅' if bot_state['use_danger_filter'] else '⬜'
    mt = bot_state['basis_type']; ml = bot_state['basis_len']; tp = bot_state['trail_points']; tof = bot_state['trail_offset']
    return {'inline_keyboard': [
        [{'text': '── MA Settings ──',         'callback_data': 'noop'}],
        [{'text': f'Use MA: {ma_i}',           'callback_data': 'toggle_use_ma'}],
        [{'text': f'Type: {mt}  (tap cycle)',  'callback_data': 'cycle_ma_type'}],
        [{'text': '−Len', 'callback_data': 'dec_ma_len'}, {'text': f'Length = {ml}', 'callback_data': 'noop'}, {'text': '+Len', 'callback_data': 'inc_ma_len'}],
        [{'text': f'MTF (per-TF): {mtf_i}',   'callback_data': 'toggle_mtf'}],
        [{'text': '→ Set HTF per TF',          'callback_data': 'menu_tfs'}],
        [{'text': '── Trailing Stop ──',       'callback_data': 'noop'}],
        [{'text': f'Trailing: {ts_i}',         'callback_data': 'toggle_trailing'}],
        [{'text': '−Pts', 'callback_data': 'dec_trail_pts'}, {'text': f'Points={tp}', 'callback_data': 'noop'}, {'text': '+Pts', 'callback_data': 'inc_trail_pts'}],
        [{'text': '−Off', 'callback_data': 'dec_trail_off'}, {'text': f'Offset={tof}', 'callback_data': 'noop'}, {'text': '+Off', 'callback_data': 'inc_trail_off'}],
        [{'text': f'Danger Zones: {dng_i}',    'callback_data': 'toggle_danger'}],
        [{'text': '← Back',                    'callback_data': 'menu_main'}],
    ]}

def get_risk_keyboard() -> dict:
    be_i  = '✅' if bot_state['use_be'] else '⬜'; spr_i = '✅' if bot_state['use_max_spread'] else '⬜'
    tgt_i = '✅' if bot_state['daily_target_enabled'] else '⬜'; tgt_v = bot_state['daily_target_usd']
    los_i = '✅' if bot_state['daily_loss_enabled']   else '⬜'; los_v = bot_state['daily_loss_usd']
    return {'inline_keyboard': [
        [{'text': f'BE 20p: {be_i}', 'callback_data': 'toggle_be'}, {'text': f'SpreadGuard {bot_state["max_spread_pips"]}p: {spr_i}', 'callback_data': 'toggle_spread'}],
        [{'text': '−', 'callback_data': 'dec_lot'}, {'text': f'Lot: {bot_state["lot_size"]:.2f}', 'callback_data': 'noop'}, {'text': '+', 'callback_data': 'inc_lot'}],
        [{'text': '── Daily Profit Target ──', 'callback_data': 'noop'}],
        [{'text': f'Target: {tgt_i}  (${tgt_v:.0f})', 'callback_data': 'toggle_daily_target'}],
        [{'text': '/target VALUE — e.g. /target 100', 'callback_data': 'noop'}],
        [{'text': '── Daily Loss Limit ──', 'callback_data': 'noop'}],
        [{'text': f'Loss Limit: {los_i}  (${los_v:.0f})', 'callback_data': 'toggle_daily_loss'}],
        [{'text': '/loss_limit VALUE — e.g. /loss_limit 100', 'callback_data': 'noop'}],
        [{'text': '📋 Edit TP/SL per TF', 'callback_data': 'view_tpsl'}],
        [{'text': '← Back', 'callback_data': 'menu_main'}],
    ]}

def get_tf_keyboard() -> dict:
    rows = [[{'text': '  TF', 'callback_data': 'noop'}, {'text': '  HTF ↻', 'callback_data': 'noop'}, {'text': '  LM −/+', 'callback_data': 'noop'}]]
    for tf in bot_state['timeframes']:
        icon = '✅' if bot_state['active_tfs'][tf] else '⬜'; htf = bot_state['htf_per_tf'].get(tf, '1h'); lm = bot_state['ma_len_per_tf'].get(tf, 10)
        rows.append([{'text': f'{icon} {tf}', 'callback_data': f'toggle_tf_{tf}'}, {'text': f'{tf} HTF:{htf}↻', 'callback_data': f'cycle_htf_{tf}'}, {'text': f'{tf} LM:{lm}±', 'callback_data': f'lm_edit_{tf}'}])
    rows.append([{'text': '/set 1m htf 15m', 'callback_data': 'noop'}, {'text': '/set 1m lm 14', 'callback_data': 'noop'}])
    rows.append([{'text': '← Back', 'callback_data': 'menu_main'}])
    return {'inline_keyboard': rows}

def get_lm_edit_keyboard(tf: str) -> dict:
    lm = bot_state['ma_len_per_tf'].get(tf, 10)
    return {'inline_keyboard': [
        [{'text': f'MA Length [{tf}] = {lm}', 'callback_data': 'noop'}],
        [{'text': '-5', 'callback_data': f'dec_lm5_{tf}'}, {'text': '-1', 'callback_data': f'dec_lm1_{tf}'}, {'text': f'Now:{lm}','callback_data': 'noop'}, {'text': '+1', 'callback_data': f'inc_lm1_{tf}'}, {'text': '+5', 'callback_data': f'inc_lm5_{tf}'}],
        [{'text': 'Presets:', 'callback_data': 'noop'}, {'text': '7', 'callback_data': f'lm_set_{tf}_7'}, {'text': '10', 'callback_data': f'lm_set_{tf}_10'}, {'text': '14', 'callback_data': f'lm_set_{tf}_14'}, {'text': '21', 'callback_data': f'lm_set_{tf}_21'}],
        [{'text': f'/set {tf} lm VALUE', 'callback_data': 'noop'}],
        [{'text': '← Back', 'callback_data': 'menu_tfs'}],
    ]}

def get_tpsl_overview_keyboard() -> dict:
    rows = [[{'text': '── Tap Edit ──', 'callback_data': 'noop'}]]
    for tf in bot_state['timeframes']:
        icon = '✅' if bot_state['active_tfs'][tf] else '⬜'; tp = bot_state['tp_pips'][tf]; sl = bot_state['sl_pips'][tf]
        rows.append([{'text': f'{icon} {tf}  TP:{tp} SL:{sl}', 'callback_data': 'noop'}, {'text': f'Edit {tf}', 'callback_data': f'tpsl_edit_{tf}'}])
    rows.append([{'text': '/set 5m sl 80', 'callback_data': 'noop'}]); rows.append([{'text': '← Back', 'callback_data': 'menu_risk'}])
    return {'inline_keyboard': rows}

def get_tpsl_edit_keyboard(tf: str) -> dict:
    tp = bot_state['tp_pips'][tf]; sl = bot_state['sl_pips'][tf]; rr = round(tp / sl, 2) if sl else 0
    return {'inline_keyboard': [
        [{'text': f'[{tf}]  TP:{tp}p  SL:{sl}p  RR:1:{rr}', 'callback_data': 'noop'}],
        [{'text': 'Take Profit', 'callback_data': 'noop'}],
        [{'text': f'-10({tp-10})', 'callback_data': f'dec_tp10_{tf}'}, {'text': f'TP={tp}', 'callback_data': 'noop'}, {'text': f'+10({tp+10})', 'callback_data': f'inc_tp10_{tf}'}],
        [{'text': f'-5({tp-5})',  'callback_data': f'dec_tp5_{tf}'}, {'text': '─', 'callback_data': 'noop'}, {'text': f'+5({tp+5})',  'callback_data': f'inc_tp5_{tf}'}],
        [{'text': 'Stop Loss', 'callback_data': 'noop'}],
        [{'text': f'-10({sl-10})', 'callback_data': f'dec_sl10_{tf}'}, {'text': f'SL={sl}', 'callback_data': 'noop'}, {'text': f'+10({sl+10})', 'callback_data': f'inc_sl10_{tf}'}],
        [{'text': f'-5({sl-5})',  'callback_data': f'dec_sl5_{tf}'}, {'text': '─', 'callback_data': 'noop'}, {'text': f'+5({sl+5})',  'callback_data': f'inc_sl5_{tf}'}],
        [{'text': f'/set {tf} tp|sl value', 'callback_data': 'noop'}],
        [{'text': '← Back', 'callback_data': 'view_tpsl'}],
    ]}

def get_scanner_keyboard() -> dict:
    return {'inline_keyboard': [[{'text': '🔍 6h', 'callback_data': 'scan_6'}, {'text': '🔍 12h', 'callback_data': 'scan_12'}, {'text': '🔍 24h', 'callback_data': 'scan_24'}], [{'text': '← Back', 'callback_data': 'menu_main'}]]}

def get_backtest_keyboard() -> dict:
    if bot_state['is_backtesting']: return {'inline_keyboard': [[{'text': 'BT running...', 'callback_data': 'bt_show_progress'}], [{'text': '⏹ Cancel', 'callback_data': 'cancel_bt'}], [{'text': '← Back', 'callback_data': 'menu_main'}]]}
    return {'inline_keyboard': [[{'text': '1 day', 'callback_data': 'bto_1'}, {'text': '3 days', 'callback_data': 'bto_3'}, {'text': '7 days', 'callback_data': 'bto_7'}], [{'text': '← Back', 'callback_data': 'menu_main'}]]}

# ─────────────────────────────────────────────────────────────
# SIGNAL SCANNER (Probe)
# ─────────────────────────────────────────────────────────────
async def run_signal_scanner(hours: int) -> None:
    await send_tg_msg(f'🔍 <b>Signal Scanner — last {hours}h</b>\nFetching 1m data...')
    start_dt = datetime.now(timezone.utc) - timedelta(hours=hours)  
    active   = [tf for tf in bot_state['timeframes'] if bot_state['active_tfs'][tf]]
    if not active: await send_tg_msg('⚠️ No active timeframes enabled.'); return
    _htf_min_map_sc = {'1m':1,'2m':2,'3m':3,'4m':4,'5m':5,'6m':6,'8m':8,'10m':10,'12m':12,'15m':15,'20m':20,'24m':24,'30m':30,'45m':45,'48m':48,'1h':60,'90m':90,'2h':120,'3h':180,'4h':240,'6h':360,'8h':480,'12h':720,'1d':1440}
    max_warmup = max(_htf_min_map_sc.get(bot_state['htf_per_tf'].get(atf,'1h'), 60) * bot_state['basis_len'] * 5 for atf in active) if active else 500
    total1m = hours * 60 + max(max_warmup, 500)
    try: candles_1m = await fetch_candles('1m', count=total1m)
    except Exception as e: await send_tg_msg(f'❌ Fetch error: {e}'); return
    if not candles_1m: await send_tg_msg('❌ No 1m data from OANDA.'); return

    lines = [f'<b>Signal Scanner — last {hours}h</b>', f'Strategy: {bot_state["basis_type"]}({bot_state["basis_len"]}) MTF:{"ON" if bot_state["use_mtf"] else "OFF"}', '']
    start_ts = _to_ts(start_dt)

    for tf in active:
        htf = bot_state['htf_per_tf'].get(tf, '1h')
        try:
            df = calculate_tema_signals(candles_1m, exec_tf=tf, htf_res=htf, use_mtf=bot_state['use_mtf'], use_ma=bot_state['use_ma'], ma_type=bot_state['basis_type'], ma_len=_get_ma_len(tf))
            if df.empty: lines.append(f'[{tf}] — not enough data'); continue
            sigs = df[(df.index >= start_ts) & (df['long_signal'] | df['short_signal'])]
            if sigs.empty: lines.append(f'[{tf}] HTF:{htf} — No signals in last {hours}h')
            else:
                lines.append(f'<b>[{tf}] HTF:{htf}  {len(sigs)} signal(s):</b>')
                for ts, row in sigs.iterrows():
                    sig = '🟢 BUY' if row['long_signal'] else '🔴 SELL'
                    tc  = f'{row["tema_close"]:.3f}' if pd.notna(row.get('tema_close')) else '-'; to  = f'{row["tema_open"]:.3f}' if pd.notna(row.get('tema_open')) else '-'
                    lines.append(f'  {sig} @ {_utc_to_dam(ts).strftime("%Y-%m-%d %H:%M")} DAM\n  TC:{tc}  TO:{to}')
        except Exception as e: lines.append(f'[{tf}] Error: {e}')
    full = '\n'.join(lines)
    for chunk in [full[i:i+4000] for i in range(0, len(full), 4000)]: await send_tg_msg(chunk)

# ─────────────────────────────────────────────────────────────
# BACKTEST HELPERS
# ─────────────────────────────────────────────────────────────
def _entry_params(open_price: float, is_buy: bool, tf: str):
    pv = bot_state['pip_value']; m = 1 if is_buy else -1
    act_ent = open_price + m * bot_state['spread_pips'] * pv
    tp_p = round(act_ent + m * bot_state['tp_pips'][tf] * pv, 2); sl_p = round(act_ent - m * bot_state['sl_pips'][tf] * pv, 2)
    return act_ent, tp_p, sl_p

def _simulate_trade(is_buy: bool, act_ent: float, tp_p: float, sl_p: float, entry_t, minute_candles: list) -> tuple:
    pv = bot_state['pip_value']; be_act = False; be_tgt = act_ent + (1 if is_buy else -1) * 20 * pv
    trail_act = False; trail_pts = bot_state['trail_points'] * pv; trail_off = bot_state['trail_offset'] * pv
    use_trail = bot_state['use_trailing']; best_price = act_ent; max_ext = entry_t + timedelta(hours=72)
    outcome = 'EXPIRED'; exit_t = max_ext

    for vc in minute_candles:
        t = vc['time']
        if isinstance(t, pd.Timestamp): t = t.to_pydatetime()
        if t.tzinfo is None: t = t.replace(tzinfo=timezone.utc)
        if t < entry_t: continue
        if t > max_ext: break

        hi = float(vc['high']); lo = float(vc['low'])
        if is_buy:
            if not trail_act and use_trail and hi >= act_ent + trail_pts: trail_act = True; best_price = hi
            if trail_act: best_price = max(best_price, hi); sl_p = best_price - trail_off
            if bot_state['use_be'] and not be_act and hi >= be_tgt: sl_p = act_ent; be_act = True
            if lo <= sl_p: outcome = 'BREAK-EVEN' if be_act else 'LOSS'; exit_t = t; break
            if hi >= tp_p: outcome = 'WIN'; exit_t = t; break
        else:
            if not trail_act and use_trail and lo <= act_ent - trail_pts: trail_act = True; best_price = lo
            if trail_act: best_price = min(best_price, lo); sl_p = best_price + trail_off
            if bot_state['use_be'] and not be_act and lo <= be_tgt: sl_p = act_ent; be_act = True
            if hi >= sl_p: outcome = 'BREAK-EVEN' if be_act else 'LOSS'; exit_t = t; break
            if lo <= tp_p: outcome = 'WIN'; exit_t = t; break
    return outcome, exit_t, sl_p

def _calc_pnl(outcome: str, act_ent: float, tp_p: float, sl_p: float) -> float:
    if outcome == 'BREAK-EVEN': return 0.0
    if outcome in ('WIN', 'LOSS'):
        exit_p = tp_p if outcome == 'WIN' else sl_p
        raw = abs(act_ent - exit_p) * 100 * bot_state['lot_size']
        return round(raw, 2) * (1 if outcome == 'WIN' else -1)
    return 0.0

def _build_trade_row(tf, is_buy, entry_t, exit_t, act_ent, tp_p, sl_p, outcome, p_usd, ma_used):
    pv = bot_state['pip_value']
    return {
        'Timeframe': tf, 'MA Used': ma_used, 'Type': 'BUY' if is_buy else 'SELL',
        'Entry (Damascus)': _fmt_dam(entry_t), 'Exit  (Damascus)': _fmt_dam(exit_t),
        'Entry (UTC)': _fmt_utc(entry_t), 'Exit  (UTC)': _fmt_utc(exit_t),
        'Entry Price': round(act_ent, 2), 'TP': tp_p, 'SL': sl_p,
        'Pips': (round(abs(act_ent - (tp_p if outcome == 'WIN' else sl_p)) / pv, 1) if outcome in ('WIN', 'LOSS') else 0),
        'Outcome': outcome, 'Profit ($)': p_usd,
    }

def _style_sheet(ws) -> None:
    from openpyxl.styles import PatternFill, Font
    green  = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
    red    = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
    yellow = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
    header = PatternFill(start_color='1F3864', end_color='1F3864', fill_type='solid')
    for cell in ws[1]: cell.fill = header; cell.font = Font(color='FFFFFF', bold=True)
    
    headers_list = [c.value for c in ws[1]]
    oc = headers_list.index('Outcome') + 1 if 'Outcome' in headers_list else 9
    rc = headers_list.index('Repaint?') + 1 if 'Repaint?' in headers_list else None

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        val = str(row[oc-1].value) if len(row) >= oc else ''
        rv  = str(row[rc-1].value) if rc and len(row) >= rc else ''
        if rv == 'YES':
            for cell in row: cell.fill = yellow
        elif val == 'WIN':
            for cell in row: cell.fill = green
        elif val == 'LOSS':
            for cell in row: cell.fill = red
    for col in ws.columns:
        mx = max((len(str(c.value or '')) for c in col), default=8)
        ws.column_dimensions[col[0].column_letter].width = min(mx + 3, 32)

# ─────────────────────────────────────────────────────────────
# BACKTEST ENGINE (Triple Engine: Pure A, Pure B, Hybrid)
# ─────────────────────────────────────────────────────────────
async def run_backtest(start_dt: datetime, end_dt: datetime = None) -> None:
    global _bt_progress
    if bot_state['is_backtesting']: return
    bot_state['is_backtesting'] = True

    now_utc = datetime.now(timezone.utc)
    if end_dt is None: end_dt = now_utc
    if end_dt > now_utc: end_dt = now_utc
    if start_dt >= end_dt:
        await send_tg_msg('❌ Backtest error: start date must be before end date.')
        bot_state['is_backtesting'] = False; return

    active_tfs = [tf for tf in bot_state['timeframes'] if bot_state['active_tfs'][tf]]
    if not active_tfs:
        await send_tg_msg('⚠️ No active timeframes. Enable at least one.')
        bot_state['is_backtesting'] = False; return

    primary_ma = bot_state['basis_type']
    fallback_ma = 'EMA'
    strategies = [primary_ma, fallback_ma] if primary_ma.upper() != fallback_ma.upper() else [primary_ma]

    desc  = f'{primary_ma}({bot_state["basis_len"]}) Triple-Comparison'
    fname = f"BT_{datetime.now(timezone.utc).strftime('%H%M%S')}.xlsx"

    prog = BtProgress(label=desc, active_tfs=active_tfs); _bt_progress = prog
    await prog.start(bot_state['chat_id'])

    all_results = {s: {
        'trade_logs': [], 'raw_trades': [], 'repaint_rows': [],
        'total_prof': 0.0, 'peak_equity': 0.0, 'max_dd': 0.0,
        'total_win': 0.0, 'total_loss': 0.0, 'win_count': 0,
        'loss_count': 0, 'be_count': 0, 'skipped_target': 0,
        'skipped_loss': 0, 'ma_switches': []
    } for s in strategies}

    try:
        await prog.set_phase('Fetching 1m base candles...')
        warmup_min = max(bot_state['basis_len'] * 3 * 60 * 4, 2000)
        total_1m   = int((end_dt - start_dt).total_seconds() / 60) + warmup_min + 72 * 60
        total_1m   = min(total_1m, 120000)  

        candles_1m = await fetch_candles('1m', count=total_1m, end_time=end_dt)
        if not candles_1m: await prog.done('❌ Could not fetch 1m data.'); return

        candles_1m_sorted = sorted(candles_1m, key=lambda x: x['time'])
        end_ts = _to_ts(end_dt); start_ts = _to_ts(start_dt)

        for tf in active_tfs:
            if prog.cancelled: break
            await asyncio.sleep(0)
            htf = bot_state['htf_per_tf'].get(tf, '1h')
            await prog.set_phase(f'Computing Matrices [{tf}] HTF:{htf}...')

            # 1. Calc Primary Matrix
            df_conf_prim = calculate_tema_signals(candles_1m_sorted, exec_tf=tf, htf_res=htf, use_mtf=bot_state['use_mtf'], use_ma=bot_state['use_ma'], ma_type=primary_ma, ma_len=_get_ma_len(tf), lookahead=False)
            df_rep_prim  = calculate_tema_signals(candles_1m_sorted, exec_tf=tf, htf_res=htf, use_mtf=bot_state['use_mtf'], use_ma=bot_state['use_ma'], ma_type=primary_ma, ma_len=_get_ma_len(tf), lookahead=True)

            # 2. Calc Fallback Matrix
            if primary_ma.upper() != fallback_ma.upper():
                df_conf_fall = calculate_tema_signals(candles_1m_sorted, exec_tf=tf, htf_res=htf, use_mtf=bot_state['use_mtf'], use_ma=bot_state['use_ma'], ma_type=fallback_ma, ma_len=_get_ma_len(tf), lookahead=False)
                df_rep_fall  = calculate_tema_signals(candles_1m_sorted, exec_tf=tf, htf_res=htf, use_mtf=bot_state['use_mtf'], use_ma=bot_state['use_ma'], ma_type=fallback_ma, ma_len=_get_ma_len(tf), lookahead=True)
            else:
                df_conf_fall, df_rep_fall = df_conf_prim, df_rep_prim

            mask_conf_prim = (df_conf_prim.index >= start_ts) & (df_conf_prim.index < end_ts)
            mask_conf_fall = (df_conf_fall.index >= start_ts) & (df_conf_fall.index < end_ts)

            # 3. Simulate Each Pure Strategy
            for strategy in strategies:
                bar_times = df_conf_prim[mask_conf_prim].index if strategy == primary_ma else df_conf_fall[mask_conf_fall].index
                open_trade_is_buy = None
                
                for loop_pos, ts in enumerate(bar_times):
                    if strategy == fallback_ma:
                        if ts not in df_conf_fall.index: continue
                        row = df_conf_fall.loc[ts]; row_rp = df_rep_fall.loc[ts] if ts in df_rep_fall.index else None
                    else:
                        if ts not in df_conf_prim.index: continue
                        row = df_conf_prim.loc[ts]; row_rp = df_rep_prim.loc[ts] if ts in df_rep_prim.index else None

                    ts_dt = ts.to_pydatetime()
                    if bot_state['use_danger_filter'] and is_blocked_time(ts_dt): continue

                    if strategy == primary_ma and row_rp is not None:
                        long_rp  = bool(row_rp.get('long_signal', False)); short_rp = bool(row_rp.get('short_signal', False))
                        if long_rp or short_rp:
                            is_repaint = not (bool(row.get('long_signal', False)) or bool(row.get('short_signal', False)))
                            if is_repaint:
                                all_results[primary_ma]['repaint_rows'].append({
                                    'Timeframe': tf, 'HTF': htf, 'MA Checked': strategy,
                                    'Signal Time (DAM)': _utc_to_dam(ts).strftime('%Y-%m-%d %H:%M'),
                                    'Signal Time (UTC)': ts.strftime('%Y-%m-%d %H:%M'),
                                    'Type': 'BUY' if long_rp else 'SELL', 'Repaint?': 'YES', 'TradingView sees': 'YES', 'Bot takes': 'NO',
                                    'TEMA_Close': round(float(row_rp.get('tema_close', 0)), 4), 'TEMA_Open': round(float(row_rp.get('tema_open', 0)), 4),
                                    'Explanation': f'TV shows signal mid-{htf}-bar (repaints when bar closes)'
                                })

                    long_sig  = bool(row.get('long_signal', False)); short_sig = bool(row.get('short_signal', False))
                    if not (long_sig or short_sig): continue

                    if strategy == primary_ma:
                        all_results[primary_ma]['repaint_rows'].append({
                            'Timeframe': tf, 'HTF': htf, 'MA Checked': strategy,
                            'Signal Time (DAM)': _utc_to_dam(ts).strftime('%Y-%m-%d %H:%M'), 'Signal Time (UTC)': ts.strftime('%Y-%m-%d %H:%M'),
                            'Type': 'BUY' if long_sig else 'SELL', 'Repaint?': 'NO', 'TradingView sees': 'YES', 'Bot takes': 'YES',
                            'TEMA_Close': round(float(row.get('tema_close', 0)), 4), 'TEMA_Open': round(float(row.get('tema_open', 0)), 4),
                            'Explanation': f'Confirmed after {htf} bar closes — matches bot entry'
                        })

                    if open_trade_is_buy is not None:
                        if (open_trade_is_buy and short_sig) or (not open_trade_is_buy and long_sig): open_trade_is_buy = None

                    df_to_use = df_conf_fall if strategy == fallback_ma else df_conf_prim
                    loc_i = df_to_use.index.get_loc(ts)
                    if loc_i + 1 >= len(df_to_use): continue
                    next_ts = df_to_use.index[loc_i + 1]; next_row = df_to_use.iloc[loc_i + 1]

                    entry_t = next_ts.to_pydatetime()
                    if entry_t.tzinfo is None: entry_t = entry_t.replace(tzinfo=timezone.utc)
                    if next_ts >= end_ts: continue

                    is_buy = bool(long_sig)
                    act_ent, tp_p, sl_p = _entry_params(float(next_row['open']), is_buy, tf)
                    outcome, exit_t, sl_p = _simulate_trade(is_buy, act_ent, tp_p, sl_p, entry_t, candles_1m_sorted)
                    p_usd = _calc_pnl(outcome, act_ent, tp_p, sl_p)

                    open_trade_is_buy = is_buy
                    _row = _build_trade_row(tf, is_buy, entry_t, exit_t, act_ent, tp_p, sl_p, outcome, p_usd, strategy)
                    all_results[strategy]['raw_trades'].append((entry_t, exit_t, p_usd, outcome, _row))

            await prog.finish_tf()

        # 4. Build Hybrid Strategy Chronologically
        hybrid_raw_trades = []
        hybrid_switches = []
        hybrid_resets = []
        if len(strategies) > 1:
            all_potentials = []
            for s in strategies:
                all_potentials.extend(all_results[s]['raw_trades'])
            all_potentials.sort(key=lambda x: x[0])
            
            active_ma = primary_ma
            bt_consec_losses = 0
            open_hybrid_trades = []
            
            for trade in all_potentials:
                t_entry = trade[0]; t_exit  = trade[1]; t_p_usd = trade[2]; t_out   = trade[3]; t_row   = trade[4]
                t_ma    = t_row['MA Used']; t_tf    = t_row['Timeframe']; t_is_buy = (t_row['Type'] == 'BUY')

                closures = [ot for ot in open_hybrid_trades if ot['exit_t'] <= t_entry]
                closures.sort(key=lambda x: x['exit_t'])
                
                for c in closures:
                    if c['outcome'] == 'WIN': 
                        bt_consec_losses = 0
                    elif c['outcome'] == 'LOSS':
                        bt_consec_losses += 1
                        if bt_consec_losses >= 3 and active_ma != fallback_ma:
                            active_ma = fallback_ma
                            bt_consec_losses = 0
                            hybrid_switches.append(_utc_to_dam(c['exit_t']).strftime('%Y-%m-%d %H:%M'))
                    open_hybrid_trades.remove(c)
                
                if t_ma == active_ma:
                    tf_open = [ot for ot in open_hybrid_trades if ot['tf'] == t_tf]
                    if tf_open:
                        pass
                    else:
                        open_hybrid_trades.append({'exit_t': t_exit, 'outcome': t_out, 'tf': t_tf, 'is_buy': t_is_buy})
                        hybrid_raw_trades.append(trade)

            all_results['Hybrid'] = {
                'trade_logs': [], 'raw_trades': hybrid_raw_trades, 'repaint_rows': [],
                'total_prof': 0.0, 'peak_equity': 0.0, 'max_dd': 0.0,
                'total_win': 0.0, 'total_loss': 0.0, 'win_count': 0,
                'loss_count': 0, 'be_count': 0, 'skipped_target': 0,
                'skipped_loss': 0, 'ma_switches': hybrid_switches, 'hybrid_resets': hybrid_resets
            }
            final_strategies = strategies + ['Hybrid']
        else:
            final_strategies = strategies

        # 5. Chronological Gating & Metrics Calculation
        target_active = bot_state['daily_target_enabled']; loss_active = bot_state['daily_loss_enabled']
        target = bot_state['daily_target_usd']; limit = bot_state['daily_loss_usd']

        for strat in final_strategies:
            res = all_results[strat]
            raw = res['raw_trades']
            raw.sort(key=lambda x: x[0])  
            day_stopped = set(); day_stopped_reason = {}; daily_profit = {}

            for entry_t, exit_t, p_usd, outcome, row in raw:
                day_key = _utc_to_dam(entry_t).date()
                if day_key in day_stopped:
                    skip_row = dict(row); skip_row['Outcome'] = day_stopped_reason.get(day_key, 'SKIP'); skip_row['Profit ($)'] = 0.0
                    res['trade_logs'].append(skip_row)
                    if skip_row['Outcome'] == 'SKIP (Target Hit)': res['skipped_target'] += 1
                    else: res['skipped_loss'] += 1
                    continue

                res['trade_logs'].append(row)
                day_pnl = daily_profit.get(day_key, 0.0) + p_usd
                daily_profit[day_key] = day_pnl

                if outcome == 'BREAK-EVEN': res['be_count'] += 1
                elif outcome == 'WIN': res['total_win'] += p_usd; res['win_count'] += 1
                elif outcome == 'LOSS': res['total_loss'] += p_usd; res['loss_count'] += 1

                res['total_prof'] += p_usd
                res['peak_equity'] = max(res['peak_equity'], res['total_prof'])
                res['max_dd'] = max(res['max_dd'], res['peak_equity'] - res['total_prof'])

                if target_active and day_pnl >= target: day_stopped.add(day_key); day_stopped_reason[day_key] = 'SKIP (Target Hit)'
                elif loss_active and day_pnl <= -limit: day_stopped.add(day_key); day_stopped_reason[day_key] = 'SKIP (Loss Limit Hit)'

        # 6. Build Telegram Report
        main_strat = 'Hybrid' if 'Hybrid' in final_strategies else primary_ma
        res_main = all_results[main_strat]
        prof_main = res_main['total_prof']
        icon_main = 'PROFIT ▲' if prof_main >= 0 else 'LOSS ▼'
        total_trades = res_main['win_count'] + res_main['loss_count']
        wr_main = round(res_main['win_count'] / max(1, total_trades) * 100, 1) if total_trades else 0
        dd_pct = round(res_main['max_dd'] / res_main['peak_equity'] * 100, 1) if res_main['peak_equity'] else 0
        
        target_line = ''
        if target_active: target_line = f'🎯 <b>Daily Target: ${target:.0f}</b>\nTrades skipped (target hit): <b>{res_main["skipped_target"]}</b>\n\n'
        loss_line = ''
        if loss_active: loss_line = f'🛑 <b>Daily Loss Limit: ${limit:.0f}</b>\nTrades skipped (loss limit): <b>{res_main["skipped_loss"]}</b>\n\n'

        repaint_count   = sum(1 for r in all_results[primary_ma]['repaint_rows'] if r['Repaint?'] == 'YES')
        confirmed_count = sum(1 for r in all_results[primary_ma]['repaint_rows'] if r['Repaint?'] == 'NO')
        mtf_status = "per-TF" if bot_state['use_mtf'] else "OFF"

        tg_lines = [
            f'<b>Backtest Complete ✅</b>', f'{primary_ma}({bot_state["basis_len"]}) MTF:{mtf_status}', f'{_fmt_dam(start_dt).split()[0]} → {_fmt_dam(end_dt).split()[0]} (DAM)\n',
            f'Net: <b>{icon_main} ${round(prof_main, 2)}</b>', f'Win:  +${round(res_main["total_win"],2)} ({res_main["win_count"]})', f'Loss: -${abs(round(res_main["total_loss"],2))} ({res_main["loss_count"]})',
            f'BE: {res_main["be_count"]}  WR: {wr_main}% ({total_trades} trades)', f'Max DD: ${round(res_main["max_dd"],2)} ({dd_pct}%)\n',
            f'{target_line}{loss_line}<b>Repaint Analysis:</b>', f'🟡 Repainting (TV shows, bot skips): <b>{repaint_count}</b>', f'🟢 Confirmed (bot takes): <b>{confirmed_count}</b>\n',
            f'<b>--- Quick Comparison ---</b>'
        ]
        
        for strat in final_strategies:
            res = all_results[strat]
            s_icon = '🏆' if strat == 'Hybrid' else '🔹'
            s_prof = res['total_prof']; s_total = res['win_count'] + res['loss_count']
            s_wr = round(res['win_count'] / max(1, s_total) * 100, 1) if s_total else 0
            tg_lines.append(f'{s_icon} <b>{strat}:</b> {"+$" if s_prof>=0 else "-$"}{abs(round(s_prof,2))} (WR: {s_wr}%)')

        if 'Hybrid' in final_strategies:
            switches = all_results['Hybrid']['ma_switches']
            resets = all_results['Hybrid'].get('hybrid_resets', [])
            tg_lines.append(f'\n🔄 <b>Hybrid Switches:</b> {len(switches)}')
            if switches:
                unique_sw = list(dict.fromkeys(switches))
                tg_lines.append(f"⏱️ Switch Times: {', '.join(unique_sw[:3])}{'...' if len(unique_sw)>3 else ''}")
            if resets:
                unique_resets = list(dict.fromkeys(resets))
                tg_lines.append(f"🌅 <b>Daily Resets to {primary_ma}:</b> {len(unique_resets)}")
                tg_lines.append(f"📅 Reset Days: {', '.join(unique_resets[:3])}{'...' if len(unique_resets)>3 else ''}")
                
        tg_lines.append('\nSending Multi-Sheet Excel...')
        await prog.done('\n'.join(tg_lines))

        # 7. Build Multi-Sheet Excel
        summary_rows = {
            'Metric': [
                'Strategy Type', 'Net Profit', 'Win Trades', 'Loss Trades', 'Win Rate', 
                'Max Drawdown', 'Break-Even', 'Daily Targets Hit', 'Daily Loss Hit', 'Adaptive Switches'
            ]
        }
        for strat in final_strategies:
            res = all_results[strat]; total = res['win_count'] + res['loss_count']
            wr = round(res['win_count'] / max(1, total) * 100, 1) if total else 0
            dd_pct = round(res['max_dd'] / res['peak_equity'] * 100, 1) if res['peak_equity'] else 0
            summary_rows[strat] = [
                f"{strat} MTF:{'ON' if bot_state['use_mtf'] else 'OFF'}", f"${round(res['total_prof'], 2)}", f"{res['win_count']} (+${round(res['total_win'],2)})", f"{res['loss_count']} (-${abs(round(res['total_loss'],2))})",
                f"{wr}% ({total})", f"${round(res['max_dd'], 2)} ({dd_pct}%)", str(res['be_count']), str(res['skipped_target']), str(res['skipped_loss']), str(len(res['ma_switches'])) if strat == 'Hybrid' else "-"
            ]

        with pd.ExcelWriter(fname, engine='openpyxl') as writer:
            pd.DataFrame(summary_rows).to_excel(writer, sheet_name='Comparison Summary', index=False)
            _style_sheet(writer.sheets['Comparison Summary'])
            
            for strat in final_strategies:
                logs = all_results[strat]['trade_logs']
                if logs:
                    sheet_name = f'{strat} Trades'
                    pd.DataFrame(logs).to_excel(writer, sheet_name=sheet_name[:31], index=False)
                    _style_sheet(writer.sheets[sheet_name[:31]])
                    
            rep_rows = all_results[primary_ma]['repaint_rows']
            if rep_rows:
                pd.DataFrame(rep_rows).to_excel(writer, sheet_name='Repaint Analysis', index=False)
                _style_sheet(writer.sheets['Repaint Analysis'])

        hybrid_res = all_results.get('Hybrid', all_results[primary_ma])
        caption = f'Backtest 3D | {start_dt.strftime("%Y-%m-%d")}→{end_dt.strftime("%Y-%m-%d")} | Hybrid Net: ${round(hybrid_res["total_prof"],2)} | Pure {primary_ma}: ${round(all_results[primary_ma]["total_prof"],2)}'
        
        import shutil
        shutil.copy(fname, '/tmp/latest_backtest.xlsx')
        
        # Format the full message for the app without HTML
        app_msg = '\n'.join(tg_lines).replace('<b>', '').replace('</b>', '').replace('<code>', '').replace('</code>', '')
        
        bot_state['last_backtest_result'] = {
            'caption': app_msg,
            'summary': summary_rows
        }
        
        await send_tg_document(fname, caption)
        os.remove(fname)

    except Exception as e:
        await prog.done(f'❌ ERROR: {e}')
        c_log(f'Backtest error: {e}')
    finally:
        bot_state['is_backtesting'] = False; _bt_progress = None
# ─────────────────────────────────────────────────────────────
# LIVE SCANNER — per timeframe
# ─────────────────────────────────────────────────────────────
async def timeframe_scanner(tf: str) -> None:
    c_log(f'Scanner [{tf}] started.')
    _htf_min_map = {
        '1m':1,'2m':2,'3m':3,'4m':4,'5m':5,'6m':6,'8m':8,'10m':10,'12m':12,'15m':15,
        '20m':20,'24m':24,'30m':30,'45m':45,'48m':48,'1h':60,'90m':90,'2h':120,'3h':180,
        '4h':240,'6h':360,'8h':480,'12h':720,'1d':1440,
    }
    def _calc_cache_size() -> int:
        _htf = bot_state['htf_per_tf'].get(tf, '1h'); _hmin = _htf_min_map.get(_htf, 60); _malen = _get_ma_len(tf)
        return max(_hmin * _malen * 5 + 500, 3000)
    cache_key = f'1m_base_{tf}'; CACHE_SIZE_1M = _calc_cache_size()

    while True:
        try:
            if not (bot_state['status'] == 'RUNNING' and bot_state['active_tfs'][tf]): await asyncio.sleep(10); continue
            if not (bot_state['live_connected'] and bot_state['account_obj']): bot_state['market_data'][tf] = 'Offline'; await asyncio.sleep(5); continue
            if bot_state['dd_triggered']: bot_state['market_data'][tf] = 'DD PAUSED'; await asyncio.sleep(30); continue

            if not bot_state.get(cache_key):
                await send_tg_msg(f'⏳ <b>[Warm-up {tf}]</b>\nOANDA: Fetching {CACHE_SIZE_1M} × 1m candles...')
                try:
                    CACHE_SIZE_1M = _calc_cache_size()
                    raw = await fetch_candles('1m', count=CACHE_SIZE_1M)
                    if not raw: await asyncio.sleep(15); continue
                    bot_state[cache_key] = raw
                    await send_tg_msg(f'✅ <b>[Warm-up {tf}]</b> {len(raw)} × 1m cached.')
                except Exception as e: c_log(f'[{tf}] WARM-UP ERROR: {e}'); await asyncio.sleep(15); continue

            try:
                raw_new = await asyncio.wait_for(fetch_candles('1m', count=5), timeout=20.0)
                if raw_new:
                    existing = {c['time'] for c in bot_state[cache_key]}
                    new_bars = [c for c in raw_new if c['time'] not in existing]
                    if new_bars:
                        bot_state[cache_key].extend(new_bars)
                        bot_state[cache_key] = sorted(bot_state[cache_key], key=lambda x: x['time'])[-CACHE_SIZE_1M:]

                htf = bot_state['htf_per_tf'].get(tf, '1h')
                df_work = calculate_tema_signals(bot_state[cache_key], exec_tf=tf, htf_res=htf, use_mtf=bot_state['use_mtf'], use_ma=bot_state['use_ma'], ma_type=bot_state['basis_type'], ma_len=_get_ma_len(tf), lookahead=False)
            except Exception as e: c_log(f'[{tf}] PULSE ERROR: {e}'); await asyncio.sleep(10); continue

            if df_work.empty or len(df_work) < 3: await asyncio.sleep(15); continue

            curr = df_work.iloc[-2]
            tc = round(float(curr.get('tema_close', curr['close'])), 3); to_v = round(float(curr.get('tema_open',  curr['open'])),  3)
            trend_icon = '🟢' if int(curr.get('trend', 0)) > 0 else ('🔴' if int(curr.get('trend', 0)) < 0 else '⚪')
            bot_state['market_data'][tf] = f'{trend_icon} {df_work.iloc[-1]["close"]:.2f}  TC:{tc}  TO:{to_v}'

            now_utc = datetime.now(timezone.utc)
            if bot_state['use_danger_filter'] and is_blocked_time(now_utc): bot_state['market_data'][tf] = f'🚫 {blocked_time_label(now_utc)}'; await asyncio.sleep(10); continue

            bar_time = df_work.index[-2]
            if bot_state['last_signal_time'][tf] == bar_time: await asyncio.sleep(10); continue

            long_sig  = bool(curr.get('long_signal', False)); short_sig = bool(curr.get('short_signal', False))
            if not (long_sig or short_sig): await asyncio.sleep(10); continue

            if bot_state['use_max_spread']:
                try:
                    tick = await bot_state['connection_obj'].get_tick(bot_state['symbol'])
                    spread_pips = (tick['ask'] - tick['bid']) / bot_state['pip_value']
                    if spread_pips > bot_state['max_spread_pips']: c_log(f'[{tf}] Spread {spread_pips:.1f}p > max. Skip.'); await asyncio.sleep(10); continue
                except Exception: pass

            bot_state['last_signal_time'][tf] = bar_time

            try:
                positions = await bot_state['connection_obj'].get_positions()
                for p in positions:
                    if p['symbol'] != bot_state['symbol']: continue
                    pos_tp = p.get('takeProfit')
                    if pos_tp is None: continue
                    pos_is_buy = float(pos_tp) > float(p['openPrice'])
                    if (pos_is_buy and short_sig) or (not pos_is_buy and long_sig):
                        await bot_state['connection_obj'].close_position(p['id'])
                        await send_tg_msg(f'↩️ <b>Close-on-Reverse [{tf}]</b>\nClosed {"BUY" if pos_is_buy else "SELL"} #{p["id"]}')
            except Exception as ce: c_log(f'[{tf}] COR error: {ce}')

            price = float(df_work.iloc[-1]['close']); pv = bot_state['pip_value']
            m = 1 if long_sig else -1; t_str = 'BUY' if long_sig else 'SELL'
            tp = round(price + m * bot_state['tp_pips'][tf] * pv, 2); sl = round(price - m * bot_state['sl_pips'][tf] * pv, 2)
            lot = bot_state['lot_size']; htf_used = bot_state['htf_per_tf'].get(tf, '1h')

            try:
                if long_sig: await bot_state['connection_obj'].create_market_buy_order(bot_state['symbol'], lot, stop_loss=sl, take_profit=tp)
                else: await bot_state['connection_obj'].create_market_sell_order(bot_state['symbol'], lot, stop_loss=sl, take_profit=tp)
                await send_tg_msg(f'<b>✅ {t_str} [{tf}]</b>  HTF:{htf_used}\nPrice:{price:.2f}  TP:{tp}  SL:{sl}  Lot:{lot}\nTC:{tc}  TO:{to_v}\nBar: {_utc_to_dam(bar_time).strftime("%H:%M")} DAM')
            except Exception as e: await send_tg_msg(f'<b>❌ Order Failed [{tf}]</b>\n{e}')

        except Exception as e: c_log(f'Scanner [{tf}] error: {e}')
        await asyncio.sleep(10)

# ─────────────────────────────────────────────────────────────
# POSITION MONITOR
# ─────────────────────────────────────────────────────────────
async def position_monitor() -> None:
    while True:
        try:
            if bot_state['live_connected'] and bot_state['connection_obj']:
                pv = bot_state['pip_value']
                positions = await bot_state['connection_obj'].get_positions()
                
                current_ids = []
                for p in positions:
                    if p['symbol'] != bot_state['symbol']: continue
                    pid = p['id']; current_ids.append(pid)
                    bot_state['tracked_positions'][pid] = float(p.get('unrealizedProfit', 0)) + float(p.get('swap', 0))

                closed_ids = [pid for pid in list(bot_state['tracked_positions'].keys()) if pid not in current_ids]
                for pid in closed_ids:
                    last_profit = bot_state['tracked_positions'][pid]
                    if last_profit < 0:
                        bot_state['consecutive_losses'] += 1
                        c_log(f'Trade #{pid} closed LOSS. Streak: {bot_state["consecutive_losses"]}')
                    elif last_profit > 0:
                        bot_state['consecutive_losses'] = 0; c_log(f'Trade #{pid} closed WIN. Streak reset.')

                    if bot_state['consecutive_losses'] >= 3:
                        if bot_state['basis_type'] != 'EMA':
                            bot_state['basis_type'] = 'EMA'; bot_state['consecutive_losses'] = 0
                            for tf in _TFS: bot_state[f'1m_base_{tf}'] = [] 
                            await send_tg_msg('⚠️ <b>تنبيه ذكي من النظام:</b>\n📉 تم تسجيل 3 خسائر متتالية.\n🔄 تم تغيير المؤشر إلى <b>EMA</b>\nتم تفريغ الكاش للبدء...')
                    del bot_state['tracked_positions'][pid]

                for p in positions:
                    if p['symbol'] != bot_state['symbol']: continue
                    op  = float(p['openPrice']); tp  = p.get('takeProfit')
                    sl  = p.get('stopLoss'); cp  = float(p['currentPrice'])
                    if tp is None: continue
                    is_buy = float(tp) > op

                    if bot_state['use_be'] and sl is not None and float(sl) != op:
                        be_tgt = op + (1 if is_buy else -1) * 20 * pv
                        if (is_buy and cp >= be_tgt) or (not is_buy and cp <= be_tgt):
                            try: await bot_state['connection_obj'].modify_position(p['id'], stop_loss=round(op, 2)); await send_tg_msg(f'🔒 BE activated — #{p["id"]}')
                            except Exception: pass

                    if bot_state['use_trailing'] and sl is not None:
                        trail_pts = bot_state['trail_points'] * pv; trail_off = bot_state['trail_offset'] * pv
                        if is_buy:
                            ideal_sl = cp - trail_off
                            if cp >= op + trail_pts and ideal_sl > float(sl):
                                try: await bot_state['connection_obj'].modify_position(p['id'], stop_loss=round(ideal_sl, 2))
                                except Exception: pass
                        else:
                            ideal_sl = cp + trail_off
                            if cp <= op - trail_pts and ideal_sl < float(sl):
                                try: await bot_state['connection_obj'].modify_position(p['id'], stop_loss=round(ideal_sl, 2))
                                except Exception: pass
        except Exception as e: c_log(f'Position monitor error: {e}')
        await asyncio.sleep(1)

# ─────────────────────────────────────────────────────────────
# COMMAND PARSERS
# ─────────────────────────────────────────────────────────────
def _parse_set_cmd(msg: str):
    parts = msg.strip().lower().split()
    if len(parts) != 4: return None
    _, tf, key, val = parts; tf = tf.strip(); key = key.strip(); val = val.strip()
    if tf not in _TFS: return None
    if key == 'htf':
        norm = 'None' if val == 'none' else val
        if norm in _HTF_OPTIONS: return {'type': 'htf', 'tf': tf, 'val': norm}
        return None
    if key == 'lm':
        try:
            v = int(val)
            if 2 <= v <= 200: return {'type': 'lm', 'tf': tf, 'val': v}
        except ValueError: pass
        return None
    if key in ('tp', 'sl'):
        try:
            v = int(val)
            if v >= 1: return {'type': key, 'tf': tf, 'val': v}
        except ValueError: pass
        return None
    return None

# ─────────────────────────────────────────────────────────────
# TELEGRAM UPDATE HANDLER
# ─────────────────────────────────────────────────────────────
async def process_tg_update(update: dict) -> None:
    if 'message' in update and 'text' in update['message']:
        msg = update['message']['text'].strip(); bot_state['chat_id'] = update['message']['chat']['id']

        if not msg.startswith('/') and msg in bot_state.get('menu_button_map', {}):
            cb = bot_state['menu_button_map'][msg]
            if cb != 'noop': await _handle_callback(cb, bot_state['chat_id'], None)
            return

        if msg == '/start':
            await send_tg_msg(
                '<b>Gold Scalper Bot v5.2 (3D Triple Engine)</b>\n<b>Strategy: Crypto World TEMA</b>\n\n'
                '• Hybrid Adaptive: Default DEMA (switches to EMA on 3 losses)\n'
                '• 3D Backtest: Compares DEMA, EMA & Hybrid at the same time!\n'
                '• Close-on-Reverse ✅  DD 3% ✅\n\n'
                '/status  /dd  /ping\n'
                '/set TF tp VALUE   — e.g. /set 5m tp 70\n/set TF sl VALUE   — e.g. /set 5m sl 100\n'
                '/set TF htf VALUE  — e.g. /set 1m htf 15m\n/set TF lm VALUE   — e.g. /set 1m lm 14\n'
                '/restart_sessions  — fix Telegram silence\n/target VALUE      — daily profit target $, e.g. /target 100\n'
                '/loss_limit VALUE  — daily loss limit $, e.g. /loss_limit 100\n'
                '/backtest YYYY-MM-DD  or\n/backtest YYYY-MM-DD YYYY-MM-DD\n\n👇 استخدم القائمة بالأسفل للتنقل...'
            )
            await send_tg_msg('Main Menu:', get_main_keyboard())

        elif msg.lower().startswith('/set'):
            result = _parse_set_cmd(msg)
            if result is None:
                await send_tg_msg('❌ Invalid /set command.\n\n<b>Supported formats:</b>\n/set TF tp VALUE\n/set TF sl VALUE\n/set TF htf VALUE\n/set TF lm VALUE\n')
            elif result['type'] in ('tp', 'sl'):
                bot_state[f'{result["type"]}_pips'][result['tf']] = result['val']
                tp = bot_state['tp_pips'][result['tf']]; sl = bot_state['sl_pips'][result['tf']]; rr = round(tp / sl, 2) if sl else 0
                await send_tg_msg(f'✅ [{result["tf"]}]: TP={tp}p  SL={sl}p  RR=1:{rr}')
            elif result['type'] in ('htf', 'lm'):
                bot_state[f'{"htf_per_tf" if result["type"]=="htf" else "ma_len_per_tf"}'][result['tf']] = result['val']
                bot_state[f'1m_base_{result["tf"]}'] = []
                await send_tg_msg(f'✅ [{result["tf"]}] {"HTF" if result["type"]=="htf" else "MA Length"} → <b>{result["val"]}</b>\nCache cleared — warm-up will restart.')

        elif msg.lower().startswith('/target'):
            parts = msg.strip().split()
            if len(parts) == 2:
                try:
                    val = float(parts[1])
                    if val > 0:
                        bot_state['daily_target_usd'] = val; status_txt = '✅ ENABLED' if bot_state['daily_target_enabled'] else '⬜ disabled (enable from Risk menu)'
                        await send_tg_msg(f'🎯 Daily profit target set to <b>${val:.2f}</b>\nStatus: {status_txt}')
                    else: await send_tg_msg('❌ Target must be greater than 0.')
                except ValueError: await send_tg_msg('Usage: /target VALUE\nExample: /target 100')
            else:
                cur = bot_state['daily_target_usd']; en = '✅ ON' if bot_state['daily_target_enabled'] else '⬜ OFF'
                await send_tg_msg(f'<b>Daily Profit Target</b>\nCurrent: ${cur:.2f}  ({en})\n\nUsage: /target VALUE\nExample: /target 100\n\nToggle on/off from Risk Settings menu.')

        elif msg.lower().startswith('/loss_limit'):
            parts = msg.strip().split()
            if len(parts) == 2:
                try:
                    val = float(parts[1])
                    if val > 0:
                        bot_state['daily_loss_usd'] = val; status_txt = '✅ ENABLED' if bot_state['daily_loss_enabled'] else '⬜ disabled (enable from Risk menu)'
                        await send_tg_msg(f'🛑 Daily loss limit set to <b>${val:.2f}</b>\nStatus: {status_txt}')
                    else: await send_tg_msg('❌ Loss limit must be greater than 0.')
                except ValueError: await send_tg_msg('Usage: /loss_limit VALUE\nExample: /loss_limit 100')
            else:
                cur = bot_state['daily_loss_usd']; en = '✅ ON' if bot_state['daily_loss_enabled'] else '⬜ OFF'
                await send_tg_msg(f'<b>Daily Loss Limit</b>\nCurrent: ${cur:.2f}  ({en})\n\nUsage: /loss_limit VALUE\nExample: /loss_limit 100\n\nToggle on/off from Risk Settings menu.')

        elif msg.lower().startswith('/backtest'):
            parts = msg.strip().split()
            try:
                start_dam = datetime.strptime(parts[1], '%Y-%m-%d'); start_dt = _dam_to_utc(start_dam)
                end_dt = _dam_to_utc(datetime.strptime(parts[2], '%Y-%m-%d')) + timedelta(days=1) if len(parts) >= 3 else start_dt + timedelta(days=1)
                if bot_state['is_backtesting']: await send_tg_msg('A backtest is already running.')
                else: asyncio.create_task(run_backtest(start_dt, end_dt))
            except (IndexError, ValueError): await send_tg_msg('Usage:\n/backtest YYYY-MM-DD\n/backtest YYYY-MM-DD YYYY-MM-DD')

        elif msg == '/cancel_bt':
            global _bt_progress
            if _bt_progress and bot_state['is_backtesting']: await _bt_progress.cancel(); await send_tg_msg('Cancel signal sent.')
            else: await send_tg_msg('No backtest running.')

        elif msg == '/dd':
            sod = bot_state['sod_balance']; date = bot_state['sod_date'] or '-'; trig = 'YES 🔴' if bot_state['dd_triggered'] else 'NO 🟢'
            if sod:
                lim = sod * (1 - DD_LIMIT_PCT)
                await send_tg_msg(f'<b>Daily DD</b>\nDate:{date}\nSOD:${sod:.2f}  Limit:${sod*DD_LIMIT_PCT:.2f}\nStop Equity:${lim:.2f}\nTriggered:{trig}')
            else: await send_tg_msg('DD: No SOD yet. Connect server first.')

        elif msg == '/restart_sessions':
            global _http, _poll_task
            if _poll_task and not _poll_task.done(): _poll_task.cancel()
            if _http and not _http.closed: await _http.close()
            _http = None; get_http()
            await send_tg_msg('✅ Sessions reset.\nPolling restarts in ~2s automatically.')

        elif msg.lower().startswith('/debug_bar'):
            parts = msg.strip().split()
            if len(parts) < 4: await send_tg_msg('📊 <b>Debug Bar</b>\nUsage: /debug_bar TF DATE TIME\nExample: /debug_bar 3m 2026-06-10 10:03\n(Time in Damascus UTC+3)\n\nCompare output with TradingView to find mismatch.')
            else:
                tf = parts[1].lower(); date = parts[2]; time_str = parts[3]
                try:
                    dam_dt = _dam_to_utc(f'{date} {time_str}'); htf = bot_state['htf_per_tf'].get(tf, '1h')
                    _db_htf_map = {'1m':1,'2m':2,'3m':3,'5m':5,'10m':10,'15m':15,'30m':30,'45m':45,'48m':48,'1h':60,'2h':120,'4h':240,'6h':360,'12h':720,'1d':1440}
                    _db_hmin = _db_htf_map.get(htf, 60); _db_warm = _db_hmin * bot_state['basis_len'] * 5
                    _db_total = min(max(_db_warm + 8*60 + 2*60, 3000), 15000)
                    await send_tg_msg(f'⏳ Fetching data for [{tf}] at {date} {time_str} DAM\n(= {dam_dt.strftime("%Y-%m-%d %H:%M")} UTC)\nHTF: {htf}  MA: {bot_state["basis_type"]}({bot_state["basis_len"]})\nFetching ~{_db_total} × 1m candles (warmup={_db_warm})...')
                    asyncio.create_task(_debug_bar_task(tf, htf, dam_dt, date, time_str))
                except ValueError as e: await send_tg_msg(f'❌ Parse error: {e}\nFormat: YYYY-MM-DD HH:MM')

        elif msg == '/ping':
            uptime = str(datetime.now(timezone.utc) - _start_time).split('.')[0]
            await send_tg_msg(f'🏓 <b>Pong!</b>\nUptime: {uptime}\nBot: {bot_state["status"]}\nServer: {"🟢" if bot_state["live_connected"] else "🔴"}')

        elif msg == '/status':
            active = [tf for tf in _TFS if bot_state['active_tfs'][tf]]
            htf_list = ', '.join(f'{tf}→{bot_state["htf_per_tf"].get(tf,"?")}' for tf in active)
            await send_tg_msg(
                f'<b>Bot Status</b>\nStatus:  {bot_state["status"]}\nServer:  {"🟢" if bot_state["live_connected"] else "🔴"}\n'
                f'Active:  {", ".join(active) or "None"}\nHTF map: {htf_list or "None"}\nMA type: {bot_state["basis_type"]}\n'
                f'LM per TF: {", ".join(f"{t}={bot_state["ma_len_per_tf"].get(t,10)}" for t in active) or "None"}\n'
                f'MTF:     {"ON" if bot_state["use_mtf"] else "OFF"}\nLot:     {bot_state["lot_size"]}\nLoss Strk:{bot_state["consecutive_losses"]}\n'
                f'DD:      {"TRIGGERED 🔴" if bot_state["dd_triggered"] else "OK 🟢"}\n'
                f'Target:  {"✅ ON ($" + str(bot_state["daily_target_usd"]) + ")" if bot_state["daily_target_enabled"] else "⬜ OFF"}{" — REACHED 🎯" if bot_state["profit_target_triggered"] else ""}\n'
                f'LossLim: {"✅ ON ($" + str(bot_state["daily_loss_usd"]) + ")" if bot_state["daily_loss_enabled"] else "⬜ OFF"}{" — HIT 🛑" if bot_state["loss_limit_triggered"] else ""}'
            )
        return

    if 'callback_query' not in update: return
    q = update['callback_query']; d = q['data']; chat_id = q['message']['chat']['id']; msg_id = q['message']['message_id']
    bot_state['chat_id'] = chat_id; c_log(f'CB: {d}')
    try: await _handle_callback(d, chat_id, msg_id)
    except Exception as e: c_log(f'CB error [{d}]: {e}')
    finally: await answer_callback(q['id'])

# ─────────────────────────────────────────────────────────────
# CALLBACK HANDLER
# ─────────────────────────────────────────────────────────────
_MA_TYPES = ['TEMA', 'EMA', 'DEMA', 'SMA', 'WMA', 'SMMA', 'HullMA', 'LSMA']

async def _handle_callback(d: str, chat_id: int, msg_id: int) -> None:
    global _bt_progress
    if d == 'noop': pass
    elif d == 'menu_main': await _show(chat_id, msg_id, 'Main Menu:', get_main_keyboard())
    elif d == 'menu_strategy': await _show(chat_id, msg_id, 'Strategy Settings:', get_strategy_keyboard())
    elif d == 'menu_risk': await _show(chat_id, msg_id, 'Risk Settings:', get_risk_keyboard())
    elif d == 'menu_tfs': await _show(chat_id, msg_id, 'Timeframes | HTF | MA Length:', get_tf_keyboard())
    elif d == 'menu_scanner': await _show(chat_id, msg_id, 'Signal Scanner:', get_scanner_keyboard())
    elif d == 'menu_backtest':
        txt = 'Backtest running...' if bot_state['is_backtesting'] else 'Select period:'
        await _show(chat_id, msg_id, txt, get_backtest_keyboard())
    elif d == 'hide_keyboard':
        bot_state['menu_button_map'] = {}
        await _show(chat_id, msg_id, 'تم إخفاء لوحة الأزرار 👁‍🗨\nلإظهار القائمة مجدداً، أرسل /start', {'remove_keyboard': True})

    elif d == 'toggle_live_conn':
        if not bot_state['live_connected']:
            await _show(chat_id, msg_id, '⏳ Connecting...', get_main_keyboard())
            try:
                api = MetaApi(METAAPI_TOKEN); bot_state['account_obj'] = await api.metatrader_account_api.get_account(ACCOUNT_ID)
                bot_state['connection_obj'] = bot_state['account_obj'].get_rpc_connection()
                await bot_state['connection_obj'].connect(); await bot_state['connection_obj'].wait_synchronized()
                bot_state['live_connected'] = True
                for tf in _TFS: bot_state[f'1m_base_{tf}'] = []
                await _capture_sod_balance()
                await _show(chat_id, msg_id, '✅ Connected!', get_main_keyboard())
            except Exception as e: await _show(chat_id, msg_id, f'❌ Failed:\n{e}', get_main_keyboard())
        else:
            bot_state['live_connected'] = False; bot_state['connection_obj'] = None; bot_state['account_obj'] = None
            await _show(chat_id, msg_id, '🔴 Disconnected.', get_main_keyboard())

    elif d == 'toggle_status':
        bot_state['status'] = 'PAUSED' if bot_state['status'] == 'RUNNING' else 'RUNNING'
        if bot_state['status'] == 'RUNNING':
            bot_state['dd_triggered'] = False; bot_state['profit_target_triggered'] = False; bot_state['loss_limit_triggered'] = False
            for tf in _TFS: bot_state[f'1m_base_{tf}'] = []
            await send_tg_msg('▶️ <b>Resumed</b> — cache cleared, warm-up starting.')
        await _show(chat_id, msg_id, 'Main Menu:', get_main_keyboard())

    elif d == 'toggle_use_ma': bot_state['use_ma'] = not bot_state['use_ma']; await _show(chat_id, msg_id, 'Strategy:', get_strategy_keyboard())
    elif d == 'cycle_ma_type':
        idx = _MA_TYPES.index(bot_state['basis_type']) if bot_state['basis_type'] in _MA_TYPES else 0
        bot_state['basis_type'] = _MA_TYPES[(idx + 1) % len(_MA_TYPES)]; bot_state['primary_basis_type'] = bot_state['basis_type']; await _show(chat_id, msg_id, 'Strategy:', get_strategy_keyboard())
    elif d == 'inc_ma_len': bot_state['basis_len'] = min(bot_state['basis_len'] + 1, 200); await _show(chat_id, msg_id, 'Strategy:', get_strategy_keyboard())
    elif d == 'dec_ma_len': bot_state['basis_len'] = max(bot_state['basis_len'] - 1, 2); await _show(chat_id, msg_id, 'Strategy:', get_strategy_keyboard())
    elif d == 'toggle_mtf': bot_state['use_mtf'] = not bot_state['use_mtf']; await _show(chat_id, msg_id, 'Strategy:', get_strategy_keyboard())
    elif d == 'toggle_trailing': bot_state['use_trailing'] = not bot_state['use_trailing']; await _show(chat_id, msg_id, 'Strategy:', get_strategy_keyboard())
    elif d == 'inc_trail_pts': bot_state['trail_points'] += 10; await _show(chat_id, msg_id, 'Strategy:', get_strategy_keyboard())
    elif d == 'dec_trail_pts': bot_state['trail_points'] = max(10, bot_state['trail_points'] - 10); await _show(chat_id, msg_id, 'Strategy:', get_strategy_keyboard())
    elif d == 'inc_trail_off': bot_state['trail_offset'] += 10; await _show(chat_id, msg_id, 'Strategy:', get_strategy_keyboard())
    elif d == 'dec_trail_off': bot_state['trail_offset'] = max(10, bot_state['trail_offset'] - 10); await _show(chat_id, msg_id, 'Strategy:', get_strategy_keyboard())
    elif d == 'toggle_danger': bot_state['use_danger_filter'] = not bot_state['use_danger_filter']; await _show(chat_id, msg_id, 'Strategy:', get_strategy_keyboard())

    elif d.startswith('cycle_htf_'):
        tf = d[len('cycle_htf_'):]
        if tf in _TFS:
            cur = bot_state['htf_per_tf'].get(tf, '1h'); idx = _HTF_OPTIONS.index(cur) if cur in _HTF_OPTIONS else 0
            bot_state['htf_per_tf'][tf] = _HTF_OPTIONS[(idx + 1) % len(_HTF_OPTIONS)]; bot_state[f'1m_base_{tf}'] = []
        await _show(chat_id, msg_id, 'Timeframes & HTF & LM:', get_tf_keyboard())

    elif d.startswith('lm_edit_'):
        tf = d[len('lm_edit_'):]
        if tf in _TFS: await _show(chat_id, msg_id, f'MA Length [{tf}]:', get_lm_edit_keyboard(tf))

    elif d.startswith('inc_lm1_') or d.startswith('inc_lm5_') or d.startswith('dec_lm1_') or d.startswith('dec_lm5_'):
        if   d.startswith('inc_lm5_'): step, tf = +5, d[len('inc_lm5_'):]
        elif d.startswith('dec_lm5_'): step, tf = -5, d[len('dec_lm5_'):]
        elif d.startswith('inc_lm1_'): step, tf = +1, d[len('inc_lm1_'):]
        else:                          step, tf = -1, d[len('dec_lm1_'):]
        if tf in _TFS:
            bot_state['ma_len_per_tf'][tf] = max(2, min(200, bot_state['ma_len_per_tf'].get(tf, 10) + step)); bot_state[f'1m_base_{tf}'] = []
        await _show(chat_id, msg_id, f'MA Length [{tf}]:', get_lm_edit_keyboard(tf))

    elif d.startswith('lm_set_'):
        rest  = d[len('lm_set_'):]; val   = rest.split('_')[-1]; tf    = rest[:-(len(val)+1)]      
        if tf in _TFS:
            try:
                v = int(val)
                if 2 <= v <= 200: bot_state['ma_len_per_tf'][tf] = v; bot_state[f'1m_base_{tf}'] = []
            except ValueError: pass
        await _show(chat_id, msg_id, f'MA Length [{tf}]:', get_lm_edit_keyboard(tf))

    elif d.startswith('toggle_tf_'):
        tf = d[len('toggle_tf_'):]
        if tf in bot_state['active_tfs']:
            bot_state['active_tfs'][tf] = not bot_state['active_tfs'][tf]
            if bot_state['active_tfs'][tf]: bot_state[f'1m_base_{tf}'] = []
        await _show(chat_id, msg_id, 'Timeframes | HTF | MA Length:', get_tf_keyboard())

    elif d == 'toggle_be': bot_state['use_be'] = not bot_state['use_be']; await _show(chat_id, msg_id, 'Risk:', get_risk_keyboard())
    elif d == 'toggle_spread': bot_state['use_max_spread'] = not bot_state['use_max_spread']; await _show(chat_id, msg_id, 'Risk:', get_risk_keyboard())
    elif d == 'toggle_daily_target':
        bot_state['daily_target_enabled'] = not bot_state['daily_target_enabled']
        if bot_state['daily_target_enabled']: bot_state['profit_target_triggered'] = False
        await _show(chat_id, msg_id, 'Risk:', get_risk_keyboard())
    elif d == 'toggle_daily_loss':
        bot_state['daily_loss_enabled'] = not bot_state['daily_loss_enabled']
        if bot_state['daily_loss_enabled']: bot_state['loss_limit_triggered'] = False
        await _show(chat_id, msg_id, 'Risk:', get_risk_keyboard())
    elif d == 'inc_lot': bot_state['lot_size'] = round(bot_state['lot_size'] + 0.01, 2); await _show(chat_id, msg_id, 'Risk:', get_risk_keyboard())
    elif d == 'dec_lot': bot_state['lot_size'] = max(0.01, round(bot_state['lot_size'] - 0.01, 2)); await _show(chat_id, msg_id, 'Risk:', get_risk_keyboard())
    elif d == 'view_tpsl': await _show(chat_id, msg_id, 'TP/SL per TF:', get_tpsl_overview_keyboard())

    elif d.startswith('tpsl_edit_'):
        tf = d[len('tpsl_edit_'):]
        if tf in _TFS: await _show(chat_id, msg_id, f'Edit [{tf}]:', get_tpsl_edit_keyboard(tf))

    elif any(d.startswith(p) for p in ('inc_tp5_','inc_tp10_','dec_tp5_','dec_tp10_')):
        if   d.startswith('inc_tp10_'): dr, st, tf = 'i', 10, d[len('inc_tp10_'):]
        elif d.startswith('dec_tp10_'): dr, st, tf = 'd', 10, d[len('dec_tp10_'):]
        elif d.startswith('inc_tp5_'):  dr, st, tf = 'i',  5, d[len('inc_tp5_'):]
        else:                           dr, st, tf = 'd',  5, d[len('dec_tp5_'):]
        if tf in _TFS:
            c = bot_state['tp_pips'][tf]; bot_state['tp_pips'][tf] = c + st if dr == 'i' else max(5, c - st)
            await _show(chat_id, msg_id, f'Edit [{tf}]:', get_tpsl_edit_keyboard(tf))

    elif any(d.startswith(p) for p in ('inc_sl5_','inc_sl10_','dec_sl5_','dec_sl10_')):
        if   d.startswith('inc_sl10_'): dr, st, tf = 'i', 10, d[len('inc_sl10_'):]
        elif d.startswith('dec_sl10_'): dr, st, tf = 'd', 10, d[len('dec_sl10_'):]
        elif d.startswith('inc_sl5_'):  dr, st, tf = 'i',  5, d[len('inc_sl5_'):]
        else:                           dr, st, tf = 'd',  5, d[len('dec_sl5_'):]
        if tf in _TFS:
            c = bot_state['sl_pips'][tf]; bot_state['sl_pips'][tf] = c + st if dr == 'i' else max(5, c - st)
            await _show(chat_id, msg_id, f'Edit [{tf}]:', get_tpsl_edit_keyboard(tf))

    elif d.startswith('scan_'):
        hours = int(d.split('_')[1])
        await _show(chat_id, msg_id, f'🔍 Scanning {hours}h...', get_scanner_keyboard()); asyncio.create_task(run_signal_scanner(hours))

    elif d == 'cancel_bt':
        if _bt_progress and bot_state['is_backtesting']: await _bt_progress.cancel(); await _show(chat_id, msg_id, 'Stopping...', get_main_keyboard())
        else: await _show(chat_id, msg_id, 'No BT running.', get_main_keyboard())

    elif d == 'bt_show_progress':
        if _bt_progress: await send_tg_msg(f'BT phase: {_bt_progress.phase}')
        else: await send_tg_msg('No BT running.')

    elif d.startswith('bto_'):
        if bot_state['is_backtesting']: await _show(chat_id, msg_id, 'BT already running.', get_backtest_keyboard())
        else:
            days = int(d.split('_')[1]); now_utc = datetime.now(timezone.utc); now_dam = _now_dam()
            end_dt = _dam_to_utc(datetime(now_dam.year, now_dam.month, now_dam.day)); start_dt = end_dt - timedelta(days=days)
            asyncio.create_task(run_backtest(start_dt, end_dt))
            await edit_tg_msg(chat_id, msg_id, f'⏳ Starting {days}-day backtest\n{_fmt_dam(start_dt).split()[0]} → {_fmt_dam(end_dt - timedelta(seconds=1)).split()[0]} (DAM)', get_backtest_keyboard())

    elif d == 'report':
        now_utc = datetime.now(timezone.utc)
        bl_str  = blocked_time_label(now_utc) if (bot_state['use_danger_filter'] and is_blocked_time(now_utc)) else '🟢 Open'
        lines = [f'<b>Market Report</b>', f'MA: {bot_state["basis_type"]}({bot_state["basis_len"]}) MTF:{"ON" if bot_state["use_mtf"] else "OFF"}', f'Filter: {bl_str}', '']
        for tf in bot_state['timeframes']:
            if bot_state['active_tfs'][tf]:
                lines.append(f'[{tf}|{bot_state["htf_per_tf"].get(tf, "?")}|L{bot_state["ma_len_per_tf"].get(tf, bot_state["basis_len"])}] {bot_state["market_data"][tf]}')
        await _show(chat_id, msg_id, '\n'.join(lines), get_main_keyboard())

    elif d == 'account':
        if not (bot_state['live_connected'] and bot_state['connection_obj']): await _show(chat_id, msg_id, '🔴 Not connected.', get_main_keyboard())
        else:
            try:
                info = await bot_state['connection_obj'].get_account_information(); pos = await bot_state['connection_obj'].get_positions()
                sod = bot_state['sod_balance']; eq = float(info.get('equity', 0))
                text = (f'<b>Account</b>\nBalance:  ${info.get("balance","?")}\nEquity:   ${eq}\nMargin:   ${info.get("freeMargin","?")}\nTrades:   {len(pos)}\nSOD:      ${sod:.2f}\nDD Used:  ${sod-eq:.2f} ({round((sod-eq)/sod*100,2)})%') if sod else f'<b>Account</b>\nBalance:{info.get("balance","?")}  Equity:{eq}  Trades:{len(pos)}'
                await _show(chat_id, msg_id, text, get_main_keyboard())
            except Exception as e: await _show(chat_id, msg_id, f'Error: {e}', get_main_keyboard())

    elif d == 'dd_status':
        sod = bot_state['sod_balance']; date = bot_state['sod_date'] or '-'; trig = '🔴 TRIGGERED' if bot_state['dd_triggered'] else '🟢 OK'
        if sod:
            lim  = sod * (1 - DD_LIMIT_PCT)
            await _show(chat_id, msg_id, f'<b>Daily DD (3%)</b>\nDate:{date}\nSOD:${sod:.2f}  Limit:${sod*DD_LIMIT_PCT:.2f}\nStop:${lim:.2f}  Status:{trig}\n\nBlocked (Damascus UTC+3):\n13:xx | 18:xx | 21:xx | 22:xx', get_main_keyboard())
        else: await _show(chat_id, msg_id, 'DD: No SOD yet.', get_main_keyboard())

    elif d == 'close_all':
        if not (bot_state['live_connected'] and bot_state['connection_obj']): await _show(chat_id, msg_id, '🔴 Not connected.', get_main_keyboard())
        else:
            try:
                positions = await bot_state['connection_obj'].get_positions()
                if not positions: await _show(chat_id, msg_id, 'No open trades.', get_main_keyboard())
                else:
                    for p in positions: await bot_state['connection_obj'].close_position(p['id'])
                    await edit_tg_msg(chat_id, msg_id, f'✅ Closed {len(positions)} trade(s).', get_main_keyboard())
            except Exception as e: await _show(chat_id, msg_id, f'Error: {e}', get_main_keyboard())

    else: c_log(f'CB unhandled: {d!r}')

# ─────────────────────────────────────────────────────────────
# TELEGRAM POLLING  +  WATCHDOG
# ─────────────────────────────────────────────────────────────
_poll_task: asyncio.Task | None = None

async def telegram_polling_loop() -> None:
    c_log('Telegram polling started.'); url = f'https://api.telegram.org/bot{TG_TOKEN}/getUpdates'
    connector = aiohttp.TCPConnector(limit=4, ttl_dns_cache=300, force_close=True); timeout = aiohttp.ClientTimeout(total=30, sock_read=22, connect=8); sess = aiohttp.ClientSession(connector=connector, timeout=timeout); backoff = 1
    try:
        while True:
            try:
                async with sess.get(url, params={'offset': bot_state['last_update_id'] + 1, 'timeout': 15}) as resp:
                    if resp.status == 200:
                        backoff = 1; bot_state['last_poll_ok'] = datetime.now(timezone.utc).timestamp(); data = await resp.json()
                        for upd in data.get('result', []):
                            bot_state['last_update_id'] = upd['update_id']; asyncio.create_task(_safe_process(upd))
                    elif resp.status == 429: retry = int(resp.headers.get('Retry-After', 5)); c_log(f'Polling 429 — waiting {retry}s'); await asyncio.sleep(retry)
                    else: c_log(f'Polling HTTP {resp.status}'); await asyncio.sleep(backoff); backoff = min(backoff * 2, 30)
            except asyncio.CancelledError: raise
            except Exception as e: c_log(f'Polling error: {e} — retry in {backoff}s'); await asyncio.sleep(backoff); backoff = min(backoff * 2, 30)
    finally:
        await sess.close(); c_log('Polling session closed.')

async def _safe_process(upd: dict) -> None:
    try: await process_tg_update(upd)
    except Exception as e: c_log(f'Update processing error (upd_id={upd.get("update_id")}): {e}')

async def telegram_watchdog() -> None:
    global _poll_task
    await asyncio.sleep(30)
    while True:
        await asyncio.sleep(20)
        try:
            last = bot_state.get('last_poll_ok', 0.0); age = datetime.now(timezone.utc).timestamp() - last
            if age > 60 and _poll_task is not None and not _poll_task.done(): c_log(f'Watchdog: polling silent {age:.0f}s — cancelling task for restart.'); _poll_task.cancel()
            elif age > 60: c_log(f'Watchdog: polling silent {age:.0f}s — task already dead, supervised will restart.')
        except Exception as e: c_log(f'Watchdog error: {e}')

# ─────────────────────────────────────────────────────────────
# TASK SUPERVISOR
# ─────────────────────────────────────────────────────────────
async def _debug_bar_task(tf: str, htf: str, target_utc: datetime, dam_date: str, dam_time: str) -> None:
    try:
        import traceback as _tb
        warmup = bot_state['basis_len'] * 3 * 60 + 500; end_time = target_utc + timedelta(hours=2)
        total_1m = min(warmup + int((end_time - (target_utc - timedelta(hours=8))).total_seconds() / 60), 8000)
        candles = await fetch_candles('1m', count=total_1m, end_time=end_time)
        if not candles: await send_tg_msg('❌ No data from OANDA. Check API key and symbol.'); return

        df = calculate_tema_signals(candles, exec_tf=tf, htf_res=htf, use_mtf=bot_state['use_mtf'], use_ma=bot_state['use_ma'], ma_type=bot_state['basis_type'], ma_len=bot_state['basis_len'], lookahead=False)
        df_rp = calculate_tema_signals(candles, exec_tf=tf, htf_res=htf, use_mtf=bot_state['use_mtf'], use_ma=bot_state['use_ma'], ma_type=bot_state['basis_type'], ma_len=bot_state['basis_len'], lookahead=True)

        if df.empty: await send_tg_msg('❌ Not enough data for TEMA calculation.'); return
        target_ts = _to_utc(target_utc); pv = bot_state['pip_value']

        def find_bar(df_in, ts):
            if ts in df_in.index: return ts, df_in.loc[ts]
            before = df_in.index[df_in.index <= ts]
            if len(before) == 0: return None, None
            return before[-1], df_in.loc[before[-1]]

        bar_ts,  bar  = find_bar(df,    target_ts)
        bar_ts2, bar2 = find_bar(df_rp, target_ts)

        if bar is None: await send_tg_msg(f'❌ No bar found at or before {target_utc}'); return

        bot_sig = '🟢 BUY' if bar.get('long_signal', False) else '🔴 SELL' if bar.get('short_signal', False) else '⚪ NO SIGNAL'
        tv_sig  = '🟢 BUY' if (bar2 is not None and bar2.get('long_signal', False)) else '🔴 SELL' if (bar2 is not None and bar2.get('short_signal', False)) else '⚪ NO SIGNAL'
        match = '✅ MATCH' if bot_sig == tv_sig else '❌ MISMATCH'

        idx = df.index.get_loc(bar_ts); ctx_bars = df.iloc[max(0, idx-2): idx+3]; ctx_lines = []
        for ts_c, r in ctx_bars.iterrows():
            dam_c = _utc_to_dam(ts_c).strftime('%H:%M'); sig_c = '→BUY' if r.get('long_signal', False) else '→SELL' if r.get('short_signal', False) else ''
            marker = ' ◄ TARGET' if ts_c == bar_ts else ''
            ctx_lines.append(f'  {dam_c}DAM  TC:{r.get("tema_close",0):.3f}  TO:{r.get("tema_open",0):.3f}  {sig_c}{marker}')

        idx_target = df.index.get_loc(bar_ts) if bar_ts in df.index else None
        nearby_signal_bar = None; nearby_signal_type = ''
        if idx_target is not None:
            for offset in [0, -1, -2, 1, 2, -3, 3]:
                check_i = idx_target + offset
                if 0 <= check_i < len(df):
                    r_check = df.iloc[check_i]
                    if r_check.get('long_signal', False) or r_check.get('short_signal', False):
                        nearby_signal_bar = df.index[check_i]; nearby_signal_type = '🟢 BUY' if r_check.get('long_signal') else '🔴 SELL'
                        entry_bar_dam = _utc_to_dam(df.index[check_i + 1] if check_i + 1 < len(df) else df.index[check_i]).strftime('%H:%M')
                        break

        nearby_note = ''
        if nearby_signal_bar is not None and bot_sig == '⚪ NO SIGNAL':
            nearby_dam = _utc_to_dam(nearby_signal_bar).strftime('%H:%M')
            nearby_note = (f'\n⚠️ <b>Signal found nearby:</b>\n  {nearby_signal_type} at {nearby_dam} DAM (signal bar)\n'
                           f'  Entry would be at {entry_bar_dam} DAM (next bar open)\n  → Query /debug_bar {tf} {dam_date} {nearby_dam} to confirm\n'
                           f'  → TV arrow appears on signal bar ({nearby_dam}), entry on next bar')
        elif bot_sig != '⚪ NO SIGNAL':
            entry_i = (df.index.get_loc(bar_ts) + 1) if bar_ts in df.index else None
            entry_dam = _utc_to_dam(df.index[entry_i]).strftime('%H:%M') if entry_i and entry_i < len(df) else '?'
            nearby_note = (f'\n✅ <b>Signal confirmed at this bar</b>\n  Entry (next bar open) = {entry_dam} DAM\n'
                           f'  TV shows arrow at {_utc_to_dam(bar_ts).strftime("%H:%M")} DAM (this bar)')

        msg_text = (
            f'📊 <b>Debug Bar [{tf}]</b>\nQuery: {dam_date} {dam_time} DAM → {bar_ts.strftime("%H:%M")} UTC\n'
            f'HTF:{htf}  MA:{bot_state["basis_type"]}({bot_state["basis_len"]})\n\n'
            f'<b>Bot (anti-repaint):</b> {bot_sig}\n<b>TV  (lookahead):</b>    {tv_sig}\n<b>{match}</b>\n{nearby_note}\n\n'
            f'TC:{bar.get("tema_close",0):.4f}  TO:{bar.get("tema_open",0):.4f}\n'
            f'Trend: {"▲ BULL" if bar.get("trend",0)>0 else ("▼ BEAR" if bar.get("trend",0)<0 else "─ FLAT")}  Close:{bar.get("close",0):.3f}\n\n'
            f'<b>Context (Damascus):</b>\n<code>{"".join(l+"\n" for l in ctx_lines)}</code>\n\n'
            f'<b>Tip:</b> TV arrow = signal bar, bot entry = next bar open\nIf TV uses Resolution=240(4h) set: /set {tf} htf 4h'
        )
        await send_tg_msg(msg_text)

    except Exception as e:
        tb = _tb.format_exc(); c_log(f'debug_bar full error:\n{tb}')
        tb_lines = [l for l in tb.strip().split('\n') if l.strip()]; short_tb = '\n'.join(tb_lines[-4:])
        await send_tg_msg(f'❌ Debug error: {e}\n\n<code>{short_tb}</code>')

async def supervised(coro_fn, *args, label: str = '') -> None:
    global _poll_task
    while True:
        try:
            task = asyncio.current_task()
            if label == 'tg_polling': _poll_task = task
            await coro_fn(*args)
        except asyncio.CancelledError: c_log(f'Task "{label}" cancelled — restarting.'); await asyncio.sleep(2)   
        except Exception as e: c_log(f'Task "{label or coro_fn.__name__}" crashed: {e} — restart in 5s'); await asyncio.sleep(5)

async def api_market_report(request: web.Request) -> web.Response:
    now_utc = datetime.now(timezone.utc)
    bl_str  = blocked_time_label(now_utc) if (bot_state['use_danger_filter'] and is_blocked_time(now_utc)) else '🟢 Open'
    lines = [f'MA: {bot_state["basis_type"]}({bot_state["basis_len"]}) MTF:{"ON" if bot_state["use_mtf"] else "OFF"}', f'Filter: {bl_str}', '']
    for tf in bot_state['timeframes']:
        if bot_state['active_tfs'][tf]:
            lines.append(f'[{tf}|{bot_state["htf_per_tf"].get(tf, "?")}|L{bot_state["ma_len_per_tf"].get(tf, bot_state["basis_len"])}] {bot_state["market_data"][tf]}')
    return web.json_response({'status': 'success', 'report': '\n'.join(lines)})

async def api_run_scanner(request: web.Request) -> web.Response:
    data = await request.json()
    hours = data.get('hours', 24)
    asyncio.create_task(run_signal_scanner(hours))
    return web.json_response({'status': 'success', 'message': f'Scanner started for {hours}h. Results will be sent to Telegram!'})

# ─────────────────────────────────────────────────────────────
# WEB SERVER
# ─────────────────────────────────────────────────────────────
_start_time = datetime.now(timezone.utc)

async def handle_ping(request: web.Request) -> web.Response:
    uptime = str(datetime.now(timezone.utc) - _start_time).split('.')[0]
    active = [tf for tf in _TFS if bot_state['active_tfs'][tf]]
    sod    = f'${bot_state["sod_balance"]:.2f}' if bot_state['sod_balance'] else 'N/A'
    last_p = datetime.now(timezone.utc).timestamp() - bot_state.get('last_poll_ok', 0)
    return web.Response(
        text=(f'Gold Scalper Bot v5.2 — OANDA + Crypto World TEMA\nMA: {bot_state["basis_type"]}({bot_state["basis_len"]}) MTF:{"ON" if bot_state["use_mtf"] else "OFF"}\n'
              f'Uptime: {uptime}\nServer: {"connected" if bot_state["live_connected"] else "disconnected"}\nBot: {bot_state["status"]}\n'
              f'Active: {", ".join(active) or "None"}\nBT: {"RUNNING" if bot_state["is_backtesting"] else "idle"}\n'
              f'SOD: {sod}\nDD: {"TRIGGERED" if bot_state["dd_triggered"] else "OK"}\nTG poll: {last_p:.0f}s ago'),
        content_type='text/plain',
    )

async def api_update_config(request: web.Request) -> web.Response:
    try:
        data = await request.json()
        for k, v in data.items():
            if k in ['active_tfs', 'tp_pips', 'sl_pips'] and isinstance(v, dict):
                bot_state[k].update(v)
            elif k in bot_state:
                bot_state[k] = v
                if k == 'basis_type':
                    bot_state['primary_basis_type'] = v
        return web.json_response({'status': 'success', 'bot_state': _get_safe_state()})
    except Exception as e:
        return web.json_response({'status': 'error', 'message': str(e)}, status=400)

def _get_safe_state() -> dict:
    return {
        'status': bot_state['status'],
        'live_connected': bot_state['live_connected'],
        'sod_balance': bot_state.get('sod_balance', 0.0) or 0.0,
        'dd_triggered': bot_state['dd_triggered'],
        'basis_type': bot_state['basis_type'],
        'basis_len': bot_state['basis_len'],
        'use_mtf': bot_state['use_mtf'],
        'use_ma': bot_state['use_ma'],
        'lot_size': bot_state['lot_size'],
        'daily_loss_enabled': bot_state['daily_loss_enabled'],
        'daily_loss_usd': bot_state['daily_loss_usd'],
        'daily_target_enabled': bot_state['daily_target_enabled'],
        'daily_target_usd': bot_state['daily_target_usd'],
        'active_tfs': bot_state['active_tfs'],
        'is_backtesting': bot_state['is_backtesting'],
        'use_be': bot_state.get('use_be', False),
        'use_trailing': bot_state.get('use_trailing', False),
        'trail_points': bot_state.get('trail_points', 200),
        'trail_offset': bot_state.get('trail_offset', 400),
        'use_danger_filter': bot_state.get('use_danger_filter', True),
        'use_max_spread': bot_state.get('use_max_spread', True),
        'max_spread_pips': bot_state.get('max_spread_pips', 3.0),
        'tp_pips': bot_state.get('tp_pips', {}),
        'sl_pips': bot_state.get('sl_pips', {}),
    }

async def api_status(request: web.Request) -> web.Response:
    return web.json_response(_get_safe_state())

async def api_engine_toggle(request: web.Request) -> web.Response:
    bot_state['status'] = 'PAUSED' if bot_state['status'] == 'RUNNING' else 'RUNNING'
    return web.json_response({'status': 'success'})

async def api_engine_live_conn(request: web.Request) -> web.Response:
    bot_state['live_connected'] = not bot_state['live_connected']
    return web.json_response({'status': 'success'})

async def api_positions_close_all(request: web.Request) -> web.Response:
    asyncio.create_task(_close_all_positions())
    return web.json_response({'status': 'success'})

async def api_backtest_start(request: web.Request) -> web.Response:
    try:
        data = await request.json()
        days = int(data.get('days', 7))
        end_dt = datetime.now(timezone.utc)
        start_dt = end_dt - timedelta(days=days)
        if not bot_state['is_backtesting']:
            bot_state.pop('last_backtest_result', None)
            asyncio.create_task(run_backtest(start_dt, end_dt))
            return web.json_response({'status': 'started', 'days': days})
        return web.json_response({'status': 'error', 'message': 'Backtest already running'}, status=400)
    except Exception as e:
        return web.json_response({'status': 'error', 'message': str(e)}, status=400)

async def api_backtest_status(request: web.Request) -> web.Response:
    global _bt_progress
    if not bot_state['is_backtesting'] or not _bt_progress:
        return web.json_response({
            'status': 'idle',
            'result': bot_state.get('last_backtest_result')
        })
    return web.json_response({
        'status': 'running',
        'phase': _bt_progress.phase,
        'progress': _bt_progress.overall_progress
    })

async def api_backtest_download(request: web.Request) -> web.Response:
    import os
    if os.path.exists('/tmp/latest_backtest.xlsx'):
        return web.FileResponse('/tmp/latest_backtest.xlsx', headers={
            'Content-Disposition': 'attachment; filename="latest_backtest.xlsx"'
        })
    return web.Response(text="No file generated yet", status=404)

_ws_clients = set()
async def ws_stream(request: web.Request):
    ws = web.WebSocketResponse()
    await ws.prepare(request)
    _ws_clients.add(ws)
    try:
        async for msg in ws: pass
    finally:
        _ws_clients.remove(ws)
    return ws

async def ws_pulse_loop():
    while True:
        await asyncio.sleep(2)
        if not _ws_clients: continue
        pulse = {tf: "Waiting..." for tf in _TFS}
        pnl = {}
        for ws in list(_ws_clients):
            try:
                await ws.send_json({"type": "market_pulse", "data": pulse})
                await ws.send_json({"type": "positions_pnl", "data": pnl})
            except Exception: pass

# ─────────────────────────────────────────────────────────────
# ENTRY POINT
# ─────────────────────────────────────────────────────────────
async def main() -> None:
    get_http()
    app = web.Application()
    app.router.add_get('/', handle_ping)
    app.router.add_get('/api/status', api_status)
    app.router.add_put('/api/config', api_update_config)
    app.router.add_post('/api/engine/toggle', api_engine_toggle)
    app.router.add_post('/api/engine/live_conn', api_engine_live_conn)
    app.router.add_post('/api/positions/close_all', api_positions_close_all)
    app.router.add_post('/api/backtest/start', api_backtest_start)
    app.router.add_get('/api/backtest/status', api_backtest_status)
    app.router.add_get('/api/backtest/download', api_backtest_download)
    app.router.add_get('/api/report', api_market_report)
    app.router.add_post('/api/scanner/start', api_run_scanner)
    app.router.add_get('/ws/stream', ws_stream)
    
    runner = web.AppRunner(app); await runner.setup()
    port = int(os.environ.get('PORT', 10000)); await web.TCPSite(runner, '0.0.0.0', port).start(); c_log(f'Web server on port {port}')

    for tf in _TFS: bot_state[f'1m_base_{tf}'] = []
    bot_state['last_poll_ok'] = datetime.now(timezone.utc).timestamp()

    tasks = [asyncio.create_task(supervised(timeframe_scanner, tf, label=f'scanner_{tf}')) for tf in bot_state['timeframes']]
    tasks += [
        asyncio.create_task(supervised(telegram_polling_loop, label='tg_polling')),
        asyncio.create_task(supervised(telegram_watchdog,     label='tg_watchdog')),
        asyncio.create_task(supervised(position_monitor,      label='pos_monitor')),
        asyncio.create_task(supervised(daily_drawdown_monitor, label='dd_monitor')),
        asyncio.create_task(supervised(daily_profit_target_monitor, label='profit_target_monitor')),
        asyncio.create_task(supervised(ws_pulse_loop, label='ws_pulse_loop')),
    ]

    c_log('Gold Scalper Bot v5.2 3D-Hybrid started.')
    try: await asyncio.gather(*tasks)
    finally:
        if _http and not _http.closed: await _http.close()
        c_log('Bot shut down.')

if __name__ == '__main__':
    asyncio.run(main())
