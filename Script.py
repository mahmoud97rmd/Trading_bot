import json
"""
Gold Scalper Bot -- v9.4 (Resilience-First Core)
Strategy : Gann Levels + Fan Angles + Break-Even Triggered by Noise Levels

v9.4 changes vs v8.9 (see PATCH_NOTES.md shipped alongside this file):
  - No hardcoded credential fallbacks; bot refuses to start without env vars.
  - No silent except-pass in execution / reconciliation / order-management paths.
  - Explicit HALT / READ_ONLY connection-state machine with Telegram escalation.
  - Persistence now captures full per-symbol cycle state, not just open trades.
  - Startup reconstructs state from disk before ANY market interaction.
"""

import asyncio
import logging
import traceback
import time
import random
import zlib
import aiohttp
import os
import sys
import pandas as pd
import numpy as np
from datetime import datetime, timedelta, timezone, time as dtime
from aiohttp import web
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from metaapi_cloud_sdk import MetaApi, SynchronizationListener

# -----------------------------------------------------------------
# LOGGING (structured, always includes tracebacks for exceptions)
# -----------------------------------------------------------------
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s [%(levelname)s] %(message)s',
    stream=sys.stdout,
)
logger = logging.getLogger('gold_scalper')

def log_exception(context: str, exc: Exception) -> None:
    """Zero-tolerance logging: every caught exception in a critical path gets
    a full traceback attached to the log line, not just str(e)."""
    logger.error("EXCEPTION in %s: %s\n%s", context, exc, traceback.format_exc())

_DIAG_LOG_MAX_ENTRIES = 50000  # ~generous cap; trimmed on every append

def _diag_log_add(entry: dict) -> None:
    """Append one row to the rolling live-scan diagnostic log (see
    bot_state['diag_log']). This is what /export_diag_excel dumps -- it's
    the ONLY place that records the *silent* skip reasons (insufficient
    candle data, cap reached, trend unknown, etc.) that never get a
    Telegram message of their own, so the operator can reconstruct exactly
    what the scanner saw/did on every (symbol, timeframe, cycle), not just
    a point-in-time snapshot."""
    log = bot_state.setdefault('diag_log', [])
    log.append(entry)
    if len(log) > _DIAG_LOG_MAX_ENTRIES:
        del log[: len(log) - _DIAG_LOG_MAX_ENTRIES]

def _record_closed_trade_history(symbol: str, tid: str, tr: dict, exit_px: float, pnl: float,
                                  outcome_label: str, close_reason: str, pnl_confirmed: bool) -> None:
    """Append one row of full detail for a just-closed real/virtual live
    trade, feeding /export_live_trades_excel. Kept deliberately rich (every
    field a human would need to judge "did this trade behave like the
    backtest expected") since this is the ONLY place live trade outcomes
    get durably recorded anywhere in the bot today."""
    hist = bot_state.setdefault('live_trade_history', [])
    entry = tr.get('entry'); is_buy = tr.get('is_buy')
    opened_at = tr.get('opened_at')
    closed_at_dt = datetime.now(timezone.utc)
    duration_min = None
    if opened_at:
        try:
            opened_dt = datetime.fromisoformat(opened_at) if isinstance(opened_at, str) else opened_at
            duration_min = round((closed_at_dt - opened_dt).total_seconds() / 60.0, 1)
        except Exception:
            duration_min = None
    intended_entry = tr.get('level_price', entry)
    entry_slip = (entry - intended_entry) if (entry is not None and intended_entry is not None) else None
    hist.append({
        'symbol': symbol, 'tid': tid, 'tf': tr.get('tf'), 'is_real': bool(tr.get('is_real')),
        'is_buy': is_buy, 'opened_at': opened_at, 'closed_at': closed_at_dt.isoformat(),
        'duration_min': duration_min, 'level_price': intended_entry, 'entry': entry,
        'entry_slippage': entry_slip, 'tp': tr.get('tp'), 'sl': tr.get('sl'), 'exit_price': exit_px,
        'outcome': outcome_label, 'pnl': pnl, 'pnl_confirmed_from_broker': pnl_confirmed,
        'close_reason': close_reason, 'be_activated': bool(tr.get('be_activated')),
        'feed_source': tr.get('feed_source'), 'feed_age_ms': tr.get('feed_age_ms'),
    })
    if len(hist) > _DIAG_LOG_MAX_ENTRIES:
        del hist[: len(hist) - _DIAG_LOG_MAX_ENTRIES]

# -----------------------------------------------------------------
# CONFIGURATION
# -----------------------------------------------------------------
def _require_env(name: str) -> str:
    val = os.environ.get(name)
    if not val:
        logger.critical("FATAL: required environment variable '%s' is not set. Refusing to start.", name)
        sys.exit(1)
    return val

METAAPI_TOKEN  = _require_env('METAAPI_TOKEN')
ACCOUNT_ID     = _require_env('ACCOUNT_ID')
TG_TOKEN       = _require_env('TG_TOKEN')
OANDA_ACCOUNT  = _require_env('OANDA_ACCOUNT')
OANDA_TOKEN    = _require_env('OANDA_TOKEN')
AVAILABLE_SYMBOLS = ['XAU_USD', 'XAU_EUR', 'XAG_USD', 'EUR_USD', 'GBP_JPY', 'GBP_AUD', 'GBP_NZD', 'AUD_JPY', 'NZD_JPY']
SYMBOL_INFO = {
    'XAU_USD': {'pip_value': 0.1,     'contract_size': 100,    'prec': 2, 'name': 'Gold (USD)'},
    'XAU_EUR': {'pip_value': 0.1,     'contract_size': 100,    'prec': 2, 'name': 'Gold (EUR)'},
    'XAG_USD': {'pip_value': 0.001,   'contract_size': 5000,   'prec': 3, 'name': 'Silver'},
    'EUR_USD': {'pip_value': 0.00001, 'contract_size': 100000, 'prec': 5, 'name': 'EUR/USD'},
    'GBP_JPY': {'pip_value': 0.01,    'contract_size': 100000, 'prec': 3, 'name': 'GBP/JPY'},
    'GBP_AUD': {'pip_value': 0.00001, 'contract_size': 100000, 'prec': 5, 'name': 'GBP/AUD'},
    'GBP_NZD': {'pip_value': 0.00001, 'contract_size': 100000, 'prec': 5, 'name': 'GBP/NZD'},
    'AUD_JPY': {'pip_value': 0.01,    'contract_size': 100000, 'prec': 3, 'name': 'AUD/JPY'},
    'NZD_JPY': {'pip_value': 0.01,    'contract_size': 100000, 'prec': 3, 'name': 'NZD/JPY'},
}
OANDA_BASE_URL = 'https://api-fxpractice.oanda.com/v3'  

_TFS = ['1m', '2m', '3m', '4m', '5m', '6m', '10m', '15m', '20m', '30m', '1h', '2h']

_http: aiohttp.ClientSession | None = None

def get_http() -> aiohttp.ClientSession:
    global _http
    if _http is None or _http.closed:
        connector = aiohttp.TCPConnector(limit=20, ttl_dns_cache=300)
        timeout   = aiohttp.ClientTimeout(total=30, connect=10)
        _http     = aiohttp.ClientSession(connector=connector, timeout=timeout)
    return _http

def c_log(msg: str) -> None:
    dam = (datetime.now(timezone.utc) + timedelta(hours=3)).strftime('%H:%M:%S')
    print(f"[{dam} DAM] {msg}", flush=True)

# -----------------------------------------------------------------
# GLOBAL STATE
# -----------------------------------------------------------------
_metaapi = None
_metaapi_account = None
_metaapi_conn = None

# ── Live-quote push cache (WebSocket/streaming feed) ──
# Populated by _GannPriceListener.on_symbol_price_updated, keyed by the
# OANDA-format symbol used everywhere else in the bot ('XAU_USD'), NOT
# the broker's own symbol name ('XAUUSD') -- _broker_to_data_symbol
# translates incoming broker-symbol ticks back to that key.
live_quotes: dict[str, dict] = {}          # {'XAU_USD': {'bid':, 'ask':, 'mid':, 'ts': monotonic}}
_broker_to_data_symbol: dict[str, str] = {}  # {'XAUUSD': 'XAU_USD', ...}
_QUOTE_STALE_SECONDS = 5.0
# Updated on EVERY tick received from MetaApi, for ANY symbol -- this is
# deliberately independent of live_quotes/_broker_to_data_symbol (which can
# be empty/wrong) and of _metaapi_account.connection_status (which the SDK
# can keep reporting 'CONNECTED' even when the underlying WS session has
# gone silent/zombie, e.g. during a broker daily-rollover freeze). The
# watchdog in gann_monitor_scanner uses ONLY this raw timestamp to decide
# whether to force a full connection teardown+reconnect.
_last_any_tick_ts = time.monotonic()
_WS_WATCHDOG_STALE_SECONDS = 60.0

# Per-symbol cache refreshed periodically by gann_monitor_scanner (levels,
# trend state, each enabled tf's closed-candle data). _gann_tick_fire_check
# reads this on every price tick -- refreshing it doesn't need tick-level
# freshness, only the live price checked against it does, and that now
# always comes straight from the tick that triggered the check.
_gann_cache: dict[str, dict] = {}


class _GannPriceListener(SynchronizationListener):
    """Pushed quotes from MetaApi's streaming connection -- this is what
    'broker-direct WebSocket price feed' actually means: MetaApi's own
    live terminal-state cache, not OANDA's REST polling. Feeds
    live_quotes[]; the scanner reads from there instead of calling out
    to OANDA for the current price."""

    async def on_symbol_price_updated(self, instance_index, price):
        global _last_any_tick_ts
        _last_any_tick_ts = time.monotonic()
        broker_sym = price.get('symbol')
        data_sym = _broker_to_data_symbol.get(broker_sym)
        if not data_sym:
            return
        bid = price.get('bid'); ask = price.get('ask')
        if bid is None or ask is None:
            return
        mid = (bid + ask) / 2
        live_quotes[data_sym] = {'bid': bid, 'ask': ask, 'mid': mid, 'ts': time.monotonic()}
        # Event-driven touch detection: react to THIS tick right now, not on
        # the next scanner cycle. create_task() so this callback returns
        # immediately -- _gann_tick_fire_check does its own (awaited) I/O,
        # but none of that blocks the SDK's socket message processing since
        # it's a separate task, not inline in this callback.
        asyncio.create_task(_gann_tick_fire_check(data_sym, mid, 0.0))

    async def on_connected(self, instance_index, replicas):
        c_log("MetaAPI streaming connection established (price feed live).")

    async def on_disconnected(self, instance_index):
        c_log("MetaAPI streaming connection lost -- reconnect loop will retry and resubscribe.")


async def _gann_tick_fire_check(symbol: str, live_px: float, feed_age_ms: float) -> None:
    """The actual touch decision, run the instant a new tick arrives (called
    via create_task from _GannPriceListener.on_symbol_price_updated). Uses
    whatever levels/trend/candle data gann_monitor_scanner's periodic
    refresh last cached for this symbol in _gann_cache -- but the PRICE
    checked against that data is always this exact tick, never a value
    read back out of a timer loop. That's what makes firing on a stale
    quote structurally impossible now, not just less likely."""
    try:
        if bot_state.get('connection_state') != CONN_RUNNING:
            return
        if bot_state.get('live_daily_hit'):
            return
        if bot_state.get('prot_dam_time_filter', True):
            dam_time = (datetime.now(timezone.utc) + timedelta(hours=3)).time()
            if any(start <= dam_time < end for start, end in _DAM_RESTRICTED_WINDOWS):
                return

        cache = _gann_cache.get(symbol)
        if not cache:
            return
        sym_state = bot_state['symbol_state'][symbol]
        if not sym_state['gann_cycle_active'] or not sym_state['gann_levels']:
            return

        max_concurrent = int(bot_state.get('prot_max_concurrent_trades', 4))
        open_count = sum(1 for v in sym_state['gann_open_trades'].values() if isinstance(v, dict))
        if open_count >= max_concurrent:
            return

        margin = cache['margin']; levels = cache['levels']; trend_up = cache['trend_up']
        entry_mode = sym_state['gann_entry_mode']
        exec_mode = bot_state.get('gann_execution_mode', 'instant')
        pv = SYMBOL_INFO[symbol]['pip_value']
        spike_limit = bot_state.get('gann_spike_limit_pts', 20) * pv
        flt_type = sym_state['trend_filter_type']; ttf = sym_state['trend_timeframe']
        detect_time = datetime.now(timezone.utc)

        for tf in cache['enabled_tfs']:
            if any(isinstance(v, dict) and v.get('tf') == tf for v in sym_state['gann_open_trades'].values()):
                continue
            tf_data = cache['tf_data'].get(tf)
            if not tf_data:
                continue
            candles = tf_data['candles']; closed_close = tf_data['closed_close']

            if entry_mode == 'touch_trend' and trend_up is None:
                continue

            # ── Which trigger channel(s) to evaluate this tick ──
            # For the 3 existing modes this is a single channel, identical
            # to the old behavior. all_concurrent runs all three
            # independently -- each is checked against the SAME level pool
            # but gets its OWN dedup key (see combo_key below), so e.g. a
            # touch and a close can both fire on the same level without
            # blocking each other, exactly as requested for the 24h
            # concurrent comparison test.
            if exec_mode == 'all_concurrent':
                channels = ['touch', 'close', 'hybrid']
            elif exec_mode == 'close':
                channels = ['close']
            elif exec_mode == 'hybrid':
                channels = ['hybrid']
            else:
                channels = ['touch']  # 'instant'

            for channel in channels:
                for lv in levels:
                    k = lv['key']; dir_ = lv['dir']
                    # Only all_concurrent suffixes the channel onto the dedup
                    # key -- the other 3 modes keep their exact original key
                    # so nothing about their existing dedup/persistence
                    # behavior changes.
                    base_combo = f"{k}_{tf}" if bot_state['prot_allow_multi_tf'] else k
                    combo_key = f"{base_combo}_{channel}" if exec_mode == 'all_concurrent' else base_combo
                    if sym_state['gann_level_status'].get(combo_key) == 'used':
                        continue
                    is_buy = (dir_ == 'dn')
                    if entry_mode == 'touch_trend':
                        if is_buy and not trend_up: continue
                        if not is_buy and trend_up: continue

                    # ── Execution mode gate ──
                    # close: fire purely off the OANDA closed-candle close, exactly
                    # like run_gann_backtest's bar_close check -- do NOT also require
                    # the current MetaApi tick to be within margin (that's a second,
                    # cross-feed condition the classic backtest never applies, and it
                    # was silently killing/mismatching trades vs the backtest).
                    if channel == 'close':
                        if abs(closed_close - lv['price']) > margin:
                            continue
                    elif channel == 'hybrid':
                        if abs(live_px - lv['price']) > margin:
                            continue
                        if abs(live_px - closed_close) > spike_limit:
                            continue
                    else:  # touch (instant)
                        if abs(live_px - lv['price']) > margin:
                            continue

                    # ── Reserve the level NOW, synchronously ──
                    # asyncio only switches tasks at an `await`, so writing this
                    # before any await here is atomic with respect to every other
                    # concurrent tick's task -- a second tick arriving a
                    # microsecond later will see 'used' immediately, even though
                    # _gann_open_trade's own internal marking (further below,
                    # after several awaits of its own) hasn't happened yet. Without
                    # this, going event-driven (many concurrent per-tick tasks
                    # instead of one serial scanner loop) would let two
                    # near-simultaneous ticks both fire on the same level.
                    sym_state['gann_level_status'][combo_key] = 'used'

                    if flt_type == 'vwap': flt_label = f"VWAP={sym_state['trend_vwap_period']}\n"
                    elif flt_type == 'ema': flt_label = f"EMA={sym_state['trend_ema_period']}\n"
                    else: flt_label = "VWAP+EMA"
                    trigger_lbl = {'touch': 'لمس مباشر ⚡', 'close': 'إغلاق شمعة ⏳', 'hybrid': 'تنفيذ هجين 🛡️'}[channel]
                    dir_word = 'BUY' if is_buy else 'SELL'
                    dir_emoji = '📈' if is_buy else '📉'
                    reason = f"{dir_word} {dir_emoji} [{symbol} - جان {tf}] {trigger_lbl} (مع {flt_label}_{ttf.upper()})"

                    t1_signal_ts = time.monotonic()
                    await _gann_open_trade(symbol, is_buy, lv, candles, reason=reason, tf=tf,
                                            initial_px=live_px, detect_time=detect_time, t1_signal_ts=t1_signal_ts,
                                            feed_source='ws', feed_age_ms=feed_age_ms, trigger_type=channel)
                    break  # this channel found its level for this tick -- other channels still get their own independent check
    except Exception as e:
        log_exception(f"_gann_tick_fire_check [{symbol}]", e)


def _is_market_hours_now() -> bool:
    """Rough broker session-hours check (from the real XAUUSD symbol spec:
    Sun 01:01 -> Fri 23:49 broker time, closed Saturday). Used only to
    decide whether a 60s tick silence is suspicious (worth a forced
    reconnect) or expected (weekend close) -- not used anywhere else."""
    offset = bot_state.get('broker_time_offset', 3)
    broker_now = datetime.now(timezone.utc) + timedelta(hours=offset)
    wd = broker_now.weekday()  # Monday=0 ... Sunday=6
    t = broker_now.time()
    if wd == 5:                                   # Saturday: fully closed
        return False
    if wd == 4 and t >= dtime(23, 49):             # Friday after close
        return False
    if wd == 6 and t < dtime(1, 1):                # Sunday before open
        return False
    return True

async def _force_full_reconnect(reason: str) -> None:
    """WebSocket watchdog escalation: unlike _lq_subscribe_symbol (which
    just re-sends a market-data subscription on the EXISTING connection
    object), this tears the streaming connection down and rebuilds it from
    scratch. Needed because a MetaApi streaming session can go "zombie" --
    connection_status still reports CONNECTED and re-subscribing throws no
    error, yet no more ticks ever arrive (observed during a broker daily
    rollover). The only reliable signal for that state is tick silence
    itself, which is what triggers this, not the SDK's own status flag."""
    global _metaapi_conn, _last_any_tick_ts
    c_log(f"WS WATCHDOG: forcing full reconnect -- {reason}")
    await set_connection_state(CONN_READ_ONLY, f"WS watchdog: {reason}")
    try:
        if _metaapi_conn is not None:
            try:
                await _metaapi_conn.close()
            except Exception as e:
                log_exception('_force_full_reconnect: close old connection', e)
        _metaapi_conn = _metaapi_account.get_streaming_connection()
        _metaapi_conn.add_synchronization_listener(_GannPriceListener())
        await _metaapi_conn.connect()
        await _metaapi_conn.wait_synchronized()
        for sym, on in bot_state['active_symbols'].items():
            if on:
                await _lq_subscribe_symbol(sym)
        _last_any_tick_ts = time.monotonic()  # don't re-trip the watchdog on the pre-reconnect silence
        c_log("WS WATCHDOG: reconnect successful, ticks should resume.")
        await set_connection_state(CONN_RUNNING, "WS watchdog: forced reconnect succeeded.")
        await send_tg_msg(f"🔁 <b>Watchdog: أعيد الاتصال تلقائياً بـ MetaApi</b>\nالسبب: {reason}")
    except Exception as e:
        log_exception('_force_full_reconnect', e)
        await send_tg_msg(f"🛑 <b>Watchdog: فشلت محاولة إعادة الاتصال التلقائي</b>\nالسبب الأصلي: {reason}\nالخطأ: {e}")


async def _lq_subscribe_symbol(symbol: str) -> None:
    """Subscribe one OANDA-format symbol's broker equivalent to live
    quotes. Safe to call multiple times (idempotent on the broker side);
    swallows failures so one bad symbol doesn't block the others."""
    if _metaapi_conn is None:
        return
    broker_sym = _resolve_broker_symbol(symbol)
    _broker_to_data_symbol[broker_sym] = symbol
    try:
        await _metaapi_conn.subscribe_to_market_data(broker_sym)
    except Exception as e:
        log_exception(f"_lq_subscribe_symbol [{symbol} -> {broker_sym}]", e)


def _lq_is_stale(symbol: str) -> bool:
    q = live_quotes.get(symbol)
    return q is None or (time.monotonic() - q['ts']) > _QUOTE_STALE_SECONDS


async def _lq_price_with_fallback(symbol: str) -> tuple[float | None, str, float | None]:
    """Returns (price, source, age_ms). Prefers the pushed WebSocket
    quote; falls back to the OANDA REST fetch (fetch_master_price) if
    the feed is missing/stale, so a temporary MetaApi hiccup degrades
    to the old behavior instead of silently blocking all touch checks."""
    q = live_quotes.get(symbol)
    if q is not None and (time.monotonic() - q['ts']) <= _QUOTE_STALE_SECONDS:
        return q['mid'], 'ws', round((time.monotonic() - q['ts']) * 1000)
    px = await fetch_master_price(symbol)
    return px, 'oanda_fallback', None

# Connection-state machine.
# RUNNING    : normal operation, new trades allowed.
# READ_ONLY  : sync with MetaAPI is degraded/unavailable. No new trades,
#              no destructive local state changes (Amnesia Prevention),
#              existing positions still managed if MT5 fallback price works.
# HALTED     : hard stop. New entries and order management both stop;
#              a human must intervene.
CONN_RUNNING   = 'RUNNING'
CONN_READ_ONLY = 'READ_ONLY'
CONN_HALTED    = 'HALTED'

_state_lock = asyncio.Lock()

async def set_connection_state(new_state: str, reason: str) -> None:
    async with _state_lock:
        old_state = bot_state.get('connection_state', CONN_RUNNING)
        if old_state == new_state:
            return
        bot_state['connection_state'] = new_state
        bot_state['connection_state_reason'] = reason
    logger.warning("Connection state: %s -> %s (%s)", old_state, new_state, reason)
    icon = {'RUNNING': '\u2705', 'READ_ONLY': '\U0001F7E1', 'HALTED': '\U0001F6D1'}.get(new_state, '\u2139')
    await send_tg_msg(f"{icon} <b>connection state changed: {old_state} -> {new_state}</b>\n{reason}")

_DAM_RESTRICTED_WINDOWS = [
    (dtime(7, 0),  dtime(9, 0)),   # European Open fakeouts
    (dtime(13, 0), dtime(14, 0)),  # Pre-US session turbulence
]

def _is_within_dam_restricted_window() -> bool:
    """DAM (Damascus / UTC+3) time-of-day filter. Based on backtest
    analysis, these windows carry enough market noise to invalidate Gann
    levels and stack losses, so new entries are skipped during them.
    Existing-position management (BE/TP/SL/closures) is NOT affected --
    this only blocks NEW trade dispatch, same scope as is_trading_allowed().
    Toggleable via bot_state['prot_dam_time_filter'] (default: on)."""
    if not bot_state.get('prot_dam_time_filter', True):
        return False
    dam_now = datetime.now(timezone.utc) + timedelta(hours=3)
    t = dam_now.time()
    return any(start <= t < end for start, end in _DAM_RESTRICTED_WINDOWS)

def is_trading_allowed() -> bool:
    """New order placement is only allowed when the connection state is
    fully healthy AND we're not inside a restricted DAM time window.
    Existing-position management (BE/TP/SL) is handled separately and is
    NOT gated by this, per the OANDA-degraded-mode rule."""
    if bot_state.get('connection_state', CONN_RUNNING) != CONN_RUNNING:
        return False
    if _is_within_dam_restricted_window():
        return False
    return True

async def _bootstrap_metaapi_connection() -> bool:
    """The actual connect-and-subscribe logic, extracted so it can be
    retried from the live scanner loop, not just called once at process
    startup. Returns True on success. This is what closes the gap where
    a transient MetaApi/broker hiccup during the ONE startup attempt left
    _metaapi_conn permanently None with no other recovery path able to
    rebuild it from scratch (both the WS tick-watchdog and the Zombie
    Singleton Heartbeat below require a connection object to already
    exist before they can do anything)."""
    global _metaapi, _metaapi_account, _metaapi_conn, _last_any_tick_ts
    try:
        _metaapi = MetaApi(METAAPI_TOKEN)
        _metaapi_account = await _metaapi.metatrader_account_api.get_account(ACCOUNT_ID)
        if _metaapi_account.state == 'DEPLOYED' and _metaapi_account.connection_status == 'CONNECTED':
            _metaapi_conn = _metaapi_account.get_streaming_connection()
            _metaapi_conn.add_synchronization_listener(_GannPriceListener())
            await _metaapi_conn.connect()
            await _metaapi_conn.wait_synchronized()
            for sym, on in bot_state['active_symbols'].items():
                if on:
                    await _lq_subscribe_symbol(sym)
            c_log("MetaAPI Streaming Connection established (live quotes subscribed).")
            _last_any_tick_ts = time.monotonic()
            await set_connection_state(CONN_RUNNING, "MetaAPI connected and synchronized.")
            return True
        else:
            c_log(f"MetaAPI account not deployed/connected (state={_metaapi_account.state}, "
                  f"conn={_metaapi_account.connection_status}).")
            await set_connection_state(CONN_READ_ONLY, "MetaAPI account is not DEPLOYED/CONNECTED.")
            return False
    except Exception as e:
        log_exception("_bootstrap_metaapi_connection", e)
        await set_connection_state(CONN_READ_ONLY, f"MetaAPI connection bootstrap failed: {e}")
        return False


async def init_metaapi():
    """Startup order is fixed:
       1) Reconstruct state from the persistence file (works even if the
          broker/API is completely unreachable).
       2) Only THEN attempt to talk to MetaAPI / the market.
    """
    load_bot_persistence()
    if bot_state.get('_persistence_load_failed'):
        await set_connection_state(
            CONN_READ_ONLY,
            "Startup persistence file was present but unreadable. Starting READ_ONLY until a human "
            "confirms the true broker state and clears this manually."
        )
    await _bootstrap_metaapi_connection()

DATA_DIR = os.environ.get('PERSISTENT_DATA_PATH', '/app/data')
os.makedirs(DATA_DIR, exist_ok=True)
PERSISTENCE_FILE = os.path.join(DATA_DIR, 'bot_persistence.json')
TEMP_PERSISTENCE_FILE = os.path.join(DATA_DIR, 'bot_persistence.tmp')
PRESETS_FILE = os.path.join(DATA_DIR, 'presets.json')
TEMP_PRESETS_FILE = os.path.join(DATA_DIR, 'presets.tmp')

# Live runtime fields that a preset must never capture or restore --
# includes gann_last_h1_time/gann_cycle_started_at (raw datetime objects,
# not JSON-serializable at all) plus other in-flight state that belongs to
# whatever is currently running, not to a saved settings snapshot.
_PRESET_EXCLUDED_KEYS = {
    'gann_levels', 'gann_level_status', 'gann_cycle_active', 'gann_open_trades',
    'gann_last_h1_time', 'gann_cycle_started_at', 'auto_trade',
}

# Event-loop I/O offloading (v9.5): json.dump + os.fsync + os.replace are
# blocking syscalls. Calling them directly from async code stalls the
# ENTIRE event loop for their duration -- every other coroutine (candle
# fetches, BE checks, Telegram alerts, callback handling) waits behind a
# single disk write. The fix: build the snapshot synchronously (cheap,
# pure in-memory dict work, no yield point, so it's still atomic w.r.t.
# bot_state), then push the actual disk I/O to a worker thread via
# asyncio.to_thread and await it there. Writes are additionally serialized
# with an asyncio.Lock so two saves in flight can't interleave writes to
# the shared .tmp path before either os.replace runs.
_persistence_write_lock = asyncio.Lock()

def _write_persistence_file_sync(data: dict) -> None:
    """Pure blocking I/O, no bot_state access -- safe to run in a thread."""
    with open(TEMP_PERSISTENCE_FILE, 'w') as f:
        json.dump(data, f)
        f.flush()
        os.fsync(f.fileno())
    os.replace(TEMP_PERSISTENCE_FILE, PERSISTENCE_FILE)

async def save_bot_persistence() -> None:
    """Atomic write: full operational state AND full settings, so a hard
    restart reconstructs the bot's world exactly -- not just open trades
    and Gann cycle state, but every user-configured setting (lot size,
    protection limits, anchor timeframe, filters, TP/SL config, etc).

    Deliberately exclude-list based, not include-whitelist based: the
    previous version only ever saved a fixed list of live trade-state
    fields, which meant lot size, protection dd/profit limits, the anchor
    timeframe, and effectively every other setting were NEVER actually
    persisted -- even though save_bot_persistence() was correctly being
    called after every mutation. An exclude-list means any new setting
    added later is persisted automatically instead of silently dropped
    until someone remembers to add it to a whitelist.
    """
    try:
        # Fields that either aren't JSON-serializable or are purely
        # transient/regenerated-on-render -- everything else in bot_state
        # is a real setting and gets saved.
        TOP_LEVEL_EXCLUDE = {'connection_obj', 'menu_button_map', 'timeframes',
                              'is_backtesting', 'live_connected', 'last_poll_ok', 'symbol_state',
                              # 'status' is a dead legacy field predating the connection_state
                              # machine -- nothing in the current code ever writes to it, but a
                              # stale value loaded from an old persistence file used to silently
                              # kill the entire scanner/cycle-manager/reconciliation loop with
                              # zero error message. Never restore it; always fixed at 'RUNNING'.
                              'status',
                              # Diagnostic-only rolling buffer -- large, purely informational,
                              # and reset-on-restart is fine. Never persist or reload it.
                              'diag_log'}

        symbol_snapshot = {}
        for sym in bot_state['active_symbols']:
            ss = bot_state['symbol_state'][sym]
            snap = {k: v for k, v in ss.items() if k not in ('gann_last_h1_time', 'gann_cycle_started_at')}
            snap['gann_last_h1_time'] = ss.get('gann_last_h1_time').isoformat() if ss.get('gann_last_h1_time') else None
            snap['gann_cycle_started_at'] = ss.get('gann_cycle_started_at').isoformat() if ss.get('gann_cycle_started_at') else None
            symbol_snapshot[sym] = snap

        data = {
            'schema_version': 3,
            'symbol_state': symbol_snapshot,
        }
        for k, v in bot_state.items():
            if k not in TOP_LEVEL_EXCLUDE:
                data[k] = v
        data['live_daily_date'] = str(bot_state.get('live_daily_date'))
    except Exception as e:
        log_exception("save_bot_persistence (snapshot phase)", e)
        return

    try:
        async with _persistence_write_lock:
            await asyncio.to_thread(_write_persistence_file_sync, data)
    except Exception as e:
        # Persistence failing is itself a critical-path failure: if we can't
        # save state, a crash right now means real, silent data loss on
        # open positions. Escalate loudly instead of swallowing it.
        log_exception("save_bot_persistence (write phase)", e)
        c_log(f"CRITICAL: Persistence Save Error -- open trade state may not survive a restart: {e}")

def load_bot_persistence():
    if not os.path.exists(PERSISTENCE_FILE):
        c_log("No persistence file found -- starting fresh (expected on first boot).")
        return
    try:
        with open(PERSISTENCE_FILE, 'r') as f:
            data = json.load(f)

        TOP_LEVEL_EXCLUDE = {'connection_obj', 'menu_button_map', 'timeframes',
                              'is_backtesting', 'live_connected', 'last_poll_ok', 'symbol_state',
                              # 'status' is a dead legacy field predating the connection_state
                              # machine -- nothing in the current code ever writes to it, but a
                              # stale value loaded from an old persistence file used to silently
                              # kill the entire scanner/cycle-manager/reconciliation loop with
                              # zero error message. Never restore it; always fixed at 'RUNNING'.
                              'status',
                              'diag_log'}
        for k, v in data.items():
            # Only restore keys that already exist in bot_state's default
            # shape -- never let a saved file inject brand-new top-level
            # keys the current code doesn't define.
            if k in bot_state and k not in TOP_LEVEL_EXCLUDE and k != 'live_daily_date':
                bot_state[k] = v

        saved_date = data.get('live_daily_date')
        if saved_date and saved_date != 'None':
            bot_state['live_daily_date'] = datetime.strptime(saved_date, '%Y-%m-%d').date()

        symbol_state_data = data.get('symbol_state')
        if symbol_state_data is not None:
            for sym, snap in symbol_state_data.items():
                if sym not in bot_state['symbol_state']:
                    continue
                ss = bot_state['symbol_state'][sym]
                for k, v in snap.items():
                    if k in ('gann_last_h1_time', 'gann_cycle_started_at'):
                        continue
                    if k in ss:  # same safety principle as top-level: only restore known fields
                        ss[k] = v
                lh1 = snap.get('gann_last_h1_time')
                ss['gann_last_h1_time'] = pd.Timestamp(lh1).to_pydatetime() if lh1 else None
                csa = snap.get('gann_cycle_started_at')
                ss['gann_cycle_started_at'] = pd.Timestamp(csa).to_pydatetime() if csa else None
        else:
            # Backward-compat with the oldest schema (open trades only).
            for sym, trades in data.get('gann_open_trades', {}).items():
                if sym in bot_state['symbol_state']:
                    bot_state['symbol_state'][sym]['gann_open_trades'] = trades

        c_log("Bot state restored from persistence file (settings, open trades, Gann cycle state, daily PnL).")
    except Exception as e:
        # If the persistence file is corrupt we must not silently pretend
        # we're starting clean while real broker positions may still be
        # open. Flag it; init_metaapi/main will use this to force READ_ONLY.
        log_exception("load_bot_persistence", e)
        c_log(f"CRITICAL: Persistence file exists but failed to load ({e}). "
              f"Bot will start in READ_ONLY to avoid trading blind.")
        bot_state['_persistence_load_failed'] = True

bot_state: dict = {
    'status':           'RUNNING',
    'connection_state': 'RUNNING',
    'connection_state_reason': '',
    'symbol':           'XAUUSD',
    'live_connected':   False,
    'connection_obj':   None,
    'chat_id':          None,
    'last_update_id':   0,
    'is_backtesting':   False,
    'is_live_twin_running': False,
    'timeframes':       _TFS,

    # ── Live execution mode (Instant / Close / Hybrid) ──
    # Default 'instant' == the scanner's existing, unchanged behavior
    # (check live_px against the level margin and fire immediately).
    # Nothing about current live trading changes until this is toggled.
    'gann_execution_mode': 'instant',
    'lt_latency_ms_min': 160,   # measured Railway-deployment -> broker round-trip ping
    'lt_latency_ms_max': 200,   # (update again if a future diagnostic run measures differently)
    'gann_spike_limit_pts': 20,   # hybrid mode: block entry if live_px has moved this many points past the last closed candle's close

    # ── Live-Twin Engine (realistic execution simulator) ──
    # Baseline spread taken from a live MT5/OANDA tick snapshot on
    # 2026-07-13 during the late-night (low-liquidity) session:
    # Bid 4112.28 / Ask 4112.62 -> 0.34 USD (34 points at tick=0.01).
    # This is the QUIET-SESSION floor; session/volatility multipliers
    # scale it up or down from here, they never invent a new baseline.
    'lt_mode': 'realistic',        # 'realistic' or 'idealized' (idealized == old run_gann_backtest, zero friction, kept as A/B baseline)
    'lt_base_spread_usd': 0.34,
    'lt_friction': {
        'spread':    True,   # dynamic session/volatility spread model
        'slippage':  True,   # asymmetric, range-scaled slippage
        'latency':   True,   # 200-800ms signal-to-fill delay
        'commission': True,  # per-lot round-turn commission
        'gaps':      True,   # weekend/rollover gap risk
        'rejection': True,   # requote/rejection probability in volatility spikes
    },
    # Calibrated from the actual XAUUSD broker spec (MT5 symbol properties
    # screenshot): "Commissions: 0-1000 -> 5 USD per lot, Instant by deal
    # volume, in deals" means 5 USD PER DEAL, i.e. per side -- a round-turn
    # trade (open + close) costs 10, not the old flat guess of 7.
    'lt_commission_per_lot': 10.0,  # USD round-turn per 1.0 lot (5 open + 5 close, per real broker spec)
    # Swap was previously a single flat value applied to both directions,
    # which is wrong: the real broker spec shows swap long/short are wildly
    # asymmetric (Swap long: -93.1728, Swap short: +21.6848, in points), and
    # Wednesday carries a 3x multiplier (standard weekend-rollover
    # compensation). Points -> USD/lot conversion uses tick size x contract
    # size from the same spec (0.01 x 100 = $1/point/lot for XAUUSD).
    # Re-derive these two numbers yourself if your broker's spec differs.
    'lt_swap_long_per_lot_night': -93.17,   # USD per lot per night held, BUY positions
    'lt_swap_short_per_lot_night': 21.68,   # USD per lot per night held, SELL positions
    'lt_swap_wednesday_multiplier': 3.0,    # applied when the rollover date is a Wednesday
    'lt_swap_per_lot_night': -6.5,  # legacy fallback only, kept for old configs; no longer used directly
    'lt_rejection_prob': 0.015,     # probability a signal is rejected/requoted during an ATR spike bar

    
    'menu_button_map': {},
    'last_poll_ok':     0.0,
    'live_daily_realized': 0.0,
    'live_daily_date': None,
    'live_daily_hit': False,

    # ── Gann Levels Engine ──
    'active_symbols': {s: (s == 'XAU_USD') for s in AVAILABLE_SYMBOLS},
    'ui_selected_symbol': 'XAU_USD',
    'symbol_state': {s: {
        'gann_levels': [],
        'gann_level_status': {},
        'gann_close_used': None,
        'gann_last_h1_time': None,
        'gann_cycle_active': False,
        'gann_cycle_started_at': None,
        'gann_open_trades': {},
        'auto_trade': False,
        'lot_size': 0.05,
        'gann_cycle_hours': 1,
        'gann_zone_filter': 'star',  
        'gann_entry_mode': 'touch_trend', 
        'trend_filter_type': 'ema',     
        'trend_vwap_period': 100,
        'trend_ema_period': 60,
        'trend_timeframe': '1h',    
        'break_even_enabled': False,
        'gann_be_trigger_points': 40,
        'gann_monitor_tfs': {tf: (tf in ['5m', '10m', '15m', '20m', '30m', '1h', '4m', '6m', '2h', '1m', '2m', '3m']) for tf in _TFS},
        'gann_touch_margin_pts': 5,       
        'gann_tpsl_mode': 'fixed', 
        'gann_tp_points': 70,
        'gann_sl_points': 110,
        'gann_tp_per_tf': {tf: 0 for tf in _TFS},
        'gann_sl_per_tf': {tf: 0 for tf in _TFS},
        'gann_atr_period': 14,
        'gann_atr_sl_mult': 1.5,
        'gann_atr_tp_mult': 2,
    } for s in AVAILABLE_SYMBOLS},
    
# ── Filter Type (star, star_fan, all) ──
    
    
    
    # ── Trend Filters ──
    
    
    
    
    
    # ── Trade Management ──
    
    'prot_daily_dd_usd':      200,
    'prot_daily_profit_usd':  150,
    'prot_true_sync': True,
    'prot_cost_be': True,
    # Max allowed execution deviation (MetaApi "slippage", in broker points)
    # for market orders. If the broker can't fill within this many points of
    # our intended price, the order is rejected (ORDER_FILLING_FOK) instead
    # of being executed 20 pips away.
    'prot_max_slippage_points': 5,
    # Hard cap on simultaneously open trades per symbol, regardless of how
    # many timeframes are enabled or how many levels got touched at once.
    # When reached, remaining timeframes are skipped for that scan cycle
    # only -- already-open trades are left alone (Option A).
    'prot_max_concurrent_trades': 4,
    # Rolling in-memory log of every (symbol, timeframe) scan decision made
    # by the live scanner -- NOT just the instantaneous /diagnose snapshot.
    # Deliberately excluded from persistence (see TOP_LEVEL_EXCLUDE) since
    # it's diagnostic-only and would otherwise bloat the save file.
    'diag_log': [],
    # Full history of every CLOSED live/real trade, rich enough to rebuild an
    # Excel report matching the backtest's own format. Unlike diag_log this
    # IS persisted (not in TOP_LEVEL_EXCLUDE) -- it's the actual trade record,
    # not throwaway diagnostics, and must survive a bot restart.
    'live_trade_history': [],
    'prot_stale_filter': True,
    'prot_cycle_inval': True,
    'prot_cycle_inval_pts': 200,
    'gann_anchor_tf': '1h',
    'prot_allow_multi_tf':    True,

    # ── Broker/display time alignment ──
    # Hours to add to raw UTC (from OANDA/MetaApi) to reach the broker's
    # own server clock (what the user's MT5 terminal displays). Used to
    # align "last closed anchor candle" boundary detection to the SAME
    # wall-clock boundaries MT5 shows, not raw UTC ones. Default 3 =
    # Damascus/EET-DST-style broker offset.
    'broker_time_offset': 3,

    # ── Gann level calculation mode ──
    # 'static_h1'   : classic behavior -- levels are (re)anchored only when
    #                 a new anchor-tf (H1/H4) candle closes.
    # 'dynamic_live': ignore candle closes; recompute levels every
    #                 GANN_DYNAMIC_RECALC_MINUTES using the current live
    #                 streamed price. Toggle in the Telegram protection menu.
    'gann_calculation_mode': 'static_h1',

    
    
    
    
    
    
    
    
    
    
    
    
}


DAM_OFF = timedelta(hours=3)
def _utc_to_dam(dt) -> datetime:
    if isinstance(dt, pd.Timestamp): dt = dt.to_pydatetime()
    if dt.tzinfo is None: dt = dt.replace(tzinfo=timezone.utc)
    return dt + DAM_OFF


# ─────────────────────────────────────────────────────────────
# UNIFIED CORE LOGIC (V9.4)
# ─────────────────────────────────────────────────────────────
def core_eval_break_even(is_buy: bool, entry: float, current_px: float, pip_value: float, be_pts: int, atr_period: int, cost_be: bool) -> float | None:
    be_dist = be_pts * pip_value
    if (is_buy and current_px >= entry + be_dist) or (not is_buy and current_px <= entry - be_dist):
        be_margin = (atr_period * 0.1 * pip_value) if cost_be else 0.0
        return (entry + be_margin) if is_buy else (entry - be_margin)
    return None

def core_eval_outcome(is_buy: bool, current_px: float, tp: float, sl: float) -> str | None:
    if is_buy:
        if current_px >= tp: return 'WIN ✅'
        if current_px <= sl: return 'LOSS ❌'
    else:
        if current_px <= tp: return 'WIN ✅'
        if current_px >= sl: return 'LOSS ❌'
    return None
    
# ─────────────────────────────────────────────────────────────
# ─────────────────────────────────────────────────────────────
# TELEGRAM HELPERS
# ─────────────────────────────────────────────────────────────
async def _tg_post(url: str, **kwargs) -> bool:
    try:
        sess = get_http()
        async with sess.post(url, **kwargs) as resp:
            if resp.status != 200:
                body = await resp.text()
                c_log(f"Telegram API call failed ({resp.status}) for {url}: {body[:300]}")
            return resp.status == 200
    except Exception as e:
        # This carries HALT/READ_ONLY escalation alerts, so a silent
        # failure here means the operator never finds out something broke.
        log_exception(f"_tg_post [{url}]", e)
        return False

def _to_reply_kbd(inline_kbd: dict):
    rows = []; bmap = {}
    for row in inline_kbd.get('inline_keyboard', []):
        new_row = []
        for btn in row:
            text = btn['text']; cb = btn.get('callback_data', 'noop')
            if text in bmap and bmap[text] != cb and cb != 'noop' and bmap[text] != 'noop':
                # This is exactly the bug class that caused the loss/profit
                # buttons to collide: two DIFFERENT actions sharing the same
                # button text, silently overwriting each other in the map
                # that resolves a tapped label back to an action. Every
                # button's text must be unique within a single keyboard.
                c_log(f"BUTTON LABEL COLLISION: '{text}' maps to both '{bmap[text]}' and '{cb}' -- "
                      f"the second silently wins and the first becomes untappable. Fix the keyboard's labels.")
            new_row.append({'text': text}); bmap[text] = cb
        rows.append(new_row)
    return {'keyboard': rows, 'resize_keyboard': True, 'is_persistent': True, 'input_field_placeholder': 'اختر من القائمة...'}, bmap

async def send_tg_msg(text: str, reply_markup: dict = None) -> None:
    if not bot_state['chat_id']: return
    if reply_markup and 'inline_keyboard' in reply_markup:
        reply_markup, bmap = _to_reply_kbd(reply_markup); bot_state['menu_button_map'] = bmap
    payload = {'chat_id': bot_state['chat_id'], 'text': text, 'parse_mode': 'HTML'}
    if reply_markup: payload['reply_markup'] = reply_markup
    await _tg_post(f'https://api.telegram.org/bot{TG_TOKEN}/sendMessage', json=payload)

async def edit_tg_msg(chat_id, message_id, text, reply_markup=None) -> None:
    payload = {'chat_id': chat_id, 'message_id': message_id, 'text': text, 'parse_mode': 'HTML'}
    if reply_markup: payload['reply_markup'] = reply_markup
    await _tg_post(f'https://api.telegram.org/bot{TG_TOKEN}/editMessageText', json=payload)

async def _show(chat_id, msg_id, text: str, reply_markup: dict = None) -> None:
    if msg_id: await edit_tg_msg(chat_id, msg_id, text, reply_markup)
    else: await send_tg_msg(text, reply_markup)

async def answer_callback(cbq_id: str) -> None:
    await _tg_post(f'https://api.telegram.org/bot{TG_TOKEN}/answerCallbackQuery', json={'callback_query_id': cbq_id})

TG_CAPTION_LIMIT = 1024  # Telegram hard limit for document/photo captions

async def send_tg_document(file_path: str, caption: str) -> None:
    if not bot_state['chat_id']: return
    try:
        # A caption over Telegram's limit doesn't get truncated by the API --
        # the whole sendDocument call is rejected, so the file itself would
        # never arrive. Keep the merged single-message intent when it fits;
        # fall back to a short caption + a separate full-text message only
        # when it doesn't.
        doc_caption = caption
        overflow_text = None
        if len(caption) > TG_CAPTION_LIMIT:
            doc_caption = caption[:TG_CAPTION_LIMIT - 20].rstrip() + "\n... (تابع أدناه)"
            overflow_text = caption

        with open(file_path, 'rb') as f:
            data = aiohttp.FormData()
            data.add_field('chat_id',  str(bot_state['chat_id']))
            data.add_field('document', f, filename=os.path.basename(file_path))
            data.add_field('caption',  doc_caption)
            await _tg_post(f'https://api.telegram.org/bot{TG_TOKEN}/sendDocument', data=data)

        if overflow_text:
            await send_tg_msg(overflow_text)
    except Exception as e:
        log_exception(f"send_tg_document [{file_path}]", e)

# ─────────────────────────────────────────────────────────────
# OANDA FETCHER 
# ─────────────────────────────────────────────────────────────
_OANDA_GRAN = {'1m':'M1','2m':'M2','3m':'M3','4m':'M4','5m':'M5','6m':'M6','10m':'M10','15m':'M15','20m':'M20','30m':'M30','1h':'H1','2h':'H2'}
_oanda_sem: asyncio.Semaphore | None = None
def _get_oanda_sem() -> asyncio.Semaphore:
    global _oanda_sem
    if _oanda_sem is None: _oanda_sem = asyncio.Semaphore(3)
    return _oanda_sem

def _safe_float(value, default: float = 0.0) -> float:
    """Closes the `.get(key, default)` null trap: dict.get()'s default only
    applies when the key is MISSING. If MetaAPI returns the key present
    with an explicit `null` (-> None in Python), .get() happily returns
    None, and a later `+=` or arithmetic op on it raises TypeError. This
    coerces None/non-numeric/NaN/inf values to `default` instead of
    raising, at every point real MetaAPI numeric fields are consumed."""
    if value is None:
        return default
    try:
        f = float(value)
    except (TypeError, ValueError):
        return default
    if f != f or f in (float('inf'), float('-inf')):  # NaN / inf guard
        return default
    return f

def _validated_candle(c: dict, symbol: str, granularity_str: str) -> dict | None:
    """Defensive boundary for external market data. OANDA/MetaAPI are not
    contractually guaranteed to always return well-typed floats -- a
    transient glitch can hand back None, a string, or a missing key.
    Returns a clean candle dict, or None if this single candle is bad.
    Never raises: a bad candle should be skipped, not take down the whole
    fetch (or the caller's while-True loop) with it."""
    try:
        mid = c.get('mid')
        if not isinstance(mid, dict):
            raise ValueError(f"missing/invalid 'mid' field: {mid!r}")
        raw_time = c.get('time')
        if not raw_time:
            raise ValueError("missing 'time' field")

        o = float(mid['o']); h = float(mid['h']); l = float(mid['l']); c_ = float(mid['c'])
        vol = float(c.get('volume', 1.0) or 1.0)

        for v in (o, h, l, c_, vol):
            if v != v or v in (float('inf'), float('-inf')):
                raise ValueError(f"non-finite value in candle: {v!r}")

        return {
            'time': pd.Timestamp(raw_time).tz_convert('UTC'),
            'open': o, 'high': h, 'low': l, 'close': c_, 'volume': vol,
        }
    except (TypeError, ValueError, KeyError) as e:
        log_exception(f"_validated_candle [{symbol} {granularity_str}] -- skipping malformed candle", e)
        return None

async def fetch_candles(symbol: str, granularity_str: str, count: int = 5000, end_time: datetime = None) -> list:
    gran_str = _OANDA_GRAN.get(granularity_str, 'M1'); fetch_count = min(count, 120000)  
    collected = []; remaining = fetch_count
    headers = {'Authorization': f'Bearer {OANDA_TOKEN}', 'Content-Type':  'application/json'}
    url = f'{OANDA_BASE_URL}/instruments/{symbol}/candles'
    current_end = end_time if end_time else datetime.now(timezone.utc)

    sem = _get_oanda_sem()
    async with sem:
        while remaining > 0:
            chunk = min(remaining, 5000)
            params = {'granularity': gran_str, 'count': chunk, 'to': current_end.strftime('%Y-%m-%dT%H:%M:%S.000000000Z'), 'price': 'M'}
            candles = []
            for attempt in range(3):
                try:
                    async with get_http().get(url, headers=headers, params=params, timeout=aiohttp.ClientTimeout(total=20)) as resp:
                        if resp.status != 200:
                            if attempt == 2: break
                            await asyncio.sleep(2 ** attempt)
                            continue
                        data = await resp.json(); candles = data.get('candles', []); break
                except Exception as e:
                    log_exception(f"fetch_candles [{symbol} {granularity_str}] attempt {attempt+1}/3", e)
                    await asyncio.sleep(2 ** attempt)

            if not candles: break
            complete = [c for c in candles if c.get('complete', True)]
            if not complete: break

            formatted = []
            for c in complete:
                vc = _validated_candle(c, symbol, granularity_str)
                if vc is not None:
                    formatted.append(vc)

            if not formatted:
                c_log(f"fetch_candles [{symbol} {granularity_str}]: entire chunk failed validation, aborting fetch.")
                break

            collected = formatted + collected; remaining -= len(complete)
            earliest = pd.Timestamp(complete[0]['time']).tz_convert('UTC')
            current_end = earliest.to_pydatetime() - timedelta(seconds=1)
            if len(complete) < chunk: break
            await asyncio.sleep(0.2)
    return collected

async def fetch_master_price(symbol: str) -> float | None:
    """Single Source of Truth for the CURRENT live price.

    Call this exactly ONCE per symbol per scanner cycle, then reuse the
    returned value for every enabled timeframe's touch-distance check.

    Why this exists: a timeframe's own last candle close is NOT "the
    current price" for anything above 1m -- a 30m candle's close can be up
    to ~30 minutes stale. Asking OANDA separately per-timeframe and using
    each tf's own close as "live price" is exactly what caused the same
    instant to read as e.g. 4067 on 1m and 4073 on 30m during a volatile
    spike, and it also multiplies OANDA requests per cycle (contributing
    to "Insufficient data from OANDA" failures under load). Timeframes
    should still be fetched separately for their own historical
    closes/EMAs/ATR -- just never for "what is the price right now".

    count=2 (not 1): OANDA's most recent candle for 'to=now' is very often
    still the in-progress (incomplete) one, and fetch_candles() drops
    incomplete candles entirely. count=1 would then frequently return an
    EMPTY list and report "insufficient data" even though OANDA itself is
    perfectly healthy. count=2 guarantees at least one genuinely
    completed, very recent candle to use.
    """
    mc = await fetch_candles(symbol, '1m', count=2)
    if not mc:
        c_log(f"fetch_master_price [{symbol}]: no 1m data from OANDA this cycle -- "
              f"skipping touch checks for this symbol rather than risk a stale/desynced price.")
        return None
    return float(mc[-1]['close'])

# ─────────────────────────────────────────────────────────────
# GANN LEVELS & FAN ENGINE (⭐ & ⭐🌀)
# ─────────────────────────────────────────────────────────────
GANN_TFC_H1 = 0.02

# تصنيف دقيق للمستويات (مطابق تماماً لتطبيق OTC-Calculator المرجعي):
# star = المستويات الأصلية القوية (AIMP في الأصل: انحصرت في 0.0833, 0.5, 1.0 فقط)
# fan  = "موازي للمروحة" -- في الأصل هذا يعني معامل الزاوية 0.125 (8x1) حصراً،
#        وليس أي معامل آخر. القائمة الأصلية (ACOEF) فيها 11 معامل فقط --
#        لا يوجد 3.0 ولا 8.0 إطلاقاً.
GANN_COEFS = [
    {'c': 0.0208, 'star': False, 'fan': False},
    {'c': 0.0417, 'star': False, 'fan': False},
    {'c': 0.0625, 'star': False, 'fan': False},
    {'c': 0.0833, 'star': True,  'fan': False},
    {'c': 0.125,  'star': False, 'fan': True},  # 8x1 -- الوحيد الذي يُعتبر "موازي للمروحة"
    {'c': 0.25,   'star': False, 'fan': False},
    {'c': 0.333,  'star': False, 'fan': False},
    {'c': 0.5,    'star': True,  'fan': False},
    {'c': 1.0,    'star': True,  'fan': False},
    {'c': 2.0,    'star': False, 'fan': False},
    {'c': 4.0,    'star': False, 'fan': False},
]

def _anchor_hours() -> int:
    """Hour-equivalent of the selected Gann anchor timeframe."""
    return 4 if bot_state.get('gann_anchor_tf', '1h') == '4h' else 1

def _anchor_label() -> str:
    """Display label ('H1'/'H4') matching the selected anchor timeframe --
    used everywhere a message previously hardcoded the literal text 'H1'."""
    return bot_state.get('gann_anchor_tf', '1h').upper()

def gann_calc_levels(symbol: str, close: float) -> list[dict]:
    levels = []
    anchor_tf = bot_state.get('gann_anchor_tf', '1h')
    multiplier = GANN_TFC_H1 * 2.0 if anchor_tf == '4h' else GANN_TFC_H1
    
    for i, item in enumerate(GANN_COEFS):
        offset = close * item['c'] * multiplier
        prec = SYMBOL_INFO[symbol]['prec']
        up = round(close + offset, prec)
        dn = round(close - offset, prec)
        
        # التسميات الدقيقة
        up_lbl = "مقاومة"
        dn_lbl = "دعم"
        if item['star'] and not item['fan']:
            up_lbl = "مقاومة ⭐"
            dn_lbl = "دعم ⭐"
        elif item['star'] and item['fan']:
            up_lbl = "مقاومة ⭐"
            dn_lbl = "دعم ⭐"
        elif item['fan']:
            up_lbl = "مقاومة موازية للمروحة 🌀"
            dn_lbl = "دعم موازي للمروحة 🌀"
            
        levels.append({'key': f'up_{i}', 'price': up, 'dir': 'up', 'star': item['star'], 'fan': item['fan'], 'label': up_lbl})
        if dn > 0:
            levels.append({'key': f'dn_{i}', 'price': dn, 'dir': 'dn', 'star': item['star'], 'fan': item['fan'], 'label': dn_lbl})
            
    levels.append({'key': 'ref', 'price': round(close, SYMBOL_INFO[symbol]['prec']), 'dir': 'ref', 'star': False, 'fan': False, 'label': f'إغلاق {_anchor_label()}'})
    levels.sort(key=lambda x: x['price'], reverse=True)
    return levels

def gann_active_levels(symbol: str) -> list[dict]:
    sym_state = bot_state['symbol_state'][symbol]
    lv = [l for l in bot_state['symbol_state'][symbol]['gann_levels'] if l['dir'] != 'ref']
    f = sym_state['gann_zone_filter']
    if f == 'star': return [l for l in lv if l['star']]
    elif f == 'star_fan': return [l for l in lv if l['star'] or l['fan']]
    return lv

def _gann_tf_tp(symbol: str, tf: str) -> int:
    sym_state = bot_state['symbol_state'][symbol]
    v = sym_state['gann_tp_per_tf'].get(tf, 0)
    return v if v > 0 else sym_state['gann_tp_points']

def _gann_tf_sl(symbol: str, tf: str) -> int:
    sym_state = bot_state['symbol_state'][symbol]
    v = sym_state['gann_sl_per_tf'].get(tf, 0)
    return v if v > 0 else sym_state['gann_sl_points']

def _gann_atr(candles: list, period: int) -> float | None:
    if len(candles) < period + 1: return None
    df = pd.DataFrame(candles[-(period + 50):])
    df['prev_close'] = df['close'].shift(1)
    tr = pd.concat([
        df['high'] - df['low'],
        (df['high'] - df['prev_close']).abs(),
        (df['low']  - df['prev_close']).abs(),
    ], axis=1).max(axis=1)
    val = tr.rolling(period).mean().iloc[-1]
    return float(val) if not pd.isna(val) else None

def _gann_calc_tpsl(symbol: str, entry: float, is_buy: bool, candles: list, tf: str = '') -> tuple[float, float]:
    sym_state = bot_state['symbol_state'][symbol]
    pv = SYMBOL_INFO[symbol]['pip_value']
    prec = SYMBOL_INFO[symbol]['prec']
    if sym_state['gann_tpsl_mode'] == 'atr':
        atr = _gann_atr(candles, sym_state['gann_atr_period'])
        if not atr: atr = _gann_tf_sl(symbol, tf) * pv
        sl_dist = atr * sym_state['gann_atr_sl_mult']
        tp_dist = atr * sym_state['gann_atr_tp_mult']
    else:
        sl_dist = _gann_tf_sl(symbol, tf) * pv
        tp_dist = _gann_tf_tp(symbol, tf) * pv
    if is_buy: return round(entry + tp_dist, prec), round(entry - sl_dist, prec)
    return round(entry - tp_dist, prec), round(entry + sl_dist, prec)

def _last_closed_anchor_time_utc(anchor_hours: int, offset_hours: float, now_utc: datetime) -> datetime:
    """UTC timestamp of the OPEN of the most recently fully-closed anchor-tf
    bucket, computed against BROKER server time (UTC + broker_time_offset)
    rather than raw UTC. For a whole-hour offset this makes no difference
    to a 1h anchor (hourly boundaries land on the same instants in any
    whole-hour-offset zone), but it matters for a 4h anchor: OANDA's raw-UTC
    4h grid (00/04/08/12/16/20 UTC) is NOT the same set of instants as the
    broker's UTC+offset 4h grid, so blindly trusting OANDA's own bucketing
    can hand the bot a Gann anchor close that doesn't match what closes on
    the user's MT5 terminal."""
    broker_now = now_utc + timedelta(hours=offset_hours)
    floored = broker_now.replace(minute=0, second=0, microsecond=0)
    bucket_start_hour = (floored.hour // anchor_hours) * anchor_hours
    bucket_start_broker = floored.replace(hour=bucket_start_hour)
    return bucket_start_broker - timedelta(hours=offset_hours)

async def _gann_fetch_last_closed_anchor(symbol: str) -> dict | None:
    anchor_tf = bot_state.get('gann_anchor_tf', '1h')
    anchor_hours = _anchor_hours()
    offset = bot_state.get('broker_time_offset', 3)
    target_close_utc = _last_closed_anchor_time_utc(anchor_hours, offset, datetime.now(timezone.utc))
    # Fetch a small pad of extra candles (not just 2) since the broker-
    # aligned boundary can fall a bucket or two behind OANDA's own most
    # recent closed candle once the offset shifts the grid.
    candles = await fetch_candles(symbol, anchor_tf, count=anchor_hours + 6)
    if not candles: return None
    candles = sorted(candles, key=lambda c: c['time'])
    # Pick the newest candle whose close does not exceed the broker-aligned
    # boundary -- NOT simply candles[-1], which is only correct when
    # OANDA's raw-UTC grid happens to coincide with the broker's grid.
    eligible = [c for c in candles if c['time'].to_pydatetime() <= target_close_utc]
    return eligible[-1] if eligible else candles[-1]

def _gann_fmt_levels_msg(symbol: str, close: float) -> str:
    sym_state = bot_state['symbol_state'][symbol]
    lines = []
    for l in bot_state['symbol_state'][symbol]['gann_levels']:
        if l['dir'] == 'ref':
            lines.append(f"➖ <b>{l['price']:.2f}</b>  (إغلاق {_anchor_label()})")
            continue
        
        icon = '🔴' if l['dir'] == 'up' else '🟢'
        lines.append(f"{icon} {l['price']:.2f}  {l['label']}")
        
    f_mode = sym_state['gann_zone_filter']
    if f_mode == 'star': filt = '⭐ المستويات الأصلية القوية فقط'
    elif f_mode == 'star_fan': filt = '⭐🌀 القوية + الموازية للمروحة'
    else: filt = '📋 كل المستويات (مخاطرة)'
    
    flt_trend = sym_state['trend_filter_type'].upper()
    if flt_trend == 'BOTH': flt_trend = 'VWAP + EMA'
    
    mode = f'لمس مباشر + فلتر ({flt_trend}_{sym_state["trend_timeframe"].upper()})' if sym_state['gann_entry_mode'] == 'touch_trend' else 'لمس أعمى (بدون فلتر)'
    return (f"📐 <b>سلّم جان (المروحة) — دورة جديدة</b>\n"
            f"إغلاق {_anchor_label()}: <b>{close:.2f}</b>\n\n"
            f"مدة المراقبة: {sym_state['gann_cycle_hours']}س  |  فلتر: {filt}\nالدخول: {mode}\n\n\n"
            + '\n'.join(lines))

_consecutive_real_order_failures = 0
_REAL_ORDER_FAILURE_HALT_THRESHOLD = 3
_last_scanner_error_alert_ts = 0.0

def _resolve_broker_symbol(symbol: str) -> str:
    """Resolve the OANDA-format data-feed symbol (e.g. 'XAU_USD') to the
    broker's actual MT5 symbol name for order execution. bot_state['symbol']
    (settable via /setsymbol) is the primary source of truth since brokers
    vary wildly in suffix conventions (XAUUSD, XAUUSDm, XAUUSD.a, GOLD...).
    A hard safety-net mapping is applied on top: if the configured value is
    missing or still looks like an unfixed raw OANDA-format symbol (i.e.
    still has the underscore), fall back to the confirmed-correct stripped
    form instead of sending a guaranteed-to-be-rejected symbol to the
    broker. The rest of the bot's data engine, fetches, and logs MUST
    continue using the OANDA-format symbol ('XAU_USD') -- this function's
    return value is ONLY for the MetaAPI execution payload."""
    configured = bot_state.get('symbol', '').strip()
    if not configured or '_' in configured:
        return symbol.replace('_', '')
    return configured

async def _gann_open_trade(symbol: str, is_buy: bool, level: dict, candles: list, reason: str, tf: str,
                            initial_px: float = None, detect_time: datetime = None, t1_signal_ts: float = None,
                            feed_source: str = None, feed_age_ms: float = None, trigger_type: str = None) -> None:
    global _consecutive_real_order_failures
    sym_state = bot_state['symbol_state'][symbol]

    # Order-management critical path: never place an order while the
    # connection state machine says we shouldn't be trading, or while
    # we're inside a restricted DAM time window.
    if not is_trading_allowed():
        if bot_state.get('connection_state', CONN_RUNNING) != CONN_RUNNING:
            c_log(f"Skipped entry [{symbol} {tf}]: connection_state={bot_state.get('connection_state')} "
                  f"({bot_state.get('connection_state_reason')})")
        else:
            c_log(f"Skipped entry [{symbol} {tf}]: inside restricted DAM trading window "
                  f"({datetime.now(timezone.utc) + timedelta(hours=3):%H:%M} DAM).")
        return

    try:
        # ── Re-verify price at execution time (Point 3) ──
        # Trades within the same scan cycle open sequentially, one timeframe
        # at a time. During a fast/volatile move, the price can drift far
        # from the level between the first and last order in the same
        # batch (this is exactly how one support touch ended up opening
        # entries $1-9 apart from each other). Re-fetch a fresh price right
        # before this specific order goes out, and re-check it's still
        # actually near the level -- not the stale master_px from when the
        # cycle started.
        #
        # This used to be a second OANDA REST call here, which was itself
        # extra latency added right before execution. Reading live_quotes
        # is an in-memory cache read (no HTTP round-trip at all) as long as
        # the WebSocket feed is fresh; REST is now only a fallback for when
        # it isn't -- so this re-check got both more accurate AND faster,
        # not just faster.
        fresh_px, fresh_feed_source, fresh_feed_age_ms = await _lq_price_with_fallback(symbol)
        margin = sym_state['gann_touch_margin_pts'] * SYMBOL_INFO[symbol]['pip_value']
        if fresh_px is None or abs(fresh_px - level['price']) > margin:
            bot_state['symbol_state'][symbol]['gann_level_status'][level['key']] = 'used'
            # Detail requested: exactly how much the price moved, over how
            # long, and what the pre-existing code threshold is (the touch
            # margin, gann_touch_margin_pts -- unchanged, no new number
            # invented) that this movement exceeded.
            elapsed_s = (datetime.now(timezone.utc) - detect_time).total_seconds() if detect_time else None
            drift = abs(fresh_px - initial_px) if (fresh_px is not None and initial_px is not None) else None
            detail_lines = []
            if drift is not None:
                margin_pts = sym_state['gann_touch_margin_pts']
                detail_lines.append(
                    f"الحركة الفعلية منذ اكتشاف اللمس: {drift:.3f} ({drift / SYMBOL_INFO[symbol]['pip_value']:.1f} نقطة)"
                )
                detail_lines.append(
                    f"الحد المسموح به مسبقاً بالكود (هامش اللمس gann_touch_margin_pts): {margin_pts} نقطة ({margin:.3f})"
                )
            if elapsed_s is not None:
                detail_lines.append(f"الفجوة الزمنية بين الاكتشاف والتنفيذ: {elapsed_s:.1f} ثانية")
            detail_block = ("\n" + "\n".join(detail_lines)) if detail_lines else ""
            await send_tg_msg(
                f"<b>⏭️ [{symbol} - جان {tf}]</b>  {reason}\n"
                f"المستوى: {level['price']:.2f}\n"
                f"تم تجاهل الفريم — السعر ابتعد عن المستوى أثناء التنفيذ "
                f"({'لا يمكن التأكد من السعر الحالي' if fresh_px is None else f'{fresh_px:.2f}'}) ولم يعد لمساً حقيقياً."
                f"{detail_block}"
            )
            return

        price = fresh_px
        tp, sl = _gann_calc_tpsl(symbol, price, is_buy, candles, tf=tf)

        # ── Pre-send sanity check (Point 4) ──
        # If price already moved past where TP or SL would sit before we
        # even send the order, the opportunity is gone (or would produce
        # nonsensical/rejected stops, e.g. "Invalid stops"). Better to skip
        # cleanly here than let the broker reject it after the fact.
        if is_buy and (price >= tp or price <= sl):
            bot_state['symbol_state'][symbol]['gann_level_status'][level['key']] = 'used'
            await send_tg_msg(
                f"<b>⏭️ [{symbol} - جان {tf}]</b>  {reason}\n"
                f"المستوى: {level['price']:.2f}\n"
                f"تم إلغاء الأمر قبل الإرسال — السعر الحالي ({price:.2f}) تجاوز فعلياً "
                f"مستوى TP/SL المحسوب (TP:{tp} SL:{sl})."
            )
            return
        if not is_buy and (price <= tp or price >= sl):
            bot_state['symbol_state'][symbol]['gann_level_status'][level['key']] = 'used'
            await send_tg_msg(
                f"<b>⏭️ [{symbol} - جان {tf}]</b>  {reason}\n"
                f"المستوى: {level['price']:.2f}\n"
                f"تم إلغاء الأمر قبل الإرسال — السعر الحالي ({price:.2f}) تجاوز فعلياً "
                f"مستوى TP/SL المحسوب (TP:{tp} SL:{sl})."
            )
            return

        lot = sym_state['lot_size']
        tp_pts = _gann_tf_tp(symbol, tf); sl_pts = _gann_tf_sl(symbol, tf)

        tpsl_lbl = (f"ATR({sym_state['gann_atr_period']})×{sym_state['gann_atr_sl_mult']}/{sym_state['gann_atr_tp_mult']}\n"
                    if sym_state['gann_tpsl_mode'] == 'atr' else f"SL:{sl_pts}p TP:{tp_pts}p")

        be_lbl = " | 🛡️ BE Active" if sym_state['break_even_enabled'] else ""

        is_real = sym_state.get('auto_trade', False)
        trade_id = f"sim_{int(datetime.now().timestamp())}_{tf}"
        real_msg = ""
        execution_failed = False
        # Real fill price (as opposed to `price`, our pre-check estimate) --
        # only ever populated for actual real orders below; stays None for
        # simulated/paper trades, where `price` IS the fill by definition.
        real_fill_price = None
        fill_price_source = 'simulated'

        if is_real:
            # Source of truth: never spin up a second, ad-hoc MetaAPI
            # connection here. If the one persistent connection created at
            # startup isn't healthy, we do not know the true account state
            # well enough to safely fire a real order.
            if _metaapi_conn is None:
                real_msg = "\n⚠️ لا يوجد اتصال MetaAPI صالح — لم يتم فتح أي صفقة."
                is_real = False
                execution_failed = True
            else:
                t2_pre_send_ts = None
                try:
                    broker_symbol = _resolve_broker_symbol(symbol)

                    # ── Slippage / Deviation control ──
                    # Cap how far from our intended price the broker is allowed
                    # to fill us. 'slippage' is in MetaApi "points" (broker tick
                    # size units, i.e. the same unit as symbol digits — for a
                    # 2-digit gold quote that's $0.01/point). ORDER_FILLING_FOK
                    # ("fill or kill") means the whole order is rejected rather
                    # than partially/badly filled if the broker can't honor it
                    # within that deviation — this is what stops us from being
                    # filled 20 pips away from our entry.
                    max_slippage_points = int(bot_state.get('prot_max_slippage_points', 5))
                    order_options = {
                        'slippage': max_slippage_points,
                        'fillingModes': ['ORDER_FILLING_FOK'],
                    }

                    # ── Latency telemetry (T1 signal -> T2 pre-send -> T3 broker ack) ──
                    t2_pre_send_ts = time.monotonic()
                    if is_buy:
                        res = await _metaapi_conn.create_market_buy_order(
                            broker_symbol, lot, stop_loss=sl, take_profit=tp, options=order_options
                        )
                    else:
                        res = await _metaapi_conn.create_market_sell_order(
                            broker_symbol, lot, stop_loss=sl, take_profit=tp, options=order_options
                        )
                    t3_ack_ts = time.monotonic()

                    code_delay_ms = round((t2_pre_send_ts - t1_signal_ts) * 1000) if t1_signal_ts else None
                    ping_ms = round((t3_ack_ts - t2_pre_send_ts) * 1000)
                    # "Quote Age at Fire" is the real-world number that
                    # actually matters for slippage: how stale was the price
                    # this order was based on, at the instant it fired --
                    # not how fast Python parsed a variable (that's Code
                    # Delay below, kept for completeness but not the headline).
                    feed_label = 'WS (MetaApi live)' if fresh_feed_source == 'ws' else 'OANDA REST (fallback — feed was stale)'
                    age_str = f"{fresh_feed_age_ms}ms" if fresh_feed_age_ms is not None else 'n/a (REST fallback has no push-age)'
                    telemetry_lbl = (
                        f"\n📡 Feed: {feed_label} | Quote Age at Fire: {age_str}"
                        + (
                            f"\n⏱ Code Delay (T2-T1): {code_delay_ms}ms | MetaApi Ping (T3-T2): {ping_ms}ms"
                            if code_delay_ms is not None else
                            f"\n⏱ MetaApi Ping (T3-T2): {ping_ms}ms (T1 unavailable)"
                        )
                    )

                    # CRITICAL: history deals/reconciliation are keyed by
                    # positionId, NOT orderId (they're different tickets in
                    # MT5 — orderId is the pending/market order ticket,
                    # positionId is the resulting open position's ticket).
                    # Previously this preferred orderId, so the reconciliation
                    # lookup below (which matches on positionId) never found
                    # a match, silently fell back to a theoretical/estimated
                    # PnL every time, and mislabeled it as the real MT5 profit.
                    # positionId must be preferred here.
                    trade_id = str(res.get('positionId', res.get('orderId', trade_id)))

                    # ── Real fill price, not our pre-check estimate ──
                    # `price` (fresh_px) is what we THOUGHT we'd get filled at,
                    # checked right before sending. The broker's actual fill
                    # can differ (that's the whole slippage question this was
                    # built to answer). res.get('price') is sometimes the
                    # order's requested price, not a confirmed fill -- the
                    # realized fill price lives on the resulting POSITION, so
                    # query that directly. MetaApi's position sync can lag the
                    # order response by a moment, so retry briefly before
                    # falling back.
                    for delay in (0, 1, 2):
                        if delay: await asyncio.sleep(delay)
                        try:
                            positions = _metaapi_conn.terminal_state.positions
                            match = next((p for p in positions if str(p.get('id')) == trade_id), None)
                            if match and match.get('openPrice') is not None:
                                real_fill_price = float(match['openPrice'])
                                fill_price_source = 'confirmed_position'
                                break
                        except Exception as pe:
                            log_exception(f"_gann_open_trade fill-price lookup [{symbol} {tf}]", pe)
                    if real_fill_price is None and res.get('price') is not None:
                        real_fill_price = float(res['price'])
                        fill_price_source = 'order_response'

                    real_msg = "\n🚀 <b>تم فتح الصفقة حقيقياً على حسابك!</b>" + telemetry_lbl
                    _consecutive_real_order_failures = 0
                except Exception as ex:
                    log_exception(f"_gann_open_trade real order [{symbol} {tf}]", ex)
                    err_str = str(ex)
                    t_fail_ts = time.monotonic()
                    ref_t2 = t2_pre_send_ts if t2_pre_send_ts is not None else t_fail_ts
                    code_delay_ms = round((ref_t2 - t1_signal_ts) * 1000) if t1_signal_ts else None
                    fail_after_ms = round((t_fail_ts - ref_t2) * 1000)
                    fail_telemetry_lbl = (
                        f"\n⏱ Code Delay (T2-T1): {code_delay_ms}ms | Failed after send: {fail_after_ms}ms"
                        if code_delay_ms is not None else
                        f"\n⏱ Failed after send: {fail_after_ms}ms (T1 unavailable)"
                    )
                    # Give an explicit signal when the rejection was caused by
                    # the deviation guard itself (requote / price moved beyond
                    # our slippage tolerance), rather than a generic failure,
                    # so it's obvious this is protective behavior, not a bug.
                    if any(code in err_str for code in ('REQUOTE', 'PRICE_CHANGED', 'OFF_QUOTES')):
                        real_msg = (f"\n🛑 <b>تم رفض الصفقة لتجاوز حد الانزلاق السعري "
                                    f"({max_slippage_points} نقاط):</b> {ex}\nلم يتم التنفيذ لحمايتك من دخول سيء."
                                    f"{fail_telemetry_lbl}")
                    else:
                        real_msg = (f"\n❌ <b>فشل فتح الصفقة حقيقياً:</b> {ex}\nلم يتم تتبعها كصفقة وهمية "
                                    f"(لا يوجد تنفيذ فعلي).{fail_telemetry_lbl}")
                    is_real = False
                    execution_failed = True
                    _consecutive_real_order_failures += 1
                    if _consecutive_real_order_failures >= _REAL_ORDER_FAILURE_HALT_THRESHOLD:
                        await set_connection_state(
                            CONN_HALTED,
                            f"{_consecutive_real_order_failures} consecutive real order failures "
                            f"(last: {ex}). Escalating to protect capital."
                        )

        # Ghost-trade fix: a FAILED real-order attempt must never enter
        # gann_open_trades -- it never had any exposure on the broker, so
        # tracking it (even as a "simulated" fallback) meant the bot would
        # later evaluate it against live price movement and report a
        # fabricated WIN/LOSS for a trade that never existed. Genuine
        # paper-trading (auto_trade was never enabled to begin with) is a
        # completely different, intentional case and is still tracked.
        if execution_failed:
            bot_state['symbol_state'][symbol]['gann_level_status'][level['key']] = 'used'
            await send_tg_msg(
                f"<b>⏭️ [{symbol} - جان {tf}]</b>  {reason}\n"
                f"المستوى: {level['price']:.2f}\n"
                f"{real_msg}"
            )
            return

        # entry_final is what actually happened: the confirmed broker fill
        # when we have one, our pre-check estimate otherwise (simulated
        # trades, or a real trade where the position lookup above failed).
        entry_final = real_fill_price if real_fill_price is not None else price

        bot_state['symbol_state'][symbol]['gann_open_trades'][trade_id] = {
            'tf': tf, 'is_buy': is_buy, 'entry': entry_final, 'is_real': is_real, 'sl': sl, 'tp': tp,
            'be_trigger': (entry_final + (tp - entry_final)/2) if is_buy else (entry_final - (entry_final - tp)/2), # simplified BE trigger
            'opened_at': datetime.now(timezone.utc).isoformat(), 'level_price': level['price'],
            'feed_source': feed_source, 'feed_age_ms': feed_age_ms, 'trigger_type': trigger_type,
        }
        bot_state['symbol_state'][symbol]['gann_level_status'][level['key']] = 'used'
        await save_bot_persistence()

        entry_note = {
            'confirmed_position': ' (مؤكد من الوسيط)',
            'order_response': ' (من استجابة الأمر)',
            'simulated': '',
        }.get(fill_price_source, ' (تقديري قبل التنفيذ — تعذّر تأكيد سعر الوسيط)')
        slippage_line = ""
        if is_real:
            actual_slippage = abs(entry_final - level['price'])
            pv = SYMBOL_INFO[symbol]['pip_value']
            slippage_line = f"الانزلاق الفعلي عن المستوى: {actual_slippage:.2f} ({actual_slippage / pv:.1f} نقطة)\n"

        await send_tg_msg(
            f"<b>✅ {reason}</b>\n\n"
            f"المستوى: {level['price']:.2f}  |  الدخول: {entry_final:.2f}{entry_note}\n\n"
            f"TP: {tp}  SL: {sl}  |  {tpsl_lbl}{be_lbl}\n"
            f"{slippage_line}"
            f"إغلاق {_anchor_label()}: {bot_state['symbol_state'][symbol]['gann_close_used']:.5f}\n"
            f"{real_msg}"
        )
    except Exception as e:
        log_exception(f"_gann_open_trade [{symbol} {tf}]", e)
        bot_state['symbol_state'][symbol]['gann_level_status'][level['key']] = 'used'
        await send_tg_msg(f"<b>❌ فشل تنفيذ الصفقة [{symbol} - جان {tf}]</b>\nالمستوى: {level['price']:.5f}\n{e}")

# ─────────────────────────────────────────────────────────────
# BACKTEST PROGRESS TRACKER
# ─────────────────────────────────────────────────────────────
class BtProgress:
    BAR_LEN = 14; HEARTBEAT = 15
    def __init__(self, label: str, active_tfs: list):
        self.label = label; self.active_tfs = active_tfs; self.cancelled = False; self.phase = 'Initialising...'
        self.tf_done = 0; self.tf_total = len(active_tfs); self.current_tf = ''
        self.bars_done = 0; self.bars_total = 0; self.win = 0; self.loss = 0; self.be = 0; self.profit = 0.0
        self.chat_id = None; self.msg_id = None; self._last_edit = 0.0; self._lock = asyncio.Lock(); self._hb_task = None; self._start_ts = 0.0

    def _bar(self, done: int, total: int) -> str:
        if total == 0: return chr(9617) * self.BAR_LEN
        filled = round(done / total * self.BAR_LEN)
        return chr(9608) * filled + chr(9617) * (self.BAR_LEN - filled)

    def _elapsed(self) -> str:
        secs = int(datetime.now(timezone.utc).timestamp() - self._start_ts); m, s = divmod(secs, 60); return f'{m}m {s:02d}s'

    def _build_text(self) -> str:
        total = self.win + self.loss; wr = f'{round(self.win / total * 100)}%' if total else '-'
        pnl = f'+${round(self.profit,2)}' if self.profit >= 0 else f'-${abs(round(self.profit,2))}'; icon = '▲' if self.profit >= 0 else '▼'
        overall = (self.tf_done + self.bars_done / self.bars_total) / max(self.tf_total, 1) if self.bars_total else self.tf_done / max(self.tf_total, 1)
        ov_bar = self._bar(round(overall * 100), 100); ov_pct = f'{round(overall * 100)}%'
        tf_bar = self._bar(self.bars_done, self.bars_total) if self.bars_total else chr(9617) * self.BAR_LEN
        tf_pct = f'{round(self.bars_done / self.bars_total * 100)}%' if self.bars_total else '-'
        lines = [f'Backtest — <b>{self.label}</b>', f'<b>Phase:</b> {self.phase}', '', f'<b>Overall</b>  {ov_pct}', f'<code>[{ov_bar}]</code>']
        if self.current_tf: lines += ['', f'<b>TF:</b> {self.current_tf}  ({self.tf_done}/{self.tf_total})', f'<code>[{tf_bar}] {tf_pct}</code>', f'Bars: {self.bars_done}/{self.bars_total}']
        lines += ['', f'W:{self.win}  L:{self.loss}  BE:{self.be}', f'{icon} {pnl}  WR:{wr}', '', f'Elapsed: {self._elapsed()}']
        if self.cancelled: lines.append('<b>CANCELLED</b>')
        return '\n'.join(lines)

    async def start(self, chat_id: int) -> None:
        self.chat_id = chat_id; self._start_ts = datetime.now(timezone.utc).timestamp(); self._last_edit = self._start_ts
        payload = {'chat_id': chat_id, 'text': self._build_text(), 'parse_mode': 'HTML', 'reply_markup': {'inline_keyboard': [[{'text': '⏹ Cancel', 'callback_data': 'cancel_bt'}]]}}
        try:
            async with aiohttp.ClientSession() as sess:
                async with sess.post(f'https://api.telegram.org/bot{TG_TOKEN}/sendMessage', json=payload) as resp:
                    if resp.status == 200: self.msg_id = (await resp.json())['result']['message_id']
        except Exception: pass
        self._hb_task = asyncio.create_task(self._heartbeat())

    async def _heartbeat(self) -> None:
        while not self.cancelled: await asyncio.sleep(self.HEARTBEAT); await self._edit(force=True)

    async def _edit(self, force: bool = False) -> None:
        now = datetime.now(timezone.utc).timestamp()
        if not force and (now - self._last_edit) < 3: return
        if not self.msg_id or not self.chat_id: return
        async with self._lock:
            self._last_edit = now; payload = {'chat_id': self.chat_id, 'message_id': self.msg_id, 'text': self._build_text(), 'parse_mode': 'HTML'}
            if not self.cancelled: payload['reply_markup'] = {'inline_keyboard': [[{'text': '⏹ Cancel', 'callback_data': 'cancel_bt'}]]}
            try: await _tg_post(f'https://api.telegram.org/bot{TG_TOKEN}/editMessageText', json=payload)
            except Exception: pass

    async def set_phase(self, phase: str) -> None: self.phase = phase; await self._edit()
    async def set_tf(self, tf: str, bars_total: int) -> None: self.current_tf = tf; self.bars_done = 0; self.bars_total = bars_total; await self._edit(force=True)
    async def tick(self, bar_n: int, win: int, loss: int, be: int, profit: float) -> None: self.bars_done = bar_n; self.win = win; self.loss = loss; self.be = be; self.profit = profit; await self._edit()
    async def done(self, final_text: str) -> None:
        if self._hb_task: self._hb_task.cancel()
        if not self.msg_id or not self.chat_id: return
        try: await edit_tg_msg(self.chat_id, self.msg_id, final_text)
        except Exception: pass
    async def cancel(self) -> None:
        self.cancelled = True; self.phase = 'Cancelling...'
        if self._hb_task: self._hb_task.cancel()
        await self._edit(force=True)

_bt_progress: BtProgress | None = None
_lt_progress: BtProgress | None = None

# ─────────────────────────────────────────────────────────────
# KEYBOARDS 
# ─────────────────────────────────────────────────────────────
def get_main_keyboard() -> dict:
    return {'inline_keyboard': [
        [{'text': '🔌 فحص حالة حساب MetaAPI', 'callback_data': 'check_metaapi_status'}],
        [{'text': '🩺 تشخيص: ليه مفيش صفقات؟', 'callback_data': 'run_diag'}],
        [{'text': '📊 تصدير سجل تشخيص تفصيلي (Excel)', 'callback_data': 'export_diag_excel'}],
        [{'text': '📒 تصدير سجل الصفقات الحية (Excel)', 'callback_data': 'export_live_trades_excel'}],
        [{'text': '🔓 استئناف يدوي بعد HALT (بعد التأكد من الحساب)', 'callback_data': 'manual_resume_step1'}],
        [{'text': '📐 محرك جان (الاستراتيجية)', 'callback_data': 'menu_gann'}],
        [{'text': '🛡️ إعدادات الحماية', 'callback_data': 'menu_protection'}],
        [{'text': '💾 إدارة الإعدادات (Presets)', 'callback_data': 'menu_presets'}],
        [{'text': '📊 بدء الباكتيست', 'callback_data': 'menu_gann_bt'}],
        [{'text': '🧪 Live-Twin Simulator (تنفيذ واقعي)', 'callback_data': 'menu_lt'}],
    ]}


def get_protection_keyboard() -> dict:
    dd = bot_state['prot_daily_dd_usd']
    profit = bot_state['prot_daily_profit_usd']
    multi_tf = '✅ مسموح' if bot_state.get('prot_allow_multi_tf', True) else '❌ ممنوع'
    
    rows = [
        [{'text': '── الحدود اليومية ──', 'callback_data': 'noop'}],
        [{'text': f'📉 أقصى تراجع يومي: ${dd}', 'callback_data': 'noop'}],
        [
            {'text': '➖ خسارة $50', 'callback_data': 'prot_dec_dd'},
            {'text': '➕ خسارة $50', 'callback_data': 'prot_inc_dd'}
        ],
        [{'text': f'💰 هدف الربح اليومي: ${profit}', 'callback_data': 'noop'}],
        [
            {'text': '➖ ربح $50', 'callback_data': 'prot_dec_profit'},
            {'text': '➕ ربح $50', 'callback_data': 'prot_inc_profit'}
        ],
        [{'text': '── الحماية المتقدمة (v9.0) ──', 'callback_data': 'noop'}],
        [{'text': f"مزامنة MT4 (Reconciliation): {'✅' if bot_state.get('prot_true_sync', True) else '🔴'}", 'callback_data': 'tg_prot_sync'}],
        [{'text': f"إلغاء الدورة وقت الانفجار: {'✅' if bot_state.get('prot_cycle_inval', True) else '🔴'}", 'callback_data': 'tg_prot_inval'}],
        [{'text': f"BE شامل التكلفة (True Cost): {'✅' if bot_state.get('prot_cost_be', True) else '🔴'}", 'callback_data': 'tg_prot_cost'}],
        [{'text': f"فلتر البيانات المتأخرة: {'✅' if bot_state.get('prot_stale_filter', True) else '🔴'}", 'callback_data': 'tg_prot_stale'}],
        [{'text': f"إطار مرجعي للجان (Anchor): {bot_state.get('gann_anchor_tf', '1h').upper()}", 'callback_data': 'tg_prot_anchor'}],
        [{'text': f"فلتر أوقات دمشق (07-09 | 13-14): {'✅' if bot_state.get('prot_dam_time_filter', True) else '🔴'}", 'callback_data': 'tg_prot_dam_time'}],
        [{'text': f"حساب جان: {'⚡ حي (كل 5 دقائق)' if bot_state.get('gann_calculation_mode', 'static_h1') == 'dynamic_live' else '📌 كلاسيكي (H1/H4)'}", 'callback_data': 'tg_gann_calc_mode'}],
        [{'text': f'تكرار الصفقات (Multi-TF): {multi_tf}', 'callback_data': 'prot_toggle_multitf'}],
        [{'text': '── ── ──', 'callback_data': 'noop'}],
        [{'text': '🔄 تصفير كل الحمايات النشطة الآن', 'callback_data': 'prot_reset_all'}],
        [{'text': '🔙 رجوع للقائمة الرئيسية', 'callback_data': 'menu_main'}],
        [{'text': '🔙 رجوع لإعدادات جان', 'callback_data': 'menu_gann'}]
    ]
    return {'inline_keyboard': rows}

def get_gann_keyboard() -> dict:
    sym = bot_state['ui_selected_symbol']
    sym_state = bot_state['symbol_state'][sym]
    zf   = sym_state['gann_zone_filter']
    em   = sym_state['gann_entry_mode']
    mg   = sym_state['gann_touch_margin_pts']
    tpsm = sym_state['gann_tpsl_mode']
    hrs  = sym_state['gann_cycle_hours']
    cyc  = '🟢 نشطة' if sym_state['gann_cycle_active'] else '⚫ غير نشطة'
    open_n = len(bot_state['symbol_state'][sym]['gann_open_trades'])
    
    flt_type = sym_state['trend_filter_type']
    
    # تحديث التسميات لزر الفلتر
    if zf == 'star': zf_lbl = '⭐ المستويات الأصلية القوية فقط'
    elif zf == 'star_fan': zf_lbl = '⭐🌀 القوية + موازية للمروحة'
    else: zf_lbl = '📋 كل المستويات (للتجارب)'
    
    if flt_type == 'ema':
        filt_btn_lbl = "📉 الفلتر المعتمد: (EMA الشامل)"
        flt_name = 'EMA'
    else:
        filt_btn_lbl = "🌊 الفلتر المعتمد: (VWAP الشامل)"
        flt_name = 'VWAP'
        
    ttf_lbl = sym_state['trend_timeframe'].upper()
    em_lbl  = f'⚡ لمس + فلتر ({flt_name}_{ttf_lbl})' if em == 'touch_trend' else '⚡ لمس أعمى (بدون فلتر)'
    tps_lbl = f'🎯 TP/SL: {"نقاط ثابتة" if tpsm == "fixed" else "حسب ATR"}'

    tp = sym_state['gann_tp_points']; sl = sym_state['gann_sl_points']
    atp = sym_state['gann_atr_tp_mult']; asp = sym_state['gann_atr_sl_mult']
    ap  = sym_state['gann_atr_period']
    be_lbl = "🟢 مفعل" if sym_state['break_even_enabled'] else "⚫ معطل"
    
    auto_t = '🟢 مفعل' if sym_state.get('auto_trade', False) else '🔴 معطل'
    
    rows = [
        [{'text': f'🤖 التداول الآلي (MetaAPI): {auto_t}', 'callback_data': 'gann_toggle_auto_trade'}],
        [{'text': '🛡️ إعدادات الحماية المتقدمة', 'callback_data': 'menu_protection'}],
        [{'text': f'📐 {sym} — دورة: {cyc}  |  صفقات: {open_n}', 'callback_data': 'noop'}],
        [{'text': '🔄 عرض الدعوم والمقاومات الحالية', 'callback_data': 'gann_show_levels'}],
        [{'text': '🕯️ تشخيص: آخر 10 شموع (وقت + إغلاق)', 'callback_data': 'gann_show_last10'}],
    ]
    
    rows.append([{'text': '── أزواج التداول والباكتيست ──', 'callback_data': 'noop'}])
    pair_row = []
    for p in AVAILABLE_SYMBOLS:
        icon = '✅' if bot_state['active_symbols'][p] else '⬜'
        pair_row.append({'text': f'{icon} {p}', 'callback_data': f'gann_toggle_pair_{p}'})
        if len(pair_row) == 2:
            rows.append(pair_row)
            pair_row = []
    if pair_row: rows.append(pair_row)

    rows.append([{'text': '── تخصيص إعدادات الزوج ──', 'callback_data': 'noop'}])
    sel_row = []
    for p in AVAILABLE_SYMBOLS:
        sel = '📌 ' if p == sym else ''
        sel_row.append({'text': f'{sel}{p}', 'callback_data': f'gann_sel_pair_{p}'})
        if len(sel_row) == 2:
            rows.append(sel_row)
            sel_row = []
    if sel_row: rows.append(sel_row)
    
    exec_mode = bot_state.get('gann_execution_mode', 'instant')
    exec_lbl = {
        'instant': '⚡ دخول لمس مباشر (Instant)',
        'close':   '⏳ انتظار إغلاق الشمعة (Close)',
        'hybrid':  '🛡️ مباشر هجين (Hybrid Spike-Limit)',
        'all_concurrent': '🔀 الثلاثة معاً (All-Concurrent — اختبار مقارن)',
    }.get(exec_mode, '⚡ دخول لمس مباشر (Instant)')

    rows += [
        [{'text': '── الاستراتيجية والفلتر ──', 'callback_data': 'noop'}],
        [{'text': f'الاستراتيجية: {em_lbl}', 'callback_data': 'gann_toggle_entry'}],
        [{'text': f'وضع التنفيذ: {exec_lbl}', 'callback_data': 'gann_toggle_exec_mode'}],
        [{'text': f'فلتر الدخول: {zf_lbl}', 'callback_data': 'gann_toggle_filter'}],
        [{'text': filt_btn_lbl, 'callback_data': 'gann_toggle_filter_type'}],
        [{'text': f'⏱️ فريم الترند: {ttf_lbl}', 'callback_data': 'gann_toggle_ttf'}],
        [{'text': f'🛡️ صمام الأمان (Break-Even): {be_lbl}', 'callback_data': 'gann_toggle_be'}],
    ]
    if sym_state.get('break_even_enabled', False):
        be_pts = sym_state.get('gann_be_trigger_points', 40)
        rows.append([
            {'text': 'BE −10p', 'callback_data': 'gann_dec_be_pts'}, 
            {'text': f'تفعيل بعد: {be_pts}p', 'callback_data': 'noop'}, 
            {'text': 'BE +10p', 'callback_data': 'gann_inc_be_pts'}
        ])
    
    if flt_type == 'vwap':
        vwap_val = sym_state['trend_vwap_period']
        rows.append([{'text': 'VWAP −10', 'callback_data': 'gann_dec_vwap'}, 
                     {'text': f'قيمة {ttf_lbl} VWAP: {vwap_val}', 'callback_data': 'noop'}, 
                     {'text': 'VWAP +10', 'callback_data': 'gann_inc_vwap'}])
                     
    if flt_type == 'ema':
        ema_val = sym_state['trend_ema_period']
        rows.append([{'text': 'EMA −10', 'callback_data': 'gann_dec_ema'}, 
                     {'text': f'قيمة {ttf_lbl} EMA: {ema_val}', 'callback_data': 'noop'}, 
                     {'text': 'EMA +10', 'callback_data': 'gann_inc_ema'}])
        
    rows += [
        [{'text': '📝 مساعدة: تغيير القيم الخاصة بالأوامر', 'callback_data': 'gann_filter_help'}],
        [{'text': '── فريمات التنفيذ ──', 'callback_data': 'noop'}],
    ]
    
    tf_items = list(sym_state['gann_monitor_tfs'].items())
    for i in range(0, len(tf_items), 4):
        rows.append([{'text': ('✅' if on else '⬜') + f' {tfk}', 'callback_data': f'gann_tf_{tfk}'} for tfk, on in tf_items[i:i+4]])
        
    rows += [
        [{'text': '── إعدادات عامة ──', 'callback_data': 'noop'}],
        [{'text': '−ساعة', 'callback_data': 'gann_dec_hours'}, {'text': f'مدة تجميد السلّم: {hrs} ساعة', 'callback_data': 'noop'}, {'text': '+ساعة', 'callback_data': 'gann_inc_hours'}],
        [{'text': 'Lot −0.01', 'callback_data': 'gann_dec_lot'}, {'text': f'حجم اللوت: {sym_state["lot_size"]}', 'callback_data': 'noop'}, {'text': 'Lot +0.01', 'callback_data': 'gann_inc_lot'}],
        [{'text': 'Margin −1', 'callback_data': 'gann_dec_margin'}, {'text': f'هامش اللمس {mg}p', 'callback_data': 'noop'}, {'text': 'Margin +1', 'callback_data': 'gann_inc_margin'}],
        [{'text': '── TP / SL ──', 'callback_data': 'noop'}],
        [{'text': tps_lbl, 'callback_data': 'gann_toggle_tpsl'}],
    ]

    if tpsm == 'fixed':
        rows += [
            [{'text': 'TP  −10', 'callback_data': 'gann_dec_tp10'}, {'text': f'TP={tp}p', 'callback_data': 'noop'}, {'text': 'TP  +10', 'callback_data': 'gann_inc_tp10'}],
            [{'text': 'SL  −10', 'callback_data': 'gann_dec_sl10'}, {'text': f'SL={sl}p', 'callback_data': 'noop'}, {'text': 'SL  +10', 'callback_data': 'gann_inc_sl10'}],
        ]
    else:
        rows += [
            [{'text': 'ATR Period −', 'callback_data': 'gann_dec_atrp'}, {'text': f'Period={ap}', 'callback_data': 'noop'}, {'text': 'ATR Period +', 'callback_data': 'gann_inc_atrp'}],
            [{'text': 'SL mult −0.5', 'callback_data': 'gann_dec_atrsl'}, {'text': f'SL×{asp}', 'callback_data': 'noop'}, {'text': 'SL mult +0.5', 'callback_data': 'gann_inc_atrsl'}],
            [{'text': 'TP mult −0.5', 'callback_data': 'gann_dec_atrtp'}, {'text': f'TP×{atp}', 'callback_data': 'noop'}, {'text': 'TP mult +0.5', 'callback_data': 'gann_inc_atrtp'}],
        ]

    rows += [
        [{'text': '⚙️ TP/SL مخصص لكل فريم', 'callback_data': 'gann_tpsl_tf'}],
        [{'text': '📊 بدء الباكتيست', 'callback_data': 'menu_gann_bt'}],
        [{'text': '← رجوع', 'callback_data': 'menu_main'}],
    ]
    return {'inline_keyboard': rows}

def get_gann_tpsl_tf_keyboard(sel_tf: str = '') -> dict:
    sym_state = bot_state['symbol_state'][bot_state['ui_selected_symbol']]
    rows = [[{'text': '⚙️ TP/SL مخصص لكل فريم', 'callback_data': 'noop'}],
            [{'text': '(0 = يرجع للقيمة العامة)', 'callback_data': 'noop'}]]
    tfs_list = list(sym_state['gann_monitor_tfs'].keys())
    tf_row = []
    for tfk in tfs_list:
        icon = '👉' if tfk == sel_tf else ''
        tf_row.append({'text': f'{icon}{tfk}', 'callback_data': f'gann_tptf_sel_{tfk}'})
        if len(tf_row) == 4: rows.append(tf_row); tf_row = []
    if tf_row: rows.append(tf_row)
    if sel_tf:
        tp_v = sym_state['gann_tp_per_tf'].get(sel_tf, 0); sl_v = sym_state['gann_sl_per_tf'].get(sel_tf, 0)
        eff_tp = tp_v if tp_v > 0 else sym_state['gann_tp_points']
        eff_sl = sl_v if sl_v > 0 else sym_state['gann_sl_points']
        rows += [
            [{'text': f'── [{sel_tf}] ──', 'callback_data': 'noop'}],
            [{'text': f'TP فعلي: {eff_tp}p {"(مخصص)" if tp_v>0 else "(عام)"}', 'callback_data': 'noop'}],
            [{'text': 'TP −10', 'callback_data': f'gann_tptf_dtp_{sel_tf}'}, {'text': f'TP={tp_v}', 'callback_data': 'noop'}, {'text': 'TP +10', 'callback_data': f'gann_tptf_itp_{sel_tf}'}],
            [{'text': f'SL فعلي: {eff_sl}p {"(مخصص)" if sl_v>0 else "(عام)"}', 'callback_data': 'noop'}],
            [{'text': 'SL −10', 'callback_data': f'gann_tptf_dsl_{sel_tf}'}, {'text': f'SL={sl_v}', 'callback_data': 'noop'}, {'text': 'SL +10', 'callback_data': f'gann_tptf_isl_{sel_tf}'}],
            [{'text': '↺ إعادة ضبط', 'callback_data': f'gann_tptf_rst_{sel_tf}'}],
        ]
    rows.append([{'text': '← رجوع', 'callback_data': 'menu_gann'}])
    return {'inline_keyboard': rows}

def get_gann_bt_keyboard() -> dict:
    if bot_state['is_backtesting']:
        return {'inline_keyboard': [[{'text': '⏳ الباكتيست يعمل...', 'callback_data': 'noop'}], [{'text': '⏹ إلغاء', 'callback_data': 'cancel_bt'}]]}
    return {'inline_keyboard': [
        [{'text': 'يوم واحد', 'callback_data': 'gbt_1'}, {'text': 'يومين', 'callback_data': 'gbt_2'}],
        [{'text': 'ثلاثة أيام', 'callback_data': 'gbt_3'}, {'text': 'أسبوع', 'callback_data': 'gbt_7'}],
        [{'text': 'شهر كامل', 'callback_data': 'gbt_30'}],
        [{'text': 'أو أرسل: /backtest YYYY-MM-DD', 'callback_data': 'noop'}],
        [{'text': '← رجوع', 'callback_data': 'menu_gann'}],
    ]}

# ─────────────────────────────────────────────────────────────
# LIVE SCANNER (VWAP / EMA / BOTH Macro)
# ─────────────────────────────────────────────────────────────
async def _close_metaapi_trade(symbol: str, tid: str, sym_state: dict) -> bool:
    """Sequential, polled closure. Caller MUST await this fully before
    moving to the next trade — never wrap calls to this in asyncio.gather."""
    if not _metaapi_conn:
        c_log(f"Cannot close {tid} ({symbol}): no live MetaAPI connection. Position remains open on broker.")
        await send_tg_msg(f"🛑 <b>تعذّر إغلاق صفقة {symbol} ({tid}):</b> لا يوجد اتصال MetaAPI. الصفقة ما زالت مفتوحة على الوسيط.")
        return False
    try:
        await _metaapi_conn.close_position(tid)
        # State-Machine Polling for confirmation — never assume success.
        # 1s interval (not 0.2s): fetching the full portfolio 5x/sec per
        # trade during a batch closure risks tripping MetaAPI's rate limit
        # (HTTP 429). The SDK already background-syncs; 1s is plenty.
        for _ in range(25):
            positions = _metaapi_conn.terminal_state.positions
            if not any(str(p.get('id')) == str(tid) for p in positions):
                await send_tg_msg(f"✅ <b>تم إغلاق صفقة {symbol} (حقيقية) بنجاح لحماية الحساب!</b>")
                if tid in sym_state['gann_open_trades']:
                    del sym_state['gann_open_trades'][tid]
                    await save_bot_persistence()
                return True
            await asyncio.sleep(1.0)
        c_log(f"Timeout waiting for {tid} to disappear from MT5 positions after close_position call.")
        await send_tg_msg(f"⚠️ <b>لم يتم تأكيد إغلاق {symbol} ({tid}) خلال المهلة.</b> يرجى التحقق يدوياً من الحساب.")
        return False
    except Exception as e:
        log_exception(f"_close_metaapi_trade [{symbol}/{tid}]", e)
        await send_tg_msg(f"⚠️ <b>فشل الإغلاق الآلي:</b> صفقة {symbol} (خطأ: {e})\nيرجى التحقق يدوياً من الحساب.")
        return False

_EMERGENCY_CLOSE_POLL_BUDGET_SECONDS = 25  # shared across the WHOLE batch, not per-trade

async def _close_metaapi_trades_batch(closures: list) -> None:
    """Emergency mass-closure path (daily DD/profit limit hit).

    Still preserves the anti-race-condition requirement: close_position()
    write requests are issued strictly one at a time, sequentially -- same
    as _close_metaapi_trade, same TRADE_CONTEXT_BUSY protection. What
    changes is confirmation polling. get_positions() is read-only; polling
    it once per second for the ENTIRE batch (instead of each trade running
    its own private 25x1s loop, one after another) turns worst-case tail
    latency from O(N x 25s) into a single shared ~25s budget regardless of
    how many trades are closing at once. Nothing here writes to the
    broker concurrently -- only the read-only status check is shared.

    `closures` is a list of (symbol, tid, sym_state, tr) tuples for real
    trades only; callers handle simulated-trade deletion/notification
    themselves. `tr` is the trade's own dict (entry/tp/sl/tf/is_buy/
    last_known_pl/last_known_px) so every outcome message here can report
    which specific trade it is, not just "a trade on this symbol closed."
    """
    def _trade_detail_line(tr: dict) -> str:
        pl = tr.get('last_known_pl', 0.0)
        px = tr.get('last_known_px', tr.get('entry'))
        outcome_lbl = 'ربح ✅' if pl >= 0 else 'خسارة ❌'
        return (f"[جان {tr.get('tf')}] {'BUY 📈' if tr.get('is_buy') else 'SELL 📉'}\n"
                f"الدخول: {tr.get('entry')}  |  آخر سعر معروف: {px}\n"
                f"TP: {tr.get('tp')}  SL: {tr.get('sl')}\n"
                f"النتيجة: {outcome_lbl} ({pl}$)")

    if not closures:
        return
    if not _metaapi_conn:
        for symbol, tid, _, tr in closures:
            c_log(f"Cannot close {tid} ({symbol}): no live MetaAPI connection. Position remains open on broker.")
        detail = "\n\n".join(f"{symbol}: {_trade_detail_line(tr)}" for symbol, _, _, tr in closures)
        await send_tg_msg(
            f"🛑 <b>تعذّر إغلاق {len(closures)} صفقة:</b> لا يوجد اتصال MetaAPI. جميعها ما زالت مفتوحة على الوسيط.\n\n{detail}"
        )
        return

    pending = {}  # tid -> (symbol, sym_state, tr)
    for symbol, tid, sym_state, tr in closures:
        try:
            await _metaapi_conn.close_position(tid)
            pending[str(tid)] = (symbol, sym_state, tr)
        except Exception as e:
            log_exception(f"_close_metaapi_trades_batch close_position [{symbol}/{tid}]", e)
            await send_tg_msg(
                f"⚠️ <b>فشل إرسال أمر إغلاق:</b> صفقة {symbol} ({tid}, خطأ: {e})\n"
                f"يرجى التحقق يدوياً من الحساب.\n\n{_trade_detail_line(tr)}"
            )

    if not pending:
        return

    for _ in range(_EMERGENCY_CLOSE_POLL_BUDGET_SECONDS):
        if not pending:
            break
        try:
            positions = _metaapi_conn.terminal_state.positions
            if not isinstance(positions, list):
                raise TypeError(f"get_positions() returned {type(positions).__name__}, expected list")
        except Exception as e:
            log_exception("_close_metaapi_trades_batch get_positions", e)
            await asyncio.sleep(1.0)
            continue

        still_open_ids = {str(p.get('id')) for p in positions}
        for tid in list(pending.keys()):
            if tid not in still_open_ids:
                symbol, sym_state, tr = pending.pop(tid)
                await send_tg_msg(
                    f"✅ <b>تم إغلاق صفقة {symbol} (حقيقية) بنجاح لحماية الحساب!</b>\n\n{_trade_detail_line(tr)}"
                )
                pl = tr.get('last_known_pl', 0.0)
                px = tr.get('last_known_px', tr.get('entry'))
                _record_closed_trade_history(
                    symbol, tid, tr, exit_px=px, pnl=pl,
                    outcome_label=('WIN' if pl > 0 else 'LOSS' if pl < 0 else 'BREAK_EVEN'),
                    close_reason='daily_capital_protection_forced_close', pnl_confirmed=False,
                )
                if tid in sym_state['gann_open_trades']:
                    del sym_state['gann_open_trades'][tid]
                    await save_bot_persistence()

        if pending:
            await asyncio.sleep(1.0)

    for tid, (symbol, sym_state, tr) in pending.items():
        c_log(f"Timeout waiting for {tid} ({symbol}) to disappear from MT5 positions after batch close.")
        await send_tg_msg(
            f"⚠️ <b>لم يتم تأكيد إغلاق {symbol} ({tid}) خلال المهلة.</b> يرجى التحقق يدوياً من الحساب.\n\n{_trade_detail_line(tr)}"
        )

async def gann_run_diagnostics() -> str:
    """Walks through every gate _gann_open_trade's callers check, per
    active symbol, and reports the exact state of each one. Read-only --
    never opens a trade, just explains why one would or wouldn't fire
    right now."""
    lines = ["<b>🩺 تشخيص أسباب عدم فتح الصفقات</b>\n"]

    # --- Global gates (apply to every symbol) ---
    legacy_status = bot_state.get('status', 'RUNNING')
    if legacy_status != 'RUNNING':
        lines.append(f"0️⃣ ⚠️ <b>bot_state['status'] = '{legacy_status}'</b> (متوقع 'RUNNING' دائماً -- "
                      f"هذا يعني أن السكانر بالكامل متوقف بصمت. أعد تشغيل البوت فوراً.)")

    conn_state = bot_state.get('connection_state', CONN_RUNNING)
    conn_ok = conn_state == CONN_RUNNING
    lines.append(f"1️⃣ حالة الاتصال: {'✅ RUNNING' if conn_ok else f'🛑 {conn_state}'}")
    if not conn_ok:
        lines.append(f"   السبب: {bot_state.get('connection_state_reason', '-')}")

    dam_blocked = _is_within_dam_restricted_window()
    dam_now = datetime.now(timezone.utc) + timedelta(hours=3)
    filter_on = bot_state.get('prot_dam_time_filter', True)
    lines.append(f"2️⃣ فلتر أوقات دمشق: {'مفعّل' if filter_on else '🔴 معطّل'} | الوقت الآن (DAM): {dam_now:%H:%M}")
    if filter_on and dam_blocked:
        lines.append("   🛑 داخل نافذة محظورة الآن -- لن تُفتح أي صفقة جديدة حتى تنتهي.")

    overall_allowed = is_trading_allowed()
    lines.append(f"3️⃣ الخلاصة العامة is_trading_allowed(): {'✅ مسموح' if overall_allowed else '🛑 ممنوع'}\n")

    if not overall_allowed:
        lines.append("↳ طالما هذه البوابة العامة مغلقة، لن تفتح أي صفقة على أي رمز مهما كانت شروط الدخول متوفرة.\n")

    active_symbols = [s for s, on in bot_state['active_symbols'].items() if on]
    if not active_symbols:
        lines.append("⚠️ لا يوجد أي رمز مفعّل حالياً في active_symbols.")
        return "\n".join(lines)

    for symbol in active_symbols:
        sym_state = bot_state['symbol_state'][symbol]
        lines.append(f"━━━━━━━━━━━━━━\n<b>{symbol}</b>")

        # ⚠️ Real entries ONLY fire from _gann_tick_fire_check, which is
        # ONLY invoked by _GannPriceListener.on_symbol_price_updated (a
        # MetaApi WS push). Everything else below (levels, trend, distance)
        # can look perfectly "ready" off OANDA data while this feed is
        # silently dead/stale -- that combination is exactly "diagnostic
        # says ready, zero trades fire for hours" with no other symptom.
        q = live_quotes.get(symbol)
        ws_age = (time.monotonic() - q['ts']) if q else None
        if q is None:
            lines.append("📡 تغذية MetaApi اللحظية (WS): 🛑 <b>لم تصل ولا تيك واحد بعد لهذا الرمز</b> -- "
                          "بدون هذه التغذية لا يمكن لأي صفقة حقيقية أن تُفتح مهما كانت المستويات جاهزة.")
        elif ws_age > _QUOTE_STALE_SECONDS:
            lines.append(f"📡 تغذية MetaApi اللحظية (WS): 🛑 <b>متوقفة منذ {ws_age:.0f} ثانية</b> "
                          f"(آخر تحديث Bid={q['bid']} Ask={q['ask']}) -- الدخول الفعلي متجمد حتى تعود.")
        else:
            lines.append(f"📡 تغذية MetaApi اللحظية (WS): ✅ حية (عمرها {ws_age:.1f}s)")

        cycle_active = sym_state.get('gann_cycle_active', False)
        n_levels = len(sym_state.get('gann_levels', []))
        lines.append(f"دورة جان نشطة: {'✅' if cycle_active else '🛑'}  |  عدد المستويات: {n_levels}")
        if not cycle_active or n_levels == 0:
            lines.append("↳ 🛑 السكانر بيتخطى هذا الرمز بالكامل (continue) لحد ما تبدأ دورة جديدة بمستويات.")
            continue

        flt_type = sym_state['trend_filter_type']
        ttf = sym_state['trend_timeframe']
        entry_mode = sym_state['gann_entry_mode']
        lines.append(f"وضع الدخول: {entry_mode}  |  فلتر الاتجاه: {flt_type} ({ttf})")

        macro_trend_up = None
        if entry_mode == 'touch_trend':
            p_vwap = sym_state['trend_vwap_period'] if flt_type == 'vwap' else 0
            p_ema  = sym_state['trend_ema_period'] if flt_type == 'ema' else 0
            max_period = max(p_vwap, p_ema, 100)
            try:
                trend_candles = await fetch_candles(symbol, ttf, count=max(max_period + 10, 120))
            except Exception as e:
                trend_candles = []
                lines.append(f"🛑 فشل جلب بيانات الاتجاه ({ttf}): {e}")
            if not trend_candles:
                lines.append(f"🛑 لا توجد بيانات اتجاه ({ttf}) -- macro_trend_up سيبقى None وستُتجاهل كل الإشارات لكل الفريمات.")
            else:
                df_trend = pd.DataFrame(trend_candles)
                current_trend_close = float(trend_candles[-1]['close'])
                if flt_type == 'vwap':
                    df_trend['Typical_Price'] = (df_trend['high'] + df_trend['low'] + df_trend['close']) / 3
                    df_trend['VWAP'] = (df_trend['Typical_Price'] * df_trend['volume']).rolling(window=p_vwap).sum() / df_trend['volume'].rolling(window=p_vwap).sum()
                    current_vwap = df_trend.iloc[-1]['VWAP']
                    if pd.isna(current_vwap): current_vwap = current_trend_close
                    macro_trend_up = current_trend_close > current_vwap
                    lines.append(f"الاتجاه (VWAP{p_vwap}): إغلاق {current_trend_close:.2f} مقابل VWAP {current_vwap:.2f} -> {'صاعد ⬆️' if macro_trend_up else 'هابط ⬇️'}")
                elif flt_type == 'ema':
                    df_trend['EMA'] = df_trend['close'].ewm(span=p_ema, adjust=False).mean()
                    current_ema = df_trend.iloc[-1]['EMA']
                    macro_trend_up = current_trend_close > current_ema
                    lines.append(f"الاتجاه (EMA{p_ema}): إغلاق {current_trend_close:.2f} مقابل EMA {current_ema:.2f} -> {'صاعد ⬆️' if macro_trend_up else 'هابط ⬇️'}")

        levels = gann_active_levels(symbol)
        margin = sym_state['gann_touch_margin_pts'] * SYMBOL_INFO[symbol]['pip_value']
        enabled_tfs = [tf for tf, on in sym_state['gann_monitor_tfs'].items() if on]
        if not enabled_tfs:
            lines.append("🛑 لا يوجد أي فريم مفعّل في gann_monitor_tfs -- لن يتم فحص أي شيء.")
            continue

        # Same single source of truth as the live scanner -- one real-time
        # price for this symbol, reused for every timeframe line below.
        # This report used to show a different "current price" per tf
        # (each one its own stale candle close), which is exactly the
        # desync users were seeing between e.g. 1m and 30m.
        master_px = await fetch_master_price(symbol)
        if master_px is None:
            lines.append("🛑 بيانات غير كافية من OANDA (تعذّر جلب السعر الحالي الموحّد) -- تم تخطي كل الفريمات لهذا الرمز.")
            continue

        for tf in enabled_tfs:
            already_open = any(isinstance(v, dict) and v.get('tf') == tf for v in sym_state['gann_open_trades'].values())
            if already_open:
                lines.append(f"[{tf}] 🛑 يوجد صفقة مفتوحة بالفعل على هذا الفريم -- لن تُفتح صفقة ثانية.")
                continue

            try:
                candles = await fetch_candles(symbol, tf, count=sym_state['gann_atr_period'] + 50)
            except Exception as e:
                lines.append(f"[{tf}] 🛑 فشل جلب الشموع: {e}")
                continue
            if not candles or len(candles) < 3:
                lines.append(f"[{tf}] 🛑 بيانات غير كافية من OANDA.")
                continue

            live_px = master_px  # unified real-time price, NOT candles[-1]['close']
            trend_up = True
            if entry_mode == 'touch_trend':
                if macro_trend_up is None:
                    lines.append(f"[{tf}] 🛑 لا يمكن التحقق من الاتجاه (انظر أعلاه) -- الإشارات متجاهَلة.")
                    continue
                trend_up = macro_trend_up

            # Only trend-compatible levels can actually produce a trade in
            # touch_trend mode -- an against-trend level being the closest
            # one is meaningless noise (it will ALWAYS show "blocked by
            # trend" and tells you nothing about whether a real opportunity
            # is nearby). Restrict "nearest" to levels the bot would
            # actually be willing to act on.
            if entry_mode == 'touch_trend' and macro_trend_up is not None:
                directional_levels = [lv for lv in levels if (lv['dir'] == 'dn') == trend_up]
            else:
                directional_levels = levels

            nearest = None
            for lv in directional_levels:
                combo_key = f"{lv['key']}_{tf}" if bot_state['prot_allow_multi_tf'] else lv['key']
                status = sym_state['gann_level_status'].get(combo_key)
                dist = abs(live_px - lv['price'])
                is_buy = (lv['dir'] == 'dn')
                if nearest is None or dist < nearest['dist']:
                    nearest = {'dist': dist, 'price': lv['price'], 'status': status, 'is_buy': is_buy}

            if nearest is None:
                lines.append(f"[{tf}] السعر: {live_px:.2f} -- لا توجد مستويات متوافقة مع الترند الحالي.")
                continue

            within_margin = nearest['dist'] <= margin
            reason_blocked = []
            if nearest['status'] == 'used':
                reason_blocked.append('المستوى مستخدم بالفعل')
            if not within_margin:
                nd = nearest['dist']
                reason_blocked.append(f"بعيد عن الهامش ({nd:.3f} > {margin:.3f})")

            # exec_mode gate: within_margin above only checks the live tick,
            # but _gann_tick_fire_check applies a SECOND, mode-specific gate
            # on top of it (see gann_execution_mode) -- this was previously
            # invisible here, so /diagnose could say "ready" for a level the
            # real entry logic would still reject in close/hybrid mode.
            exec_mode = bot_state.get('gann_execution_mode', 'instant')
            closed_close = float(candles[-1]['close'])
            spike_limit = bot_state.get('gann_spike_limit_pts', 20) * SYMBOL_INFO[symbol]['pip_value']

            if exec_mode == 'all_concurrent':
                # Each channel is fully independent (own dedup key), so report
                # each one's own ready/blocked state rather than a single verdict.
                chan_status = []
                for chan, chan_lbl in (('touch', 'لمس'), ('close', 'إغلاق'), ('hybrid', 'هجين')):
                    combo = f"{nearest['price']}_{chan}"  # display purposes only
                    if not within_margin:
                        chan_status.append(f"{chan_lbl}:🛑بعيد")
                    elif chan == 'close' and abs(closed_close - nearest['price']) > margin:
                        chan_status.append(f"{chan_lbl}:🛑إغلاق_بعيد")
                    elif chan == 'hybrid' and abs(live_px - closed_close) > spike_limit:
                        chan_status.append(f"{chan_lbl}:🛑قفزة")
                    else:
                        chan_status.append(f"{chan_lbl}:✅")
                dir_lbl = 'دعم/شراء 🟢' if nearest['is_buy'] else 'مقاومة/بيع 🔴'
                lines.append(f"[{tf}] السعر: {live_px:.2f} (all_concurrent)  |  أقرب مستوى [{dir_lbl}]: "
                              f"{nearest['price']:.2f} (فرق {nearest['dist']:.3f})  |  " + '  '.join(chan_status))
                continue

            if within_margin and exec_mode == 'close':
                if abs(closed_close - nearest['price']) > margin:
                    reason_blocked.append(
                        f"وضع التنفيذ Close: إغلاق آخر شمعة {tf} ({closed_close:.2f}) بعيد عن المستوى "
                        f"({abs(closed_close - nearest['price']):.3f} > {margin:.3f})")
            elif within_margin and exec_mode == 'hybrid':
                if abs(live_px - closed_close) > spike_limit:
                    reason_blocked.append(
                        f"وضع التنفيذ Hybrid: قفزة سعرية عن آخر إغلاق ({abs(live_px - closed_close):.3f} > {spike_limit:.3f})")

            status_icon = '✅ جاهز للدخول' if (within_margin and not reason_blocked) else ('🛑 ' + ' | '.join(reason_blocked) if reason_blocked else '🟡 خارج الهامش')
            dir_lbl = 'دعم/شراء 🟢' if nearest['is_buy'] else 'مقاومة/بيع 🔴'
            lines.append(f"[{tf}] السعر: {live_px:.2f} (وضع: {exec_mode})  |  أقرب مستوى موافق للترند [{dir_lbl}]: {nearest['price']:.2f} (فرق {nearest['dist']:.3f})  |  {status_icon}")

    return "\n".join(lines)

async def export_diag_log_excel() -> None:
    """Export the FULL rolling diagnostic log (bot_state['diag_log']) to an
    .xlsx file and send it via Telegram. Unlike /diagnose (a single
    point-in-time snapshot), this covers every (symbol, timeframe) decision
    the live scanner made since the bot last restarted -- including the
    previously-silent skip reasons (insufficient OANDA candles, trend
    undetermined, cap reached, already open) that never got their own
    Telegram message.
    """
    log = list(bot_state.get('diag_log', []))
    if not log:
        await send_tg_msg("لا يوجد سجل تشخيص محفوظ بعد (السجل يبدأ بالتجمع فور بدء تشغيل البوت).")
        return

    df = pd.DataFrame(log)
    if 'ts' in df.columns:
        df['الوقت (DAM)'] = df['ts'].apply(lambda t: _utc_to_dam(t).strftime('%Y-%m-%d %H:%M:%S') if pd.notna(t) else '')
        df = df.drop(columns=['ts'])

    # Friendlier column order/names, but keep every raw field -- nothing summarized away.
    preferred_order = ['الوقت (DAM)', 'symbol', 'tf', 'master_px', 'trend_up', 'margin',
                        'nearest_compatible_level', 'nearest_dist', 'within_margin',
                        'touch_attempted', 'skip_reason']
    cols = [c for c in preferred_order if c in df.columns] + [c for c in df.columns if c not in preferred_order]
    df = df[cols]

    fname = f"DiagLog_{datetime.now(timezone.utc).strftime('%Y%m%d_%H%M%S')}.xlsx"
    try:
        with pd.ExcelWriter(fname, engine='openpyxl') as writer:
            if 'symbol' in df.columns:
                for sym in sorted(df['symbol'].dropna().unique()):
                    sheet_name = str(sym)[:31]  # Excel sheet name hard limit
                    df[df['symbol'] == sym].to_excel(writer, sheet_name=sheet_name, index=False)
            else:
                df.to_excel(writer, sheet_name='diag_log', index=False)

        first_ts = _utc_to_dam(log[0]['ts']).strftime('%Y-%m-%d %H:%M') if log[0].get('ts') else '?'
        last_ts = _utc_to_dam(log[-1]['ts']).strftime('%Y-%m-%d %H:%M') if log[-1].get('ts') else '?'
        await send_tg_document(
            fname,
            f"📊 <b>سجل تشخيص تفصيلي كامل</b>\n"
            f"{len(log)} سطر (قرار فحص) — من {first_ts} إلى {last_ts} (توقيت دمشق)\n"
            f"كل سطر = قرار واحد للسكانر الحي لكل (رمز، فريم) بكل دورة فحص، بما فيها أسباب "
            f"التجاهل التي لم تُرسَل كرسالة تيليجرام من قبل."
        )
    finally:
        if os.path.exists(fname):
            os.remove(fname)

async def export_live_trades_excel() -> None:
    """Export every CLOSED real/live trade (bot_state['live_trade_history'])
    to an .xlsx styled to match the backtest reports exactly (same headers
    where they overlap, same WIN/LOSS/BREAK_EVEN color fills, same borders/
    column widths) -- but for actual broker trades, with extra columns a
    backtest doesn't need: entry slippage vs the intended level, whether the
    close price/pnl is broker-confirmed or an estimate, why it closed early
    (TP/SL hit vs daily capital-protection force-close), duration, and the
    feed/latency this specific trade fired under."""
    hist = list(bot_state.get('live_trade_history', []))
    if not hist:
        await send_tg_msg("لا يوجد سجل صفقات حية مغلقة بعد (يبدأ التسجيل تلقائياً من أول صفقة حقيقية تُغلق بعد هذا التحديث).")
        return

    gray_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
    red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    be_fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")
    yellow_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")  # unconfirmed PnL estimate

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "الصفقات الحية"
    headers = ["الزوج", "TF", "حقيقية/وهمية", "اتجاه", "وقت الفتح (DAM)", "وقت الإغلاق (DAM)",
               "المدة (د)", "مستوى الدخول", "الدخول الفعلي", "انزلاق الدخول", "TP", "SL",
               "سعر الإغلاق", "النتيجة", "ربح ($)", "مؤكد من الوسيط؟", "سبب الإغلاق",
               "BE مفعّل؟", "مصدر التغذية", "عمر التغذية (ms)"]
    ws.append(headers)
    for cell in ws[1]:
        cell.fill = gray_fill
        cell.font = Font(bold=True)

    _OUTCOME_DISPLAY = {'WIN': 'WIN ✅', 'LOSS': 'LOSS ❌', 'BREAK_EVEN': 'BREAK_EVEN ⚖️'}
    _REASON_DISPLAY = {
        'tp_sl_or_manual_broker_close': 'TP/SL (مؤكد من الوسيط)',
        'tp_sl_hit': 'TP/SL (تقديري)',
        'daily_capital_protection_forced_close': '⏹️ إغلاق مبكر (حماية رأس المال اليومية)',
    }
    running_bal = 0.0
    n_win = n_loss = n_be = 0
    for tr in hist:
        pnl = tr.get('pnl') or 0.0
        running_bal += pnl
        outcome = tr.get('outcome')
        if outcome == 'WIN': n_win += 1
        elif outcome == 'LOSS': n_loss += 1
        elif outcome == 'BREAK_EVEN': n_be += 1

        def _dam(iso):
            if not iso: return ''
            try:
                dt = datetime.fromisoformat(iso)
                return _utc_to_dam(dt).strftime('%Y-%m-%d %H:%M:%S')
            except Exception:
                return str(iso)

        row = [
            tr.get('symbol'), tr.get('tf'), 'حقيقية' if tr.get('is_real') else 'وهمية (Paper)',
            'BUY 📈' if tr.get('is_buy') else 'SELL 📉', _dam(tr.get('opened_at')), _dam(tr.get('closed_at')),
            tr.get('duration_min'), tr.get('level_price'), tr.get('entry'), tr.get('entry_slippage'),
            tr.get('tp'), tr.get('sl'), tr.get('exit_price'), _OUTCOME_DISPLAY.get(outcome, outcome), pnl,
            '✅' if tr.get('pnl_confirmed_from_broker') else '⚠️ تقديري', _REASON_DISPLAY.get(tr.get('close_reason'), tr.get('close_reason')),
            '✅' if tr.get('be_activated') else '—', tr.get('feed_source') or '—', tr.get('feed_age_ms'),
        ]
        ws.append(row)
        row_idx = ws.max_row
        fill = green_fill if outcome == 'WIN' else red_fill if outcome == 'LOSS' else be_fill if outcome == 'BREAK_EVEN' else None
        if fill:
            for col in range(1, len(headers) + 1):
                ws.cell(row=row_idx, column=col).fill = fill
        if not tr.get('pnl_confirmed_from_broker') and tr.get('is_real'):
            ws.cell(row=row_idx, column=16).fill = yellow_fill

    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    center_align = Alignment(horizontal='center', vertical='center')
    for row in ws.iter_rows():
        for cell in row:
            cell.border = thin_border
            cell.alignment = center_align
    from openpyxl.utils import get_column_letter
    for i in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(i)].width = 20.0

    fname = f"LiveTrades_{datetime.now(timezone.utc).strftime('%Y%m%d_%H%M%S')}.xlsx"
    wb.save(fname)
    try:
        total = len(hist)
        wr = round(100 * n_win / max(1, n_win + n_loss), 1)
        summary = (
            f"📒 <b>سجل الصفقات الحية الكامل</b>\n"
            f"{total} صفقة مغلقة  |  WR: {wr}% ({n_win} ربح / {n_loss} خسارة / {n_be} تعادل)\n"
            f"صافي: {running_bal:+.2f}$\n\n"
            f"⚠️ الصفوف الصفراء = ربح تقديري لم يتأكد بعد من سجل الوسيط."
        )
        await send_tg_document(fname, summary)
    finally:
        if os.path.exists(fname):
            os.remove(fname)

async def gann_monitor_scanner() -> None:
    global _last_scanner_error_alert_ts, _last_any_tick_ts
    c_log('Gann live scanner started.')
    while True:
        try:
            # ── Cold-start self-heal ──
            # If _metaapi_conn (or _metaapi_account) is still None, neither
            # of the two watchdogs below can do anything -- both require a
            # connection object to already exist before they'll act. This
            # is exactly the gap that left the bot in silent, permanent
            # READ_ONLY when the ONE startup connection attempt lost a race
            # with a transient MetaApi/broker hiccup. Retry from scratch
            # here, every scanner tick, until it succeeds.
            if _metaapi_conn is None or _metaapi_account is None:
                await _bootstrap_metaapi_connection()

            # ── WS tick-silence watchdog (runs BEFORE and INDEPENDENTLY of the
            # connection_status check below) ──
            # This is the fix for the 6-hour rollover freeze: connection_status
            # can keep reporting CONNECTED while the WS session is actually a
            # zombie (no more ticks, no on_disconnected fired). The only thing
            # that can't lie here is "how long since the last real tick
            # arrived" -- _last_any_tick_ts, stamped directly in the price
            # listener. >60s of silence during market hours forces a full
            # connection teardown+rebuild, not just a symbol re-subscribe.
            active_syms_now = [s for s, on in bot_state['active_symbols'].items() if on]
            if (_metaapi_conn is not None and active_syms_now and _is_market_hours_now()
                    and (time.monotonic() - _last_any_tick_ts) > _WS_WATCHDOG_STALE_SECONDS):
                await _force_full_reconnect(
                    f"لا تيك واحد وصل منذ {time.monotonic() - _last_any_tick_ts:.0f}s "
                    f"(الحد: {_WS_WATCHDOG_STALE_SECONDS:.0f}s)"
                )

            # MT5 Zombie Singleton Heartbeat
            if _metaapi_account and _metaapi_account.connection_status != 'CONNECTED':
                await set_connection_state(CONN_READ_ONLY, "MetaAPI connection lost — attempting reconnect.")
                reconnected = False
                for attempt in range(5):
                    try:
                        await _metaapi_conn.connect()
                        await _metaapi_conn.wait_synchronized()
                        for sym, on in bot_state['active_symbols'].items():
                            if on:
                                await _lq_subscribe_symbol(sym)
                        c_log("MetaAPI Reconnected successfully (live quotes resubscribed).")
                        reconnected = True
                        _last_any_tick_ts = time.monotonic()
                        break
                    except Exception as e:
                        log_exception(f"MetaAPI reconnect attempt {attempt+1}/5", e)
                        await asyncio.sleep(2 ** attempt)
                if reconnected:
                    await set_connection_state(CONN_RUNNING, "MetaAPI reconnected and synchronized.")
                else:
                    # Do not spin forever inside this loop iteration; stay
                    # READ_ONLY, log it, and let the next scanner tick retry.
                    # If this persists, an operator will see the escalation
                    # message and the repeated READ_ONLY state in logs.
                    c_log("MetaAPI reconnect exhausted 5 attempts this tick; will retry next cycle.")

            # Feed-level staleness watchdog: connection_status can still say
            # CONNECTED while a symbol's subscription silently dropped (no
            # more ticks arriving). This is caught independently of the
            # connection-level check above, and just re-subscribes rather
            # than tearing down the whole connection.
            elif _metaapi_conn is not None:
                for sym, on in bot_state['active_symbols'].items():
                    if on and _lq_is_stale(sym):
                        c_log(f"Live quote feed stale for {sym} -- resubscribing.")
                        await _lq_subscribe_symbol(sym)

            now_dt = datetime.now(timezone.utc)

            today_date = now_dt.date()
            if bot_state.get('live_daily_date') != today_date:
                c_log(f"New trading day detected ({bot_state.get('live_daily_date')} -> {today_date}). "
                      f"Resetting daily PnL counters.")
                bot_state['live_daily_date'] = today_date
                bot_state['live_daily_realized'] = 0.0
                bot_state['live_daily_hit'] = False
                # Save immediately — do not wait for the next trade event.
                # A crash between this reset and the next save would
                # otherwise reload yesterday's PnL/hit-flag on restart.
                await save_bot_persistence()

            if bot_state.get('live_daily_hit'):
                # New entries stay blocked either way. But don't let a failed/
                # incomplete mass closure go un-retried forever just because
                # the flag that blocks new entries is also what this gate
                # checks -- those are two different concerns.
                stale_active_symbols = [s for s, on in bot_state['active_symbols'].items() if on]
                stale_real_closures = []
                for symbol in stale_active_symbols:
                    sym_state = bot_state['symbol_state'][symbol]
                    for tid, tr in list(sym_state['gann_open_trades'].items()):
                        if tr.get('is_real') and _metaapi_conn:
                            stale_real_closures.append((symbol, tid, sym_state, tr))
                if stale_real_closures:
                    c_log(f"live_daily_hit is set but {len(stale_real_closures)} real trade(s) are still open -- "
                          f"retrying the mass closure (previous attempt may have crashed/errored mid-way).")
                    await _close_metaapi_trades_batch(stale_real_closures)
                await asyncio.sleep(60)
                continue
                
            active_symbols = [s for s, on in bot_state['active_symbols'].items() if on]
            total_floating = 0.0
            
            # --- First pass: track open trades ---
            for symbol in active_symbols:
                sym_state = bot_state['symbol_state'][symbol]
                
                if bot_state.get('prot_cycle_inval', True) and sym_state.get('gann_close_used'):
                    mc = await fetch_candles(symbol, '1m', count=2)
                    if mc:
                        live_px = float(mc[-1]['close'])
                        dist = abs(live_px - sym_state['gann_close_used'])
                        pv = SYMBOL_INFO[symbol]['pip_value']
                        inval_pts = bot_state.get('prot_cycle_inval_pts', 200) * pv
                        if dist > inval_pts:
                            sym_state['gann_levels'] = []
                            sym_state['gann_close_used'] = None
                            await send_tg_msg(f"🚨 <b>إلغاء دورة {symbol}:</b> السعر تحرك بحدة! تم تجميد التداول بانتظار الدورة القادمة للحماية.")
                
                if sym_state['gann_open_trades']:
                    # --- MetaAPI Strict Reconciliation (Per Symbol, Just-In-Time) ---
                    actual_positions = {}
                    sync_failed = False
                    if bot_state.get('prot_true_sync', True) and _metaapi_conn:
                        try:
                            positions = _metaapi_conn.terminal_state.positions
                            for p in positions: actual_positions[str(p.get('id'))] = p
                            # Sync succeeded — if we were previously degraded because of
                            # sync failures specifically, this is our signal to recover.
                            if bot_state.get('connection_state') == CONN_READ_ONLY and \
                               'sync' in bot_state.get('connection_state_reason', '').lower():
                                await set_connection_state(CONN_RUNNING, "MetaAPI get_positions() succeeded again.")
                        except Exception as e:
                            log_exception(f"MetaAPI get_positions [{symbol}]", e)
                            sync_failed = True
                            await set_connection_state(
                                CONN_READ_ONLY,
                                f"MetaAPI get_positions() sync failed for {symbol}: {e}. "
                                f"Halting new trades and skipping reconciliation this tick (Amnesia Prevention)."
                            )

                    if sync_failed:
                        continue # DO NOT proceed with reconciliation or risk Amnesia Wipe

                    mc = await fetch_candles(symbol, '1m', count=2)
                    
                    # Drawdown Blindspot Fix: Fallback to MT5 prices if Oanda fails
                    live_px = None
                    oanda_failed = False
                    if not mc:
                        oanda_failed = True
                    else:
                        candle_age = (now_dt - mc[-1]['time']).total_seconds()
                        if bot_state.get('prot_stale_filter', True) and candle_age > 120:
                            oanda_failed = True
                    
                    live_px = None
                    if not oanda_failed:
                        live_px = float(mc[-1]['close'])
                    else:
                        c_log(f"Oanda failed for {symbol}. Decoupled Mode: using MT5 currentPrice for open trade management.")
                        


                    closed_ids = []
                    
                    # Pre-fetch history if there are missing real trades to prevent DDoS
                    history_deals_cache = None
                    missing_tids = [t for t, v in sym_state['gann_open_trades'].items() if v.get('is_real') and t not in actual_positions]
                    if missing_tids and _metaapi_conn:
                        # Retry with backoff (Point 6): MetaAPI's own history sync can
                        # lag a few seconds behind the actual broker-side close,
                        # especially when several positions close in the same tick
                        # (like a mass level-touch closing many trades at once).
                        # Try immediately first (no delay -- the common/fast case
                        # where sync already caught up costs nothing extra), then
                        # back off 3s, then 5s, re-checking each time whether every
                        # missing trade's closing deal has shown up yet. Only after
                        # all three attempts do we fall back to the estimate.
                        from datetime import timedelta
                        start_time = datetime.now(timezone.utc) - timedelta(days=2)
                        for attempt_i, delay in enumerate((0, 3, 5)):
                            if delay:
                                await asyncio.sleep(delay)
                            try:
                                history_deals_cache = await _metaapi_conn.get_history_deals_by_time_range(start_time, datetime.now(timezone.utc))
                            except Exception as e:
                                log_exception(f"get_history_deals_by_time_range [{symbol}] attempt {attempt_i+1}/3", e)
                                continue
                            found_now = {
                                str(d.get('positionId')) for d in history_deals_cache
                                if d.get('entryType') in ('DEAL_ENTRY_OUT', 'DEAL_ENTRY_OUT_BY')
                            }
                            if all(str(t) in found_now for t in missing_tids):
                                break  # every missing trade's closing deal is visible -- no need to keep waiting
                    
                    for tid, tr in list(sym_state['gann_open_trades'].items()):
                        is_buy = tr.get('is_buy')
                        tp = tr.get('tp')
                        sl = tr.get('sl')
                        entry = tr.get('entry')
                        tf = tr.get('tf')
                        is_real = tr.get('is_real')
                        
                        active_px = live_px
                        if active_px is None:
                            if tid in actual_positions:
                                active_px = _safe_float(actual_positions[tid].get('currentPrice'), entry)
                            else:
                                active_px = tr.get('last_known_px') # Use last known, never artificially force entry
                                
                        if active_px is None:
                            # Completely blind, skip risk evaluation for this specific trade to avoid corrupting limits
                            continue
                            
                        tr['last_known_px'] = active_px
                        
                        diff = (active_px - entry) if is_buy else (entry - active_px)
                        cs = SYMBOL_INFO[symbol]['contract_size']
                        trade_pl = round(diff * sym_state['lot_size'] * cs, 2)
                        tr['last_known_pl'] = trade_pl
                        
                        if is_real and bot_state.get('prot_true_sync', True) and _metaapi_conn:
                            if tid not in actual_positions:
                                # Pre-fetched history deals (outside the loop to prevent Rate Limit Suicide)
                                exact_pnl = trade_pl  # Estimate fallback only — NOT the real MT5 profit
                                found_deal = False
                                if history_deals_cache is not None:
                                    deal_pnl = 0.0
                                    # DEAL_ENTRY_OUT covers a normal full close; DEAL_ENTRY_OUT_BY
                                    # covers "close by" an opposite position. Both are genuine
                                    # closing deals and both must count, or partial-close /
                                    # close-by trades will silently fall back to the estimate too.
                                    for d in history_deals_cache:
                                        if (str(d.get('positionId')) == str(tid)
                                                and d.get('entryType') in ('DEAL_ENTRY_OUT', 'DEAL_ENTRY_OUT_BY')):
                                            deal_pnl += _safe_float(d.get('profit')) + _safe_float(d.get('swap')) + _safe_float(d.get('commission'))
                                            found_deal = True
                                    if found_deal:
                                        exact_pnl = deal_pnl
                                        c_log(f"Reconciliation: Exact PnL for {tid} fetched from MT5: {exact_pnl}$")

                                closed_ids.append(tid)
                                bot_state['live_daily_realized'] += exact_pnl

                                if found_deal:
                                    # This mirrors the true, realized, closed profit from MT5's
                                    # own history (includes slippage, swap, commission) — never
                                    # a cached floating/estimated value.
                                    msg = f"🔔 <b>مزامنة: إغلاق صفقة [{symbol} - {tf}]</b>\nالربح الفعلي (MT5): {exact_pnl:.2f}$"
                                else:
                                    # MT5 history hasn't synced this deal yet (or lookup failed).
                                    # Never present an unconfirmed estimate as "الربح الفعلي" —
                                    # that's exactly how a fake/incorrect profit gets reported.
                                    log_exception(f"Reconciliation MISS [{symbol}/{tid}]",
                                                  Exception("closing deal not found in MT5 history; reporting estimate"))
                                    msg = (f"🔔 <b>مزامنة: إغلاق صفقة [{symbol} - {tf}]</b>\n"
                                           f"⚠️ ربح تقديري (لم تُؤكَّد بعد من سجل MT5): ~{exact_pnl:.2f}$\n"
                                           f"سيتم تصحيح الرقم تلقائياً عند تأكيد الصفقة من السجل.")

                                await send_tg_msg(msg)
                                _record_closed_trade_history(
                                    symbol, tid, tr, exit_px=active_px, pnl=exact_pnl,
                                    outcome_label=('WIN' if exact_pnl > 0 else 'LOSS' if exact_pnl < 0 else 'BREAK_EVEN'),
                                    close_reason='tp_sl_or_manual_broker_close', pnl_confirmed=found_deal,
                                )
                                continue
                            else:
                                trade_pl = _safe_float(actual_positions[tid].get('unrealizedProfit'), trade_pl)
                        
                        outcome = core_eval_outcome(is_buy, active_px, tp, sl)
                            
                        if bot_state.get('prot_cost_be', True) and sym_state.get('break_even_enabled') and not tr.get('be_activated'):
                            be_pts = sym_state.get('gann_be_trigger_points', 40)
                            net_be = core_eval_break_even(is_buy, entry, active_px, SYMBOL_INFO[symbol]['pip_value'], be_pts, sym_state.get('gann_atr_period', 14), bot_state.get('prot_cost_be', True))
                            if net_be is not None:
                                if is_real and _metaapi_conn:
                                    try:
                                        await _metaapi_conn.modify_position(tid, stop_loss=net_be)
                                        tr['sl'] = net_be
                                        tr['be_activated'] = True # Only set if successful!
                                        await save_bot_persistence()
                                        await send_tg_msg(f"🛡️ تم تفعيل Break-Even لـ {symbol}!")
                                    except Exception as e:
                                        log_exception(f"BE modify_position [{symbol}/{tid}]", e)
                                        # be_activated stays False so we retry next tick; the
                                        # user is told immediately since capital protection failed.
                                        await send_tg_msg(f"⚠️ <b>فشل تفعيل Break-Even لـ {symbol} ({tid}):</b> {e}\nسيُعاد المحاولة تلقائياً.")
                                else:
                                    tr['sl'] = net_be
                                    tr['be_activated'] = True
                                    await save_bot_persistence()

                        if outcome:
                            closed_ids.append(tid)
                            bot_state['live_daily_realized'] += trade_pl
                            msg = f"🔔 <b>تحديث صفقة [{symbol} - جان {tf}]</b>\n\nالنتيجة: {outcome} ({trade_pl}$)\nسعر الإغلاق: {live_px:.2f}"
                            await send_tg_msg(msg)
                            _record_closed_trade_history(
                                symbol, tid, tr, exit_px=live_px, pnl=trade_pl, outcome_label=outcome,
                                close_reason='tp_sl_hit', pnl_confirmed=False,
                            )
                        else:
                            total_floating += trade_pl
                            
                    for tid in closed_ids:
                        if tid in sym_state['gann_open_trades']:
                            del sym_state['gann_open_trades'][tid]
                            await save_bot_persistence()

            # --- Evaluate Daily Limits ---
            total_daily = bot_state['live_daily_realized'] + total_floating
            dd_limit = -float(bot_state.get('prot_daily_dd_usd', 220))
            profit_limit = float(bot_state.get('prot_daily_profit_usd', 150))
            
            if (dd_limit < 0 and total_daily <= dd_limit) or (profit_limit > 0 and total_daily >= profit_limit):
                bot_state['live_daily_hit'] = True
                limit_type = '🛑 تراجع عائم' if total_daily <= dd_limit else '✅ هدف يومي عائم'
                await send_tg_msg(f"{limit_type} تم الوصول إليه! ({total_daily:.2f}$)\nسيتم إغلاق جميع الصفقات المفتوحة بالتسلسل.")
                
                # Batch closure: sequential close requests, one shared
                # confirmation loop (see _close_metaapi_trades_batch) --
                # avoids O(N x 25s) tail latency during a mass closure.
                real_closures = []
                for symbol in active_symbols:
                    sym_state = bot_state['symbol_state'][symbol]
                    for tid, tr in list(sym_state['gann_open_trades'].items()):
                        if tr.get('is_real') and _metaapi_conn:
                            real_closures.append((symbol, tid, sym_state, tr))
                        else:
                            pl = tr.get('last_known_pl', 0.0)
                            px = tr.get('last_known_px', tr.get('entry'))
                            outcome_lbl = 'ربح ✅' if pl >= 0 else 'خسارة ❌'
                            await send_tg_msg(
                                f"⏹️ <b>إغلاق (وهمي) [{symbol} - جان {tr.get('tf')}]</b>\n"
                                f"سبب الإغلاق: حماية رأس المال (تراجع/هدف يومي)\n\n"
                                f"الاتجاه: {'BUY 📈' if tr.get('is_buy') else 'SELL 📉'}\n"
                                f"الدخول: {tr.get('entry')}  |  الإغلاق: {px}\n"
                                f"TP: {tr.get('tp')}  SL: {tr.get('sl')}\n"
                                f"النتيجة: {outcome_lbl} ({pl}$)"
                            )
                            _record_closed_trade_history(
                                symbol, tid, tr, exit_px=px, pnl=pl,
                                outcome_label=('WIN' if pl > 0 else 'LOSS' if pl < 0 else 'BREAK_EVEN'),
                                close_reason='daily_capital_protection_forced_close', pnl_confirmed=False,
                            )
                            del sym_state['gann_open_trades'][tid]
                            await save_bot_persistence()
                await _close_metaapi_trades_batch(real_closures)
                continue

            for symbol in active_symbols:
                try:
                    sym_state = bot_state['symbol_state'][symbol]

                    flt_type = sym_state['trend_filter_type']
                    ttf = sym_state['trend_timeframe']
                    enabled_tfs = [tf for tf, on in sym_state['gann_monitor_tfs'].items() if on]
                    if not sym_state['gann_cycle_active'] or not sym_state['gann_levels']:
                        continue

                    macro_trend_up = None
                    if sym_state['gann_entry_mode'] == 'touch_trend':
                        p_vwap = sym_state['trend_vwap_period'] if flt_type == 'vwap' else 0
                        p_ema  = sym_state['trend_ema_period'] if flt_type == 'ema' else 0
                        max_period = max(p_vwap, p_ema, 100)
                    
                        trend_candles = await fetch_candles(symbol, ttf, count=max(max_period+10, 120))
                        if trend_candles:
                            df_trend = pd.DataFrame(trend_candles)
                            current_trend_close = float(trend_candles[-1]['close'])
                        
                            if flt_type == 'vwap':
                                df_trend['Typical_Price'] = (df_trend['high'] + df_trend['low'] + df_trend['close']) / 3
                                df_trend['VWAP'] = (df_trend['Typical_Price'] * df_trend['volume']).rolling(window=p_vwap).sum() / df_trend['volume'].rolling(window=p_vwap).sum()
                                current_vwap = df_trend.iloc[-1]['VWAP']
                                if pd.isna(current_vwap): current_vwap = current_trend_close
                            
                            if flt_type == 'ema':
                                df_trend['EMA'] = df_trend['close'].ewm(span=p_ema, adjust=False).mean()
                                current_ema = df_trend.iloc[-1]['EMA']

                            if flt_type == 'vwap':
                                macro_trend_up = (current_trend_close > current_vwap)
                            elif flt_type == 'ema':
                                macro_trend_up = (current_trend_close > current_ema)

                    levels      = gann_active_levels(symbol)
                    margin      = sym_state['gann_touch_margin_pts'] * SYMBOL_INFO[symbol]['pip_value']
                    detect_time = datetime.now(timezone.utc)

                    # ── Event-driven touch detection (structural fix) ──
                    # This block used to fetch a price and immediately check
                    # it against every level/timeframe right here, once per
                    # scan cycle -- which is exactly how a quote could sit
                    # around for seconds before actually being acted on if
                    # anything upstream (reconciliation, a slow OANDA call)
                    # made this cycle run long.
                    #
                    # It no longer fires anything. It ONLY refreshes
                    # _gann_cache[symbol] -- the levels, trend state, and
                    # each enabled tf's closed-candle data. The actual touch
                    # decision now happens inside _gann_tick_fire_check(),
                    # invoked directly from _GannPriceListener.on_symbol_
                    # price_updated the INSTANT a new tick arrives, using
                    # that exact tick's price. There is no longer a "wait
                    # for the next cycle to notice" step between a tick
                    # landing and a decision being made -- acting on a stale
                    # quote is no longer something a slow cycle can cause.
                    #
                    # Refreshing this cache on the existing ~cycle cadence is
                    # fine: levels/trend/candle data doesn't need tick-level
                    # freshness, only the live price used against it does,
                    # and that now always comes straight from the tick.
                    tf_data = {}
                    for tf in enabled_tfs:
                        need = sym_state['gann_atr_period'] + 50
                        candles = await fetch_candles(symbol, tf, count=need)
                        if not candles or len(candles) < 3:
                            # This used to be a fully SILENT skip -- no Telegram
                            # message, nothing. It's exactly the kind of failure
                            # that can make a whole batch of timeframes go quiet
                            # with zero visibility. Now at least it's captured in
                            # the diagnostic log for /export_diag_excel.
                            _diag_log_add({'ts': detect_time, 'symbol': symbol, 'tf': tf,
                                           'skip_reason': f'insufficient_oanda_candles(got={len(candles) if candles else 0})'})
                            continue
                        tf_data[tf] = {'candles': candles, 'closed_close': float(candles[-1]['close'])}

                    _gann_cache[symbol] = {
                        'levels': levels, 'margin': margin, 'trend_up': macro_trend_up,
                        'enabled_tfs': list(tf_data.keys()), 'tf_data': tf_data,
                        'refreshed_at': detect_time,
                    }

                    # Diagnostics: same visibility as before (nearest compatible
                    # level + distance, whether it was in margin), now logged
                    # EVERY cycle regardless of WS quote health -- this used to
                    # log NOTHING at all whenever live_quotes[symbol] was empty/
                    # stale, which is exactly the situation most worth capturing:
                    # _gann_tick_fire_check (the only thing that ever opens a
                    # real trade) is driven purely by WS ticks, so a dead/starved
                    # WS feed silently stops all entries AND silently emptied
                    # this very log at the same time, making both symptoms
                    # ("ready per /diagnose, zero trades" and "no diag log yet")
                    # look unrelated when they're actually the same root cause.
                    q = live_quotes.get(symbol)
                    ws_age_s = round(time.monotonic() - q['ts'], 1) if q else None
                    ws_status = 'live' if (q and ws_age_s <= _QUOTE_STALE_SECONDS) else ('stale' if q else 'never_received')
                    diag_px, price_source, _age_ms = await _lq_price_with_fallback(symbol)
                    entry_mode = sym_state['gann_entry_mode']
                    directional_levels = (
                        [l for l in levels if (l['dir'] == 'dn') == macro_trend_up]
                        if entry_mode == 'touch_trend' and macro_trend_up is not None else levels
                    )
                    nearest_dist = None; nearest_price = None
                    if diag_px is not None:
                        for l in directional_levels:
                            d = abs(diag_px - l['price'])
                            if nearest_dist is None or d < nearest_dist:
                                nearest_dist = d; nearest_price = l['price']
                    _diag_log_add({'ts': detect_time, 'symbol': symbol, 'master_px': diag_px,
                                   'price_source': price_source, 'ws_status': ws_status, 'ws_quote_age_s': ws_age_s,
                                   'trend_up': macro_trend_up, 'margin': margin,
                                   'nearest_compatible_level': nearest_price, 'nearest_dist': nearest_dist,
                                   'within_margin': (nearest_dist is not None and nearest_dist <= margin),
                                   'skip_reason': ('no_price_available' if diag_px is None else
                                                   'cache_refresh_only(firing_is_now_tick_driven)')})

                except Exception as sym_exc:
                    log_exception(f"gann_monitor_scanner per-symbol [{symbol}]", sym_exc)
                    now_mono_sym = time.monotonic()
                    if now_mono_sym - _last_scanner_error_alert_ts > 300:
                        _last_scanner_error_alert_ts = now_mono_sym
                        await send_tg_msg(
                            f"🛑 <b>[{symbol}]</b> خطأ غير متوقع أثناء فحص هذا الرمز -- تم تخطيه لهذه الدورة فقط "
                            f"(باقي الرموز تستمر بشكل طبيعي):\n{sym_exc}"
                        )
                    continue
        except Exception as e:
            log_exception('gann_monitor_scanner main loop', e)
            # This top-level catch wraps EVERY symbol's processing for the
            # whole cycle -- an exception anywhere (even for just one
            # symbol/timeframe) previously aborted the ENTIRE cycle
            # completely silently (server-side log only, no Telegram
            # message), which could look exactly like "the bot just isn't
            # opening trades" with zero clue why. Surface it, rate-limited
            # so a persistent failure doesn't spam every 15s.
            now_mono = time.monotonic()
            if now_mono - _last_scanner_error_alert_ts > 300:  # at most once per 5 min
                _last_scanner_error_alert_ts = now_mono
                await send_tg_msg(
                    f"🛑 <b>خطأ غير متوقع بدورة الفحص الحية (gann_monitor_scanner):</b>\n{e}\n"
                    f"تم تخطي بقية هذه الدورة بالكامل بسببه. سيُعاد المحاولة بالدورة القادمة (~15 ثانية). "
                    f"إذا تكرر هذا الخطأ، راجع السجل الكامل (traceback) على السيرفر."
                )
        await asyncio.sleep(15)

# ─────────────────────────────────────────────────────────────
# PRO BACKTEST ENGINE (Macro Trend & Smart Break-Even)
# ─────────────────────────────────────────────────────────────

# Recalculation cadence for gann_calculation_mode == 'dynamic_live'. Shared
# constant so the live scanner (gann_cycle_manager) and both backtest
# engines (_build_gann_cycle_defs) use the exact same cadence -- required
# for backtest/live parity.
GANN_DYNAMIC_RECALC_MINUTES = 5

async def gann_cycle_manager() -> None:
    c_log('Gann cycle manager started.')
    while True:
        try:
            now_utc = datetime.now(timezone.utc)
            calc_mode = bot_state.get('gann_calculation_mode', 'static_h1')
            active_symbols = [s for s, on in bot_state['active_symbols'].items() if on]
            for symbol in active_symbols:
                sym_state = bot_state['symbol_state'][symbol]
                if not sym_state['gann_cycle_active']:
                    continue

                if calc_mode == 'dynamic_live':
                    # Classic H1-close anchoring is bypassed entirely: levels
                    # are recomputed every GANN_DYNAMIC_RECALC_MINUTES off the
                    # current live streamed price. 'gann_cycle_started_at'
                    # doubles as "last dynamic recalc timestamp" in this mode
                    # (it's not otherwise load-bearing outside display/logging).
                    last_recalc = sym_state['gann_cycle_started_at']
                    if last_recalc and (now_utc - last_recalc).total_seconds() < GANN_DYNAMIC_RECALC_MINUTES * 60:
                        continue
                    live_px, _src, _age = await _lq_price_with_fallback(symbol)
                    if live_px is None:
                        continue
                    sym_state['gann_levels'] = gann_calc_levels(symbol, live_px)
                    sym_state['gann_close_used'] = live_px
                    sym_state['gann_last_h1_time'] = now_utc
                    sym_state['gann_cycle_started_at'] = now_utc
                    sym_state['gann_level_status'] = {}
                    c_log(f'[{symbol}] Dynamic Gann recalculation at live_px={live_px}')
                    continue

                cycle_h = sym_state['gann_cycle_hours']
                last_h1 = await _gann_fetch_last_closed_anchor(symbol)
                
                if last_h1:
                    h1_time = last_h1['time']
                    # Check if this new H1 candle is newer than our currently tracked one
                    if not sym_state['gann_last_h1_time'] or h1_time > sym_state['gann_last_h1_time']:
                        # Only trigger if the difference in hours is >= cycle_h
                        if not sym_state['gann_last_h1_time'] or (h1_time - sym_state['gann_last_h1_time']).total_seconds() / 3600.0 >= cycle_h:
                            h1_close = float(last_h1['close'])
                            sym_state['gann_levels'] = gann_calc_levels(symbol, h1_close)
                            sym_state['gann_close_used'] = h1_close
                            sym_state['gann_last_h1_time'] = h1_time
                            sym_state['gann_cycle_started_at'] = now_utc
                            
                            # Clear used levels so we can take trades again!
                            sym_state['gann_level_status'] = {}
                            
                            c_log(f'[{symbol}] New {cycle_h}h cycle started at {h1_close}')
                            await send_tg_msg(f"🔄 <b>تحديث دورة جان ({cycle_h}h)</b>\nالزوج: {symbol}\nإغلاق {_anchor_label()}: {h1_close:.5f}\nتم تصفير المستويات لتبدأ من جديد!")
                            
        except Exception as e:
            log_exception('gann_cycle_manager main loop', e)

        await asyncio.sleep(60)

# -----------------------------------------------------------------
# INDEPENDENT GLOBAL LEDGER RECONCILIATION
# -----------------------------------------------------------------
# Defense-in-depth on top of the per-tick reconciliation already inside
# gann_monitor_scanner. That reconciliation only runs per-symbol, only
# when a symbol has locally-tracked open trades, and shares its process
# and state with the rest of the bot -- if bot_state itself is what's
# wrong (e.g. a trade never got recorded locally in the first place),
# the in-loop check can't catch it because it only checks trades IT
# already knows about.
#
# This task instead starts from the broker's side: fetch every open
# position that MetaAPI reports for this account, and check whether
# each one is accounted for in bot_state. A position on the broker that
# the bot has NO record of at all is the "ghost position" / unmanaged-
# exposure scenario, and it's undetectable from inside
# gann_monitor_scanner's per-symbol loop by construction.
RECONCILIATION_INTERVAL_SECONDS = 300  # every 5 minutes
_recon_consecutive_mismatches = 0
_RECON_MISMATCH_HALT_THRESHOLD = 3

async def global_ledger_reconciliation() -> None:
    global _recon_consecutive_mismatches
    c_log('Global ledger reconciliation started (independent broker cross-check).')
    while True:
        try:
            await asyncio.sleep(RECONCILIATION_INTERVAL_SECONDS)

            if bot_state.get('connection_state', CONN_RUNNING) != CONN_RUNNING or not _metaapi_conn:
                continue

            try:
                broker_positions = _metaapi_conn.terminal_state.positions
                if not isinstance(broker_positions, list):
                    raise TypeError(f"get_positions() returned {type(broker_positions).__name__}, expected list")
            except Exception as e:
                log_exception('global_ledger_reconciliation get_positions', e)
                continue

            broker_ids = {str(p.get('id')) for p in broker_positions}

            known_ids = set()
            for sym, ss in bot_state['symbol_state'].items():
                for tid, tr in ss.get('gann_open_trades', {}).items():
                    if tr.get('is_real'):
                        known_ids.add(str(tid))

            ghost_ids = broker_ids - known_ids
            missing_ids = known_ids - broker_ids

            if ghost_ids:
                _recon_consecutive_mismatches += 1
                c_log(f"RECONCILIATION MISMATCH: {len(ghost_ids)} broker position(s) with NO matching bot "
                      f"record: {ghost_ids}. Consecutive mismatches: {_recon_consecutive_mismatches}")
                await send_tg_msg(
                    f"🚨 <b>تحذير مطابقة الحساب المستقل:</b>\n"
                    f"يوجد {len(ghost_ids)} صفقة مفتوحة على الوسيط لا يعرفها البوت إطلاقاً.\n"
                    f"هذا يعني احتمال وجود تعرض غير مُدار (ghost position). يرجى التحقق يدوياً فوراً.\n"
                    f"IDs: {ghost_ids}"
                )
                if _recon_consecutive_mismatches >= _RECON_MISMATCH_HALT_THRESHOLD:
                    await set_connection_state(
                        CONN_HALTED,
                        f"{_recon_consecutive_mismatches} consecutive independent reconciliation checks found "
                        f"unmanaged broker positions. Halting all trading until a human confirms account state."
                    )
            else:
                if _recon_consecutive_mismatches > 0:
                    c_log("Reconciliation recovered: no ghost positions found this check.")
                _recon_consecutive_mismatches = 0

            if missing_ids:
                c_log(f"Reconciliation note: {len(missing_ids)} bot-tracked trade(s) not currently on "
                      f"broker (expected if closed this cycle): {missing_ids}")

        except Exception as e:
            log_exception('global_ledger_reconciliation main loop', e)

def _build_gann_cycle_defs(sym_state: dict, valid_h1: list, mc_1m: list) -> list[dict]:
    """Single source of truth for 'when do Gann levels get (re)anchored and
    from what price', shared by BOTH backtest engines (run_gann_backtest and
    run_live_twin_simulation) so a change to bot_state['gann_calculation_mode']
    always simulates identically to what the live scanner (gann_cycle_manager)
    actually does -- this is what satisfies the backtest/live parity
    requirement, rather than keeping two separately-maintained copies of the
    same logic that could silently drift apart.

    Returns a list of {'t_start', 't_end', 'close'} dicts:
      - static_h1   : one entry per closed anchor-tf candle (legacy/unchanged).
      - dynamic_live: one entry every GANN_DYNAMIC_RECALC_MINUTES, priced off
        the most recent 1m close at that instant -- the backtest's proxy for
        "current live_px", since historical tick data isn't available.
    """
    mode = bot_state.get('gann_calculation_mode', 'static_h1')
    cycle_h = sym_state['gann_cycle_hours']

    if mode != 'dynamic_live':
        out = []
        for h1 in valid_h1:
            t_start = h1['time'] + timedelta(hours=1)
            out.append({'t_start': t_start, 't_end': t_start + timedelta(hours=cycle_h), 'close': float(h1['close'])})
        return out

    px_series = sorted(mc_1m, key=lambda c: c['time']) if mc_1m else []
    if not px_series:
        return []
    out = []
    bucket = px_series[0]['time'].floor(f'{GANN_DYNAMIC_RECALC_MINUTES}min')
    last_t = px_series[-1]['time']
    i = 0; n = len(px_series)
    while bucket <= last_t:
        while i + 1 < n and px_series[i + 1]['time'] <= bucket:
            i += 1
        if px_series[i]['time'] <= bucket:
            out.append({'t_start': bucket, 't_end': bucket + timedelta(minutes=GANN_DYNAMIC_RECALC_MINUTES),
                        'close': float(px_series[i]['close'])})
        bucket += timedelta(minutes=GANN_DYNAMIC_RECALC_MINUTES)
    return out

def _add_concurrent_analysis_sheets(wb, trade_logs: list, pnl_key: str, outcome_key: str, slippage_key: str = None) -> None:
    """Only called when gann_execution_mode == 'all_concurrent'. Splits
    trade_logs (each row already tagged with 'trigger_type' in
    {'touch','close','hybrid'} from the signal-generation stage of whichever
    engine called this) into 3 per-channel sheets plus a side-by-side
    Performance_Comparison sheet, so a 24h concurrent run directly answers
    "which of the 3 execution methods actually wins" without the user having
    to run and diff 3 separate backtests by hand."""
    from openpyxl.utils import get_column_letter
    gray_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
    channel_sheets = {'touch': 'Touch_Trades', 'close': 'Close_Trades', 'hybrid': 'Hybrid_Trades'}
    channel_rows = {'touch': [], 'close': [], 'hybrid': []}
    for tr in trade_logs:
        ch = tr.get('trigger_type', 'touch')
        if ch in channel_rows:
            channel_rows[ch].append(tr)

    internal_keys = {'cycle_ts', 'cycle_time_str', 'cycle_close', 'trigger_type'}
    cols = [k for k in trade_logs[0].keys() if k not in internal_keys] if trade_logs else []

    made_sheets = []
    for ch, sheet_name in channel_sheets.items():
        ws = wb.create_sheet(sheet_name)
        made_sheets.append(ws)
        ws.append(cols)
        for cell in ws[1]:
            cell.fill = gray_fill; cell.font = Font(bold=True)
        for tr in channel_rows[ch]:
            ws.append([tr.get(c) for c in cols])
        for i in range(1, len(cols) + 1):
            ws.column_dimensions[get_column_letter(i)].width = 20.0

    ws_cmp = wb.create_sheet("Performance_Comparison")
    made_sheets.append(ws_cmp)
    ws_cmp.append(["Metric", "Touch (لمس مباشر)", "Close (إغلاق شمعة)", "Hybrid (هجين)"])
    for cell in ws_cmp[1]:
        cell.fill = gray_fill; cell.font = Font(bold=True)

    def metrics_for(rows):
        wins = [r for r in rows if r.get(outcome_key) == 'WIN']
        losses = [r for r in rows if r.get(outcome_key) == 'LOSS']
        gross_profit = sum((r.get(pnl_key) or 0) for r in wins)
        gross_loss = sum((r.get(pnl_key) or 0) for r in losses)
        net = sum((r.get(pnl_key) or 0) for r in rows)
        wr = round(100 * len(wins) / max(1, len(wins) + len(losses)), 1)
        rows_sorted = sorted(rows, key=lambda r: r.get('cycle_ts', 0))
        eq = 0.0; peak = 0.0; mdd = 0.0
        for r in rows_sorted:
            eq += (r.get(pnl_key) or 0)
            peak = max(peak, eq)
            mdd = min(mdd, eq - peak)
        avg_slip = None
        if slippage_key:
            slips = [r.get(slippage_key) for r in rows if r.get(slippage_key) is not None]
            avg_slip = round(sum(slips) / len(slips), 2) if slips else None
        return dict(total=len(rows), win=len(wins), loss=len(losses), wr=wr,
                    gp=round(gross_profit, 2), gl=round(gross_loss, 2), net=round(net, 2),
                    mdd=round(mdd, 2), avg_slip=avg_slip)

    m = {ch: metrics_for(channel_rows[ch]) for ch in channel_rows}
    rows_spec = [
        ("Total Trades", 'total'), ("Winning Trades", 'win'), ("Losing Trades", 'loss'),
        ("Win Rate (%)", 'wr'), ("Gross Profit ($)", 'gp'), ("Gross Loss ($)", 'gl'),
        ("Net PnL ($)", 'net'), ("Max Drawdown ($)", 'mdd'),
    ]
    if slippage_key:
        rows_spec.append(("Average Slippage (Pips)", 'avg_slip'))
    for label, key in rows_spec:
        ws_cmp.append([label, m['touch'][key], m['close'][key], m['hybrid'][key]])
    for i in range(1, 5):
        ws_cmp.column_dimensions[get_column_letter(i)].width = 26.0

    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    center_align = Alignment(horizontal='center', vertical='center')
    for ws in made_sheets:
        for row in ws.iter_rows():
            for cell in row:
                cell.border = thin_border
                cell.alignment = center_align

async def run_gann_backtest(start_dt: datetime, end_dt: datetime) -> None:
    global _bt_progress
    bot_state['is_backtesting'] = True
    fname = f"GannBT_{datetime.now(timezone.utc).strftime('%H%M%S')}.xlsx"
    exec_mode = bot_state.get('gann_execution_mode', 'instant')  # hoisted: must exist even if no trades are ever scanned
    
    active_symbols = [s for s, on in bot_state['active_symbols'].items() if on]
    if not active_symbols:
        bot_state['is_backtesting'] = False
        return
        
    first_sym_state = bot_state['symbol_state'][active_symbols[0]]
    
    enabled_tfs = [tf for tf, on in first_sym_state['gann_monitor_tfs'].items() if on] or ['5m']
    flt_type = first_sym_state['trend_filter_type']
    ttf = first_sym_state['trend_timeframe']
    desc_ttf = ttf.upper()
    
    if first_sym_state['gann_entry_mode'] == 'touch_trend':
        if flt_type == 'vwap': desc_mode = f"Touch(VWAP{first_sym_state['trend_vwap_period']}_{desc_ttf})\n"
        elif flt_type == 'ema': desc_mode = f"Touch(EMA{first_sym_state['trend_ema_period']}_{desc_ttf})\n"
        else: desc_mode = f"Touch(VWAP+EMA_{desc_ttf})\n"
    else:
        desc_mode = "Pure Touch"
        
    desc_be = " | 🛡️ BE" if first_sym_state['break_even_enabled'] else ""
    
    zf = first_sym_state['gann_zone_filter']
    if zf == 'star': desc_star = "⭐ الأصلية"
    elif zf == 'star_fan': desc_star = "⭐🌀 الأصلية والمروحة"
    else: desc_star = "📋 الكل"
        
    desc_tfs = "+".join(enabled_tfs)
    syms_label = "+".join(active_symbols)
    
    prog = BtProgress(label=f"{syms_label} جان H1→[{desc_tfs}] | {desc_mode} | {desc_star}{desc_be}", active_tfs=['H1']); _bt_progress = prog
    await prog.start(bot_state['chat_id'])

    res = {'win': 0, 'loss': 0, 'be': 0, 'total_prof': 0.0, 'total_win_usd': 0.0, 'total_loss_usd': 0.0, 'peak_equity': 0.0, 'max_dd': 0.0, 'trade_logs': [], 'cycle_logs': []}
    
    try:
        delta_hours = int((end_dt - start_dt).total_seconds() / 3600)
        
        # PHASE 1: Data Gathering & Signal Generation
        all_signals = []
        all_candles_events = []
        
        for symbol in active_symbols:
            sym_state = bot_state['symbol_state'][symbol]
            cycle_h = sym_state['gann_cycle_hours']; tpsl_mode = sym_state['gann_tpsl_mode']
            pv  = SYMBOL_INFO[symbol]['pip_value']; lot = sym_state['lot_size']; margin = sym_state['gann_touch_margin_pts'] * pv
            cs  = SYMBOL_INFO[symbol]['contract_size'];
            prec = SYMBOL_INFO[symbol]['prec'];
            
            quote = symbol.split('_')[1] if '_' in symbol else 'USD'
            quote_conv = {'USD': 1.0, 'JPY': 1/150.0, 'AUD': 0.66, 'NZD': 0.61, 'EUR': 1.08, 'GBP': 1.27, 'CAD': 0.73, 'CHF': 1.11}.get(quote, 1.0)

            await prog.set_phase(f'جلب بيانات الترند ({desc_ttf})...')
            max_period = max(sym_state['trend_vwap_period'], sym_state['trend_ema_period'], 100)
            trend_count = (delta_hours * (2 if ttf == '30m' else 1)) + max_period + 10
            candles_trend = await fetch_candles(symbol, ttf, count=trend_count, end_time=end_dt)
            if not candles_trend: continue

            df_trend = pd.DataFrame(candles_trend)
            if flt_type == 'vwap':
                p_vwap = sym_state['trend_vwap_period']
                df_trend['Typical_Price'] = (df_trend['high'] + df_trend['low'] + df_trend['close']) / 3
                df_trend['VWAP'] = (df_trend['Typical_Price'] * df_trend['volume']).rolling(window=p_vwap).sum() / df_trend['volume'].rolling(window=p_vwap).sum()
            if flt_type == 'ema':
                p_ema = sym_state['trend_ema_period']
                df_trend['EMA'] = df_trend['close'].ewm(span=p_ema, adjust=False).mean()

            df_trend.set_index('time', inplace=True)
            if flt_type == 'vwap': df_trend['macro_trend_up'] = df_trend['close'] > df_trend['VWAP']
            elif flt_type == 'ema': df_trend['macro_trend_up'] = df_trend['close'] > df_trend['EMA']
            elif flt_type == 'both':
                c1_up = df_trend['close'] > df_trend['VWAP']; c2_up = df_trend['close'] > df_trend['EMA']
                c1_dn = df_trend['close'] < df_trend['VWAP']; c2_dn = df_trend['close'] < df_trend['EMA']
                df_trend['macro_trend_up'] = np.where(c1_up & c2_up, True, np.where(c1_dn & c2_dn, False, None))

            anchor_gran = bot_state.get('gann_anchor_tf', '1h')
            await prog.set_phase(f'جلب بيانات {_anchor_label()}...')
            candles_h1 = await fetch_candles(symbol, anchor_gran, count=(delta_hours // _anchor_hours()) + 10, end_time=end_dt)
            if not candles_h1: continue

            await prog.set_phase('جلب شموع الفريمات الصغيرة...')
            monitor_tfs_data = {}
            days_diff = (end_dt - start_dt).days or 1
            # Always fetch 1m for high-resolution price tracking during simulation
            need_1m = days_diff * 24 * 60 + 300
            mc_1m = await fetch_candles(symbol, '1m', count=need_1m, end_time=end_dt)
            if mc_1m:
                for c in mc_1m:
                    all_candles_events.append({'time': c['time'], 'symbol': symbol, 'high': float(c['high']), 'low': float(c['low']), 'close': float(c['close']), 'tf': '1m_track'})

            for btf in enabled_tfs:
                bmin = int(''.join(filter(str.isdigit, btf)))
                if 'h' in btf: bmin *= 60
                need_m = days_diff * 24 * (60 // max(bmin, 1)) + 300
                mc = await fetch_candles(symbol, btf, count=need_m, end_time=end_dt)
                if mc: 
                    monitor_tfs_data[btf] = sorted(mc, key=lambda c: c['time'])
                    # We only need to add these to events if they might trigger signals at times not covered by 1m
                    # But 1m covers everything. Still, let's keep them in events for safety.
                    for c in mc:
                        all_candles_events.append({'time': c['time'], 'symbol': symbol, 'high': float(c['high']), 'low': float(c['low']), 'close': float(c['close']), 'tf': btf})

            start_ts = start_dt.timestamp(); end_ts = end_dt.timestamp()
            valid_h1 = [c for c in candles_h1 if start_ts <= (c['time'].timestamp() + 3600) <= end_ts]
            
            trend_freq = '30min' if ttf == '30m' else '1h'

            # gann_calculation_mode-aware: static_h1 walks closed anchor
            # candles (legacy); dynamic_live recomputes every
            # GANN_DYNAMIC_RECALC_MINUTES off 1m closes -- see
            # _build_gann_cycle_defs docstring for why this is shared with
            # run_live_twin_simulation instead of duplicated.
            cycle_defs = _build_gann_cycle_defs(sym_state, valid_h1, mc_1m)

            for idx, cdef in enumerate(cycle_defs):
                if prog.cancelled: return
                await asyncio.sleep(0)
                t_start = cdef['t_start']
                t_end   = cdef['t_end']
                close   = cdef['close']
                levels = gann_calc_levels(symbol, close)
                f_mode = sym_state['gann_zone_filter']
                active_lv = [l for l in levels if l['dir'] != 'ref' and (f_mode == 'all' or (f_mode == 'star' and l['star']) or (f_mode == 'star_fan' and (l['star'] or l['fan'])))]
                
                res['cycle_logs'].append({'symbol': symbol, 'time_ts': t_start.timestamp(), 'time_dt': t_start, 'close': close, 'levels': len(active_lv)})
                
                level_used = set()
                exec_mode = bot_state.get('gann_execution_mode', 'instant')
                spike_limit = bot_state.get('gann_spike_limit_pts', 20) * pv

                for btf, candles_m in monitor_tfs_data.items():
                    m_window = [c for c in candles_m if t_start <= c['time'] < t_end]
                    m_before = [c for c in candles_m if c['time'] < t_start]
                    atr_val  = _gann_atr(m_before, sym_state['gann_atr_period']) if tpsl_mode == 'atr' else None

                    for bar in m_window:
                        bar_close = float(bar['close']); bar_time = bar['time']
                        bar_high = float(bar['high']); bar_low = float(bar['low'])
                        trend_up = True
                        if sym_state['gann_entry_mode'] == 'touch_trend':
                            trend_time = bar_time.floor(trend_freq)
                            if trend_time in df_trend.index:
                                val = df_trend.loc[trend_time, 'macro_trend_up']
                                if isinstance(val, pd.Series): val = val.iloc[-1]
                                macro_trend_up = None if pd.isna(val) else bool(val)
                            else: macro_trend_up = None
                            if macro_trend_up is None: continue
                            trend_up = macro_trend_up

                        if bot_state.get('prot_cycle_inval', True):
                            inval_pts = bot_state.get('prot_cycle_inval_pts', 200) * pv
                            if abs(bar_close - close) > inval_pts:
                                active_lv = [] # Wipe levels for the rest of this cycle on this TF
                                break
                                
                        for lv in active_lv:
                            k = lv['key']; dir = lv['dir']
                            is_buy = (dir == 'dn')
                            if sym_state['gann_entry_mode'] == 'touch_trend':
                                if is_buy and not trend_up: continue
                                if not is_buy and trend_up: continue

                            channels = ['touch', 'close', 'hybrid'] if exec_mode == 'all_concurrent' else [
                                'close' if exec_mode == 'close' else 'hybrid' if exec_mode == 'hybrid' else 'touch']

                            for channel in channels:
                                base_combo = f'{k}_{btf}' if bot_state.get('prot_allow_multi_tf', True) else k
                                combo_key = f"{base_combo}_{channel}" if exec_mode == 'all_concurrent' else base_combo
                                if combo_key in level_used: continue

                                if channel == 'close':
                                    if abs(bar_close - lv['price']) > margin: continue
                                elif channel == 'hybrid':
                                    if not (bar_low - margin <= lv['price'] <= bar_high + margin): continue
                                    if abs(bar_close - lv['price']) > spike_limit: continue
                                else:  # touch (instant): any part of the bar's range counts (intrabar touch)
                                    if not (bar_low - margin <= lv['price'] <= bar_high + margin): continue

                                entry = lv['price']
                                be_trigger_px = None
                                if sym_state['break_even_enabled']:
                                    be_trigger_px = 'dynamic'

                                tf_tp = _gann_tf_tp(symbol, btf); tf_sl = _gann_tf_sl(symbol, btf)
                                if tpsl_mode == 'atr' and atr_val:
                                    sl_d = atr_val * sym_state['gann_atr_sl_mult']
                                    tp_d = atr_val * sym_state['gann_atr_tp_mult']
                                else:
                                    sl_d = tf_sl * pv; tp_d = tf_tp * pv

                                tp_px = entry + tp_d if is_buy else entry - tp_d
                                sl_px = entry - sl_d if is_buy else entry + sl_d

                                all_signals.append({
                                    'time': bar_time, 'symbol': symbol, 'is_buy': is_buy, 'entry': entry,
                                    'tp_px': tp_px, 'sl_px': sl_px, 'sl_d': sl_d, 'tp_d': tp_d, 'be_trigger_px': be_trigger_px,
                                    'lot': lot, 'cs': cs, 'quote_conv': quote_conv, 'tf': btf, 'combo_key': combo_key,
                                    'cycle_time': t_start, 'cycle_close': close, 'level_key': k, 'trigger_type': channel,
                                })
                                level_used.add(combo_key)
        
        # PHASE 2: Chronological Event-Driven Simulation
        await prog.set_phase('محاكاة الصفقات الزمنية (تقييم الأرباح العائمة)...')
        c_log(f'BT: Sorting {len(all_signals)} signals')
        all_signals.sort(key=lambda x: x['time'])
        c_log(f'BT: Sorting {len(all_candles_events)} events')
        all_candles_events.sort(key=lambda x: x['time'])
        
        open_trades = []
        closed_trades = []
        suspended_days = {}
        suspend_trigger_time = {}
        daily_pl = 0.0
        current_day = None
        latest_price = {}
        
        signal_idx = 0
        total_signals = len(all_signals)
        
        dd_limit = - float(bot_state['prot_daily_dd_usd'])
        profit_limit = float(bot_state['prot_daily_profit_usd'])
        
        total_events = len(all_candles_events)
        await prog.set_tf('محاكاة عائمة', total_events)
        
        for i, event in enumerate(all_candles_events):
            if i % 5000 == 0:
                await asyncio.sleep(0)
            if prog.cancelled: break
            t = event['time']; sym = event['symbol']; h = event['high']; l = event['low']; c = event['close']
            day_str = _utc_to_dam(t).strftime('%Y-%m-%d')
            latest_price[sym] = c
            
            if day_str != current_day:
                current_day = day_str
                daily_pl = 0.0
            
            # Check floating PnL against limits
            if current_day not in suspended_days:
                floating_pl = 0.0        # close-based -- used for the PROFIT check (unchanged, a late trigger costs nothing)
                floating_pl_worst = 0.0  # intrabar-worst-case -- used for the LOSS check only (tight, no overshoot)
                for tr in open_trades:
                    lp = latest_price.get(tr['symbol'], tr['entry'])
                    diff = (lp - tr['entry']) if tr['is_buy'] else (tr['entry'] - lp)
                    floating_pl += round(diff * tr['lot'] * tr['cs'] * tr['quote_conv'], 2)

                    # For the trade's own symbol, we have this candle's full
                    # high/low right now -- use the worst excursion within
                    # the bar (low for a long, high for a short) instead of
                    # only the close. For other symbols we only have their
                    # last close at this instant (an inherent limit of
                    # single-symbol event-driven iteration), so fall back
                    # to the same close-based price there.
                    worst_px = (l if tr['is_buy'] else h) if tr['symbol'] == sym else lp
                    diff_worst = (worst_px - tr['entry']) if tr['is_buy'] else (tr['entry'] - worst_px)
                    floating_pl_worst += round(diff_worst * tr['lot'] * tr['cs'] * tr['quote_conv'], 2)

                total_daily = daily_pl + floating_pl
                total_daily_worst = daily_pl + floating_pl_worst
                if dd_limit < 0 and total_daily_worst <= dd_limit:
                    suspended_days[current_day] = f'🛑 تراجع عائم (الحد {dd_limit}$ | المحقق: {round(daily_pl, 2)}$ + العائم (أسوأ لحظة داخل الشمعة): {round(floating_pl_worst, 2)}$ = {round(total_daily_worst, 2)}$)'
                elif profit_limit > 0 and total_daily >= profit_limit:
                    suspended_days[current_day] = f'✅ هدف عائم (الحد {profit_limit}$ | المحقق: {round(daily_pl, 2)}$ + العائم: {round(floating_pl, 2)}$ = {round(total_daily, 2)}$)'

                if current_day in suspended_days and current_day not in suspend_trigger_time:
                    suspend_trigger_time[current_day] = t

                if current_day in suspended_days:
                    # Close all open trades. For the loss-triggering
                    # symbol, fill at the same worst-case intrabar price
                    # that tripped the check (tight to the limit) rather
                    # than the candle's close; other symbols still fill at
                    # their last known close, same as before.
                    was_loss_trigger = dd_limit < 0 and total_daily_worst <= dd_limit
                    for tr in open_trades:
                        if was_loss_trigger and tr['symbol'] == sym:
                            lp = l if tr['is_buy'] else h
                        else:
                            lp = latest_price.get(tr['symbol'], tr['entry'])
                        diff = (lp - tr['entry']) if tr['is_buy'] else (tr['entry'] - lp)
                        p_usd = round(diff * tr['lot'] * tr['cs'] * tr['quote_conv'], 2)
                        tr['outcome'] = 'DAILY_LIMIT'
                        tr['p_usd'] = p_usd
                        tr['close_time'] = t
                        closed_trades.append(tr)
                        daily_pl += p_usd
                    open_trades.clear()
            
            # Process Exits for open trades (if not suspended)
            if current_day not in suspended_days:
                surviving_trades = []
                for tr in open_trades:
                    if tr['symbol'] != sym:
                        surviving_trades.append(tr)
                        continue
                        
                    is_buy = tr['is_buy']; sl_current = tr['sl_current']; entry = tr['entry']
                    be_trigger_px = tr['be_trigger_px']; tp_px = tr['tp_px']; sl_d = tr['sl_d']
                    lot = tr['lot']; cs = tr['cs']; quote_conv = tr['quote_conv']
                    
                    closed = False
                    tp_dist = abs(tp_px - entry)
                    pv = SYMBOL_INFO[sym]['pip_value']
                    be_pts = sym_state.get('gann_be_trigger_points', 40)
                    atr_per = sym_state.get('gann_atr_period', 14)
                    cost_be = bot_state.get('prot_cost_be', True)
                    
                    if not tr['be_activated'] and be_trigger_px is not None:
                        # For BE trigger in backtest, we test against High for Buy, Low for Sell
                        test_px = h if is_buy else l
                        net_be = core_eval_break_even(is_buy, entry, test_px, pv, be_pts, atr_per, cost_be)
                        if net_be is not None:
                            tr['sl_current'] = net_be
                            tr['be_activated'] = True

                    # Outcome check uses h/l for extreme boundary testing
                    if is_buy:
                        if l <= sl_current:
                            tr['outcome'] = 'BREAK_EVEN' if sl_current > entry - 0.01 else 'LOSS'
                            tr['p_usd'] = round(abs(sl_current - entry) * lot * cs * quote_conv, 2) if tr['outcome'] == 'BREAK_EVEN' else -round(sl_d * lot * cs * quote_conv, 2)
                            closed = True
                        elif not closed and h >= tp_px:
                            tr['outcome'] = 'WIN'
                            tr['p_usd'] = round(tr['tp_d'] * lot * cs * quote_conv, 2)
                            closed = True
                    else:
                        if h >= sl_current:
                            tr['outcome'] = 'BREAK_EVEN' if sl_current < entry + 0.01 else 'LOSS'
                            tr['p_usd'] = round(abs(entry - sl_current) * lot * cs * quote_conv, 2) if tr['outcome'] == 'BREAK_EVEN' else -round(sl_d * lot * cs * quote_conv, 2)
                            closed = True
                        elif not closed and l <= tp_px:
                            tr['outcome'] = 'WIN'
                            tr['p_usd'] = round(tr['tp_d'] * lot * cs * quote_conv, 2)
                            closed = True
                            
                    if closed:
                        tr['close_time'] = t
                        daily_pl += tr['p_usd']
                        closed_trades.append(tr)
                    else:
                        surviving_trades.append(tr)
                open_trades = surviving_trades
            
            # Process Entries
            while signal_idx < total_signals and all_signals[signal_idx]['time'] <= t:
                sig = all_signals[signal_idx]
                signal_idx += 1
                if current_day not in suspended_days:
                    # DAM time-window filter (07:00-09:00, 13:00-14:00) --
                    # this backtest engine has its own signal-admission
                    # path and never calls is_trading_allowed()/_gann_open_
                    # trade (that gate only exists in the live engine), so
                    # the filter has to be re-applied here explicitly,
                    # checked against the SIGNAL's own historical
                    # timestamp rather than wall-clock time.
                    if bot_state.get('prot_dam_time_filter', True):
                        sig_dam_time = (sig['time'] + timedelta(hours=3)).time()
                        if any(start <= sig_dam_time < end for start, end in _DAM_RESTRICTED_WINDOWS):
                            continue
                    # Max concurrent trades cap -- mirrors the live bot's
                    # prot_max_concurrent_trades. Without this, the backtest
                    # can open one trade per enabled timeframe off a single
                    # level touch with no limit (the exact multi-tf stacking
                    # that caused real losses live), which is NOT what the
                    # live bot actually does anymore, and inflates backtest
                    # trade counts relative to what live can ever produce.
                    max_concurrent_bt = int(bot_state.get('prot_max_concurrent_trades', 4))
                    open_count_bt = sum(1 for tr in open_trades if tr['symbol'] == sig['symbol'])
                    if open_count_bt >= max_concurrent_bt:
                        continue
                    sig['sl_current'] = sig['sl_px']
                    sig['be_activated'] = False
                    open_trades.append(sig)
                    
            await prog.tick(i, res['win'], res['loss'], res['be'], res['total_prof'])
            
        # Post-process closed trades to match old format
        c_log(f'BT: Post-processing {len(closed_trades)} closed trades')
        for tr in closed_trades:
            if tr['outcome'] == 'WIN' or (tr['outcome'] == 'DAILY_LIMIT' and tr['p_usd'] > 0): 
                res['win'] += 1; res['total_win_usd'] += tr['p_usd']
            elif tr['outcome'] == 'LOSS' or (tr['outcome'] == 'DAILY_LIMIT' and tr['p_usd'] < 0): 
                res['loss'] += 1; res['total_loss_usd'] += abs(tr['p_usd'])
            elif tr['outcome'] == 'BREAK_EVEN' or (tr['outcome'] == 'DAILY_LIMIT' and tr['p_usd'] == 0): 
                res['be'] += 1
            
            res['total_prof'] += tr['p_usd']
            dir_str = 'BUY 📈' if tr['is_buy'] else 'SELL 📉'
            res['trade_logs'].append({
                'الزوج': tr['symbol'], 
                'وقت الصفقة (DAM)': _utc_to_dam(tr['time']).strftime('%Y-%m-%d %H:%M'),
                'TF': tr['tf'], 
                'اتجاه': dir_str, 
                'المستوى (الدخول)': f"{tr['entry']:.2f} ({tr['level_key']})",
                'الهدف (TP)': round(tr['tp_px'], 2),
                'الوقف (SL)': round(tr['sl_px'], 2),
                'النتيجة': tr['outcome'], 
                'ربح ($)': tr['p_usd'],
                'cycle_ts': tr['cycle_time'].timestamp(),
                'cycle_time_str': _utc_to_dam(tr['cycle_time']).strftime('%Y-%m-%d %H:%M'),
                'cycle_close': tr['cycle_close'],
                'trigger_type': tr.get('trigger_type', 'touch'),
            })
            
        res['trade_logs'].sort(key=lambda x: x['وقت الصفقة (DAM)'])
        
        running_eq = 5000.0
        peak_eq = 5000.0
        max_dd = 0.0
        for t_log in res['trade_logs']:
            running_eq += t_log['ربح ($)']
            t_log['رصيد تراكمي ($)'] = round(running_eq, 2)
            if running_eq > peak_eq: peak_eq = running_eq
            dd = peak_eq - running_eq
            if dd > max_dd: max_dd = dd
            
        res['peak_equity'] = peak_eq
        res['max_dd'] = max_dd

        if not res['trade_logs']:
            await prog.done('<b>باكتيست اكتمل ✅</b>\nلا توجد صفقات في هذا النطاق.')
            bot_state['is_backtesting'] = False; return

        await prog.set_phase('إنشاء ملف Excel المنسق...')
        
        c_log('BT: Generating Excel')
        
        sum_text = (
            f"<b>باكتيست جان اكتمل ✅</b>\n"
            f"{syms_label} H1→[{desc_tfs}] | {desc_mode} | {desc_star}{desc_be}\n"
            f"{start_dt.strftime('%Y-%m-%d')} → {end_dt.strftime('%Y-%m-%d')}\n\n"
            f"Net: {'PROFIT ▲' if res['total_prof']>=0 else 'LOSS ▼'} ${round(res['total_prof'], 2)}\n"
            f"Win:  +${round(res['total_win_usd'], 2)} ({res['win']})\n"
            f"Loss: -${round(res['total_loss_usd'], 2)} ({res['loss']})\n"
            f"Break-Even: $0 ({res['be']})\n"
            f"WR: {round(res['win']/max(1, res['win']+res['loss'])*100)}% ({len(res['trade_logs'])} صفقة)\n"
            f"Max DD: ${round(res['max_dd'],2)} ({round((res['max_dd']/max(1,res['peak_equity']))*100)}%)\n"
        )
        
        if suspended_days:
            sum_text += "\nالتعليق بسبب حماية رأس المال:\n"
            for d_str, rsn in suspended_days.items():
                sum_text += f"- {d_str}: {rsn}\n"
                
        sum_text += f"\nدورات H1: {len(res['cycle_logs'])}  |  TP/SL: {str('ATR' if tpsl_mode=='atr' else 'نقاط ثابتة')} | Lot: {lot}"

        wb = openpyxl.Workbook()
        ws_trades = wb.active
        ws_trades.title = "الصفقات"
        
        headers = ["الزوج", "وقت الصفقة (DAM)", "TF", "اتجاه", "المستوى (الدخول)", "الهدف (TP)", "الوقف (SL)", "النتيجة", "ربح ($)", "رصيد تراكمي ($)"]
        ws_trades.append(headers)
        
        gray_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
        header_fill = PatternFill(start_color="E2E3E5", end_color="E2E3E5", fill_type="solid")
        red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        be_fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")
        
        for cell in ws_trades[1]:
            cell.fill = gray_fill
            cell.font = Font(bold=True)
            
        current_cycle = None
        for tr in res['trade_logs']:
            if tr['cycle_ts'] != current_cycle:
                current_cycle = tr['cycle_ts']
                ws_trades.append([f"دورة {_anchor_label()}: {tr['cycle_time_str']}  |  إغلاق {_anchor_label()}: {tr['cycle_close']:.2f}"] + [""]*9)
                row_idx = ws_trades.max_row
                ws_trades.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=10)
                ws_trades.cell(row=row_idx, column=1).fill = header_fill
                ws_trades.cell(row=row_idx, column=1).font = Font(bold=True)
                
            _OUTCOME_DISPLAY = {'WIN': 'WIN ✅', 'LOSS': 'LOSS ❌', 'BREAK_EVEN': 'BREAK_EVEN ⚖️', 'DAILY_LIMIT': 'DAILY_LIMIT ⏹️'}
            row_data = [
                tr['الزوج'], tr['وقت الصفقة (DAM)'], tr['TF'], tr['اتجاه'], tr['المستوى (الدخول)'],
                tr['الهدف (TP)'], tr['الوقف (SL)'], _OUTCOME_DISPLAY.get(tr['النتيجة'], tr['النتيجة']), tr['ربح ($)'], tr['رصيد تراكمي ($)']
            ]
            ws_trades.append(row_data)
            row_idx = ws_trades.max_row
            
            fill = None
            if tr['النتيجة'] == 'WIN': fill = green_fill
            elif tr['النتيجة'] == 'LOSS': fill = red_fill
            elif tr['النتيجة'] == 'BREAK_EVEN': fill = be_fill
            
            if fill:
                for col in range(1, 11):
                    ws_trades.cell(row=row_idx, column=col).fill = fill

        ws_cycles = wb.create_sheet("دورات H1")
        ws_cycles.append(["الزوج", "الدورة (DAM)", f"إغلاق {_anchor_label()}", "عدد الصفقات", "ملاحظة"])
        for cell in ws_cycles[1]: cell.fill = gray_fill; cell.font = Font(bold=True)
        
        for cycle in res['cycle_logs']:
            num_trades = len([t for t in res['trade_logs'] if t['cycle_ts'] == cycle['time_ts']])
            cycle_day = _utc_to_dam(cycle['time_dt']).strftime('%Y-%m-%d')
            if num_trades > 0:
                note = f"تم تنفيذ {num_trades} صفقة"
            elif cycle_day in suspend_trigger_time and cycle['time_dt'] >= suspend_trigger_time[cycle_day]:
                # Distinguish "day was already halted by capital protection"
                # from "price genuinely never reached a level" -- these are
                # very different situations and were previously reported
                # identically, which made it look like the strategy just
                # wasn't triggering when actually trading had stopped.
                note = "🛑 اليوم متوقف (تم تفعيل حماية رأس المال مسبقاً)"
            else:
                note = "لم يلمس السعر أي مستوى"
            ws_cycles.append([cycle['symbol'], _utc_to_dam(cycle['time_dt']).strftime('%Y-%m-%d %H:%M'), cycle['close'], num_trades, note])
            
        ws_susp = wb.create_sheet("أيام الإيقاف")
        ws_susp.append(["التاريخ", "السبب (النتيجة)"])
        for cell in ws_susp[1]: cell.fill = gray_fill; cell.font = Font(bold=True)
        for d_str, rsn in suspended_days.items():
            ws_susp.append([d_str, rsn])
            

        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        center_align = Alignment(horizontal='center', vertical='center')

        for ws in [ws_trades, ws_cycles, ws_susp]:
            for row in ws.iter_rows():
                for cell in row:
                    cell.border = thin_border
                    cell.alignment = center_align

        from openpyxl.utils import get_column_letter
        for i in range(1, 11): ws_trades.column_dimensions[get_column_letter(i)].width = 22.0
        for i in range(1, 6): ws_cycles.column_dimensions[get_column_letter(i)].width = 22.0
        for i in range(1, 3): ws_susp.column_dimensions[get_column_letter(i)].width = 25.0

        if exec_mode == 'all_concurrent':
            _add_concurrent_analysis_sheets(wb, res['trade_logs'], pnl_key='ربح ($)', outcome_key='النتيجة')

        wb.save(fname)
        
        await prog.done(f'<b>باكتيست جان اكتمل ✅</b>\n{syms_label} — {len(res["trade_logs"])} صفقة\nجاري إرسال التقرير والملف...')
        await send_tg_document(fname, sum_text)
        os.remove(fname)

    except Exception as e:
        c_log(f'BT Error: {e}'); bot_state['is_backtesting'] = False
        if _bt_progress:
            import html
            try: await _bt_progress.done(f'❌ خطأ داخلي في الباكتيست:\n{html.escape(str(e))}')
            except Exception as inner_e: log_exception('backtest error notification', inner_e)
    finally:
        bot_state['is_backtesting'] = False

# ═════════════════════════════════════════════════════════════
# LIVE-TWIN ENGINE — realistic execution simulator
# ═════════════════════════════════════════════════════════════
# Replaces run_gann_backtest's zero-friction assumption (perfect fill
# at the exact level price, no cost, no ambiguity about which of
# SL/TP was touched first) with a market-friction model: dynamic
# spread, asymmetric slippage, signal-to-fill latency, commission/
# swap, weekend gap risk, and a Brownian-bridge intrabar path used
# only to resolve SL-vs-TP ordering when a single 1m bar's range
# contains both (OHLC alone can't answer that; assuming the worst
# or the best every time is its own bias, so we reconstruct a
# plausible-but-randomized path instead).
#
# run_gann_backtest is left completely untouched and reachable from
# its own menu -- with lt_mode == 'idealized' this engine calls it
# directly, so it doubles as the zero-friction A/B baseline.
#
# Spread baseline is hardcoded from a live MT5/OANDA XAUUSD tick
# snapshot taken 2026-07-13 in the late-night/low-liquidity session:
#   Bid 4112.28 / Ask 4112.62 -> 0.34 USD (34 points @ tick size 0.01)
# That reading IS the quiet-session floor. Every multiplier below is
# defined as a ratio against it, never as an independent guess:
#   - Asian / dead-zone hours (where the snapshot was taken): 1.00x
#   - London session                                         : 0.70x
#   - London/NY overlap (deepest liquidity)                  : 0.55x
#   - NY session (post-overlap)                              : 0.75x
#   - Broker rollover window (21:55-22:05 UTC)                : up to 3.5x
#   - High-ATR bars (volatility spike, stacks on top of session): up to +2.5x more
# ═════════════════════════════════════════════════════════════

def _lt_session_multiplier(dt_utc: datetime) -> tuple[float, bool]:
    """Returns (spread_multiplier, is_rollover) for a UTC timestamp."""
    hm = dt_utc.hour + dt_utc.minute / 60.0
    if (21 + 55/60) <= hm <= (22 + 5/60):
        return 3.5, True                # broker rollover window -- spreads spike hard
    if 12.0 <= hm < 16.0:
        return 0.55, False              # London/NY overlap -- tightest liquidity
    if 7.0 <= hm < 12.0:
        return 0.70, False              # London session
    if 16.0 <= hm < 20.0:
        return 0.75, False              # NY session (post-overlap)
    return 1.00, False                  # Asian / dead-zone -- matches the live snapshot session


def _lt_volatility_multiplier(bar_range: float, atr_val: float | None) -> float:
    """Extra spread widening when a bar's range blows past its recent ATR."""
    if not atr_val or atr_val <= 0:
        return 1.0
    ratio = bar_range / atr_val
    if ratio <= 1.2:
        return 1.0
    return min(1.0 + (ratio - 1.2) * 0.9, 3.5)


def _lt_current_spread(base_spread: float, dt_utc: datetime, bar_range: float, atr_val: float | None) -> tuple[float, bool]:
    sess_mult, is_rollover = _lt_session_multiplier(dt_utc)
    vol_mult = _lt_volatility_multiplier(bar_range, atr_val)
    spread = base_spread * sess_mult * vol_mult
    return max(spread, base_spread * 0.45), is_rollover


def _lt_bridge_path(o: float, h: float, l: float, c: float, steps: int, rng: random.Random) -> np.ndarray:
    """
    Reconstructs a plausible intrabar tick path as a scaled Brownian
    bridge from open to close, clipped into [low, high]. Used only to
    decide the ORDER in which SL/TP thresholds would have been crossed
    inside a bar where raw OHLC can't tell us -- not claimed as the
    literal historical tick path, just a principled stand-in for one.
    """
    incs = np.array([rng.gauss(0, 1) for _ in range(steps)])
    w = np.concatenate(([0.0], np.cumsum(incs)))
    t = np.linspace(0.0, 1.0, steps + 1)
    bridge = w - t * w[-1]
    bstd = float(np.std(bridge))
    rng_size = max(h - l, 1e-6)
    scale = (rng_size * 0.5) / bstd if bstd > 1e-9 else 0.0
    path = o + (c - o) * t + bridge * scale
    return np.clip(path, l, h)


def _lt_first_hit(path: np.ndarray, is_buy: bool, sl_px: float, tp_px: float) -> str | None:
    """Walk the reconstructed path and return which of 'sl'/'tp' is crossed first, or None."""
    for px in path:
        if is_buy:
            if px <= sl_px: return 'sl'
            if px >= tp_px: return 'tp'
        else:
            if px >= sl_px: return 'sl'
            if px <= tp_px: return 'tp'
    return None


def _lt_slippage(bar_range: float, atr_val: float | None, rng: random.Random) -> float:
    """Asymmetric, range-scaled slippage magnitude -- always adverse (models cost, not luck)."""
    ref = atr_val if atr_val and atr_val > 0 else max(bar_range, 0.05)
    base = ref * 0.06
    tail = abs(rng.gauss(0, base))
    return min(tail, ref * 0.5)


def _lt_latency_shift(path: np.ndarray, steps: int, rng: random.Random) -> float:
    """Signal-to-fill delay expressed as a fractional shift along the intrabar path.
    Bounds default to a rough guess (200-800ms) but should be set from the bot's
    OWN measured MetaApi ping (see the 'Code Delay'/'MetaApi Ping' fields logged on
    every real fill) via bot_state['lt_latency_ms_min']/['lt_latency_ms_max'] so the
    simulation reflects this account's actual broker/network latency, not a guess."""
    lo = bot_state.get('lt_latency_ms_min', 160)
    hi = bot_state.get('lt_latency_ms_max', 200)
    latency_ms = rng.randint(lo, hi)
    frac = min(latency_ms / 60000.0, 1.0)  # fraction of a 1-minute bar consumed by the delay
    idx = min(int(frac * steps), steps)
    return float(path[idx])


async def run_live_twin_simulation(start_dt: datetime, end_dt: datetime) -> None:
    """Realistic-execution counterpart to run_gann_backtest. Same signal
    logic (Gann level touches + trend filter), but fills, spreads,
    slippage, latency, commission, swap, and SL/TP ordering are all run
    through the friction model above instead of assumed perfect."""
    global _lt_progress
    if bot_state.get('lt_mode') == 'idealized':
        # A/B baseline: reuse the existing zero-friction engine untouched.
        await run_gann_backtest(start_dt, end_dt)
        return

    bot_state['is_live_twin_running'] = True
    fname = f"LiveTwin_{datetime.now(timezone.utc).strftime('%H%M%S')}.xlsx"
    exec_mode = bot_state.get('gann_execution_mode', 'instant')  # hoisted: must exist even if no trades are ever scanned
    fric = bot_state['lt_friction']
    base_spread = float(bot_state['lt_base_spread_usd'])
    comm_per_lot = float(bot_state['lt_commission_per_lot'])
    swap_long_per_lot = float(bot_state.get('lt_swap_long_per_lot_night', -93.17))
    swap_short_per_lot = float(bot_state.get('lt_swap_short_per_lot_night', 21.68))
    swap_wed_mult = float(bot_state.get('lt_swap_wednesday_multiplier', 3.0))
    rej_prob = float(bot_state['lt_rejection_prob'])

    active_symbols = [s for s, on in bot_state['active_symbols'].items() if on]
    if not active_symbols:
        bot_state['is_live_twin_running'] = False
        return

    # Deterministic seed: identical config (symbols + date range + exec mode +
    # friction toggles) => identical random slippage/latency/rejection draws
    # every run. Before this, rng = random.Random() re-seeded from system
    # entropy/time on every call, so re-running the SAME backtest with NO
    # settings changed could still swing net P&L wildly (pure noise, not signal)
    # -- that alone can dwarf real differences between exec modes or code fixes.
    # Set bot_state['lt_seed'] to an int to override; None keeps this auto-seed.
    override_seed = bot_state.get('lt_seed')
    if override_seed is not None:
        seed_val = int(override_seed)
    else:
        seed_key = (tuple(sorted(active_symbols)), start_dt.isoformat(), end_dt.isoformat(),
                    bot_state.get('gann_execution_mode', 'instant'), tuple(sorted(fric.items())))
        seed_val = zlib.crc32(str(seed_key).encode())
    rng = random.Random(seed_val)

    first_sym_state = bot_state['symbol_state'][active_symbols[0]]
    enabled_tfs = [tf for tf, on in first_sym_state['gann_monitor_tfs'].items() if on] or ['5m']
    flt_type = first_sym_state['trend_filter_type']
    ttf = first_sym_state['trend_timeframe']
    syms_label = "+".join(active_symbols)
    on_tags = "+".join(k for k, v in fric.items() if v) or "none"

    prog = BtProgress(label=f"Live-Twin {syms_label} | friction:[{on_tags}]", active_tfs=['H1'])
    _lt_progress = prog
    await prog.start(bot_state['chat_id'])

    res = {'win': 0, 'loss': 0, 'be': 0, 'total_prof': 0.0, 'total_win_usd': 0.0, 'total_loss_usd': 0.0,
           'peak_equity': 0.0, 'max_dd': 0.0, 'trade_logs': [], 'total_commission': 0.0, 'total_swap': 0.0,
           'rejected': 0, 'gap_events': 0}

    try:
        delta_hours = int((end_dt - start_dt).total_seconds() / 3600)
        all_signals = []
        m1_by_symbol = {}

        # ── PHASE 1: signal generation (identical strategy logic to the idealized engine) ──
        for symbol in active_symbols:
            sym_state = bot_state['symbol_state'][symbol]
            cycle_h = sym_state['gann_cycle_hours']; tpsl_mode = sym_state['gann_tpsl_mode']
            pv = SYMBOL_INFO[symbol]['pip_value']; lot = sym_state['lot_size']; margin = sym_state['gann_touch_margin_pts'] * pv
            cs = SYMBOL_INFO[symbol]['contract_size']

            quote = symbol.split('_')[1] if '_' in symbol else 'USD'
            quote_conv = {'USD': 1.0, 'JPY': 1/150.0, 'AUD': 0.66, 'NZD': 0.61, 'EUR': 1.08, 'GBP': 1.27, 'CAD': 0.73, 'CHF': 1.11}.get(quote, 1.0)

            await prog.set_phase(f'جلب بيانات الترند ({ttf.upper()})...')
            max_period = max(sym_state['trend_vwap_period'], sym_state['trend_ema_period'], 100)
            trend_count = (delta_hours * (2 if ttf == '30m' else 1)) + max_period + 10
            candles_trend = await fetch_candles(symbol, ttf, count=trend_count, end_time=end_dt)
            if not candles_trend: continue

            df_trend = pd.DataFrame(candles_trend)
            if flt_type == 'vwap':
                p_vwap = sym_state['trend_vwap_period']
                df_trend['Typical_Price'] = (df_trend['high'] + df_trend['low'] + df_trend['close']) / 3
                df_trend['VWAP'] = (df_trend['Typical_Price'] * df_trend['volume']).rolling(window=p_vwap).sum() / df_trend['volume'].rolling(window=p_vwap).sum()
            if flt_type == 'ema':
                p_ema = sym_state['trend_ema_period']
                df_trend['EMA'] = df_trend['close'].ewm(span=p_ema, adjust=False).mean()

            df_trend.set_index('time', inplace=True)
            if flt_type == 'vwap': df_trend['macro_trend_up'] = df_trend['close'] > df_trend['VWAP']
            elif flt_type == 'ema': df_trend['macro_trend_up'] = df_trend['close'] > df_trend['EMA']
            elif flt_type == 'both':
                c1_up = df_trend['close'] > df_trend['VWAP']; c2_up = df_trend['close'] > df_trend['EMA']
                c1_dn = df_trend['close'] < df_trend['VWAP']; c2_dn = df_trend['close'] < df_trend['EMA']
                df_trend['macro_trend_up'] = np.where(c1_up & c2_up, True, np.where(c1_dn & c2_dn, False, None))

            anchor_gran = bot_state.get('gann_anchor_tf', '1h')
            await prog.set_phase(f'جلب بيانات {_anchor_label()}...')
            candles_h1 = await fetch_candles(symbol, anchor_gran, count=(delta_hours // _anchor_hours()) + 10, end_time=end_dt)
            if not candles_h1: continue

            await prog.set_phase('جلب شموع الدقيقة الواحدة (تنفيذ واقعي)...')
            days_diff = (end_dt - start_dt).days or 1
            need_1m = days_diff * 24 * 60 + 300
            mc_1m = await fetch_candles(symbol, '1m', count=need_1m, end_time=end_dt)
            if not mc_1m: continue
            m1_by_symbol[symbol] = sorted(mc_1m, key=lambda c: c['time'])

            monitor_tfs_data = {}
            for btf in enabled_tfs:
                bmin = int(''.join(filter(str.isdigit, btf))); bmin = bmin * 60 if 'h' in btf else bmin
                need_m = days_diff * 24 * (60 // max(bmin, 1)) + 300
                mc = await fetch_candles(symbol, btf, count=need_m, end_time=end_dt)
                if mc: monitor_tfs_data[btf] = sorted(mc, key=lambda c: c['time'])

            start_ts = start_dt.timestamp(); end_ts = end_dt.timestamp()
            valid_h1 = [c for c in candles_h1 if start_ts <= (c['time'].timestamp() + 3600) <= end_ts]
            trend_freq = '30min' if ttf == '30m' else '1h'

            # Same shared builder as run_gann_backtest -- required so a
            # 'dynamic_live' backtest here matches the idealized engine and
            # the live scanner exactly instead of drifting from them.
            cycle_defs = _build_gann_cycle_defs(sym_state, valid_h1, mc_1m)

            for cdef in cycle_defs:
                if prog.cancelled: return
                await asyncio.sleep(0)
                t_start = cdef['t_start']
                t_end = cdef['t_end']
                close = cdef['close']
                levels = gann_calc_levels(symbol, close)
                f_mode = sym_state['gann_zone_filter']
                active_lv = [l for l in levels if l['dir'] != 'ref' and (f_mode == 'all' or (f_mode == 'star' and l['star']) or (f_mode == 'star_fan' and (l['star'] or l['fan'])))]
                level_used = set()

                for btf, candles_m in monitor_tfs_data.items():
                    m_window = [c for c in candles_m if t_start <= c['time'] < t_end]
                    m_before = [c for c in candles_m if c['time'] < t_start]
                    atr_val = _gann_atr(m_before, sym_state['gann_atr_period']) if tpsl_mode == 'atr' else None
                    # Execution-mode state: 'prev_bar_close' tracks bar-over-bar
                    # momentum for hybrid's spike check, since OHLC data has no
                    # sub-bar tick stream to check live momentum against.
                    prev_bar_close = float(m_before[-1]['close']) if m_before else None
                    exec_mode = bot_state.get('gann_execution_mode', 'instant')
                    spike_limit = bot_state.get('gann_spike_limit_pts', 20) * pv

                    for bar in m_window:
                        bar_open = float(bar['open']); bar_close = float(bar['close']); bar_time = bar['time']
                        bar_high = float(bar['high']); bar_low = float(bar['low'])
                        trend_up = True
                        if sym_state['gann_entry_mode'] == 'touch_trend':
                            trend_time = bar_time.floor(trend_freq)
                            if trend_time in df_trend.index:
                                val = df_trend.loc[trend_time, 'macro_trend_up']
                                if isinstance(val, pd.Series): val = val.iloc[-1]
                                macro_trend_up = None if pd.isna(val) else bool(val)
                            else: macro_trend_up = None
                            if macro_trend_up is None:
                                prev_bar_close = bar_close; continue
                            trend_up = macro_trend_up

                        if bot_state.get('prot_cycle_inval', True):
                            inval_pts = bot_state.get('prot_cycle_inval_pts', 200) * pv
                            if abs(bar_close - close) > inval_pts:
                                active_lv = []; break

                        for lv in active_lv:
                            k = lv['key']; dir = lv['dir']
                            is_buy = (dir == 'dn')
                            if sym_state['gann_entry_mode'] == 'touch_trend':
                                if is_buy and not trend_up: continue
                                if not is_buy and trend_up: continue

                            # ── Execution-mode gate (mirrors the live scanner) ──
                            # close   : only the bar's CLOSE has to be within margin
                            #           (this was the engine's only behavior before
                            #           execution modes existed, and stays the default).
                            # instant : any part of the bar's range (intrabar high/low,
                            #           not just its close) touching the level counts --
                            #           a live tick could have fired mid-bar.
                            # hybrid  : same intrabar touch as instant, but rejected if
                            #           this bar's close has already moved more than the
                            #           spike limit from the PREVIOUS bar's close (the
                            #           backtest's only available proxy for "live_px ran
                            #           away from the last print" since OHLC has no
                            #           sub-bar ticks to check momentum against directly).
                            # all_concurrent: check all three independently, each with
                            #           its own dedup key, so a touch/close/hybrid can
                            #           each fire their own trade on the same level.
                            channels = ['touch', 'close', 'hybrid'] if exec_mode == 'all_concurrent' else [
                                'close' if exec_mode == 'close' else 'hybrid' if exec_mode == 'hybrid' else 'touch']

                            for channel in channels:
                                base_combo = f'{k}_{btf}' if bot_state.get('prot_allow_multi_tf', True) else k
                                combo_key = f"{base_combo}_{channel}" if exec_mode == 'all_concurrent' else base_combo
                                if combo_key in level_used: continue

                                if channel == 'close':
                                    if abs(bar_close - lv['price']) > margin: continue
                                elif channel == 'hybrid':
                                    if not (bar_low - margin <= lv['price'] <= bar_high + margin): continue
                                    if prev_bar_close is not None and abs(bar_close - prev_bar_close) > spike_limit: continue
                                else:  # touch (instant)
                                    if not (bar_low - margin <= lv['price'] <= bar_high + margin): continue

                                entry = lv['price']
                                tf_tp = _gann_tf_tp(symbol, btf); tf_sl = _gann_tf_sl(symbol, btf)
                                if tpsl_mode == 'atr' and atr_val:
                                    sl_d = atr_val * sym_state['gann_atr_sl_mult']; tp_d = atr_val * sym_state['gann_atr_tp_mult']
                                else:
                                    sl_d = tf_sl * pv; tp_d = tf_tp * pv

                                all_signals.append({
                                    'time': bar_time, 'symbol': symbol, 'is_buy': is_buy, 'intended_entry': entry,
                                    'sl_d': sl_d, 'tp_d': tp_d, 'be_enabled': sym_state['break_even_enabled'],
                                    'lot': lot, 'cs': cs, 'quote_conv': quote_conv, 'tf': btf, 'combo_key': combo_key,
                                    'cycle_time': t_start, 'cycle_close': close, 'level_key': k, 'trigger_type': channel,
                                    'bar_o': bar_open, 'bar_h': bar_high, 'bar_l': bar_low, 'bar_c': bar_close,
                                })
                                level_used.add(combo_key)

                        prev_bar_close = bar_close

        # ── PHASE 2: chronological, friction-aware, 1-minute-bar simulation ──
        await prog.set_phase('محاكاة التنفيذ الواقعي (سبريد/انزلاق/تأخير/عمولة)...')
        all_signals.sort(key=lambda x: x['time'])
        all_1m_events = sorted(
            [{'time': c['time'], 'symbol': sym, 'open': float(c['open']), 'high': float(c['high']),
              'low': float(c['low']), 'close': float(c['close'])}
             for sym, candles in m1_by_symbol.items() for c in candles],
            key=lambda x: x['time']
        )
        m1_lookup = {
            sym: {c['time']: {'open': float(c['open']), 'high': float(c['high']), 'low': float(c['low']), 'close': float(c['close'])}
                  for c in candles}
            for sym, candles in m1_by_symbol.items()
        }

        open_trades = []
        closed_trades = []
        suspended_days = {}
        suspend_trigger_time = {}
        daily_pl = 0.0
        current_day = None
        latest_price = {}

        signal_idx = 0
        total_signals = len(all_signals)
        dd_limit = -float(bot_state['prot_daily_dd_usd'])
        profit_limit = float(bot_state['prot_daily_profit_usd'])
        max_concurrent = int(bot_state.get('prot_max_concurrent_trades', 4))

        total_events = len(all_1m_events)
        await prog.set_tf('محاكاة 1m واقعية', total_events)

        for i, ev in enumerate(all_1m_events):
            if i % 5000 == 0: await asyncio.sleep(0)
            if prog.cancelled: break
            t = ev['time']; sym = ev['symbol']; o = ev['open']; h = ev['high']; l = ev['low']; c = ev['close']
            bar_range = h - l
            day_str = _utc_to_dam(t).strftime('%Y-%m-%d')
            latest_price[sym] = c
            if day_str != current_day:
                current_day = day_str; daily_pl = 0.0

            atr_ref = bar_range if bar_range > 0 else 0.1
            spread_now, is_rollover = _lt_current_spread(base_spread, t, bar_range, atr_ref) if fric['spread'] else (base_spread, False)
            half_spread = spread_now / 2.0

            if fric['gaps'] and is_rollover:
                res['gap_events'] += 1

            # -- capital-protection daily limit check (worst-case intrabar) --
            if current_day not in suspended_days:
                floating_pl = 0.0; floating_pl_worst = 0.0
                for tr in open_trades:
                    lp = latest_price.get(tr['symbol'], tr['entry'])
                    diff = (lp - tr['entry']) if tr['is_buy'] else (tr['entry'] - lp)
                    floating_pl += round(diff * tr['lot'] * tr['cs'] * tr['quote_conv'], 2)
                    worst_px = (l if tr['is_buy'] else h) if tr['symbol'] == sym else lp
                    diff_worst = (worst_px - tr['entry']) if tr['is_buy'] else (tr['entry'] - worst_px)
                    floating_pl_worst += round(diff_worst * tr['lot'] * tr['cs'] * tr['quote_conv'], 2)

                total_daily = daily_pl + floating_pl
                total_daily_worst = daily_pl + floating_pl_worst
                if dd_limit < 0 and total_daily_worst <= dd_limit:
                    suspended_days[current_day] = f'🛑 تراجع عائم (الحد {dd_limit}$)'
                elif profit_limit > 0 and total_daily >= profit_limit:
                    suspended_days[current_day] = f'✅ هدف عائم (الحد {profit_limit}$)'
                if current_day in suspended_days and current_day not in suspend_trigger_time:
                    suspend_trigger_time[current_day] = t
                if current_day in suspended_days:
                    for tr in open_trades:
                        lp = (l if tr['is_buy'] else h) if tr['symbol'] == sym else latest_price.get(tr['symbol'], tr['entry'])
                        exit_spread = _lt_current_spread(base_spread, t, bar_range, atr_ref)[0] if fric['spread'] else base_spread
                        lp_adj = lp - (exit_spread/2.0 if tr['is_buy'] else -exit_spread/2.0) if fric['spread'] else lp
                        diff = (lp_adj - tr['entry']) if tr['is_buy'] else (tr['entry'] - lp_adj)
                        p_usd = round(diff * tr['lot'] * tr['cs'] * tr['quote_conv'], 2)
                        tr['outcome'] = 'DAILY_LIMIT'; tr['p_usd'] = p_usd; tr['close_time'] = t
                        closed_trades.append(tr); daily_pl += p_usd
                    open_trades.clear()

            # -- exits: reconstruct an intrabar path per open trade on this symbol's bar --
            if current_day not in suspended_days:
                surviving = []
                for tr in open_trades:
                    if tr['symbol'] != sym:
                        surviving.append(tr); continue

                    is_buy = tr['is_buy']; entry = tr['entry']; sl_current = tr['sl_current']; tp_px = tr['tp_px']
                    lot = tr['lot']; cs = tr['cs']; quote_conv = tr['quote_conv']
                    closed = False

                    # Break-even arm (tested against the bar's favorable extreme, same as before)
                    if tr['be_enabled'] and not tr['be_activated']:
                        test_px = h if is_buy else l
                        be_pts = bot_state['symbol_state'][sym].get('gann_be_trigger_points', 40)
                        pv_sym = SYMBOL_INFO[sym]['pip_value']
                        atr_per = bot_state['symbol_state'][sym].get('gann_atr_period', 14)
                        cost_be = bot_state.get('prot_cost_be', True)
                        net_be = core_eval_break_even(is_buy, entry, test_px, pv_sym, be_pts, atr_per, cost_be)
                        if net_be is not None:
                            tr['sl_current'] = net_be; tr['be_activated'] = True; sl_current = net_be

                    hits_sl = (l <= sl_current) if is_buy else (h >= sl_current)
                    hits_tp = (h >= tp_px) if is_buy else (l <= tp_px)

                    outcome = None
                    if hits_sl and hits_tp:
                        # Ambiguous bar -- reconstruct a plausible path instead of always assuming one side.
                        path = _lt_bridge_path(o, h, l, c, steps=20, rng=rng)
                        outcome = _lt_first_hit(path, is_buy, sl_current, tp_px) or 'sl'
                    elif hits_sl:
                        outcome = 'sl'
                    elif hits_tp:
                        outcome = 'tp'

                    if outcome:
                        exit_spread = spread_now if fric['spread'] else 0.0
                        slip = _lt_slippage(bar_range, atr_ref, rng) if fric['slippage'] else 0.0
                        if outcome == 'sl':
                            raw_px = sl_current
                            fill_px = raw_px - (exit_spread/2.0 + slip) if is_buy else raw_px + (exit_spread/2.0 + slip)
                            tr['outcome'] = 'BREAK_EVEN' if sl_current > entry - 0.01 and is_buy or (not is_buy and sl_current < entry + 0.01) else 'LOSS'
                        else:
                            raw_px = tp_px
                            fill_px = raw_px - (exit_spread/2.0 + slip) if is_buy else raw_px + (exit_spread/2.0 + slip)
                            tr['outcome'] = 'WIN'
                        diff = (fill_px - entry) if is_buy else (entry - fill_px)
                        p_usd = round(diff * lot * cs * quote_conv, 2)
                        commission = comm_per_lot * lot if fric['commission'] else 0.0
                        nights = max((t.date() - tr['time'].date()).days, 0)
                        swap = 0.0
                        if fric['gaps'] and nights > 0:
                            per_night = swap_long_per_lot if is_buy else swap_short_per_lot
                            # Each night held may itself be a Wednesday (tripled) or not --
                            # walk the actual calendar days rather than assuming a flat rate.
                            for i in range(nights):
                                d = tr['time'].date() + timedelta(days=i)
                                mult = swap_wed_mult if d.weekday() == 2 else 1.0  # Monday=0 .. Wednesday=2
                                swap += per_night * mult
                            swap *= lot
                        p_usd_net = round(p_usd - commission + swap, 2)
                        if tr['outcome'] == 'WIN' and p_usd_net < 0:
                            tr['outcome'] = 'LOSS'  # friction ate the whole win -- report it honestly
                        tr['p_usd'] = p_usd_net
                        res['total_commission'] += commission; res['total_swap'] += swap
                        tr['close_time'] = t; daily_pl += p_usd_net
                        closed_trades.append(tr)
                        closed = True
                    if not closed:
                        surviving.append(tr)
                open_trades = surviving

            # -- entries --
            # NOTE: all_1m_events interleaves every active symbol's 1m bars
            # chronologically, so the loop's *current* sym/o/h/l/c belong to
            # whichever symbol's bar happens to land at this timestamp --
            # NOT necessarily the signal's own symbol. Entry fills must be
            # priced off the signal's OWN symbol's bar via m1_lookup, never
            # off the loop's current bar, or cross-symbol signals get
            # silently dropped/mispriced whenever two symbols interleave.
            while signal_idx < total_signals and all_signals[signal_idx]['time'] <= t:
                sig = all_signals[signal_idx]; signal_idx += 1
                if current_day in suspended_days:
                    continue
                if bot_state.get('prot_dam_time_filter', True):
                    sig_dam_time = (sig['time'] + timedelta(hours=3)).time()
                    if any(start <= sig_dam_time < end for start, end in _DAM_RESTRICTED_WINDOWS):
                        continue
                open_count = sum(1 for tr in open_trades if tr['symbol'] == sig['symbol'])
                if open_count >= max_concurrent:
                    continue
                if fric['rejection'] and rng.random() < rej_prob:
                    res['rejected'] += 1; continue

                # NOTE: previously this looked up m1_lookup[symbol][sig['time']],
                # i.e. only the FIRST 1-minute slice of the signal's own bar --
                # for a 5m/3m/2m signal that silently truncated the real bar
                # range down to a 1-minute one, understating slippage sizing
                # and shortening the reconstructed intrabar path. The touched
                # bar's true OHLC is now carried on the signal itself from
                # Phase 1, so use that directly.
                so, sh, sl_, sc = sig['bar_o'], sig['bar_h'], sig['bar_l'], sig['bar_c']
                sig_bar_range = sh - sl_

                entry_spread, _ = _lt_current_spread(base_spread, sig['time'], sig_bar_range, sig_bar_range or 0.1) if fric['spread'] else (base_spread, False)
                path = _lt_bridge_path(so, sh, sl_, sc, steps=20, rng=rng)
                shifted_px = _lt_latency_shift(path, 20, rng) if fric['latency'] else sig['intended_entry']
                slip = _lt_slippage(sig_bar_range, sig_bar_range or 0.1, rng) if fric['slippage'] else 0.0
                fill_entry = shifted_px + (entry_spread/2.0 + slip) if sig['is_buy'] else shifted_px - (entry_spread/2.0 + slip)

                is_buy = sig['is_buy']
                tp_px = fill_entry + sig['tp_d'] if is_buy else fill_entry - sig['tp_d']
                sl_px = fill_entry - sig['sl_d'] if is_buy else fill_entry + sig['sl_d']

                open_trades.append({
                    **sig, 'entry': fill_entry, 'tp_px': tp_px, 'sl_px': sl_px, 'sl_current': sl_px,
                    'be_activated': False, 'tp_d': sig['tp_d'], 'sl_d': sig['sl_d'],
                })

            await prog.tick(i, res['win'], res['loss'], res['be'], res['total_prof'])

        for tr in closed_trades:
            if tr['outcome'] == 'WIN' or (tr['outcome'] == 'DAILY_LIMIT' and tr['p_usd'] > 0):
                res['win'] += 1; res['total_win_usd'] += tr['p_usd']
            elif tr['outcome'] == 'LOSS' or (tr['outcome'] == 'DAILY_LIMIT' and tr['p_usd'] < 0):
                res['loss'] += 1; res['total_loss_usd'] += abs(tr['p_usd'])
            elif tr['outcome'] == 'BREAK_EVEN' or (tr['outcome'] == 'DAILY_LIMIT' and tr['p_usd'] == 0):
                res['be'] += 1
            res['total_prof'] += tr['p_usd']
            dir_str = 'BUY 📈' if tr['is_buy'] else 'SELL 📉'
            slip_px = tr['entry'] - tr['intended_entry'] if tr['is_buy'] else tr['intended_entry'] - tr['entry']
            slip_pips = round(slip_px / SYMBOL_INFO[tr['symbol']]['pip_value'], 2)
            res['trade_logs'].append({
                'الزوج': tr['symbol'], 'وقت الصفقة (DAM)': _utc_to_dam(tr['time']).strftime('%Y-%m-%d %H:%M'),
                'TF': tr['tf'], 'اتجاه': dir_str, 'المستوى (الدخول الفعلي)': f"{tr['entry']:.2f} ({tr['level_key']})",
                'الهدف (TP)': round(tr['tp_px'], 2), 'الوقف (SL)': round(tr['sl_px'], 2),
                'النتيجة': tr['outcome'], 'ربح صافي ($)': tr['p_usd'], 'cycle_ts': tr['cycle_time'].timestamp(),
                'trigger_type': tr.get('trigger_type', 'touch'), 'انزلاق (نقطة)': slip_pips,
            })

        res['trade_logs'].sort(key=lambda x: x['وقت الصفقة (DAM)'])
        running_eq = 5000.0; peak_eq = 5000.0; max_dd = 0.0
        for t_log in res['trade_logs']:
            running_eq += t_log['ربح صافي ($)']; t_log['رصيد تراكمي ($)'] = round(running_eq, 2)
            if running_eq > peak_eq: peak_eq = running_eq
            dd = peak_eq - running_eq
            if dd > max_dd: max_dd = dd
        res['peak_equity'] = peak_eq; res['max_dd'] = max_dd

        if not res['trade_logs']:
            await prog.done('<b>Live-Twin اكتمل ✅</b>\nلا توجد صفقات في هذا النطاق.')
            bot_state['is_live_twin_running'] = False; return

        await prog.set_phase('إنشاء ملف Excel المنسق...')
        sum_text = (
            f"<b>Live-Twin Engine اكتمل ✅ (واقعي)</b>\n"
            f"{syms_label} | friction: [{on_tags}]\n"
            f"{start_dt.strftime('%Y-%m-%d')} → {end_dt.strftime('%Y-%m-%d')}\n\n"
            f"Net: {'PROFIT ▲' if res['total_prof']>=0 else 'LOSS ▼'} ${round(res['total_prof'], 2)}\n"
            f"Win:  +${round(res['total_win_usd'], 2)} ({res['win']})\n"
            f"Loss: -${round(res['total_loss_usd'], 2)} ({res['loss']})\n"
            f"Break-Even: ({res['be']})\n"
            f"WR: {round(res['win']/max(1, res['win']+res['loss'])*100)}% ({len(res['trade_logs'])} صفقة)\n"
            f"Max DD: ${round(res['max_dd'],2)} ({round((res['max_dd']/max(1,res['peak_equity']))*100)}%)\n\n"
            f"عمولة إجمالية: -${round(res['total_commission'],2)} | سواب: ${round(res['total_swap'],2)}\n"
            f"صفقات مرفوضة (Requote): {res['rejected']} | نوافذ Rollover: {res['gap_events']}\n"
            f"Spread الأساسي: ${base_spread} (34pt من تيك حي)"
        )

        wb = openpyxl.Workbook()
        ws = wb.active; ws.title = "Live-Twin Trades"
        headers = ["الزوج", "وقت الصفقة (DAM)", "TF", "اتجاه", "الدخول الفعلي", "TP", "SL", "النتيجة", "ربح صافي ($)", "رصيد تراكمي ($)"]
        ws.append(headers)
        gray_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
        red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        be_fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")
        for cell in ws[1]: cell.fill = gray_fill; cell.font = Font(bold=True)
        for t_log in res['trade_logs']:
            row = [t_log['الزوج'], t_log['وقت الصفقة (DAM)'], t_log['TF'], t_log['اتجاه'], t_log['المستوى (الدخول الفعلي)'],
                   t_log['الهدف (TP)'], t_log['الوقف (SL)'], t_log['النتيجة'], t_log['ربح صافي ($)'], t_log['رصيد تراكمي ($)']]
            ws.append(row)
            fill = {'WIN': green_fill, 'LOSS': red_fill, 'BREAK_EVEN': be_fill}.get(t_log['النتيجة'])
            if fill:
                for col in range(1, 11): ws.cell(row=ws.max_row, column=col).fill = fill

        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        center_align = Alignment(horizontal='center', vertical='center')
        for row in ws.iter_rows():
            for cell in row: cell.border = thin_border; cell.alignment = center_align
        from openpyxl.utils import get_column_letter
        for i in range(1, 11): ws.column_dimensions[get_column_letter(i)].width = 22.0

        if exec_mode == 'all_concurrent':
            _add_concurrent_analysis_sheets(wb, res['trade_logs'], pnl_key='ربح صافي ($)',
                                             outcome_key='النتيجة', slippage_key='انزلاق (نقطة)')

        wb.save(fname)

        await prog.done(f'<b>Live-Twin اكتمل ✅</b>\n{syms_label} — {len(res["trade_logs"])} صفقة\nجاري إرسال التقرير...')
        await send_tg_document(fname, sum_text)
        os.remove(fname)

    except Exception as e:
        c_log(f'Live-Twin Error: {e}'); bot_state['is_live_twin_running'] = False
        if _lt_progress:
            import html
            try: await _lt_progress.done(f'❌ خطأ داخلي في Live-Twin:\n{html.escape(str(e))}')
            except Exception as inner_e: log_exception('live-twin error notification', inner_e)
    finally:
        bot_state['is_live_twin_running'] = False


def get_live_twin_keyboard() -> dict:
    if bot_state['is_live_twin_running']:
        return {'inline_keyboard': [[{'text': '⏳ Live-Twin يعمل...', 'callback_data': 'noop'}], [{'text': '⏹ إلغاء', 'callback_data': 'cancel_lt'}]]}
    mode = bot_state.get('lt_mode', 'realistic')
    mode_label = '🧪 واقعي (Live-Twin)' if mode == 'realistic' else '🧊 مثالي (Idealized A/B)'
    return {'inline_keyboard': [
        [{'text': f'الوضع: {mode_label}', 'callback_data': 'lt_toggle_mode'}],
        [{'text': '⚙️ إعدادات الاحتكاك (Friction)', 'callback_data': 'menu_lt_friction'}],
        [{'text': 'يوم واحد', 'callback_data': 'lt_1'}, {'text': 'يومين', 'callback_data': 'lt_2'}],
        [{'text': 'ثلاثة أيام', 'callback_data': 'lt_3'}, {'text': 'أسبوع', 'callback_data': 'lt_7'}],
        [{'text': 'شهر كامل', 'callback_data': 'lt_30'}],
        [{'text': '← رجوع', 'callback_data': 'menu_main'}],
    ]}


def get_live_twin_friction_keyboard() -> dict:
    fric = bot_state['lt_friction']
    def tag(key, label):
        return {'text': f"{label}: {'✅' if fric.get(key) else '🔴'}", 'callback_data': f'lt_fric_{key}'}
    return {'inline_keyboard': [
        [{'text': f"Spread أساسي: ${bot_state['lt_base_spread_usd']} (34pt/تيك حي)", 'callback_data': 'noop'}],
        [tag('spread', '📶 سبريد ديناميكي')],
        [tag('slippage', '⚡ انزلاق (Slippage)')],
        [tag('latency', '⏱ تأخير التنفيذ (200-800ms)')],
        [tag('commission', '💵 عمولة')],
        [tag('gaps', '📉 فجوات نهاية الأسبوع/Rollover')],
        [tag('rejection', '🚫 رفض/Requote')],
        [{'text': '← رجوع', 'callback_data': 'menu_lt'}],
    ]}


async def check_metaapi_status_command(chat_id: int):
    if not METAAPI_TOKEN or METAAPI_TOKEN == 'YOUR_METAAPI_TOKEN':
        await send_tg_msg("❌ MetaAPI Token غير مهيأ.")
        return
    if not ACCOUNT_ID or ACCOUNT_ID == 'YOUR_ACCOUNT_ID':
        await send_tg_msg("❌ Account ID غير مهيأ.")
        return
        
    await send_tg_msg("⏳ جاري فحص حالة الحساب من MetaAPI...")
    api = MetaApi(METAAPI_TOKEN)
    try:
        account = await api.metatrader_account_api.get_account(ACCOUNT_ID)
        state = account.state
        conn_status = account.connection_status
        
        msg = f"<b>حالة الحساب (MetaAPI)</b>\n"
        msg += f"الاسم: {account.name}\n"
        msg += f"الحالة: {state}\n"
        msg += f"الاتصال: {conn_status}\n\n"
        
        if state == 'DEPLOYED' and conn_status == 'CONNECTED':
            conn = account.get_rpc_connection()
            await conn.connect()
            await conn.wait_synchronized()
            
            acc_info = await conn.get_account_information()
            msg += f"<b>الرصيد:</b> {acc_info.get('balance', 0):.2f}\n"
            msg += f"<b>الاكويتي:</b> {acc_info.get('equity', 0):.2f}\n"
            msg += f"<b>الهامش المتاح:</b> {acc_info.get('freeMargin', 0):.2f}\n\n"
            
            positions = await conn.get_positions()
            msg += f"<b>الصفقات المفتوحة:</b> {len(positions)}\n"
            for p in positions:
                msg += f"🔸 {p['symbol']} | {p['type']} | {p['volume']} | Profit: {p.get('profit', 0):.2f}\n"
                
        else:
            msg += "⚠️ الحساب غير متصل حالياً لجلب تفاصيل الرصيد والصفقات."
            
        await send_tg_msg(msg)
    except Exception as e:
        import html
        await send_tg_msg(f"❌ خطأ في الاتصال بـ MetaAPI:\n{html.escape(str(e))}")

async def _handle_callback(d: str, chat_id: int, msg_id: int) -> None:
    if d == 'check_metaapi_status':
        asyncio.create_task(check_metaapi_status_command(chat_id))
        return
    if d == 'run_diag':
        async def _run_diag_task():
            try:
                report = await gann_run_diagnostics()
                # Telegram hard-caps messages at 4096 chars; a multi-symbol,
                # multi-timeframe report can exceed that easily. Split on
                # line boundaries rather than sending an oversized message
                # that would just fail outright.
                lines = report.split('\n')
                chunk = ""
                for line in lines:
                    if len(chunk) + len(line) + 1 > 3500:
                        await send_tg_msg(chunk)
                        chunk = ""
                    chunk += line + "\n"
                if chunk.strip():
                    await send_tg_msg(chunk)
            except Exception as e:
                log_exception('gann_run_diagnostics', e)
                await send_tg_msg(f"❌ فشل التشخيص: {e}")
        asyncio.create_task(_run_diag_task())
        return
    if d == 'export_diag_excel':
        async def _export_diag_task():
            try:
                await export_diag_log_excel()
            except Exception as e:
                log_exception('export_diag_log_excel', e)
                await send_tg_msg(f"❌ فشل تصدير سجل التشخيص: {e}")
        asyncio.create_task(_export_diag_task())
        return
    if d == 'export_live_trades_excel':
        async def _export_live_trades_task():
            try:
                await export_live_trades_excel()
            except Exception as e:
                log_exception('export_live_trades_excel', e)
                await send_tg_msg(f"❌ فشل تصدير سجل الصفقات الحية: {e}")
        asyncio.create_task(_export_live_trades_task())
        return
    if d == 'manual_resume_step1':
        current_state = bot_state.get('connection_state', CONN_RUNNING)
        if current_state == CONN_RUNNING:
            await send_tg_msg("✅ البوت أصلاً في حالة RUNNING -- لا حاجة لأي استئناف.")
            return
        await send_tg_msg(
            f"⚠️ <b>تأكيد الاستئناف اليدوي</b>\n"
            f"الحالة الحالية: {current_state}\n"
            f"السبب: {bot_state.get('connection_state_reason', '-')}\n\n"
            f"هل تأكدت فعلياً من حساب الوسيط (MT5) ومقارنته بما يتتبعه البوت؟ "
            f"الضغط على تأكيد سيعيد البوت للعمل فوراً بافتراض أن الحساب سليم.",
            reply_markup={'inline_keyboard': [
                [{'text': '✅ نعم، تأكدت -- استأنف الآن', 'callback_data': 'manual_resume_confirm'}],
                [{'text': '❌ إلغاء', 'callback_data': 'menu_main'}],
            ]}
        )
        return
    if d == 'manual_resume_confirm':
        global _recon_consecutive_mismatches, _consecutive_real_order_failures
        prior_state = bot_state.get('connection_state', CONN_RUNNING)
        _recon_consecutive_mismatches = 0
        _consecutive_real_order_failures = 0
        await set_connection_state(
            CONN_RUNNING,
            f"Manually resumed by operator via Telegram after verifying account state "
            f"(was {prior_state})."
        )
        await send_tg_msg("✅ تم الاستئناف اليدوي. البوت الآن RUNNING وسيقبل صفقات جديدة من التحديث القادم.")
        return

    sym = bot_state['ui_selected_symbol']
    sym_state = bot_state['symbol_state'][sym]
    if d == 'menu_main': await _show(chat_id, msg_id, 'القائمة الرئيسية:', get_main_keyboard())
    elif d == 'menu_main':
        await _show(chat_id, msg_id, '<b>مرحباً بك في Gold Scalper Bot v8.9</b>', get_main_keyboard())

    elif d == 'menu_presets':
        kbd = {'inline_keyboard': [
            [{'text': '💾 حفظ كـ Preset 1', 'callback_data': 'save_preset_1'}, {'text': '📂 تحميل Preset 1', 'callback_data': 'load_preset_1'}],
            [{'text': '💾 حفظ كـ Preset 2', 'callback_data': 'save_preset_2'}, {'text': '📂 تحميل Preset 2', 'callback_data': 'load_preset_2'}],
            [{'text': '💾 حفظ كـ Preset 3', 'callback_data': 'save_preset_3'}, {'text': '📂 تحميل Preset 3', 'callback_data': 'load_preset_3'}],
            [{'text': '🔙 رجوع', 'callback_data': 'menu_main'}]
        ]}
        await _show(chat_id, msg_id, '<b>إدارة الإعدادات (Presets):</b>\nهنا يمكنك حفظ إعدادات جميع الأزواج واستعادتها لاحقاً.', kbd)
    elif d.startswith('save_preset_'):
        p_num = d.split('_')[-1]
        data = {}
        if os.path.exists(PRESETS_FILE):
            try:
                with open(PRESETS_FILE, 'r') as f: data = json.load(f)
            except Exception as e:
                # Corrupt presets file -- log it instead of silently
                # discarding whatever else was saved in there.
                log_exception(f"save_preset_{p_num} (reading existing presets)", e)
                await send_tg_msg(f"⚠️ ملف الـ Presets الحالي تالف، سيتم إنشاء ملف جديد. (الخطأ: {e})")
                data = {}

        # A preset should only ever capture settings, never live runtime
        # state -- and critically, gann_last_h1_time/gann_cycle_started_at
        # are live datetime objects once any cycle has run (which is
        # almost immediately after startup). json.dump() cannot serialize
        # a raw datetime at all, so saving used to raise TypeError as soon
        # as this state existed. _PRESET_EXCLUDED_KEYS matches exactly
        # what load_preset already refuses to restore, so nothing is lost
        # by leaving them out of what gets saved in the first place.
        data[f'preset_{p_num}'] = {
            s_name: {k: v for k, v in s_data.items() if k not in _PRESET_EXCLUDED_KEYS}
            for s_name, s_data in bot_state['symbol_state'].items()
        }
        try:
            with open(TEMP_PRESETS_FILE, 'w') as f:
                json.dump(data, f)
                f.flush()
                os.fsync(f.fileno())
            os.replace(TEMP_PRESETS_FILE, PRESETS_FILE)
            await send_tg_msg(f"✅ تم حفظ الإعدادات الحالية في Preset {p_num}")
        except Exception as e:
            log_exception(f"save_preset_{p_num} (writing)", e)
            await send_tg_msg(f"❌ فشل حفظ Preset {p_num}: {e}")
    elif d.startswith('load_preset_'):
        p_num = d.split('_')[-1]
        if not os.path.exists(PRESETS_FILE):
            await send_tg_msg(
                "❌ لا يوجد ملف Presets محفوظ بعد.\n"
                "ملاحظة: كانت الإصدارات السابقة تحفظ هذا الملف في مسار مؤقت يُمسح عند إعادة التشغيل -- "
                "تم إصلاح ذلك الآن، فأي Preset تحفظه من الآن فصاعداً سيبقى بعد إعادة التشغيل."
            )
        else:
            try:
                with open(PRESETS_FILE, 'r') as f: data = json.load(f)
                if f'preset_{p_num}' in data:
                    # Load settings, but keep live data like open_trades and gann_levels untouched
                    for s_name, s_data in data[f'preset_{p_num}'].items():
                        if s_name in bot_state['symbol_state']:
                            for k, v in s_data.items():
                                if k not in _PRESET_EXCLUDED_KEYS:
                                    bot_state['symbol_state'][s_name][k] = v
                    await send_tg_msg(f"✅ تم تحميل الإعدادات من Preset {p_num} بنجاح!")
                else:
                    await send_tg_msg("❌ لا يوجد إعدادات محفوظة في هذا الـ Preset.")
            except Exception as e:
                log_exception(f"load_preset_{p_num}", e)
                await send_tg_msg(f"❌ حدث خطأ أثناء التحميل: {e}")

    elif d == 'menu_protection':
        await _show(chat_id, msg_id, 'إعدادات الحماية:', get_protection_keyboard())
    elif d == 'prot_toggle_multitf':
        bot_state['prot_allow_multi_tf'] = not bot_state['prot_allow_multi_tf']
        await _show(chat_id, msg_id, 'إعدادات الحماية:', get_protection_keyboard())
    elif d == 'prot_dec_dd':
        bot_state['prot_daily_dd_usd'] = max(50, bot_state['prot_daily_dd_usd'] - 50)
        await _show(chat_id, msg_id, 'إعدادات الحماية:', get_protection_keyboard())
    elif d == 'prot_inc_dd':
        bot_state['prot_daily_dd_usd'] = min(5000, bot_state['prot_daily_dd_usd'] + 50)
        await _show(chat_id, msg_id, 'إعدادات الحماية:', get_protection_keyboard())
    elif d == 'prot_dec_profit':
        bot_state['prot_daily_profit_usd'] = max(0, bot_state['prot_daily_profit_usd'] - 50)
        await _show(chat_id, msg_id, 'إعدادات الحماية:', get_protection_keyboard())
    elif d == 'prot_inc_profit':
        bot_state['prot_daily_profit_usd'] = min(10000, bot_state['prot_daily_profit_usd'] + 50)
        await _show(chat_id, msg_id, 'إعدادات الحماية:', get_protection_keyboard())
    elif d == 'menu_gann': await _show(chat_id, msg_id, 'إعدادات جان:', get_gann_keyboard())
    elif d == 'menu_protection': await _show(chat_id, msg_id, '🛡️ إعدادات الحماية:', get_protection_keyboard())
    elif d == 'tg_prot_sync': bot_state['prot_true_sync'] = not bot_state.get('prot_true_sync', True); await _show(chat_id, msg_id, '🛡️ إعدادات الحماية:', get_protection_keyboard())
    elif d == 'tg_prot_inval': bot_state['prot_cycle_inval'] = not bot_state.get('prot_cycle_inval', True); await _show(chat_id, msg_id, '🛡️ إعدادات الحماية:', get_protection_keyboard())
    elif d == 'tg_prot_cost': bot_state['prot_cost_be'] = not bot_state.get('prot_cost_be', True); await _show(chat_id, msg_id, '🛡️ إعدادات الحماية:', get_protection_keyboard())
    elif d == 'tg_prot_stale': bot_state['prot_stale_filter'] = not bot_state.get('prot_stale_filter', True); await _show(chat_id, msg_id, '🛡️ إعدادات الحماية:', get_protection_keyboard())
    elif d == 'tg_prot_anchor':
        bot_state['gann_anchor_tf'] = '4h' if bot_state.get('gann_anchor_tf', '1h') == '1h' else '1h'
        for sym, ss in bot_state['symbol_state'].items():
            # Force every symbol to recompute its levels from the NEW
            # anchor timeframe on the very next scanner tick, rather than
            # waiting for whatever boundary the OLD anchor would have
            # used next. Without this, "switching to 4h" would silently
            # keep running on the old H1-derived levels until the current
            # cycle happened to expire.
            #
            # gann_cycle_hours (ladder freeze / monitoring duration) is
            # intentionally left untouched here -- it stays fully manual,
            # adjustable via its own +/- buttons (e.g. 1h -> 2h/3h/4h),
            # independent of whatever the anchor timeframe is set to.
            ss['gann_last_h1_time'] = None
            ss['gann_cycle_started_at'] = None
        await save_bot_persistence()
        await _show(chat_id, msg_id, '🛡️ إعدادات الحماية:', get_protection_keyboard())
        await send_tg_msg(
            f"✅ <b>تم تغيير الإطار المرجعي إلى {_anchor_label()}</b>\n"
            f"سيتم استخراج مستويات جديدة بالكامل من إغلاق شمعة {_anchor_label()} في التحديث القادم.\n"
            f"ملاحظة: مدة تجميد السلم (مدة المراقبة) لم تتغيّر — عدّلها يدوياً من أزرارها الخاصة لو حبيت."
        )

    elif d == 'prot_reset_all':
        # Clears every "stuck until next natural trigger" protection state:
        # - live_daily_hit: capital-protection daily DD/profit lock (normally
        #   only clears at midnight broker time)
        # - per-symbol cycle invalidation from prot_cycle_inval (spike >200pts
        #   freezes that symbol until its NEXT H1 close -- this forces an
        #   immediate rebuild instead of waiting)
        # Does NOT touch gann_level_status (already-used levels this cycle) --
        # that's a normal trading-logic guard, not a "protection freeze", and
        # clearing it would let the bot re-enter a level it already traded.
        was_daily_hit = bot_state.get('live_daily_hit', False)
        bot_state['live_daily_hit'] = False
        frozen_symbols = []
        for sym, ss in bot_state['symbol_state'].items():
            if ss.get('gann_close_used') is None and not ss.get('gann_levels'):
                continue  # wasn't actually frozen
            frozen_symbols.append(sym)
            ss['gann_levels'] = []
            ss['gann_close_used'] = None
            ss['gann_last_h1_time'] = None
            ss['gann_cycle_started_at'] = None
        await save_bot_persistence()
        await _show(chat_id, msg_id, '🛡️ إعدادات الحماية:', get_protection_keyboard())
        summary = []
        if was_daily_hit:
            summary.append("• قفل حماية رأس المال اليومي (ربح/تراجع) — تم فكّه، الدخول مسموح من الآن")
        if frozen_symbols:
            summary.append(f"• تجميد الدورة بسبب انفجار سعري — تم فكّه لـ: {', '.join(frozen_symbols)} (سيُعاد بناء المستويات فوراً)")
        if not summary:
            summary.append("لا توجد حمايات نشطة حالياً لتصفيرها — كل شيء طبيعي.")
        await send_tg_msg("🔄 <b>تصفير الحمايات</b>\n\n" + "\n".join(summary))

    elif d == 'tg_prot_dam_time':
        bot_state['prot_dam_time_filter'] = not bot_state.get('prot_dam_time_filter', True)
        # Was previously never persisted -- the in-memory toggle worked
        # immediately, but reverted to the default on any restart or reload,
        # which looked exactly like "toggling does nothing."
        await save_bot_persistence()
        await _show(chat_id, msg_id, '🛡️ إعدادات الحماية:', get_protection_keyboard())

    elif d == 'tg_gann_calc_mode':
        new_mode = 'static_h1' if bot_state.get('gann_calculation_mode', 'static_h1') == 'dynamic_live' else 'dynamic_live'
        bot_state['gann_calculation_mode'] = new_mode
        for sym, ss in bot_state['symbol_state'].items():
            # Force an immediate recompute on gann_cycle_manager's very next
            # tick instead of waiting for whatever boundary the OLD mode
            # would have used next -- matches the anchor-tf toggle's
            # "revert immediately" behavior.
            ss['gann_last_h1_time'] = None
            ss['gann_cycle_started_at'] = None
        await save_bot_persistence()
        await _show(chat_id, msg_id, '🛡️ إعدادات الحماية:', get_protection_keyboard())
        mode_lbl = '⚡ حي (Dynamic Live -- كل 5 دقائق حسب السعر اللحظي)' if new_mode == 'dynamic_live' else '📌 كلاسيكي (Static -- عند إغلاق شمعة الأنكر فقط)'
        await send_tg_msg(f"✅ <b>وضع حساب جان: {mode_lbl}</b>\nسيتم تطبيق هذا فوراً في الدورة القادمة (~60 ثانية) وفي أي باكتيست جديد.")

    elif d == 'gann_show_levels':
        sym = bot_state['ui_selected_symbol']
        if not bot_state['symbol_state'][sym]['gann_levels'] or not bot_state['symbol_state'][sym]['gann_close_used']:
            await send_tg_msg(f'⏳ لا يوجد سلّم نشط لـ {sym}، جاري جلب آخر شمعة H1...')
            last_h1 = await _gann_fetch_last_closed_anchor(sym)
            if last_h1:
                h1_close = float(last_h1['close'])
                bot_state['symbol_state'][sym]['gann_levels']          = gann_calc_levels(sym, h1_close)
                bot_state['symbol_state'][sym]['gann_close_used']       = h1_close
                bot_state['symbol_state'][sym]['gann_last_h1_time']     = last_h1['time']
                bot_state['symbol_state'][sym]['gann_cycle_started_at'] = datetime.now(timezone.utc)
                bot_state['symbol_state'][sym]['gann_cycle_active']     = True
            else:
                await send_tg_msg('❌ تعذّر جلب البيانات.'); return
        await send_tg_msg(_gann_fmt_levels_msg(sym, bot_state['symbol_state'][sym]['gann_close_used']))
        await _show(chat_id, msg_id, 'إعدادات جان:', get_gann_keyboard())
    elif d == 'gann_show_last10':
        sym = bot_state['ui_selected_symbol']
        anchor_tf = bot_state.get('gann_anchor_tf', '1h')
        anchor_hours = _anchor_hours()
        offset = bot_state.get('broker_time_offset', 3)
        await send_tg_msg(f'⏳ جاري جلب آخر 10 شموع {_anchor_label()} لـ {sym} من اواندا...')
        candles = await fetch_candles(sym, anchor_tf, count=10)
        if not candles:
            await send_tg_msg('❌ تعذّر جلب الشموع.')
            return
        candles = sorted(candles, key=lambda c: c['time'])[-10:]
        # What the bot itself currently treats as "last closed" -- shown
        # separately so a mismatch between this and the chart is obvious.
        bot_pick = await _gann_fetch_last_closed_anchor(sym)
        bot_pick_time = bot_pick['time'] if bot_pick else None

        lines = [
            f'🕯️ <b>آخر 10 شموع {_anchor_label()} — {sym}</b>',
            f'(المصدر: OANDA | التوقيت المعروض: دمشق UTC+{offset} — وبين قوسين UTC الخام)',
            ''
        ]
        for i, c in enumerate(candles, 1):
            t_utc = c['time'].to_pydatetime()
            t_dam_start = t_utc + timedelta(hours=offset)
            t_dam_end = t_dam_start + timedelta(hours=anchor_hours)
            t_utc_end = t_utc + timedelta(hours=anchor_hours)
            marker = ' ✅ ← يعتمدها البوت الآن' if bot_pick_time and t_utc == bot_pick_time else ''
            lines.append(
                f"{i}) {t_dam_start.strftime('%m-%d %H:%M')} → {t_dam_end.strftime('%H:%M')} دمشق  "
                f"({t_utc.strftime('%H:%M')}-{t_utc_end.strftime('%H:%M')} UTC)\n"
                f"    إغلاق: {float(c['close']):.5f}{marker}"
            )
        if not bot_pick_time:
            lines.append('\n⚠️ لم يتمكن البوت من تحديد آخر شمعة مغلقة حالياً.')
        await send_tg_msg('\n'.join(lines))
        await _show(chat_id, msg_id, 'إعدادات جان:', get_gann_keyboard())
    elif d == 'gann_toggle_entry':
        sym_state['gann_entry_mode'] = 'pure_touch' if sym_state['gann_entry_mode'] == 'touch_trend' else 'touch_trend'
        await _show(chat_id, msg_id, 'إعدادات جان:', get_gann_keyboard())
    elif d == 'gann_toggle_filter':
        current = sym_state['gann_zone_filter']
        if current == 'star': sym_state['gann_zone_filter'] = 'star_fan'
        elif current == 'star_fan': sym_state['gann_zone_filter'] = 'all'
        else: sym_state['gann_zone_filter'] = 'star'
        await _show(chat_id, msg_id, 'إعدادات جان:', get_gann_keyboard())
    elif d == 'gann_toggle_filter_type':
        current = sym_state['trend_filter_type']
        if current == 'vwap': sym_state['trend_filter_type'] = 'ema'
        else: sym_state['trend_filter_type'] = 'vwap'
        await _show(chat_id, msg_id, 'إعدادات جان:', get_gann_keyboard())
    elif d == 'gann_toggle_exec_mode':
        order = ['instant', 'close', 'hybrid', 'all_concurrent']
        current = bot_state.get('gann_execution_mode', 'instant')
        bot_state['gann_execution_mode'] = order[(order.index(current) + 1) % len(order)]
        await _show(chat_id, msg_id, 'إعدادات جان:', get_gann_keyboard())
    elif d == 'gann_toggle_auto_trade':
        sym_state['auto_trade'] = not sym_state.get('auto_trade', False)
        await _show(chat_id, msg_id, 'إعدادات جان:', get_gann_keyboard())
    elif d == 'gann_toggle_ttf':
        sym_state['trend_timeframe'] = '30m' if sym_state['trend_timeframe'] == '1h' else '1h'
        await _show(chat_id, msg_id, 'إعدادات جان:', get_gann_keyboard())
    elif d == 'gann_toggle_be':
        sym_state['break_even_enabled'] = not sym_state['break_even_enabled']
        await _show(chat_id, msg_id, 'إعدادات جان:', get_gann_keyboard())
    elif d == 'gann_dec_be_pts':
        sym_state['gann_be_trigger_points'] = max(10, sym_state.get('gann_be_trigger_points', 40) - 10)
        await _show(chat_id, msg_id, 'إعدادات جان:', get_gann_keyboard())
    elif d == 'gann_inc_be_pts':
        sym_state['gann_be_trigger_points'] = min(200, sym_state.get('gann_be_trigger_points', 40) + 10)
        await _show(chat_id, msg_id, 'إعدادات جان:', get_gann_keyboard())
    elif d == 'gann_dec_vwap': 
        sym_state['trend_vwap_period'] = max(10, sym_state['trend_vwap_period'] - 10)
        await _show(chat_id, msg_id, 'إعدادات جان:', get_gann_keyboard())
    elif d == 'gann_inc_vwap': 
        sym_state['trend_vwap_period'] = min(500, sym_state['trend_vwap_period'] + 10)
        await _show(chat_id, msg_id, 'إعدادات جان:', get_gann_keyboard())
    elif d == 'gann_dec_ema': 
        sym_state['trend_ema_period'] = max(10, sym_state['trend_ema_period'] - 10)
        await _show(chat_id, msg_id, 'إعدادات جان:', get_gann_keyboard())
    elif d == 'gann_inc_ema': 
        sym_state['trend_ema_period'] = min(500, sym_state['trend_ema_period'] + 10)
        await _show(chat_id, msg_id, 'إعدادات جان:', get_gann_keyboard())
    elif d == 'gann_filter_help':
        help_txt = ("<b>⚙️ دليل تخصيص القيم:</b>\n\n"
                    "أرسل أمراً مباشراً في الدردشة لتغيير أي قيمة بالصيغة التالية:\n\n"
                    "<b>تغيير فلاتر الترند الشامل:</b>\n"
                    "<code>/set ema 50</code>\n"
                    "<code>/set vwap 100</code>\n\n"
                    "<b>تخصيص الأهداف والوقف لكل فريم:</b>\n"
                    "<code>/set 5m tp 40</code>\n"
                    "<code>/set 15m sl 25</code>\n\n"
                    "سيتم حفظ القيمة وتطبيقها فوراً.")
        await _show(chat_id, msg_id, help_txt, get_gann_keyboard())
    elif d == 'gann_dec_margin': 
        sym_state['gann_touch_margin_pts'] = max(1, sym_state['gann_touch_margin_pts'] - 1)
        await _show(chat_id, msg_id, 'إعدادات جان:', get_gann_keyboard())
    elif d == 'gann_inc_margin': 
        sym_state['gann_touch_margin_pts'] = min(50, sym_state['gann_touch_margin_pts'] + 1)
        await _show(chat_id, msg_id, 'إعدادات جان:', get_gann_keyboard())
    elif d == 'gann_dec_lot':
        sym_state['lot_size'] = round(max(0.01, sym_state['lot_size'] - 0.01), 2)
        await _show(chat_id, msg_id, 'إعدادات جان:', get_gann_keyboard())
    elif d == 'gann_inc_lot':
        sym_state['lot_size'] = round(min(50.0, sym_state['lot_size'] + 0.01), 2)
        await _show(chat_id, msg_id, 'إعدادات جان:', get_gann_keyboard())
    elif d == 'gann_dec_hours': 
        sym_state['gann_cycle_hours'] = max(1, sym_state['gann_cycle_hours'] - 1)
        await _show(chat_id, msg_id, 'إعدادات جان:', get_gann_keyboard())
    elif d == 'gann_inc_hours': 
        sym_state['gann_cycle_hours'] = min(24, sym_state['gann_cycle_hours'] + 1)
        await _show(chat_id, msg_id, 'إعدادات جان:', get_gann_keyboard())
    elif d == 'gann_toggle_tpsl':
        sym_state['gann_tpsl_mode'] = 'atr' if sym_state['gann_tpsl_mode'] == 'fixed' else 'fixed'
        await _show(chat_id, msg_id, 'إعدادات جان:', get_gann_keyboard())
    elif d == 'gann_dec_tp10': sym_state['gann_tp_points'] = max(10, sym_state['gann_tp_points'] - 10); await _show(chat_id, msg_id, 'إعدادات جان:', get_gann_keyboard())
    elif d == 'gann_inc_tp10': sym_state['gann_tp_points'] = min(1000, sym_state['gann_tp_points'] + 10); await _show(chat_id, msg_id, 'إعدادات جان:', get_gann_keyboard())
    elif d == 'gann_dec_sl10': sym_state['gann_sl_points'] = max(10, sym_state['gann_sl_points'] - 10); await _show(chat_id, msg_id, 'إعدادات جان:', get_gann_keyboard())
    elif d == 'gann_inc_sl10': sym_state['gann_sl_points'] = min(1000, sym_state['gann_sl_points'] + 10); await _show(chat_id, msg_id, 'إعدادات جان:', get_gann_keyboard())
    elif d == 'gann_dec_atrp':  sym_state['gann_atr_period'] = max(5,   sym_state['gann_atr_period'] - 1); await _show(chat_id, msg_id, 'إعدادات جان:', get_gann_keyboard())
    elif d == 'gann_inc_atrp':  sym_state['gann_atr_period'] = min(50,  sym_state['gann_atr_period'] + 1); await _show(chat_id, msg_id, 'إعدادات جان:', get_gann_keyboard())
    elif d == 'gann_dec_atrsl': sym_state['gann_atr_sl_mult'] = max(0.5, round(sym_state['gann_atr_sl_mult'] - 0.5, 1)); await _show(chat_id, msg_id, 'إعدادات جان:', get_gann_keyboard())
    elif d == 'gann_inc_atrsl': sym_state['gann_atr_sl_mult'] = min(5.0, round(sym_state['gann_atr_sl_mult'] + 0.5, 1)); await _show(chat_id, msg_id, 'إعدادات جان:', get_gann_keyboard())
    elif d == 'gann_dec_atrtp': sym_state['gann_atr_tp_mult'] = max(0.5, round(sym_state['gann_atr_tp_mult'] - 0.5, 1)); await _show(chat_id, msg_id, 'إعدادات جان:', get_gann_keyboard())
    elif d == 'gann_inc_atrtp': sym_state['gann_atr_tp_mult'] = min(8.0, round(sym_state['gann_atr_tp_mult'] + 0.5, 1)); await _show(chat_id, msg_id, 'إعدادات جان:', get_gann_keyboard())
    elif d.startswith('gann_toggle_pair_'):
        pair = d[len('gann_toggle_pair_'):]
        bot_state['active_symbols'][pair] = not bot_state['active_symbols'][pair]
        if bot_state['active_symbols'][pair]:
            await _lq_subscribe_symbol(pair)
        await _show(chat_id, msg_id, 'إعدادات جان:', get_gann_keyboard())
    elif d.startswith('gann_sel_pair_'):
        pair = d[len('gann_sel_pair_'):]
        bot_state['ui_selected_symbol'] = pair
        await _show(chat_id, msg_id, 'إعدادات جان:', get_gann_keyboard())
    elif d.startswith('gann_tf_'):
        tfk = d[len('gann_tf_'):]
        if tfk in sym_state['gann_monitor_tfs']: sym_state['gann_monitor_tfs'][tfk] = not sym_state['gann_monitor_tfs'][tfk]
        await _show(chat_id, msg_id, 'إعدادات جان:', get_gann_keyboard())
    elif d == 'gann_tpsl_tf': await _show(chat_id, msg_id, '⚙️ TP/SL مخصص لكل فريم:', get_gann_tpsl_tf_keyboard())
    elif d.startswith('gann_tptf_sel_'):
        sel_tf = d[len('gann_tptf_sel_'):]; await _show(chat_id, msg_id, f'⚙️ TP/SL [{sel_tf}]:', get_gann_tpsl_tf_keyboard(sel_tf))
    elif d.startswith('gann_tptf_itp_'):
        tf = d[len('gann_tptf_itp_'):]; sym_state['gann_tp_per_tf'][tf] = sym_state['gann_tp_per_tf'].get(tf, 0) + 10; await _show(chat_id, msg_id, f'⚙️ TP/SL [{tf}]:', get_gann_tpsl_tf_keyboard(tf))
    elif d.startswith('gann_tptf_dtp_'):
        tf = d[len('gann_tptf_dtp_'):]; sym_state['gann_tp_per_tf'][tf] = max(0, sym_state['gann_tp_per_tf'].get(tf, 0) - 10); await _show(chat_id, msg_id, f'⚙️ TP/SL [{tf}]:', get_gann_tpsl_tf_keyboard(tf))
    elif d.startswith('gann_tptf_isl_'):
        tf = d[len('gann_tptf_isl_'):]; sym_state['gann_sl_per_tf'][tf] = sym_state['gann_sl_per_tf'].get(tf, 0) + 10; await _show(chat_id, msg_id, f'⚙️ TP/SL [{tf}]:', get_gann_tpsl_tf_keyboard(tf))
    elif d.startswith('gann_tptf_dsl_'):
        tf = d[len('gann_tptf_dsl_'):]; sym_state['gann_sl_per_tf'][tf] = max(0, sym_state['gann_sl_per_tf'].get(tf, 0) - 10); await _show(chat_id, msg_id, f'⚙️ TP/SL [{tf}]:', get_gann_tpsl_tf_keyboard(tf))
    elif d.startswith('gann_tptf_rst_'):
        tf = d[len('gann_tptf_rst_'):]; sym_state['gann_tp_per_tf'][tf] = 0; sym_state['gann_sl_per_tf'][tf] = 0; await _show(chat_id, msg_id, f'⚙️ تمت إعادة الضبط:', get_gann_tpsl_tf_keyboard(tf))
    elif d == 'menu_gann_bt':
        await _show(chat_id, msg_id, 'اختر مدة الباكتيست:', get_gann_bt_keyboard())
    elif d.startswith('gbt_'):
        days = int(d.split('_')[1])
        end_dt = datetime.now(timezone.utc)
        start_dt = end_dt - timedelta(days=days)
        if not bot_state['is_backtesting']: asyncio.create_task(run_gann_backtest(start_dt, end_dt))
        await _show(chat_id, msg_id, f'⏳ باكتيست يعمل...', get_gann_bt_keyboard())
    elif d == 'cancel_bt':
        global _bt_progress
        if _bt_progress and bot_state['is_backtesting']: await _bt_progress.cancel()
        bot_state['is_backtesting'] = False
        await _show(chat_id, msg_id, 'إعدادات جان:', get_gann_keyboard())
    elif d == 'menu_lt':
        await _show(chat_id, msg_id, '🧪 Live-Twin Simulator:', get_live_twin_keyboard())
    elif d == 'menu_lt_friction':
        await _show(chat_id, msg_id, '⚙️ إعدادات الاحتكاك:', get_live_twin_friction_keyboard())
    elif d == 'lt_toggle_mode':
        bot_state['lt_mode'] = 'idealized' if bot_state.get('lt_mode', 'realistic') == 'realistic' else 'realistic'
        await _show(chat_id, msg_id, '🧪 Live-Twin Simulator:', get_live_twin_keyboard())
    elif d.startswith('lt_fric_'):
        key = d[len('lt_fric_'):]
        if key in bot_state['lt_friction']: bot_state['lt_friction'][key] = not bot_state['lt_friction'][key]
        await _show(chat_id, msg_id, '⚙️ إعدادات الاحتكاك:', get_live_twin_friction_keyboard())
    elif d.startswith('lt_'):
        days = int(d.split('_')[1])
        end_dt = datetime.now(timezone.utc)
        start_dt = end_dt - timedelta(days=days)
        if not bot_state['is_live_twin_running']: asyncio.create_task(run_live_twin_simulation(start_dt, end_dt))
        await _show(chat_id, msg_id, '⏳ Live-Twin يعمل...', get_live_twin_keyboard())
    elif d == 'cancel_lt':
        global _lt_progress
        if _lt_progress and bot_state['is_live_twin_running']: await _lt_progress.cancel()
        bot_state['is_live_twin_running'] = False
        await _show(chat_id, msg_id, '🧪 Live-Twin Simulator:', get_live_twin_keyboard())
    else: c_log(f'Unhandled callback: {d}')

    # UI Settings Amnesia fix: every branch above except the early-return
    # status check (which mutates nothing) falls through to here. Save
    # once, after the mutation has landed in bot_state, so a restart never
    # reverts a toggle/setting change back to the last trade's snapshot.
    await save_bot_persistence()

# ─────────────────────────────────────────────────────────────
# TELEGRAM POLLING & WATCHDOG
# ─────────────────────────────────────────────────────────────
async def process_tg_update(update: dict) -> None:
    if 'message' in update and 'text' in update['message']:
        msg = update['message']['text'].strip(); bot_state['chat_id'] = update['message']['chat']['id']
        
        parts = msg.lower().split()
        if parts[0] == '/set':
            sym_state = bot_state['symbol_state'][bot_state['ui_selected_symbol']]
            if len(parts) == 3 and parts[1] in ['ema', 'vwap'] and parts[2].isdigit():
                val = int(parts[2])
                if parts[1] == 'ema': sym_state['trend_ema_period'] = val
                elif parts[1] == 'vwap': sym_state['trend_vwap_period'] = val
                await save_bot_persistence()
                await send_tg_msg(f"✅ <b>تم التحديث بنجاح!</b>\n⚙️ {parts[1].upper()} الشامل: {val}")
                return
            elif len(parts) == 4:
                _, tf, param, val = parts
                if tf in _TFS and param in ['tp', 'sl'] and val.isdigit():
                    val = int(val)
                    if param == 'tp': sym_state['gann_tp_per_tf'][tf] = val
                    elif param == 'sl': sym_state['gann_sl_per_tf'][tf] = val
                    await save_bot_persistence()
                    await send_tg_msg(f"✅ <b>تم التحديث بنجاح!</b>\n📌 الفريم: {tf}\n⚙️ {param.upper()}: {val}")
                    return
            await send_tg_msg("❌ <b>صيغة خاطئة!</b>\n<b>أمثلة صحيحة:</b>\n<code>/set ema 50</code>\n<code>/set vwap 100</code>\n<code>/set 5m tp 40</code>\n<code>/set 15m sl 25</code>")
            return

        if parts[0] == '/backtest':
            try:
                if len(parts) == 2:
                    dt = datetime.strptime(parts[1], "%Y-%m-%d").replace(tzinfo=timezone.utc)
                    if not bot_state['is_backtesting']: asyncio.create_task(run_gann_backtest(dt, dt + timedelta(days=1)))
                    await send_tg_msg(f"⏳ جاري باكتيست ليوم {parts[1]}...")
                    return
                elif len(parts) == 3:
                    dt1 = datetime.strptime(parts[1], "%Y-%m-%d").replace(tzinfo=timezone.utc)
                    dt2 = datetime.strptime(parts[2], "%Y-%m-%d").replace(tzinfo=timezone.utc) + timedelta(days=1)
                    if not bot_state['is_backtesting']: asyncio.create_task(run_gann_backtest(dt1, dt2))
                    await send_tg_msg(f"⏳ جاري باكتيست من {parts[1]} إلى {parts[2]}...")
                    return
            except Exception:
                await send_tg_msg("❌ <b>خطأ في التاريخ!</b>\nالصيغة: <code>/backtest 2026-06-24</code>\nأو <code>/backtest 2026-06-24 2026-06-26</code>")
                return

        if parts[0] == '/livetwin':
            try:
                if len(parts) == 2:
                    dt = datetime.strptime(parts[1], "%Y-%m-%d").replace(tzinfo=timezone.utc)
                    if not bot_state['is_live_twin_running']: asyncio.create_task(run_live_twin_simulation(dt, dt + timedelta(days=1)))
                    await send_tg_msg(f"⏳ جاري Live-Twin ليوم {parts[1]}...")
                    return
                elif len(parts) == 3:
                    dt1 = datetime.strptime(parts[1], "%Y-%m-%d").replace(tzinfo=timezone.utc)
                    dt2 = datetime.strptime(parts[2], "%Y-%m-%d").replace(tzinfo=timezone.utc) + timedelta(days=1)
                    if not bot_state['is_live_twin_running']: asyncio.create_task(run_live_twin_simulation(dt1, dt2))
                    await send_tg_msg(f"⏳ جاري Live-Twin من {parts[1]} إلى {parts[2]}...")
                    return
            except Exception:
                await send_tg_msg("❌ <b>خطأ في التاريخ!</b>\nالصيغة: <code>/livetwin 2026-06-24</code>\nأو <code>/livetwin 2026-06-24 2026-06-26</code>")
                return

        if not msg.startswith('/') and msg in bot_state.get('menu_button_map', {}):
            cb = bot_state['menu_button_map'][msg]
            if cb != 'noop': await _handle_callback(cb, bot_state['chat_id'], None)
            return

        if msg.startswith('/setsymbol '):
            new_sym = msg.split(' ')[1].strip()
            bot_state['symbol'] = new_sym
            await save_bot_persistence()
            await send_tg_msg(f"✅ تم تغيير الرمز الخاص بـ MetaTrader إلى: <b>{new_sym}</b>")
        elif msg == '/start': await send_tg_msg('<b>مرحباً بك في Gold Scalper Bot v8.9</b>', get_main_keyboard())
        return

    if 'callback_query' not in update: return
    q = update['callback_query']; d = q['data']; chat_id = q['message']['chat']['id']; msg_id = q['message']['message_id']
    bot_state['chat_id'] = chat_id
    asyncio.create_task(answer_callback(q['id']))
    try: await _handle_callback(d, chat_id, msg_id)
    except Exception as e: log_exception(f'callback dispatch [{d}]', e)

_poll_task: asyncio.Task | None = None

async def telegram_polling_loop() -> None:
    c_log('Telegram polling started.'); url = f'https://api.telegram.org/bot{TG_TOKEN}/getUpdates'
    backoff = 1
    # Single persistent session for the lifetime of this task. Recreating
    # a ClientSession + TCPConnector on every backoff cycle leaked sockets
    # into TIME_WAIT over long uptimes. We only ever tear this down once,
    # in the finally block below, on task cancellation/shutdown.
    connector = aiohttp.TCPConnector(limit=4, ttl_dns_cache=300, force_close=True)
    timeout = aiohttp.ClientTimeout(total=None, connect=10, sock_read=28)
    sess = aiohttp.ClientSession(connector=connector, timeout=timeout)
    try:
        while True:
            try:
                async with sess.get(url, params={'offset': bot_state['last_update_id'] + 1, 'timeout': 20}) as resp:
                    if resp.status == 200:
                        backoff = 1; bot_state['last_poll_ok'] = datetime.now(timezone.utc).timestamp()
                        data = await resp.json()
                        for upd in data.get('result', []):
                            bot_state['last_update_id'] = upd['update_id']
                            asyncio.create_task(process_tg_update(upd))
                    else:
                        await asyncio.sleep(backoff); backoff = min(backoff * 2, 30)
            except asyncio.CancelledError:
                raise
            except Exception as e:
                log_exception('telegram_polling_loop', e)
                await asyncio.sleep(backoff); backoff = min(backoff * 2, 30)
    finally:
        await sess.close()

async def telegram_watchdog() -> None:
    global _poll_task
    await asyncio.sleep(30)
    while True:
        await asyncio.sleep(20)
        last = bot_state.get('last_poll_ok', 0.0); age = datetime.now(timezone.utc).timestamp() - last
        if age > 60 and _poll_task is not None and not _poll_task.done(): _poll_task.cancel()

async def supervised(coro_fn, *args, label: str = '') -> None:
    global _poll_task
    while True:
        try:
            task = asyncio.current_task()
            if label == 'tg_polling': _poll_task = task
            await coro_fn(*args)
        except asyncio.CancelledError: await asyncio.sleep(2)   
        except Exception as e:
            log_exception(f'supervised task "{label}"', e)
            await asyncio.sleep(5)

# ─────────────────────────────────────────────────────────────
# ENTRY POINT & WEB SERVER
# ─────────────────────────────────────────────────────────────
async def handle_ping(request: web.Request) -> web.Response:
    return web.Response(text="Bot is running smoothly!")

async def main() -> None:
    get_http()
    await init_metaapi()
    app = web.Application()
    app.router.add_get('/', handle_ping)
    
    runner = web.AppRunner(app); await runner.setup()
    port = int(os.environ.get('PORT', 10000))
    await web.TCPSite(runner, '0.0.0.0', port).start()
    c_log(f'Web server started on port {port}')

    bot_state['last_poll_ok'] = datetime.now(timezone.utc).timestamp()

    tasks = [
        asyncio.create_task(supervised(telegram_polling_loop, label='tg_polling')),
        asyncio.create_task(supervised(telegram_watchdog,     label='tg_watchdog')),
        asyncio.create_task(supervised(gann_monitor_scanner,  label='gann_monitor')),
        asyncio.create_task(supervised(gann_cycle_manager,    label='gann_cycle')),
        asyncio.create_task(supervised(global_ledger_reconciliation, label='global_reconciliation')),
    ]
    
    c_log('Gold Scalper Bot v9.4 (Resilience-First Core) started successfully.')
    try: await asyncio.gather(*tasks)
    finally:
        if _http and not _http.closed: await _http.close()

if __name__ == '__main__':
    asyncio.run(main())
