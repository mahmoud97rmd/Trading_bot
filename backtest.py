"""
backtest.py — Backtest and Live-Twin simulation engines.

Owns:
  - run_gann_backtest (idealized zero-friction engine)
  - run_live_twin_simulation (friction-aware historical engine)
  - run_live_twin_forward (real-time forward paper trader)
  - BtProgress (Telegram progress bar)
  - _build_gann_cycle_defs (shared cycle-anchoring logic)
  - Live-Twin friction model (_lt_bridge_path, _lt_slippage, etc.)
  - _forward_portfolio / _forward_stats (forward paper-trade state)
  - Excel export formatting
"""

import asyncio
import os
import random
import time
import zlib
from datetime import datetime, timedelta, timezone

import numpy as np
import openpyxl
import pandas as pd
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from state import (
    bot_state, SYMBOL_INFO, _TFS, DAM_OFF, _safe_float, log_exception, c_log,
    _safe_task, get_http, TG_TOKEN,
)
from market_data import fetch_candles, _lq_price_with_fallback, live_quotes
from strategy import (
    gann_calc_levels, gann_active_levels, _gann_atr, _gann_tf_tp,
    _gann_tf_sl, core_eval_break_even, _anchor_hours, _anchor_label,
    _last_closed_anchor_time_utc, _is_within_dam_restricted_window,
    _DAM_RESTRICTED_WINDOWS,
)
from state import DAM_OFF


# ── Gann Dynamic Recalc Cadence ──
GANN_DYNAMIC_RECALC_MINUTES = 5


def _build_gann_cycle_defs(sym_state: dict, valid_h1: list, mc_1m: list) -> list[dict]:
    mode = bot_state.get('gann_calculation_mode', 'static_h1')
    cycle_h = sym_state['gann_cycle_hours']
    if mode != 'dynamic_live':
        out = []
        for h1 in valid_h1:
            t_start = h1['time'] + timedelta(hours=1)
            out.append({'t_start': t_start, 't_end': t_start + timedelta(hours=cycle_h),
                         'close': float(h1['close'])})
        return out
    px_series = sorted(mc_1m, key=lambda c: c['time']) if mc_1m else []
    if not px_series:
        return []
    out = []
    bucket = px_series[0]['time'].floor(f'{GANN_DYNAMIC_RECALC_MINUTES}min')
    last_t = px_series[-1]['time']
    i = 0; n = len(px_series)
    while bucket <= last_t:
        while i + 1 < n and px_series[i + 1]['time'] <= bucket:
            i += 1
        if px_series[i]['time'] <= bucket:
            out.append({'t_start': bucket,
                         't_end': bucket + timedelta(minutes=GANN_DYNAMIC_RECALC_MINUTES),
                         'close': float(px_series[i]['close'])})
        bucket += timedelta(minutes=GANN_DYNAMIC_RECALC_MINUTES)
    return out


# ── Backtest Progress Tracker ──
class BtProgress:
    BAR_LEN = 14; HEARTBEAT = 15

    def __init__(self, label: str, active_tfs: list):
        self.label = label; self.active_tfs = active_tfs; self.cancelled = False
        self.phase = 'Initialising...'; self.tf_done = 0; self.tf_total = len(active_tfs)
        self.current_tf = ''; self.bars_done = 0; self.bars_total = 0
        self.win = 0; self.loss = 0; self.be = 0; self.profit = 0.0
        self.chat_id = None; self.msg_id = None; self._last_edit = 0.0
        self._lock = asyncio.Lock(); self._hb_task = None; self._start_ts = 0.0

    def _bar(self, done: int, total: int) -> str:
        if total == 0: return chr(9617) * self.BAR_LEN
        filled = round(done / total * self.BAR_LEN)
        return chr(9608) * filled + chr(9617) * (self.BAR_LEN - filled)

    def _elapsed(self) -> str:
        secs = int(datetime.now(timezone.utc).timestamp() - self._start_ts)
        m, s = divmod(secs, 60); return f'{m}m {s:02d}s'

    def _build_text(self) -> str:
        total = self.win + self.loss
        wr = f'{round(self.win / total * 100)}%' if total else '-'
        pnl = f'+${round(self.profit,2)}' if self.profit >= 0 else f'-${abs(round(self.profit,2))}'
        icon = '▲' if self.profit >= 0 else '▼'
        overall = (self.tf_done + self.bars_done / self.bars_total) / max(self.tf_total, 1) if self.bars_total else self.tf_done / max(self.tf_total, 1)
        ov_bar = self._bar(round(overall * 100), 100); ov_pct = f'{round(overall * 100)}%'
        tf_bar = self._bar(self.bars_done, self.bars_total) if self.bars_total else chr(9617) * self.BAR_LEN
        tf_pct = f'{round(self.bars_done / self.bars_total * 100)}%' if self.bars_total else '-'
        lines = [f'Backtest — <b>{self.label}</b>', f'<b>Phase:</b> {self.phase}', '',
                 f'<b>Overall</b>  {ov_pct}', f'<code>[{ov_bar}]</code>']
        if self.current_tf:
            lines += ['', f'<b>TF:</b> {self.current_tf}  ({self.tf_done}/{self.tf_total})',
                      f'<code>[{tf_bar}] {tf_pct}</code>', f'Bars: {self.bars_done}/{self.bars_total}']
        lines += ['', f'W:{self.win}  L:{self.loss}  BE:{self.be}', f'{icon} {pnl}  WR:{wr}', '',
                  f'Elapsed: {self._elapsed()}']
        if self.cancelled: lines.append('<b>CANCELLED</b>')
        return '\n'.join(lines)

    async def start(self, chat_id: int) -> None:
        self.chat_id = chat_id
        self._start_ts = datetime.now(timezone.utc).timestamp()
        self._last_edit = self._start_ts
        payload = {'chat_id': chat_id, 'text': self._build_text(), 'parse_mode': 'HTML',
                   'reply_markup': {'inline_keyboard': [[{'text': '⏹ Cancel', 'callback_data': 'cancel_bt'}]]}}
        try:
            import aiohttp
            from state import TG_TOKEN
            async with aiohttp.ClientSession() as sess:
                async with sess.post(f'https://api.telegram.org/bot{TG_TOKEN}/sendMessage', json=payload) as resp:
                    if resp.status == 200:
                        self.msg_id = (await resp.json())['result']['message_id']
        except Exception:
            pass
        self._hb_task = asyncio.create_task(self._heartbeat())

    async def _heartbeat(self) -> None:
        while not self.cancelled:
            await asyncio.sleep(self.HEARTBEAT)
            await self._edit(force=True)

    async def _edit(self, force: bool = False) -> None:
        now = datetime.now(timezone.utc).timestamp()
        if not force and (now - self._last_edit) < 3:
            return
        if not self.msg_id or not self.chat_id:
            return
        async with self._lock:
            self._last_edit = now
            import aiohttp
            from state import TG_TOKEN
            payload = {'chat_id': self.chat_id, 'message_id': self.msg_id,
                       'text': self._build_text(), 'parse_mode': 'HTML'}
            if not self.cancelled:
                payload['reply_markup'] = {'inline_keyboard': [[{'text': '⏹ Cancel', 'callback_data': 'cancel_bt'}]]}
            try:
                async with aiohttp.ClientSession() as sess:
                    await sess.post(f'https://api.telegram.org/bot{TG_TOKEN}/editMessageText', json=payload)
            except Exception:
                pass

    async def set_phase(self, phase: str) -> None: self.phase = phase; await self._edit()
    async def set_tf(self, tf: str, bars_total: int) -> None:
        self.current_tf = tf; self.bars_done = 0; self.bars_total = bars_total; await self._edit(force=True)
    async def tick(self, bar_n: int, win: int, loss: int, be: int, profit: float) -> None:
        self.bars_done = bar_n; self.win = win; self.loss = loss; self.be = be
        self.profit = profit; await self._edit()
    async def done(self, final_text: str) -> None:
        if self._hb_task: self._hb_task.cancel()
        if not self.msg_id or not self.chat_id: return
        try:
            import aiohttp
            from state import TG_TOKEN
            payload = {'chat_id': self.chat_id, 'message_id': self.msg_id,
                       'text': final_text, 'parse_mode': 'HTML'}
            async with aiohttp.ClientSession() as sess:
                await sess.post(f'https://api.telegram.org/bot{TG_TOKEN}/editMessageText', json=payload)
        except Exception:
            pass
    async def cancel(self) -> None:
        self.cancelled = True; self.phase = 'Cancelling...'
        if self._hb_task: self._hb_task.cancel()
        await self._edit(force=True)


_bt_progress: BtProgress | None = None
_lt_progress: BtProgress | None = None


# ── Utility ──
def _utc_to_dam(dt) -> datetime:
    if isinstance(dt, pd.Timestamp): dt = dt.to_pydatetime()
    if dt.tzinfo is None: dt = dt.replace(tzinfo=timezone.utc)
    return dt + DAM_OFF


# ── Concurrent Analysis Sheets ──
def _add_concurrent_analysis_sheets(wb, trade_logs: list, pnl_key: str,
                                     outcome_key: str, slippage_key: str = None) -> None:
    gray_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
    channel_sheets = {'touch': 'Touch_Trades', 'close': 'Close_Trades', 'hybrid': 'Hybrid_Trades'}
    channel_rows = {'touch': [], 'close': [], 'hybrid': []}
    for tr in trade_logs:
        ch = tr.get('trigger_type', 'touch')
        if ch in channel_rows:
            channel_rows[ch].append(tr)
    internal_keys = {'cycle_ts', 'cycle_time_str', 'cycle_close', 'trigger_type'}
    cols = [k for k in trade_logs[0].keys() if k not in internal_keys] if trade_logs else []
    made_sheets = []
    for ch, sheet_name in channel_sheets.items():
        ws = wb.create_sheet(sheet_name); made_sheets.append(ws)
        ws.append(cols)
        for cell in ws[1]: cell.fill = gray_fill; cell.font = Font(bold=True)
        for tr in channel_rows[ch]:
            ws.append([tr.get(c) for c in cols])
        for i in range(1, len(cols) + 1):
            ws.column_dimensions[get_column_letter(i)].width = 20.0
    ws_cmp = wb.create_sheet("Performance_Comparison"); made_sheets.append(ws_cmp)
    ws_cmp.append(["Metric", "Touch (لمس مباشر)", "Close (إغلاق شمعة)", "Hybrid (هجين)"])
    for cell in ws_cmp[1]: cell.fill = gray_fill; cell.font = Font(bold=True)

    def metrics_for(rows):
        wins = [r for r in rows if r.get(outcome_key) == 'WIN']
        losses = [r for r in rows if r.get(outcome_key) == 'LOSS']
        gross_profit = sum((r.get(pnl_key) or 0) for r in wins)
        gross_loss = sum((r.get(pnl_key) or 0) for r in losses)
        net = sum((r.get(pnl_key) or 0) for r in rows)
        wr = round(100 * len(wins) / max(1, len(wins) + len(losses)), 1)
        rows_sorted = sorted(rows, key=lambda r: r.get('cycle_ts', 0))
        eq = 0.0; peak = 0.0; mdd = 0.0
        for r in rows_sorted:
            eq += (r.get(pnl_key) or 0); peak = max(peak, eq); mdd = min(mdd, eq - peak)
        avg_slip = None
        if slippage_key:
            slips = [r.get(slippage_key) for r in rows if r.get(slippage_key) is not None]
            avg_slip = round(sum(slips) / len(slips), 2) if slips else None
        return dict(total=len(rows), win=len(wins), loss=len(losses), wr=wr,
                    gp=round(gross_profit, 2), gl=round(gross_loss, 2), net=round(net, 2),
                    mdd=round(mdd, 2), avg_slip=avg_slip)
    m = {ch: metrics_for(channel_rows[ch]) for ch in channel_rows}
    rows_spec = [
        ("Total Trades", 'total'), ("Winning Trades", 'win'), ("Losing Trades", 'loss'),
        ("Win Rate (%)", 'wr'), ("Gross Profit ($)", 'gp'), ("Gross Loss ($)", 'gl'),
        ("Net PnL ($)", 'net'), ("Max Drawdown ($)", 'mdd'),
    ]
    if slippage_key: rows_spec.append(("Average Slippage (Pips)", 'avg_slip'))
    for label, key in rows_spec:
        ws_cmp.append([label, m['touch'][key], m['close'][key], m['hybrid'][key]])
    for i in range(1, 5): ws_cmp.column_dimensions[get_column_letter(i)].width = 26.0
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                          top=Side(style='thin'), bottom=Side(style='thin'))
    center_align = Alignment(horizontal='center', vertical='center')
    for ws in made_sheets:
        for row in ws.iter_rows():
            for cell in row: cell.border = thin_border; cell.alignment = center_align


# ── Idealized Backtest Engine ──
async def run_gann_backtest(start_dt: datetime, end_dt: datetime) -> None:
    """Zero-friction Gann backtest (original engine, untouched logic)."""
    global _bt_progress
    bot_state['is_backtesting'] = True
    fname = f"GannBT_{datetime.now(timezone.utc).strftime('%H%M%S')}.xlsx"
    exec_mode = bot_state.get('gann_execution_mode', 'instant')

    active_symbols = [s for s, on in bot_state['active_symbols'].items() if on]
    if not active_symbols:
        bot_state['is_backtesting'] = False
        return

    from state import _PRESET_EXCLUDED_KEYS
    from telegram_ui import send_tg_document, send_tg_msg

    first_sym_state = bot_state['symbol_state'][active_symbols[0]]
    enabled_tfs = [tf for tf, on in first_sym_state['gann_monitor_tfs'].items() if on] or ['5m']
    flt_type = first_sym_state['trend_filter_type']
    ttf = first_sym_state['trend_timeframe']
    desc_ttf = ttf.upper()

    if first_sym_state['gann_entry_mode'] == 'touch_trend':
        if flt_type == 'vwap': desc_mode = f"Touch(VWAP{first_sym_state['trend_vwap_period']}_{desc_ttf})\n"
        elif flt_type == 'ema': desc_mode = f"Touch(EMA{first_sym_state['trend_ema_period']}_{desc_ttf})\n"
        else: desc_mode = f"Touch(VWAP+EMA_{desc_ttf})\n"
    else:
        desc_mode = "Pure Touch"
    desc_be = " | 🛡️ BE" if first_sym_state['break_even_enabled'] else ""
    zf = first_sym_state['gann_zone_filter']
    if zf == 'star': desc_star = "⭐ الأصلية"
    elif zf == 'star_fan': desc_star = "⭐🌀 الأصلية والمروحة"
    else: desc_star = "📋 الكل"
    desc_tfs = "+".join(enabled_tfs); syms_label = "+".join(active_symbols)

    prog = BtProgress(label=f"{syms_label} جان H1→[{desc_tfs}] | {desc_mode} | {desc_star}{desc_be}", active_tfs=['H1'])
    _bt_progress = prog
    await prog.start(bot_state['chat_id'])

    res = {'win': 0, 'loss': 0, 'be': 0, 'total_prof': 0.0, 'total_win_usd': 0.0,
           'total_loss_usd': 0.0, 'peak_equity': 0.0, 'max_dd': 0.0,
           'trade_logs': [], 'cycle_logs': []}
    _earliest_1m_seen = {}

    try:
        delta_hours = int((end_dt - start_dt).total_seconds() / 3600)
        all_signals = []; all_candles_events = []

        for symbol in active_symbols:
            sym_state = bot_state['symbol_state'][symbol]
            cycle_h = sym_state['gann_cycle_hours']; tpsl_mode = sym_state['gann_tpsl_mode']
            pv = SYMBOL_INFO[symbol]['pip_value']; lot = sym_state['lot_size']
            margin = sym_state['gann_touch_margin_pts'] * pv
            cs = SYMBOL_INFO[symbol]['contract_size']
            prec = SYMBOL_INFO[symbol]['prec']

            quote = symbol.split('_')[1] if '_' in symbol else 'USD'
            _QUOTE_RATES = {'USD': 1.0, 'JPY': 1/150.0, 'AUD': 0.66, 'NZD': 0.61,
                            'EUR': 1.08, 'GBP': 1.27, 'CAD': 0.73, 'CHF': 1.11}
            quote_conv = _QUOTE_RATES.get(quote)
            if quote_conv is None:
                c_log(f"WARNING: unknown quote currency '{quote}' in {symbol}")
                quote_conv = 1.0

            await prog.set_phase(f'جلب بيانات الترند ({desc_ttf})...')
            max_period = max(sym_state['trend_vwap_period'], sym_state['trend_ema_period'], 100)
            trend_count = (delta_hours * (2 if ttf == '30m' else 1)) + max_period + 10
            candles_trend = await fetch_candles(symbol, ttf, count=trend_count, end_time=end_dt)
            if not candles_trend: continue

            df_trend = pd.DataFrame(candles_trend)
            if flt_type == 'vwap':
                p_vwap = sym_state['trend_vwap_period']
                df_trend['Typical_Price'] = (df_trend['high'] + df_trend['low'] + df_trend['close']) / 3
                df_trend['VWAP'] = (df_trend['Typical_Price'] * df_trend['volume']).rolling(window=p_vwap).sum() / df_trend['volume'].rolling(window=p_vwap).sum()
            if flt_type == 'ema':
                p_ema = sym_state['trend_ema_period']
                df_trend['EMA'] = df_trend['close'].ewm(span=p_ema, adjust=False).mean()

            df_trend.set_index('time', inplace=True)
            if flt_type == 'vwap': df_trend['macro_trend_up'] = df_trend['close'] > df_trend['VWAP']
            elif flt_type == 'ema': df_trend['macro_trend_up'] = df_trend['close'] > df_trend['EMA']
            elif flt_type == 'both':
                c1_up = df_trend['close'] > df_trend['VWAP']; c2_up = df_trend['close'] > df_trend['EMA']
                c1_dn = df_trend['close'] < df_trend['VWAP']; c2_dn = df_trend['close'] < df_trend['EMA']
                df_trend['macro_trend_up'] = np.where(c1_up & c2_up, True, np.where(c1_dn & c2_dn, False, None))

            anchor_gran = bot_state.get('gann_anchor_tf', '1h')
            await prog.set_phase(f'جلب بيانات {_anchor_label()}...')
            candles_h1 = await fetch_candles(symbol, anchor_gran, count=(delta_hours // _anchor_hours()) + 10, end_time=end_dt)
            if not candles_h1: continue

            await prog.set_phase('جلب شموع الفريمات الصغيرة...')
            monitor_tfs_data = {}
            days_diff = (end_dt - start_dt).days or 1
            need_1m = days_diff * 24 * 60 + 300
            mc_1m = await fetch_candles(symbol, '1m', count=need_1m, end_time=end_dt)
            if mc_1m:
                _earliest_1m_seen[symbol] = min(c['time'] for c in mc_1m)
                for c in mc_1m:
                    all_candles_events.append({'time': c['time'], 'symbol': symbol,
                                                'high': float(c['high']), 'low': float(c['low']),
                                                'close': float(c['close']), 'tf': '1m_track'})

            for btf in enabled_tfs:
                bmin = int(''.join(filter(str.isdigit, btf)))
                if 'h' in btf: bmin *= 60
                need_m = days_diff * 24 * (60 // max(bmin, 1)) + 300
                mc = await fetch_candles(symbol, btf, count=need_m, end_time=end_dt)
                if mc:
                    monitor_tfs_data[btf] = sorted(mc, key=lambda c: c['time'])
                    for c in mc:
                        all_candles_events.append({'time': c['time'], 'symbol': symbol,
                                                    'high': float(c['high']), 'low': float(c['low']),
                                                    'close': float(c['close']), 'tf': btf})

            start_ts = start_dt.timestamp(); end_ts = end_dt.timestamp()
            valid_h1 = [c for c in candles_h1 if start_ts <= (c['time'].timestamp() + 3600) <= end_ts]
            trend_freq = '30min' if ttf == '30m' else '1h'
            cycle_defs = _build_gann_cycle_defs(sym_state, valid_h1, mc_1m)

            for idx, cdef in enumerate(cycle_defs):
                if prog.cancelled: return
                await asyncio.sleep(0)
                t_start = cdef['t_start']; t_end = cdef['t_end']; close = cdef['close']
                levels = gann_calc_levels(symbol, close)
                f_mode = sym_state['gann_zone_filter']
                active_lv = [l for l in levels if l['dir'] != 'ref' and (
                    f_mode == 'all' or (f_mode == 'star' and l['star']) or
                    (f_mode == 'star_fan' and (l['star'] or l['fan'])))]
                res['cycle_logs'].append({'symbol': symbol, 'time_ts': t_start.timestamp(),
                                           'time_dt': t_start, 'close': close, 'levels': len(active_lv)})
                level_used = set()
                exec_mode_bt = bot_state.get('gann_execution_mode', 'instant')
                spike_limit = bot_state.get('gann_spike_limit_pts', 20) * pv

                for btf, candles_m in monitor_tfs_data.items():
                    m_window = [c for c in candles_m if t_start <= c['time'] < t_end]
                    m_before = [c for c in candles_m if c['time'] < t_start]
                    atr_val = _gann_atr(m_before, sym_state['gann_atr_period']) if tpsl_mode == 'atr' else None
                    for bar in m_window:
                        bar_close = float(bar['close']); bar_time = bar['time']
                        bar_high = float(bar['high']); bar_low = float(bar['low'])
                        trend_up = True
                        if sym_state['gann_entry_mode'] == 'touch_trend':
                            trend_time = bar_time.floor(trend_freq)
                            if trend_time in df_trend.index:
                                val = df_trend.loc[trend_time, 'macro_trend_up']
                                if isinstance(val, pd.Series): val = val.iloc[-1]
                                macro_trend_up = None if pd.isna(val) else bool(val)
                            else: macro_trend_up = None
                            if macro_trend_up is None: continue
                            trend_up = macro_trend_up

                        if bot_state.get('prot_cycle_inval', True):
                            inval_pts = bot_state.get('prot_cycle_inval_pts', 200) * pv
                            if abs(bar_close - close) > inval_pts:
                                active_lv = []; break

                        for lv in active_lv:
                            k = lv['key']; dir = lv['dir']; is_buy = (dir == 'dn')
                            if sym_state['gann_entry_mode'] == 'touch_trend':
                                if is_buy and not trend_up: continue
                                if not is_buy and trend_up: continue

                            channels = ['touch', 'close', 'hybrid'] if exec_mode_bt == 'all_concurrent' else [
                                'close' if exec_mode_bt == 'close' else 'hybrid' if exec_mode_bt == 'hybrid' else 'touch']
                            for channel in channels:
                                base_combo = f'{k}_{btf}' if bot_state.get('prot_allow_multi_tf', True) else k
                                combo_key = f"{base_combo}_{channel}" if exec_mode_bt == 'all_concurrent' else base_combo
                                if combo_key in level_used: continue
                                if channel == 'close':
                                    if abs(bar_close - lv['price']) > margin: continue
                                elif channel == 'hybrid':
                                    if not (bar_low - margin <= lv['price'] <= bar_high + margin): continue
                                    if bot_state.get('prot_spike_filter', True) and abs(bar_close - lv['price']) > spike_limit: continue
                                else:
                                    if not (bar_low - margin <= lv['price'] <= bar_high + margin): continue

                                entry = lv['price']
                                tf_tp = _gann_tf_tp(symbol, btf); tf_sl = _gann_tf_sl(symbol, btf)
                                if tpsl_mode == 'atr' and atr_val:
                                    sl_d = atr_val * sym_state['gann_atr_sl_mult']
                                    tp_d = atr_val * sym_state['gann_atr_tp_mult']
                                else:
                                    sl_d = tf_sl * pv; tp_d = tf_tp * pv
                                tp_px = entry + tp_d if is_buy else entry - tp_d
                                sl_px = entry - sl_d if is_buy else entry + sl_d
                                all_signals.append({
                                    'time': bar_time, 'symbol': symbol, 'is_buy': is_buy,
                                    'entry': entry, 'tp_px': tp_px, 'sl_px': sl_px,
                                    'sl_d': sl_d, 'tp_d': tp_d,
                                    'be_trigger_px': 'dynamic' if sym_state['break_even_enabled'] else None,
                                    'lot': lot, 'cs': cs, 'quote_conv': quote_conv,
                                    'tf': btf, 'combo_key': combo_key,
                                    'cycle_time': t_start, 'cycle_close': close,
                                    'level_key': k, 'trigger_type': channel,
                                })
                                level_used.add(combo_key)

        # PHASE 2: Event-driven simulation
        await prog.set_phase('محاكاة الصفقات الزمنية...')
        all_signals.sort(key=lambda x: x['time'])
        all_candles_events.sort(key=lambda x: x['time'])
        open_trades = []; closed_trades = []
        suspended_days = {}; suspend_trigger_time = {}
        daily_pl = 0.0; current_day = None; latest_price = {}
        signal_idx = 0; total_signals = len(all_signals)
        dd_limit = -float(bot_state['prot_daily_dd_usd'])
        profit_limit = float(bot_state['prot_daily_profit_usd'])
        total_events = len(all_candles_events)
        await prog.set_tf('محاكاة عائمة', total_events)

        for i, event in enumerate(all_candles_events):
            if i % 5000 == 0: await asyncio.sleep(0)
            if prog.cancelled: break
            t = event['time']; sym = event['symbol']
            h = event['high']; l = event['low']; c = event['close']
            day_str = _utc_to_dam(t).strftime('%Y-%m-%d')
            latest_price[sym] = c
            if day_str != current_day: current_day = day_str; daily_pl = 0.0

            if current_day not in suspended_days:
                floating_pl = 0.0; floating_pl_worst = 0.0
                for tr in open_trades:
                    lp = latest_price.get(tr['symbol'], tr['entry'])
                    diff = (lp - tr['entry']) if tr['is_buy'] else (tr['entry'] - lp)
                    floating_pl += round(diff * tr['lot'] * tr['cs'] * tr['quote_conv'], 2)
                    worst_px = (l if tr['is_buy'] else h) if tr['symbol'] == sym else lp
                    diff_worst = (worst_px - tr['entry']) if tr['is_buy'] else (tr['entry'] - worst_px)
                    floating_pl_worst += round(diff_worst * tr['lot'] * tr['cs'] * tr['quote_conv'], 2)
                total_daily = daily_pl + floating_pl
                total_daily_worst = daily_pl + floating_pl_worst
                if dd_limit < 0 and total_daily_worst <= dd_limit:
                    suspended_days[current_day] = f'🛑 تراجع عائم (الحد {dd_limit}$)'
                elif profit_limit > 0 and total_daily >= profit_limit:
                    suspended_days[current_day] = f'✅ هدف عائم (الحد {profit_limit}$)'
                if current_day in suspended_days and current_day not in suspend_trigger_time:
                    suspend_trigger_time[current_day] = t
                if current_day in suspended_days:
                    for tr in open_trades:
                        lp = (l if tr['is_buy'] else h) if tr['symbol'] == sym else latest_price.get(tr['symbol'], tr['entry'])
                        diff = (lp - tr['entry']) if tr['is_buy'] else (tr['entry'] - lp)
                        p_usd = round(diff * tr['lot'] * tr['cs'] * tr['quote_conv'], 2)
                        tr['outcome'] = 'DAILY_LIMIT'; tr['p_usd'] = p_usd
                        tr['close_time'] = t; closed_trades.append(tr); daily_pl += p_usd
                    open_trades.clear()

            if current_day not in suspended_days:
                surviving = []
                for tr in open_trades:
                    if tr['symbol'] != sym: surviving.append(tr); continue
                    is_buy = tr['is_buy']; sl_current = tr['sl_current']; entry = tr['entry']
                    tp_px = tr['tp_px']; sl_d = tr['sl_d']; lot = tr['lot']
                    cs = tr['cs']; quote_conv = tr['quote_conv']
                    closed = False
                    pv_sym = SYMBOL_INFO[sym]['pip_value']
                    be_pts = bot_state['symbol_state'][sym].get('gann_be_trigger_points', 40)
                    atr_per = bot_state['symbol_state'][sym].get('gann_atr_period', 14)
                    cost_be_val = bot_state.get('prot_cost_be', True)

                    if not tr['be_activated'] and tr['be_trigger_px'] is not None:
                        test_px = h if is_buy else l
                        net_be = core_eval_break_even(is_buy, entry, test_px, pv_sym, be_pts, atr_per, cost_be_val)
                        if net_be is not None: tr['sl_current'] = net_be; tr['be_activated'] = True
                    _be_thresh = pv_sym * 2
                    if is_buy:
                        if l <= sl_current:
                            tr['outcome'] = 'BREAK_EVEN' if sl_current > entry - _be_thresh else 'LOSS'
                            tr['p_usd'] = round(abs(sl_current - entry) * lot * cs * quote_conv, 2) if tr['outcome'] == 'BREAK_EVEN' else -round(sl_d * lot * cs * quote_conv, 2)
                            closed = True
                        elif h >= tp_px:
                            tr['outcome'] = 'WIN'
                            tr['p_usd'] = round(tr['tp_d'] * lot * cs * quote_conv, 2); closed = True
                    else:
                        if h >= sl_current:
                            tr['outcome'] = 'BREAK_EVEN' if sl_current < entry + _be_thresh else 'LOSS'
                            tr['p_usd'] = round(abs(entry - sl_current) * lot * cs * quote_conv, 2) if tr['outcome'] == 'BREAK_EVEN' else -round(sl_d * lot * cs * quote_conv, 2)
                            closed = True
                        elif l <= tp_px:
                            tr['outcome'] = 'WIN'
                            tr['p_usd'] = round(tr['tp_d'] * lot * cs * quote_conv, 2); closed = True
                    if closed: tr['close_time'] = t; daily_pl += tr['p_usd']; closed_trades.append(tr)
                    else: surviving.append(tr)
                open_trades = surviving

            while signal_idx < total_signals and all_signals[signal_idx]['time'] <= t:
                sig = all_signals[signal_idx]; signal_idx += 1
                if current_day in suspended_days: continue
                if bot_state.get('prot_dam_time_filter', True):
                    sig_dam_time = (sig['time'] + timedelta(hours=3)).time()
                    if any(start <= sig_dam_time < end for start, end in _DAM_RESTRICTED_WINDOWS):
                        continue
                max_concurrent_bt = max(1, int(bot_state.get('prot_max_concurrent_trades', 4)))
                open_count_bt = sum(1 for tr in open_trades if tr['symbol'] == sig['symbol'])
                if open_count_bt >= max_concurrent_bt: continue
                sig['sl_current'] = sig['sl_px']; sig['be_activated'] = False
                open_trades.append(sig)
            await prog.tick(i, res['win'], res['loss'], res['be'], res['total_prof'])

        # Post-process
        for tr in closed_trades:
            if tr['outcome'] == 'WIN' or (tr['outcome'] == 'DAILY_LIMIT' and tr['p_usd'] > 0):
                res['win'] += 1; res['total_win_usd'] += tr['p_usd']
            elif tr['outcome'] == 'LOSS' or (tr['outcome'] == 'DAILY_LIMIT' and tr['p_usd'] < 0):
                res['loss'] += 1; res['total_loss_usd'] += abs(tr['p_usd'])
            elif tr['outcome'] == 'BREAK_EVEN' or (tr['outcome'] == 'DAILY_LIMIT' and tr['p_usd'] == 0):
                res['be'] += 1
            res['total_prof'] += tr['p_usd']
            dir_str = 'BUY 📈' if tr['is_buy'] else 'SELL 📉'
            res['trade_logs'].append({
                'الزوج': tr['symbol'],
                'وقت الصفقة (DAM)': _utc_to_dam(tr['time']).strftime('%Y-%m-%d %H:%M'),
                'TF': tr['tf'], 'اتجاه': dir_str,
                'المستوى (الدخول)': f"{tr['entry']:.2f} ({tr['level_key']})",
                'الهدف (TP)': round(tr['tp_px'], 2),
                'الوقف (SL)': round(tr['sl_px'], 2),
                'النتيجة': tr['outcome'], 'ربح ($)': tr['p_usd'],
                'cycle_ts': tr['cycle_time'].timestamp(),
                'cycle_time_str': _utc_to_dam(tr['cycle_time']).strftime('%Y-%m-%d %H:%M'),
                'cycle_close': tr['cycle_close'],
                'trigger_type': tr.get('trigger_type', 'touch'),
            })
        res['trade_logs'].sort(key=lambda x: x['وقت الصفقة (DAM)'])
        running_eq = 5000.0; peak_eq = 5000.0; max_dd = 0.0
        for t_log in res['trade_logs']:
            running_eq += t_log['ربح ($)']; t_log['رصيد تراكمي ($)'] = round(running_eq, 2)
            if running_eq > peak_eq: peak_eq = running_eq
            dd = peak_eq - running_eq
            if dd > max_dd: max_dd = dd
        res['peak_equity'] = peak_eq; res['max_dd'] = max_dd

        if not res['trade_logs']:
            await prog.done('<b>باكتيست اكتمل ✅</b>\nلا توجد صفقات في هذا النطاق.')
            bot_state['is_backtesting'] = False; return

        await prog.set_phase('إنشاء ملف Excel...')
        sum_text = (
            f"<b>باكتيست جان اكتمل ✅</b>\n{syms_label} H1→[{desc_tfs}] | {desc_mode} | {desc_star}{desc_be}\n"
            f"{start_dt.strftime('%Y-%m-%d')} → {end_dt.strftime('%Y-%m-%d')}\n\n"
            f"Net: {'PROFIT ▲' if res['total_prof']>=0 else 'LOSS ▼'} ${round(res['total_prof'], 2)}\n"
            f"Win: +${round(res['total_win_usd'], 2)} ({res['win']})\n"
            f"Loss: -${round(res['total_loss_usd'], 2)} ({res['loss']})\n"
            f"Break-Even: $0 ({res['be']})\n"
            f"WR: {round(res['win']/max(1, res['win']+res['loss'])*100)}% ({len(res['trade_logs'])} صفقة)\n"
            f"Max DD: ${round(res['max_dd'],2)}\n"
            f"دورات H1: {len(res['cycle_logs'])} | Lot: {lot}"
        )
        # Excel generation
        wb = openpyxl.Workbook(); ws_trades = wb.active; ws_trades.title = "الصفقات"
        headers = ["الزوج", "وقت الصفقة (DAM)", "TF", "اتجاه", "المستوى (الدخول)",
                   "الهدف (TP)", "الوقف (SL)", "النتيجة", "ربح ($)", "رصيد تراكمي ($)"]
        ws_trades.append(headers)
        gray_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
        green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        be_fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")
        for cell in ws_trades[1]: cell.fill = gray_fill; cell.font = Font(bold=True)
        for tr in res['trade_logs']:
            _OUTCOME_DISPLAY = {'WIN': 'WIN ✅', 'LOSS': 'LOSS ❌', 'BREAK_EVEN': 'BREAK_EVEN ⚖️', 'DAILY_LIMIT': 'DAILY_LIMIT ⏹️'}
            row_data = [tr['الزوج'], tr['وقت الصفقة (DAM)'], tr['TF'], tr['اتجاه'],
                        tr['المستوى (الدخول)'], tr['الهدف (TP)'], tr['الوقف (SL)'],
                        _OUTCOME_DISPLAY.get(tr['النتيجة'], tr['النتيجة']), tr['ربح ($)'], tr['رصيد تراكمي ($)']]
            ws_trades.append(row_data)
            fill = green_fill if tr['النتيجة'] == 'WIN' else red_fill if tr['النتيجة'] == 'LOSS' else be_fill if tr['النتيجة'] == 'BREAK_EVEN' else None
            if fill:
                for col in range(1, 11): ws_trades.cell(row=ws_trades.max_row, column=col).fill = fill
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                              top=Side(style='thin'), bottom=Side(style='thin'))
        center_align = Alignment(horizontal='center', vertical='center')
        for row in ws_trades.iter_rows():
            for cell in row: cell.border = thin_border; cell.alignment = center_align
        for i in range(1, 11): ws_trades.column_dimensions[get_column_letter(i)].width = 22.0
        if exec_mode == 'all_concurrent':
            _add_concurrent_analysis_sheets(wb, res['trade_logs'], pnl_key='ربح ($)', outcome_key='النتيجة')
        wb.save(fname)
        await prog.done(f'<b>باكتيست جان اكتمل ✅</b>\n{syms_label} — {len(res["trade_logs"])} صفقة\nجاري إرسال التقرير...')
        await send_tg_document(fname, sum_text)
        if os.path.exists(fname): os.remove(fname)
    except Exception as e:
        c_log(f'BT Error: {e}'); bot_state['is_backtesting'] = False
        if _bt_progress:
            import html
            try: await _bt_progress.done(f'❌ خطأ داخلي:\n{html.escape(str(e))}')
            except Exception: pass
    finally:
        bot_state['is_backtesting'] = False


# ═══════════════════════════════════════════════════════════════
# LIVE-TWIN FRICTION MODEL
# ═══════════════════════════════════════════════════════════════

def _lt_session_multiplier(dt_utc: datetime) -> tuple[float, bool]:
    hm = dt_utc.hour + dt_utc.minute / 60.0
    if (21 + 55/60) <= hm <= (22 + 5/60):
        return 3.5, True
    if 12.0 <= hm < 16.0:
        return 0.55, False
    if 7.0 <= hm < 12.0:
        return 0.70, False
    if 16.0 <= hm < 20.0:
        return 0.75, False
    return 1.00, False


def _lt_volatility_multiplier(bar_range: float, atr_val: float | None) -> float:
    if not atr_val or atr_val <= 0:
        return 1.0
    ratio = bar_range / atr_val
    if ratio <= 1.2:
        return 1.0
    return min(1.0 + (ratio - 1.2) * 0.9, 3.5)


def _lt_current_spread(base_spread: float, dt_utc: datetime, bar_range: float, atr_val: float | None) -> tuple[float, bool]:
    sess_mult, is_rollover = _lt_session_multiplier(dt_utc)
    vol_mult = _lt_volatility_multiplier(bar_range, atr_val)
    spread = base_spread * sess_mult * vol_mult
    return max(spread, base_spread * 0.45), is_rollover


def _lt_bridge_path(o: float, h: float, l: float, c: float, steps: int, rng: random.Random) -> np.ndarray:
    incs = np.array([rng.gauss(0, 1) for _ in range(steps)])
    w = np.concatenate(([0.0], np.cumsum(incs)))
    t = np.linspace(0.0, 1.0, steps + 1)
    bridge = w - t * w[-1]
    bstd = float(np.std(bridge))
    rng_size = max(h - l, 1e-6)
    scale = (rng_size * 0.5) / bstd if bstd > 1e-9 else 0.0
    path = o + (c - o) * t + bridge * scale
    return np.clip(path, l, h)


def _lt_first_hit(path: np.ndarray, is_buy: bool, sl_px: float, tp_px: float) -> str | None:
    for px in path:
        if is_buy:
            if px <= sl_px: return 'sl'
            if px >= tp_px: return 'tp'
        else:
            if px >= sl_px: return 'sl'
            if px <= tp_px: return 'tp'
    return None


def _lt_slippage(bar_range: float, atr_val: float | None, rng: random.Random) -> float:
    # Realistic retail slippage for a small XAU order is typically a few
    # points, occasionally more in a fast market -- NOT a double-digit-dollar
    # move. The previous constants (base=ref*0.06, cap=ref*0.5) meant slippage
    # scaled almost linearly with, and routinely saturated at HALF of, the
    # current single 1-minute candle's range. Any genuinely volatile bar
    # (which a session like this one had many of) then produced $6-7+ of
    # "slippage" on nearly every trade -- 10-50x larger than realistic
    # execution slippage. Anchor the estimate to a small fixed floor (a
    # realistic worst-case tick-level slippage) plus a much smaller fraction
    # of volatility, and cap it far below the full bar range.
    ref = atr_val if atr_val and atr_val > 0 else max(bar_range, 0.05)
    base = max(ref * 0.008, 0.015)
    tail = abs(rng.gauss(0, base))
    return min(tail, ref * 0.08, 0.60)


def _lt_latency_shift(path: np.ndarray, steps: int, rng: random.Random) -> float:
    lo = bot_state.get('lt_latency_ms_min', 160)
    hi = bot_state.get('lt_latency_ms_max', 200)
    latency_ms = rng.randint(lo, hi)
    frac = min(latency_ms / 60000.0, 1.0)
    idx_f = frac * steps
    idx = int(idx_f)
    if idx >= steps:
        return float(path[-1])
    t = idx_f - idx
    return float(path[idx] * (1.0 - t) + path[idx + 1] * t)


# ═══════════════════════════════════════════════════════════════
# HISTORICAL LIVE-TWIN SIMULATOR (friction-aware backtest)
# ═══════════════════════════════════════════════════════════════

async def run_live_twin_simulation(start_dt: datetime, end_dt: datetime) -> None:
    global _lt_progress
    if bot_state.get('lt_mode') == 'idealized':
        await run_gann_backtest(start_dt, end_dt)
        return

    bot_state['is_live_twin_running'] = True
    fname = f"LiveTwin_{datetime.now(timezone.utc).strftime('%H%M%S')}.xlsx"
    exec_mode = bot_state.get('gann_execution_mode', 'instant')
    fric = bot_state['lt_friction']
    base_spread = float(bot_state['lt_base_spread_usd'])
    comm_per_lot = float(bot_state['lt_commission_per_lot'])
    swap_long_per_lot = float(bot_state.get('lt_swap_long_per_lot_night', -93.17))
    swap_short_per_lot = float(bot_state.get('lt_swap_short_per_lot_night', 21.68))
    swap_wed_mult = float(bot_state.get('lt_swap_wednesday_multiplier', 3.0))
    rej_prob = float(bot_state['lt_rejection_prob'])

    active_symbols = [s for s, on in bot_state['active_symbols'].items() if on]
    if not active_symbols:
        bot_state['is_live_twin_running'] = False
        return

    override_seed = bot_state.get('lt_seed')
    if override_seed is not None:
        seed_val = int(override_seed)
    else:
        seed_key = (tuple(sorted(active_symbols)), start_dt.isoformat(), end_dt.isoformat(),
                    bot_state.get('gann_execution_mode', 'instant'), tuple(sorted(fric.items())))
        seed_val = zlib.crc32(str(seed_key).encode())
    rng = random.Random(seed_val)

    first_sym_state = bot_state['symbol_state'][active_symbols[0]]
    enabled_tfs = [tf for tf, on in first_sym_state['gann_monitor_tfs'].items() if on] or ['5m']
    flt_type = first_sym_state['trend_filter_type']
    ttf = first_sym_state['trend_timeframe']
    syms_label = "+".join(active_symbols)
    on_tags = "+".join(k for k, v in fric.items() if v) or "none"

    prog = BtProgress(label=f"Live-Twin {syms_label} | friction:[{on_tags}]", active_tfs=['H1'])
    _lt_progress = prog
    await prog.start(bot_state['chat_id'])

    res = {'win': 0, 'loss': 0, 'be': 0, 'total_prof': 0.0, 'total_win_usd': 0.0, 'total_loss_usd': 0.0,
           'peak_equity': 0.0, 'max_dd': 0.0, 'trade_logs': [], 'total_commission': 0.0, 'total_swap': 0.0,
           'rejected': 0, 'gap_events': 0}
    _earliest_1m_seen = {}

    try:
        delta_hours = int((end_dt - start_dt).total_seconds() / 3600)
        all_signals = []
        m1_by_symbol = {}

        for symbol in active_symbols:
            sym_state = bot_state['symbol_state'][symbol]
            cycle_h = sym_state['gann_cycle_hours']; tpsl_mode = sym_state['gann_tpsl_mode']
            pv = SYMBOL_INFO[symbol]['pip_value']; lot = sym_state['lot_size']; margin = sym_state['gann_touch_margin_pts'] * pv
            cs = SYMBOL_INFO[symbol]['contract_size']

            quote = symbol.split('_')[1] if '_' in symbol else 'USD'
            _QUOTE_RATES = {'USD': 1.0, 'JPY': 1/150.0, 'AUD': 0.66, 'NZD': 0.61, 'EUR': 1.08, 'GBP': 1.27, 'CAD': 0.73, 'CHF': 1.11}
            quote_conv = _QUOTE_RATES.get(quote)
            if quote_conv is None:
                c_log(f"WARNING: unknown quote currency '{quote}' in {symbol}")
                quote_conv = 1.0

            await prog.set_phase(f'جلب بيانات الترند ({ttf.upper()})...')
            max_period = max(sym_state['trend_vwap_period'], sym_state['trend_ema_period'], 100)
            trend_count = (delta_hours * (2 if ttf == '30m' else 1)) + max_period + 10
            candles_trend = await fetch_candles(symbol, ttf, count=trend_count, end_time=end_dt)
            if not candles_trend: continue

            df_trend = pd.DataFrame(candles_trend)
            if flt_type == 'vwap':
                p_vwap = sym_state['trend_vwap_period']
                df_trend['Typical_Price'] = (df_trend['high'] + df_trend['low'] + df_trend['close']) / 3
                df_trend['VWAP'] = (df_trend['Typical_Price'] * df_trend['volume']).rolling(window=p_vwap).sum() / df_trend['volume'].rolling(window=p_vwap).sum()
            if flt_type == 'ema':
                p_ema = sym_state['trend_ema_period']
                df_trend['EMA'] = df_trend['close'].ewm(span=p_ema, adjust=False).mean()

            df_trend.set_index('time', inplace=True)
            if flt_type == 'vwap': df_trend['macro_trend_up'] = df_trend['close'] > df_trend['VWAP']
            elif flt_type == 'ema': df_trend['macro_trend_up'] = df_trend['close'] > df_trend['EMA']
            elif flt_type == 'both':
                c1_up = df_trend['close'] > df_trend['VWAP']; c2_up = df_trend['close'] > df_trend['EMA']
                c1_dn = df_trend['close'] < df_trend['VWAP']; c2_dn = df_trend['close'] < df_trend['EMA']
                df_trend['macro_trend_up'] = np.where(c1_up & c2_up, True, np.where(c1_dn & c2_dn, False, None))

            anchor_gran = bot_state.get('gann_anchor_tf', '1h')
            await prog.set_phase(f'جلب بيانات {_anchor_label()}...')
            candles_h1 = await fetch_candles(symbol, anchor_gran, count=(delta_hours // _anchor_hours()) + 10, end_time=end_dt)
            if not candles_h1: continue

            await prog.set_phase('جلب شموع الدقيقة الواحدة...')
            days_diff = (end_dt - start_dt).days or 1
            need_1m = days_diff * 24 * 60 + 300
            mc_1m = await fetch_candles(symbol, '1m', count=need_1m, end_time=end_dt)
            if not mc_1m: continue
            _earliest_1m_seen[symbol] = min(c['time'] for c in mc_1m)
            m1_by_symbol[symbol] = sorted(mc_1m, key=lambda c: c['time'])

            monitor_tfs_data = {}
            for btf in enabled_tfs:
                bmin = int(''.join(filter(str.isdigit, btf)))
                bmin = bmin * 60 if 'h' in btf else bmin
                need_m = days_diff * 24 * (60 // max(bmin, 1)) + 300
                mc = await fetch_candles(symbol, btf, count=need_m, end_time=end_dt)
                if mc: monitor_tfs_data[btf] = sorted(mc, key=lambda c: c['time'])

            start_ts = start_dt.timestamp(); end_ts = end_dt.timestamp()
            valid_h1 = [c for c in candles_h1 if start_ts <= (c['time'].timestamp() + 3600) <= end_ts]
            trend_freq = '30min' if ttf == '30m' else '1h'

            cycle_defs = _build_gann_cycle_defs(sym_state, valid_h1, mc_1m)

            for cdef in cycle_defs:
                if prog.cancelled: return
                await asyncio.sleep(0)
                t_start = cdef['t_start']
                t_end = cdef['t_end']
                close = cdef['close']
                levels = gann_calc_levels(symbol, close)
                f_mode = sym_state['gann_zone_filter']
                active_lv = [l for l in levels if l['dir'] != 'ref' and (f_mode == 'all' or (f_mode == 'star' and l['star']) or (f_mode == 'star_fan' and (l['star'] or l['fan'])))]
                level_used = set()

                for btf, candles_m in monitor_tfs_data.items():
                    m_window = [c for c in candles_m if t_start <= c['time'] < t_end]
                    m_before = [c for c in candles_m if c['time'] < t_start]
                    atr_val = _gann_atr(m_before, sym_state['gann_atr_period']) if tpsl_mode == 'atr' else None
                    prev_bar_close = float(m_before[-1]['close']) if m_before else None
                    exec_mode = bot_state.get('gann_execution_mode', 'instant')
                    spike_limit = bot_state.get('gann_spike_limit_pts', 20) * pv

                    for bar in m_window:
                        bar_open = float(bar['open']); bar_close = float(bar['close']); bar_time = bar['time']
                        bar_high = float(bar['high']); bar_low = float(bar['low'])
                        trend_up = True
                        if sym_state['gann_entry_mode'] == 'touch_trend':
                            trend_time = bar_time.floor(trend_freq)
                            if trend_time in df_trend.index:
                                val = df_trend.loc[trend_time, 'macro_trend_up']
                                if isinstance(val, pd.Series): val = val.iloc[-1]
                                macro_trend_up = None if pd.isna(val) else bool(val)
                            else: macro_trend_up = None
                            if macro_trend_up is None:
                                prev_bar_close = bar_close; continue
                            trend_up = macro_trend_up

                        if bot_state.get('prot_cycle_inval', True):
                            inval_pts = bot_state.get('prot_cycle_inval_pts', 200) * pv
                            if abs(bar_close - close) > inval_pts:
                                active_lv = []; break

                        for lv in active_lv:
                            k = lv['key']; dir = lv['dir']
                            is_buy = (dir == 'dn')
                            if sym_state['gann_entry_mode'] == 'touch_trend':
                                if is_buy and not trend_up: continue
                                if not is_buy and trend_up: continue

                            channels = ['touch', 'close', 'hybrid'] if exec_mode == 'all_concurrent' else [
                                'close' if exec_mode == 'close' else 'hybrid' if exec_mode == 'hybrid' else 'touch']

                            for channel in channels:
                                base_combo = f'{k}_{btf}' if bot_state.get('prot_allow_multi_tf', True) else k
                                combo_key = f"{base_combo}_{channel}" if exec_mode == 'all_concurrent' else base_combo
                                if combo_key in level_used: continue

                                if channel == 'close':
                                    if abs(bar_close - lv['price']) > margin: continue
                                elif channel == 'hybrid':
                                    if not (bar_low - margin <= lv['price'] <= bar_high + margin): continue
                                    if bot_state.get('prot_spike_filter', True) and prev_bar_close is not None and abs(bar_close - prev_bar_close) > spike_limit: continue
                                else:
                                    if not (bar_low - margin <= lv['price'] <= bar_high + margin): continue

                                entry = lv['price']
                                tf_tp = _gann_tf_tp(symbol, btf); tf_sl = _gann_tf_sl(symbol, btf)
                                if tpsl_mode == 'atr' and atr_val:
                                    sl_d = atr_val * sym_state['gann_atr_sl_mult']; tp_d = atr_val * sym_state['gann_atr_tp_mult']
                                else:
                                    sl_d = tf_sl * pv; tp_d = tf_tp * pv

                                all_signals.append({
                                    'time': bar_time, 'symbol': symbol, 'is_buy': is_buy, 'intended_entry': entry,
                                    'sl_d': sl_d, 'tp_d': tp_d, 'be_enabled': sym_state['break_even_enabled'],
                                    'lot': lot, 'cs': cs, 'quote_conv': quote_conv, 'tf': btf, 'combo_key': combo_key,
                                    'cycle_time': t_start, 'cycle_close': close, 'level_key': k, 'trigger_type': channel,
                                    'bar_o': bar_open, 'bar_h': bar_high, 'bar_l': bar_low, 'bar_c': bar_close,
                                })
                                level_used.add(combo_key)

                        prev_bar_close = bar_close

        await prog.set_phase('محاكاة التنفيذ الواقعي...')
        all_signals.sort(key=lambda x: x['time'])
        all_1m_events = sorted(
            [{'time': c['time'], 'symbol': sym, 'open': float(c['open']), 'high': float(c['high']),
              'low': float(c['low']), 'close': float(c['close'])}
             for sym, candles in m1_by_symbol.items() for c in candles],
            key=lambda x: x['time']
        )
        m1_lookup = {
            sym: {c['time']: {'open': float(c['open']), 'high': float(c['high']), 'low': float(c['low']), 'close': float(c['close'])}
                  for c in candles}
            for sym, candles in m1_by_symbol.items()
        }

        open_trades = []
        closed_trades = []
        suspended_days = {}
        suspend_trigger_time = {}
        daily_pl = 0.0
        current_day = None
        latest_price = {}

        signal_idx = 0
        total_signals = len(all_signals)
        dd_limit = -float(bot_state['prot_daily_dd_usd'])
        profit_limit = float(bot_state['prot_daily_profit_usd'])
        max_concurrent = max(1, int(bot_state.get('prot_max_concurrent_trades', 4)))

        total_events = len(all_1m_events)
        await prog.set_tf('محاكاة 1m واقعية', total_events)

        for i, ev in enumerate(all_1m_events):
            if i % 5000 == 0: await asyncio.sleep(0)
            if prog.cancelled: break
            t = ev['time']; sym = ev['symbol']; o = ev['open']; h = ev['high']; l = ev['low']; c = ev['close']
            bar_range = h - l
            day_str = _utc_to_dam(t).strftime('%Y-%m-%d')
            latest_price[sym] = c
            if day_str != current_day:
                current_day = day_str; daily_pl = 0.0

            atr_ref = bar_range if bar_range > 0 else 0.1
            spread_now, is_rollover = _lt_current_spread(base_spread, t, bar_range, atr_ref) if fric['spread'] else (base_spread, False)
            half_spread = spread_now / 2.0

            if fric['gaps'] and is_rollover:
                res['gap_events'] += 1

            if current_day not in suspended_days:
                floating_pl = 0.0; floating_pl_worst = 0.0
                for tr in open_trades:
                    lp = latest_price.get(tr['symbol'], tr['entry'])
                    diff = (lp - tr['entry']) if tr['is_buy'] else (tr['entry'] - lp)
                    floating_pl += round(diff * tr['lot'] * tr['cs'] * tr['quote_conv'], 2)
                    worst_px = (l if tr['is_buy'] else h) if tr['symbol'] == sym else lp
                    diff_worst = (worst_px - tr['entry']) if tr['is_buy'] else (tr['entry'] - worst_px)
                    floating_pl_worst += round(diff_worst * tr['lot'] * tr['cs'] * tr['quote_conv'], 2)

                total_daily = daily_pl + floating_pl
                total_daily_worst = daily_pl + floating_pl_worst
                if dd_limit < 0 and total_daily_worst <= dd_limit:
                    suspended_days[current_day] = f'🛑 تراجع عائم (الحد {dd_limit}$)'
                elif profit_limit > 0 and total_daily >= profit_limit:
                    suspended_days[current_day] = f'✅ هدف عائم (الحد {profit_limit}$)'
                if current_day in suspended_days and current_day not in suspend_trigger_time:
                    suspend_trigger_time[current_day] = t
                if current_day in suspended_days:
                    for tr in open_trades:
                        lp = (l if tr['is_buy'] else h) if tr['symbol'] == sym else latest_price.get(tr['symbol'], tr['entry'])
                        exit_spread = _lt_current_spread(base_spread, t, bar_range, atr_ref)[0] if fric['spread'] else base_spread
                        lp_adj = lp - (exit_spread/2.0 if tr['is_buy'] else -exit_spread/2.0) if fric['spread'] else lp
                        diff = (lp_adj - tr['entry']) if tr['is_buy'] else (tr['entry'] - lp_adj)
                        p_usd = round(diff * tr['lot'] * tr['cs'] * tr['quote_conv'], 2)
                        tr['outcome'] = 'DAILY_LIMIT'; tr['p_usd'] = p_usd; tr['close_time'] = t
                        closed_trades.append(tr); daily_pl += p_usd
                    open_trades.clear()

            if current_day not in suspended_days:
                surviving = []
                for tr in open_trades:
                    if tr['symbol'] != sym:
                        surviving.append(tr); continue

                    is_buy = tr['is_buy']; entry = tr['entry']; sl_current = tr['sl_current']; tp_px = tr['tp_px']
                    lot = tr['lot']; cs = tr['cs']; quote_conv = tr['quote_conv']
                    closed = False

                    if tr['be_enabled'] and not tr['be_activated']:
                        test_px = h if is_buy else l
                        be_pts = bot_state['symbol_state'][sym].get('gann_be_trigger_points', 40)
                        pv_sym = SYMBOL_INFO[sym]['pip_value']
                        atr_per = bot_state['symbol_state'][sym].get('gann_atr_period', 14)
                        cost_be = bot_state.get('prot_cost_be', True)
                        net_be = core_eval_break_even(is_buy, entry, test_px, pv_sym, be_pts, atr_per, cost_be)
                        if net_be is not None:
                            tr['sl_current'] = net_be; tr['be_activated'] = True; sl_current = net_be

                    hits_sl = (l <= sl_current) if is_buy else (h >= sl_current)
                    hits_tp = (h >= tp_px) if is_buy else (l <= tp_px)

                    outcome = None
                    if hits_sl and hits_tp:
                        path = _lt_bridge_path(o, h, l, c, steps=20, rng=rng)
                        outcome = _lt_first_hit(path, is_buy, sl_current, tp_px) or 'sl'
                    elif hits_sl:
                        outcome = 'sl'
                    elif hits_tp:
                        outcome = 'tp'

                    if outcome:
                        exit_spread = spread_now if fric['spread'] else 0.0
                        slip = _lt_slippage(bar_range, atr_ref, rng) if fric['slippage'] else 0.0
                        if outcome == 'sl':
                            raw_px = sl_current
                            fill_px = raw_px - (exit_spread/2.0 + slip) if is_buy else raw_px + (exit_spread/2.0 + slip)
                            _be_thresh = SYMBOL_INFO[tr['symbol']]['pip_value'] * 2
                            tr['outcome'] = 'BREAK_EVEN' if (is_buy and sl_current > entry - _be_thresh) or (not is_buy and sl_current < entry + _be_thresh) else 'LOSS'
                        else:
                            raw_px = tp_px
                            fill_px = raw_px - (exit_spread/2.0 + slip) if is_buy else raw_px + (exit_spread/2.0 + slip)
                            tr['outcome'] = 'WIN'
                        diff = (fill_px - entry) if is_buy else (entry - fill_px)
                        p_usd = round(diff * lot * cs * quote_conv, 2)
                        commission = comm_per_lot * lot if fric['commission'] else 0.0
                        nights = max((t.date() - tr['time'].date()).days, 0)
                        swap = 0.0
                        if fric['gaps'] and nights > 0:
                            per_night = swap_long_per_lot if is_buy else swap_short_per_lot
                            for i in range(nights):
                                d = tr['time'].date() + timedelta(days=i)
                                mult = swap_wed_mult if d.weekday() == 2 else 1.0
                                swap += per_night * mult
                            swap *= lot
                        p_usd_net = round(p_usd - commission + swap, 2)
                        if tr['outcome'] == 'WIN' and p_usd_net < 0:
                            tr['outcome'] = 'LOSS'
                        tr['p_usd'] = p_usd_net
                        res['total_commission'] += commission; res['total_swap'] += swap
                        tr['close_time'] = t; daily_pl += p_usd_net
                        closed_trades.append(tr)
                        closed = True
                    if not closed:
                        surviving.append(tr)
                open_trades = surviving

            while signal_idx < total_signals and all_signals[signal_idx]['time'] <= t:
                sig = all_signals[signal_idx]; signal_idx += 1
                if current_day in suspended_days:
                    continue
                if bot_state.get('prot_dam_time_filter', True):
                    sig_dam_time = (sig['time'] + timedelta(hours=3)).time()
                    if any(start <= sig_dam_time < end for start, end in _DAM_RESTRICTED_WINDOWS):
                        continue
                open_count = sum(1 for tr in open_trades if tr['symbol'] == sig['symbol'])
                if open_count >= max_concurrent:
                    continue
                if fric['rejection'] and rng.random() < rej_prob:
                    res['rejected'] += 1; continue

                so, sh, sl_, sc = sig['bar_o'], sig['bar_h'], sig['bar_l'], sig['bar_c']
                sig_bar_range = sh - sl_

                entry_spread, _ = _lt_current_spread(base_spread, sig['time'], sig_bar_range, sig_bar_range or 0.1) if fric['spread'] else (base_spread, False)
                path = _lt_bridge_path(so, sh, sl_, sc, steps=20, rng=rng)
                shifted_px = _lt_latency_shift(path, 20, rng) if fric['latency'] else sig['intended_entry']
                slip = _lt_slippage(sig_bar_range, sig_bar_range or 0.1, rng) if fric['slippage'] else 0.0
                fill_entry = shifted_px + (entry_spread/2.0 + slip) if sig['is_buy'] else shifted_px - (entry_spread/2.0 + slip)

                is_buy = sig['is_buy']
                tp_px = fill_entry + sig['tp_d'] if is_buy else fill_entry - sig['tp_d']
                sl_px = fill_entry - sig['sl_d'] if is_buy else fill_entry + sig['sl_d']

                open_trades.append({
                    **sig, 'entry': fill_entry, 'tp_px': tp_px, 'sl_px': sl_px, 'sl_current': sl_px,
                    'be_activated': False, 'tp_d': sig['tp_d'], 'sl_d': sig['sl_d'],
                })

            await prog.tick(i, res['win'], res['loss'], res['be'], res['total_prof'])

        for tr in closed_trades:
            if tr['outcome'] == 'WIN' or (tr['outcome'] == 'DAILY_LIMIT' and tr['p_usd'] > 0):
                res['win'] += 1; res['total_win_usd'] += tr['p_usd']
            elif tr['outcome'] == 'LOSS' or (tr['outcome'] == 'DAILY_LIMIT' and tr['p_usd'] < 0):
                res['loss'] += 1; res['total_loss_usd'] += abs(tr['p_usd'])
            elif tr['outcome'] == 'BREAK_EVEN' or (tr['outcome'] == 'DAILY_LIMIT' and tr['p_usd'] == 0):
                res['be'] += 1
            res['total_prof'] += tr['p_usd']
            dir_str = 'BUY 📈' if tr['is_buy'] else 'SELL 📉'
            slip_px = tr['entry'] - tr['intended_entry'] if tr['is_buy'] else tr['intended_entry'] - tr['entry']
            slip_pips = round(slip_px / SYMBOL_INFO[tr['symbol']]['pip_value'], 2)
            res['trade_logs'].append({
                'الزوج': tr['symbol'], 'وقت الصفقة (DAM)': _utc_to_dam(tr['time']).strftime('%Y-%m-%d %H:%M'),
                'TF': tr['tf'], 'اتجاه': dir_str, 'المستوى (الدخول الفعلي)': f"{tr['entry']:.2f} ({tr['level_key']})",
                'الهدف (TP)': round(tr['tp_px'], 2), 'الوقف (SL)': round(tr['sl_px'], 2),
                'النتيجة': tr['outcome'], 'ربح صافي ($)': tr['p_usd'], 'cycle_ts': tr['cycle_time'].timestamp(),
                'trigger_type': tr.get('trigger_type', 'touch'), 'انزلاق (نقطة)': slip_pips,
            })

        res['trade_logs'].sort(key=lambda x: x['وقت الصفقة (DAM)'])
        running_eq = 5000.0; peak_eq = 5000.0; max_dd = 0.0
        for t_log in res['trade_logs']:
            running_eq += t_log['ربح صافي ($)']; t_log['رصيد تراكمي ($)'] = round(running_eq, 2)
            if running_eq > peak_eq: peak_eq = running_eq
            dd = peak_eq - running_eq
            if dd > max_dd: max_dd = dd
        res['peak_equity'] = peak_eq; res['max_dd'] = max_dd

        if not res['trade_logs']:
            await prog.done('<b>Live-Twin اكتمل ✅</b>\nلا توجد صفقات في هذا النطاق.')
            bot_state['is_live_twin_running'] = False; return

        await prog.set_phase('إنشاء ملف Excel...')
        short_syms = []
        for sym, earliest in _earliest_1m_seen.items():
            earliest_dt = earliest if earliest.tzinfo else earliest.replace(tzinfo=timezone.utc)
            gap_days = (earliest_dt - start_dt).total_seconds() / 86400
            if gap_days > 1.0:
                short_syms.append(f"{sym}: البيانات الفعلية بدأت من {earliest_dt.strftime('%Y-%m-%d %H:%M')} "
                                   f"بدل {start_dt.strftime('%Y-%m-%d %H:%M')} (نقص {gap_days:.1f} يوم)")
        coverage_warning = ""
        if short_syms:
            coverage_warning = ("\n⚠️ <b>تحذير: تغطية بيانات ناقصة</b>\n"
                                 "الفترة المطلوبة أكبر مما استطعنا جلبه فعلياً من أوندا:\n"
                                 + "\n".join(short_syms) + "\n")
        sum_text = (
            f"<b>Live-Twin Engine اكتمل ✅ (واقعي)</b>\n"
            f"{syms_label} | friction: [{on_tags}]\n"
            f"{start_dt.strftime('%Y-%m-%d')} → {end_dt.strftime('%Y-%m-%d')}\n\n"
            f"Net: {'PROFIT ▲' if res['total_prof']>=0 else 'LOSS ▼'} ${round(res['total_prof'], 2)}\n"
            f"Win:  +${round(res['total_win_usd'], 2)} ({res['win']})\n"
            f"Loss: -${round(res['total_loss_usd'], 2)} ({res['loss']})\n"
            f"Break-Even: ({res['be']})\n"
            f"WR: {round(res['win']/max(1, res['win']+res['loss'])*100)}% ({len(res['trade_logs'])} صفقة)\n"
            f"Max DD: ${round(res['max_dd'],2)} ({round((res['max_dd']/max(1,res['peak_equity']))*100)}%)\n\n"
            f"عمولة إجمالية: -${round(res['total_commission'],2)} | سواب: ${round(res['total_swap'],2)}\n"
            f"صفقات مرفوضة (Requote): {res['rejected']} | نوافذ Rollover: {res['gap_events']}\n"
            f"Spread الأساسي: ${base_spread}"
            f"{coverage_warning}"
        )

        wb = openpyxl.Workbook()
        ws = wb.active; ws.title = "Live-Twin Trades"
        headers = ["الزوج", "وقت الصفقة (DAM)", "TF", "اتجاه", "الدخول الفعلي", "TP", "SL", "النتيجة", "ربح صافي ($)", "رصيد تراكمي ($)"]
        ws.append(headers)
        gray_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
        red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        be_fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")
        for cell in ws[1]: cell.fill = gray_fill; cell.font = Font(bold=True)
        for t_log in res['trade_logs']:
            row = [t_log['الزوج'], t_log['وقت الصفقة (DAM)'], t_log['TF'], t_log['اتجاه'], t_log['المستوى (الدخول الفعلي)'],
                   t_log['الهدف (TP)'], t_log['الوقف (SL)'], t_log['النتيجة'], t_log['ربح صافي ($)'], t_log['رصيد تراكمي ($)']]
            ws.append(row)
            fill = {'WIN': green_fill, 'LOSS': red_fill, 'BREAK_EVEN': be_fill}.get(t_log['النتيجة'])
            if fill:
                for col in range(1, 11): ws.cell(row=ws.max_row, column=col).fill = fill

        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        center_align = Alignment(horizontal='center', vertical='center')
        for row in ws.iter_rows():
            for cell in row: cell.border = thin_border; cell.alignment = center_align
        for i in range(1, 11): ws.column_dimensions[get_column_letter(i)].width = 22.0

        if exec_mode == 'all_concurrent':
            _add_concurrent_analysis_sheets(wb, res['trade_logs'], pnl_key='ربح صافي ($)',
                                             outcome_key='النتيجة', slippage_key='انزلاق (نقطة)')

        wb.save(fname)
        await prog.done(f'<b>Live-Twin اكتمل ✅</b>\n{syms_label} — {len(res["trade_logs"])} صفقة\nجاري إرسال التقرير...')
        from telegram_ui import send_tg_document
        await send_tg_document(fname, sum_text)
        os.remove(fname)

    except Exception as e:
        c_log(f'Live-Twin Error: {e}'); bot_state['is_live_twin_running'] = False
        if _lt_progress:
            import html
            try: await _lt_progress.done(f'❌ خطأ داخلي في Live-Twin:\n{html.escape(str(e))}')
            except Exception: pass
    finally:
        bot_state['is_live_twin_running'] = False


# ═══════════════════════════════════════════════════════════════
# REAL-TIME FORWARD PAPER-TRADING LISTENER (Live-Twin Forward)
# ═══════════════════════════════════════════════════════════════
# Consumes ticks from the shared market_data._live_twin_queue and
# simulates paper trades using the friction model above.  Runs as a
# background asyncio task.  Launched via telegram_ui when the user
# starts the Live-Twin "forward" mode.

_forward_portfolio: dict[str, list[dict]] = {}  # symbol -> list of open paper trades
_forward_rng = random.Random()
_forward_stats: dict = {'total_signals': 0, 'entered': 0, 'win': 0, 'loss': 0, 'be': 0,
                        'rejected': 0, 'total_pnl': 0.0, 'total_slippage_pips': 0.0}


async def run_live_twin_forward() -> None:
    """Real-time forward Live-Twin.  Reads ticks from the shared asyncio
    queue (fed by the MetaApi price listener), detects Gann level touches
    against LIVE levels cached by the scanner, and simulates paper trades
    with the friction model.  Does NOT touch MetaApi at all."""
    from market_data import _live_twin_queue
    global _forward_portfolio, _forward_rng, _forward_stats

    _forward_rng = random.Random()
    _forward_portfolio.clear()
    _forward_stats = {k: 0 for k in _forward_stats}

    bot_state['is_live_twin_running'] = True
    c_log('Live-Twin Forward: started — consuming ticks from shared queue')

    from telegram_ui import send_tg_msg as _fwd_tg
    await _fwd_tg('🧪 <b>Live-Twin Forward Online</b>\n'
                  'المراقبة الورقية الفورية مفعلة. يتم محاكاة الصفقات مع الاحتكاك في الخلفية.')

    try:
        while bot_state.get('is_live_twin_running', False):
            try:
                tick = await asyncio.wait_for(_live_twin_queue.get(), timeout=2.0)
            except asyncio.TimeoutError:
                continue

            symbol = tick['symbol']
            sym_state = bot_state['symbol_state'].get(symbol)
            if not sym_state:
                continue

            # ── Entries: check against live cached levels ──
            if (sym_state['gann_cycle_active'] and sym_state['gann_levels']
                    and sym_state.get('auto_trade', False)):

                from market_data import _gann_cache
                cache = _gann_cache.get(symbol)
                if not cache:
                    continue

                margin = cache['margin']
                levels = cache['levels']
                trend_up = cache['trend_up']
                entry_mode = sym_state['gann_entry_mode']
                pv = SYMBOL_INFO[symbol]['pip_value']
                lot = sym_state['lot_size']

                open_sym = _forward_portfolio.get(symbol, [])
                max_concurrent = max(1, int(bot_state.get('prot_max_concurrent_trades', 4)))
                if len(open_sym) >= max_concurrent:
                    continue

                for tf in cache.get('enabled_tfs', []):
                    if any(p.get('tf') == tf for p in open_sym):
                        continue

                    tf_data = cache['tf_data'].get(tf)
                    if not tf_data:
                        continue
                    candles = tf_data['candles']

                    for lv in levels:
                        k = lv['key']; dir_ = lv['dir']
                        is_buy = (dir_ == 'dn')

                        if entry_mode == 'touch_trend':
                            if is_buy and not trend_up: continue
                            if not is_buy and trend_up: continue

                        check_px = tick['bid'] if is_buy else tick['ask']
                        if abs(check_px - lv['price']) > margin:
                            continue

                        _forward_stats['total_signals'] += 1

                        # Friction model for entry
                        bar_range = SYMBOL_INFO[symbol]['pip_value'] * 10
                        atr_ref = bar_range
                        entry_spread, _ = _lt_current_spread(
                            float(bot_state['lt_base_spread_usd']),
                            datetime.now(timezone.utc), bar_range, atr_ref) if bot_state['lt_friction']['spread'] else (float(bot_state['lt_base_spread_usd']), False)

                        slip = _lt_slippage(bar_range, atr_ref, _forward_rng) if bot_state['lt_friction']['slippage'] else 0.0

                        if bot_state['lt_friction']['rejection'] and _forward_rng.random() < float(bot_state['lt_rejection_prob']):
                            _forward_stats['rejected'] += 1
                            continue

                        fill_entry = lv['price'] + (entry_spread/2.0 + slip) if is_buy else lv['price'] - (entry_spread/2.0 + slip)

                        tf_tp = _gann_tf_tp(symbol, tf); tf_sl = _gann_tf_sl(symbol, tf)
                        sl_d = tf_sl * pv; tp_d = tf_tp * pv

                        tp_px = fill_entry + tp_d if is_buy else fill_entry - tp_d
                        sl_px = fill_entry - sl_d if is_buy else fill_entry + sl_d

                        trade_id = f"fw_{int(time.time()*1000)}_{symbol}_{tf}"
                        paper = {
                            'trade_id': trade_id, 'symbol': symbol, 'tf': tf, 'is_buy': is_buy,
                            'entry': fill_entry, 'level_price': lv['price'], 'tp_px': tp_px, 'sl_px': sl_px,
                            'sl_current': sl_px, 'lot': lot, 'pv': pv,
                            'be_activated': False, 'be_enabled': sym_state.get('break_even_enabled', False),
                            'cs': SYMBOL_INFO[symbol]['contract_size'],
                            'open_time': datetime.now(timezone.utc), 'close_time': None,
                            'slippage': slip, 'spread_paid': entry_spread,
                            'closed': False, 'outcome': None, 'pnl': 0.0,
                        }
                        if symbol not in _forward_portfolio:
                            _forward_portfolio[symbol] = []
                        _forward_portfolio[symbol].append(paper)
                        _forward_stats['entered'] += 1
                        slip_pips = round(slip / pv, 2) if pv else 0
                        c_log(f'LT-FWD ENTER [{symbol} {tf}] {"BUY" if is_buy else "SELL"} '
                              f'level={lv["price"]:.2f} fill={fill_entry:.2f} slip={slip_pips}p')
                        break  # one trade per tick

            # ── Exits: check SL/TP for open paper trades on this symbol ──
            bid, ask = tick['bid'], tick['ask']
            open_list = _forward_portfolio.get(symbol, [])
            surviving = []
            for paper in open_list:
                if paper['closed']:
                    continue
                is_buy = paper['is_buy']
                px = bid if is_buy else ask

                # Break-even check
                if paper['be_enabled'] and not paper['be_activated']:
                    be_pts = sym_state.get('gann_be_trigger_points', 40) if sym_state else 40
                    net_be = core_eval_break_even(is_buy, paper['entry'], px,
                                                   paper['pv'], be_pts,
                                                   sym_state.get('gann_atr_period', 14) if sym_state else 14,
                                                   bot_state.get('prot_cost_be', True))
                    if net_be is not None:
                        paper['sl_current'] = net_be
                        paper['be_activated'] = True

                hits_sl = (px <= paper['sl_current']) if is_buy else (px >= paper['sl_current'])
                hits_tp = (px >= paper['tp_px']) if is_buy else (px <= paper['tp_px'])

                if hits_sl or hits_tp:
                    paper['close_time'] = datetime.now(timezone.utc)
                    if hits_sl and hits_tp:
                        outcome = 'sl'
                    elif hits_sl:
                        outcome = 'sl'
                    else:
                        outcome = 'tp'

                    bar_range = max(abs(ask - bid) * 10, paper['pv'])
                    atr_ref = bar_range
                    exit_spread, _ = _lt_current_spread(
                        float(bot_state['lt_base_spread_usd']),
                        paper['close_time'], bar_range, atr_ref) if bot_state['lt_friction']['spread'] else (0.0, False)
                    exit_slip = _lt_slippage(bar_range, atr_ref, _forward_rng) if bot_state['lt_friction']['slippage'] else 0.0

                    if outcome == 'sl':
                        raw_px = paper['sl_current']
                        fill_px = raw_px - (exit_spread/2.0 + exit_slip) if is_buy else raw_px + (exit_spread/2.0 + exit_slip)
                        be_thresh = paper['pv'] * 2
                        paper['outcome'] = 'BREAK_EVEN' if (is_buy and paper['sl_current'] > paper['entry'] - be_thresh) or (not is_buy and paper['sl_current'] < paper['entry'] + be_thresh) else 'LOSS'
                    else:
                        raw_px = paper['tp_px']
                        fill_px = raw_px - (exit_spread/2.0 + exit_slip) if is_buy else raw_px + (exit_spread/2.0 + exit_slip)
                        paper['outcome'] = 'WIN'

                    diff = (fill_px - paper['entry']) if is_buy else (paper['entry'] - fill_px)
                    pnl_usd = round(diff * paper['lot'] * paper['cs'], 2)

                    commission = float(bot_state['lt_commission_per_lot']) * paper['lot'] if bot_state['lt_friction']['commission'] else 0.0
                    pnl_net = round(pnl_usd - commission, 2)
                    if paper['outcome'] == 'WIN' and pnl_net < 0:
                        paper['outcome'] = 'LOSS'
                    paper['pnl'] = pnl_net
                    paper['closed'] = True

                    sip_pips = round((paper['entry'] - paper['level_price']) / paper['pv'], 2) if paper['pv'] else 0
                    _forward_stats['total_slippage_pips'] += sip_pips
                    if paper['outcome'] == 'WIN':
                        _forward_stats['win'] += 1
                    elif paper['outcome'] == 'LOSS':
                        _forward_stats['loss'] += 1
                    else:
                        _forward_stats['be'] += 1
                    _forward_stats['total_pnl'] += pnl_net

                    c_log(f'LT-FWD EXIT [{symbol} {paper["tf"]}] {paper["outcome"]} pnl={pnl_net:.2f}')
                    continue

                surviving.append(paper)

            if surviving:
                _forward_portfolio[symbol] = surviving
            elif symbol in _forward_portfolio and not any(not p['closed'] for p in _forward_portfolio[symbol]):
                del _forward_portfolio[symbol]

    except asyncio.CancelledError:
        pass
    except Exception as e:
        log_exception('run_live_twin_forward', e)
    finally:
        bot_state['is_live_twin_running'] = False
        # Send summary
        s = _forward_stats
        summary = (
            f"🧪 <b>Live-Twin Forward Summary</b>\n\n"
            f"Total signals detected: {s['total_signals']}\n"
            f"Paper trades entered: {s['entered']} (rejected: {s['rejected']})\n"
            f"Wins: {s['win']}  Losses: {s['loss']}  BE: {s['be']}\n"
            f"Net P&L: ${s['total_pnl']:.2f}\n"
            f"Avg slippage: {round(s['total_slippage_pips'] / max(s['entered'], 1), 2)} pips"
        )
        await _fwd_tg(summary)
        c_log('Live-Twin Forward: stopped')
