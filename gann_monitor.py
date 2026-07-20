"""
gann_monitor.py — Live scanner, cycle manager, diagnostics, and reconciliation.

Owns:
  - gann_monitor_scanner (main live trading loop)
  - gann_cycle_manager (Gann level anchoring)
  - global_ledger_reconciliation (broker cross-check)
  - gann_run_diagnostics (Telegram /diagnose)
  - export_diag_log_excel, export_live_trades_excel, export_execution_details_report
"""

import asyncio
import os
import time
from datetime import datetime, timedelta, timezone

import numpy as np
import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from state import (
    bot_state, SYMBOL_INFO, _TFS, DAM_OFF,
    CONN_RUNNING, CONN_READ_ONLY, CONN_HALTED,
    _daily_state_lock, _fail_counter_lock, _safe_float, _safe_task,
    _debounced_persist_save, save_bot_persistence, _diag_log_add,
    log_exception, c_log, is_trading_allowed, set_connection_state,
    _record_closed_trade_history, _trade_history_lock,
)
import market_data
from market_data import (
    live_quotes,
    _lq_price_with_fallback, _force_full_reconnect, _bootstrap_metaapi_connection,
    _lq_is_stale, _lq_subscribe_symbol,
    _gann_cache, fetch_candles, fetch_master_price,
    _QUOTE_STALE_SECONDS, _WS_WATCHDOG_STALE_SECONDS,
)
# NOTE: _metaapi_conn / _metaapi_account / _last_any_tick_ts are deliberately
# NOT imported with `from market_data import ...` here. That syntax copies the
# value that name pointed to AT IMPORT TIME into this module's namespace.
# market_data.py later reassigns them with `global _metaapi_conn; _metaapi_conn = ...`
# inside its own functions -- that only rebinds market_data's own copy, it never
# updates the frozen copy that would live here. The result: this module would
# see _metaapi_conn as permanently None and _last_any_tick_ts as whatever it was
# the instant this module was first imported (long before the connection was
# ever established) -- so the watchdog below would never fire.
# Always read `market_data._metaapi_conn` / `market_data._last_any_tick_ts`
# fresh at the point of use instead.
from strategy import (
    gann_calc_levels, gann_active_levels, _gann_atr, _gann_fetch_last_closed_anchor,
    _gann_tf_tp, _gann_tf_sl, _anchor_label, _anchor_hours,
    _is_market_hours_now, _is_within_dam_restricted_window,
    _last_closed_anchor_time_utc,
)
from backtest import GANN_DYNAMIC_RECALC_MINUTES, _utc_to_dam


# ── Scanner error alert rate limit ──
_last_scanner_error_alert_ts = 0.0

# ── Reconciliation state ──
_recon_consecutive_mismatches = 0
_RECON_MISMATCH_HALT_THRESHOLD = 3
RECONCILIATION_INTERVAL_SECONDS = 300


# ── Live Scanner ──
async def gann_monitor_scanner() -> None:
    global _last_scanner_error_alert_ts
    c_log('Gann live scanner started.')
    while True:
        try:
            # ── Cold-start self-heal ──
            # If _metaapi_conn (or _metaapi_account) is still None, neither of
            # the watchdogs below can do anything -- both require a connection
            # object to already exist. This closes the gap where a transient
            # MetaApi/broker hiccup during the ONE startup attempt would leave
            # the bot in silent, permanent READ_ONLY forever. Retry from
            # scratch here, every scanner tick, until it succeeds.
            if market_data._metaapi_conn is None or market_data._metaapi_account is None:
                await _bootstrap_metaapi_connection()

            # Stale tick watchdog: if no tick for >60s, trigger full reconnect.
            # Connection management is in market_data.py (init_metaapi / _bootstrap).
            # Scanner reads from the shared live_quotes cache ONLY.
            # Read these fresh from the market_data module every iteration --
            # see the comment at the top of this file for why a plain
            # `from market_data import _metaapi_conn` would freeze them.
            _metaapi_conn = market_data._metaapi_conn
            _last_any_tick_ts = market_data._last_any_tick_ts
            active_syms_now = [s for s, on in bot_state['active_symbols'].items() if on]
            if (_metaapi_conn is not None and active_syms_now and _is_market_hours_now()
                    and (time.monotonic() - _last_any_tick_ts) > _WS_WATCHDOG_STALE_SECONDS):
                await _force_full_reconnect(
                    f"لا تيك واحد وصل منذ {time.monotonic() - _last_any_tick_ts:.0f}s "
                    f"(الحد: {_WS_WATCHDOG_STALE_SECONDS:.0f}s)"
                )

            # ── MT5 Zombie Singleton Heartbeat ──
            # Catches the case where connection_state is stuck away from
            # CONN_RUNNING (e.g. _force_full_reconnect above keeps timing out)
            # and escalates to a FULL client+account rebuild via
            # _bootstrap_metaapi_connection(), with up to 5 attempts and
            # exponential backoff (1s, 2s, 4s, 8s, 16s) within this one tick.
            # This is intentionally heavier and rarer than the tick-silence
            # watchdog above -- it only runs while genuinely stuck, not every
            # 15s scan cycle, so it does not add meaningful extra MetaApi API
            # traffic during a normal transient blip.
            _metaapi_account = market_data._metaapi_account
            if _metaapi_account and bot_state.get('connection_state') != CONN_RUNNING:
                await set_connection_state(CONN_READ_ONLY, "MetaAPI connection lost — attempting reconnect.")
                reconnected = False
                for attempt in range(5):
                    try:
                        reconnected = await _bootstrap_metaapi_connection()
                        if reconnected:
                            c_log("MetaAPI Reconnected successfully (live quotes resubscribed).")
                            break
                    except Exception as e:
                        log_exception(f"MetaAPI reconnect attempt {attempt+1}/5", e)
                    await asyncio.sleep(2 ** attempt)
                if not reconnected:
                    c_log("MetaAPI reconnect exhausted 5 attempts this tick; will retry next cycle.")
            # Feed-level staleness watchdog: connection_status can still say
            # CONNECTED while a symbol's subscription silently dropped (no
            # more ticks arriving). Caught independently of the connection-
            # level check above, and just re-subscribes rather than tearing
            # down the whole connection.
            elif _metaapi_conn is not None:
                for sym, on in bot_state['active_symbols'].items():
                    if on and _lq_is_stale(sym):
                        c_log(f"Live quote feed stale for {sym} -- resubscribing.")
                        await _lq_subscribe_symbol(sym)

            now_dt = datetime.now(timezone.utc)
            today_date = now_dt.date()
            if bot_state.get('live_daily_date') != today_date:
                c_log(f"New trading day detected ({bot_state.get('live_daily_date')} -> {today_date}).")
                async with _daily_state_lock:
                    bot_state['live_daily_date'] = today_date
                    bot_state['live_daily_realized'] = 0.0
                    bot_state['live_daily_hit'] = False
                await save_bot_persistence()

            if bot_state.get('live_daily_hit'):
                stale_active_symbols = [s for s, on in bot_state['active_symbols'].items() if on]
                stale_real_closures = []
                for symbol in stale_active_symbols:
                    sym_state = bot_state['symbol_state'][symbol]
                    for tid, tr in list(sym_state['gann_open_trades'].items()):
                        if tr.get('is_real') and _metaapi_conn:
                            stale_real_closures.append((symbol, tid, sym_state, tr))
                if stale_real_closures:
                    from execution import _close_metaapi_trades_batch
                    await _close_metaapi_trades_batch(stale_real_closures)
                await asyncio.sleep(60)
                continue

            active_symbols = [s for s, on in bot_state['active_symbols'].items() if on]
            total_floating = 0.0

            for symbol in active_symbols:
                sym_state = bot_state['symbol_state'][symbol]

                if bot_state.get('prot_cycle_inval', True) and sym_state.get('gann_close_used'):
                    mc = await fetch_candles(symbol, '1m', count=2)
                    if mc:
                        live_px = float(mc[-1]['close'])
                        dist = abs(live_px - sym_state['gann_close_used'])
                        pv = SYMBOL_INFO[symbol]['pip_value']
                        inval_pts = bot_state.get('prot_cycle_inval_pts', 200) * pv
                        if dist > inval_pts:
                            sym_state['gann_levels'] = []
                            sym_state['gann_close_used'] = None
                            from telegram_ui import send_tg_msg
                            await send_tg_msg(f"🚨 <b>إلغاء دورة {symbol}:</b> السعر تحرك بحدة!")

                if sym_state['gann_open_trades']:
                    actual_positions = {}; sync_failed = False
                    if bot_state.get('prot_true_sync', True) and _metaapi_conn:
                        try:
                            positions = _metaapi_conn.terminal_state.positions
                            for p in positions:
                                actual_positions[str(p.get('id'))] = p
                            if (bot_state.get('connection_state') == CONN_READ_ONLY
                                    and 'sync' in bot_state.get('connection_state_reason', '').lower()):
                                await set_connection_state(CONN_RUNNING, "MetaAPI get_positions() succeeded again.")
                        except Exception as e:
                            log_exception(f"MetaAPI get_positions [{symbol}]", e)
                            sync_failed = True
                            await set_connection_state(CONN_READ_ONLY,
                                f"MetaAPI get_positions() sync failed for {symbol}: {e}.")
                    if sync_failed:
                        continue

                    mc = await fetch_candles(symbol, '1m', count=2)
                    live_px = None; oanda_failed = False
                    if not mc:
                        oanda_failed = True
                    else:
                        candle_age = (now_dt - mc[-1]['time']).total_seconds()
                        if bot_state.get('prot_stale_filter', True) and candle_age > 120:
                            oanda_failed = True
                    if not oanda_failed:
                        live_px = float(mc[-1]['close'])
                    else:
                        c_log(f"Oanda failed for {symbol}. Decoupled Mode: using MT5 currentPrice.")

                    closed_ids = []
                    history_deals_cache = None
                    missing_tids = [t for t, v in sym_state['gann_open_trades'].items()
                                    if v.get('is_real') and t not in actual_positions]
                    if missing_tids and _metaapi_conn:
                        start_time = datetime.now(timezone.utc) - timedelta(days=2)
                        for attempt_i, delay in enumerate((0, 3, 5)):
                            if delay:
                                await asyncio.sleep(delay)
                            try:
                                history_deals_cache = await _metaapi_conn.get_history_deals_by_time_range(
                                    start_time, datetime.now(timezone.utc))
                            except Exception as e:
                                log_exception(f"get_history_deals [{symbol}] attempt {attempt_i+1}/3", e)
                                continue
                            found_now = {str(d.get('positionId')) for d in history_deals_cache
                                         if d.get('entryType') in ('DEAL_ENTRY_OUT', 'DEAL_ENTRY_OUT_BY')}
                            if all(str(t) in found_now for t in missing_tids):
                                break

                    for tid, tr in list(sym_state['gann_open_trades'].items()):
                        is_buy = tr.get('is_buy'); tp = tr.get('tp'); sl = tr.get('sl')
                        entry = tr.get('entry'); tf = tr.get('tf'); is_real = tr.get('is_real')
                        active_px = live_px
                        if active_px is None:
                            if tid in actual_positions:
                                active_px = _safe_float(actual_positions[tid].get('currentPrice'), entry)
                            else:
                                active_px = tr.get('last_known_px')
                        if active_px is None:
                            continue
                        tr['last_known_px'] = active_px
                        diff = (active_px - entry) if is_buy else (entry - active_px)
                        cs = SYMBOL_INFO[symbol]['contract_size']
                        trade_pl = round(diff * sym_state['lot_size'] * cs, 2)
                        tr['last_known_pl'] = trade_pl

                        if is_real and bot_state.get('prot_true_sync', True) and _metaapi_conn:
                            if tid not in actual_positions:
                                exact_pnl = trade_pl; found_deal = False
                                if history_deals_cache is not None:
                                    deal_pnl = 0.0
                                    for d in history_deals_cache:
                                        if (str(d.get('positionId')) == str(tid)
                                                and d.get('entryType') in ('DEAL_ENTRY_OUT', 'DEAL_ENTRY_OUT_BY')):
                                            deal_pnl += _safe_float(d.get('profit')) + _safe_float(d.get('swap')) + _safe_float(d.get('commission'))
                                            found_deal = True
                                    if found_deal:
                                        exact_pnl = deal_pnl
                                closed_ids.append(tid)
                                async with _daily_state_lock:
                                    bot_state['live_daily_realized'] += exact_pnl
                                from telegram_ui import send_tg_msg
                                if found_deal:
                                    msg = f"🔔 <b>مزامنة: إغلاق صفقة [{symbol} - {tf}]</b>\nالربح الفعلي (MT5): {exact_pnl:.2f}$"
                                else:
                                    msg = f"🔔 <b>مزامنة: إغلاق صفقة [{symbol} - {tf}]</b>\n⚠️ ربح تقديري: ~{exact_pnl:.2f}$"
                                await send_tg_msg(msg)
                                await _record_closed_trade_history(
                                    symbol, tid, tr, exit_px=active_px, pnl=exact_pnl,
                                    outcome_label='WIN' if exact_pnl > 0 else 'LOSS' if exact_pnl < 0 else 'BREAK_EVEN',
                                    close_reason='tp_sl_or_manual_broker_close', pnl_confirmed=found_deal)
                                continue
                            else:
                                trade_pl = _safe_float(actual_positions[tid].get('unrealizedProfit'), trade_pl)

                        from strategy import core_eval_outcome, core_eval_break_even
                        outcome = core_eval_outcome(is_buy, active_px, tp, sl)
                        if bot_state.get('prot_cost_be', True) and sym_state.get('break_even_enabled') and not tr.get('be_activated'):
                            be_pts = sym_state.get('gann_be_trigger_points', 40)
                            net_be = core_eval_break_even(is_buy, entry, active_px,
                                SYMBOL_INFO[symbol]['pip_value'], be_pts, sym_state.get('gann_atr_period', 14),
                                bot_state.get('prot_cost_be', True))
                            if net_be is not None:
                                if is_real and _metaapi_conn:
                                    try:
                                        await _metaapi_conn.modify_position(tid, stop_loss=net_be)
                                        tr['sl'] = net_be; tr['be_activated'] = True
                                        await save_bot_persistence()
                                        from telegram_ui import send_tg_msg
                                        await send_tg_msg(f"🛡️ تم تفعيل Break-Even لـ {symbol}!")
                                    except Exception as e:
                                        log_exception(f"BE modify_position [{symbol}/{tid}]", e)
                                        from telegram_ui import send_tg_msg
                                        await send_tg_msg(f"⚠️ <b>فشل تفعيل Break-Even لـ {symbol}:</b> {e}")
                                else:
                                    tr['sl'] = net_be; tr['be_activated'] = True
                                    await save_bot_persistence()

                        if outcome:
                            closed_ids.append(tid)
                            async with _daily_state_lock:
                                bot_state['live_daily_realized'] += trade_pl
                            from telegram_ui import send_tg_msg
                            await send_tg_msg(f"🔔 <b>تحديث صفقة [{symbol} - جان {tf}]</b>\n\nالنتيجة: {outcome} ({trade_pl}$)")
                            await _record_closed_trade_history(
                                symbol, tid, tr, exit_px=live_px, pnl=trade_pl,
                                outcome_label=outcome, close_reason='tp_sl_hit', pnl_confirmed=False)
                        else:
                            total_floating += trade_pl

                    for tid in closed_ids:
                        if tid in sym_state['gann_open_trades']:
                            del sym_state['gann_open_trades'][tid]
                            await save_bot_persistence()

            # Daily limits
            total_daily = bot_state['live_daily_realized'] + total_floating
            dd_limit = -float(bot_state.get('prot_daily_dd_usd', 220))
            profit_limit = float(bot_state.get('prot_daily_profit_usd', 150))

            if (dd_limit < 0 and total_daily <= dd_limit) or (profit_limit > 0 and total_daily >= profit_limit):
                async with _daily_state_lock:
                    bot_state['live_daily_hit'] = True
                limit_type = '🛑 تراجع عائم' if total_daily <= dd_limit else '✅ هدف يومي عائم'
                from telegram_ui import send_tg_msg
                await send_tg_msg(f"{limit_type} تم الوصول إليه! ({total_daily:.2f}$)\nسيتم إغلاق جميع الصفقات.")
                real_closures = []
                for symbol in active_symbols:
                    sym_state = bot_state['symbol_state'][symbol]
                    for tid, tr in list(sym_state['gann_open_trades'].items()):
                        if tr.get('is_real') and _metaapi_conn:
                            real_closures.append((symbol, tid, sym_state, tr))
                        else:
                            pl = tr.get('last_known_pl', 0.0)
                            px = tr.get('last_known_px', tr.get('entry'))
                            outcome_lbl = 'ربح ✅' if pl >= 0 else 'خسارة ❌'
                            from telegram_ui import send_tg_msg as stm
                            await stm(f"⏹️ <b>إغلاق (وهمي) [{symbol} - جان {tr.get('tf')}]</b>\nالنتيجة: {outcome_lbl} ({pl}$)")
                            await _record_closed_trade_history(
                                symbol, tid, tr, exit_px=px, pnl=pl,
                                outcome_label='WIN' if pl > 0 else 'LOSS' if pl < 0 else 'BREAK_EVEN',
                                close_reason='daily_capital_protection_forced_close', pnl_confirmed=False)
                            del sym_state['gann_open_trades'][tid]
                            await save_bot_persistence()
                from execution import _close_metaapi_trades_batch
                await _close_metaapi_trades_batch(real_closures)
                continue

            for symbol in active_symbols:
                try:
                    sym_state = bot_state['symbol_state'][symbol]
                    flt_type = sym_state['trend_filter_type']; ttf = sym_state['trend_timeframe']
                    enabled_tfs = [tf for tf, on in sym_state['gann_monitor_tfs'].items() if on]
                    if not sym_state['gann_cycle_active'] or not sym_state['gann_levels']:
                        continue

                    macro_trend_up = None
                    if sym_state['gann_entry_mode'] == 'touch_trend':
                        flt_is_vwap = flt_type in ('vwap', 'both'); flt_is_ema = flt_type in ('ema', 'both')
                        p_vwap = sym_state['trend_vwap_period'] if flt_is_vwap else 0
                        p_ema = sym_state['trend_ema_period'] if flt_is_ema else 0
                        max_period = max(p_vwap, p_ema, 100)
                        trend_candles = await fetch_candles(symbol, ttf, count=max(max_period+10, 120))
                        if trend_candles:
                            df_trend = pd.DataFrame(trend_candles)
                            current_trend_close = float(trend_candles[-1]['close'])
                            if flt_is_vwap:
                                df_trend['Typical_Price'] = (df_trend['high'] + df_trend['low'] + df_trend['close']) / 3
                                df_trend['VWAP'] = (df_trend['Typical_Price'] * df_trend['volume']).rolling(window=p_vwap).sum() / df_trend['volume'].rolling(window=p_vwap).sum()
                                current_vwap = df_trend.iloc[-1]['VWAP']
                                if pd.isna(current_vwap): current_vwap = current_trend_close
                                vwap_up = current_trend_close > current_vwap
                            if flt_is_ema:
                                df_trend['EMA'] = df_trend['close'].ewm(span=p_ema, adjust=False).mean()
                                current_ema = df_trend.iloc[-1]['EMA']
                                ema_up = current_trend_close > current_ema
                            if flt_type == 'vwap': macro_trend_up = vwap_up
                            elif flt_type == 'ema': macro_trend_up = ema_up
                            elif flt_type == 'both':
                                macro_trend_up = vwap_up if vwap_up == ema_up else None

                    levels = gann_active_levels(symbol)
                    margin = sym_state['gann_touch_margin_pts'] * SYMBOL_INFO[symbol]['pip_value']
                    detect_time = datetime.now(timezone.utc)

                    tf_data = {}
                    for tf in enabled_tfs:
                        need = sym_state['gann_atr_period'] + 50
                        candles = await fetch_candles(symbol, tf, count=need)
                        if not candles or len(candles) < 3:
                            _diag_log_add({'ts': detect_time, 'symbol': symbol, 'tf': tf,
                                           'skip_reason': f'insufficient_oanda_candles(got={len(candles) if candles else 0})'})
                            continue
                        tf_data[tf] = {'candles': candles, 'closed_close': float(candles[-1]['close'])}
                    _gann_cache[symbol] = {'levels': levels, 'margin': margin, 'trend_up': macro_trend_up,
                                            'enabled_tfs': list(tf_data.keys()), 'tf_data': tf_data,
                                            'refreshed_at': detect_time}
                    q = live_quotes.get(symbol)
                    ws_age_s = round(time.monotonic() - q['ts'], 1) if q else None
                    ws_status = 'live' if (q and ws_age_s <= _QUOTE_STALE_SECONDS) else ('stale' if q else 'never_received')
                    diag_px, price_source, _age_ms = await _lq_price_with_fallback(symbol)
                    entry_mode = sym_state['gann_entry_mode']
                    directional_levels = (
                        [l for l in levels if (l['dir'] == 'dn') == macro_trend_up]
                        if entry_mode == 'touch_trend' and macro_trend_up is not None else levels)
                    nearest_dist = None; nearest_price = None
                    if diag_px is not None:
                        for l in directional_levels:
                            d = abs(diag_px - l['price'])
                            if nearest_dist is None or d < nearest_dist:
                                nearest_dist = d; nearest_price = l['price']
                    _diag_log_add({'ts': detect_time, 'symbol': symbol, 'master_px': diag_px,
                                   'price_source': price_source, 'ws_status': ws_status,
                                   'ws_quote_age_s': ws_age_s, 'trend_up': macro_trend_up,
                                   'margin': margin, 'nearest_compatible_level': nearest_price,
                                   'nearest_dist': nearest_dist,
                                   'within_margin': (nearest_dist is not None and nearest_dist <= margin),
                                   'skip_reason': ('no_price_available' if diag_px is None else
                                                   'cache_refresh_only(firing_is_now_tick_driven)')})
                except Exception as sym_exc:
                    log_exception(f"gann_monitor_scanner per-symbol [{symbol}]", sym_exc)
                    now_mono_sym = time.monotonic()
                    if now_mono_sym - _last_scanner_error_alert_ts > 300:
                        _last_scanner_error_alert_ts = now_mono_sym
                        from telegram_ui import send_tg_msg
                        await send_tg_msg(f"🛑 <b>[{symbol}]</b> خطأ غير متوقع: {sym_exc}")
                    continue
        except Exception as e:
            log_exception('gann_monitor_scanner main loop', e)
            now_mono = time.monotonic()
            if now_mono - _last_scanner_error_alert_ts > 300:
                _last_scanner_error_alert_ts = now_mono
                from telegram_ui import send_tg_msg
                await send_tg_msg(f"🛑 <b>خطأ غير متوقع بدورة الفحص الحية:</b>\n{e}")
        await asyncio.sleep(15)


# ── Cycle Manager ──
async def gann_cycle_manager() -> None:
    from backtest import GANN_DYNAMIC_RECALC_MINUTES
    c_log('Gann cycle manager started.')
    while True:
        try:
            now_utc = datetime.now(timezone.utc)
            calc_mode = bot_state.get('gann_calculation_mode', 'static_h1')
            active_symbols = [s for s, on in bot_state['active_symbols'].items() if on]
            for symbol in active_symbols:
                sym_state = bot_state['symbol_state'][symbol]
                if not sym_state['gann_cycle_active']:
                    continue
                if calc_mode == 'dynamic_live':
                    last_recalc = sym_state['gann_cycle_started_at']
                    if last_recalc and (now_utc - last_recalc).total_seconds() < GANN_DYNAMIC_RECALC_MINUTES * 60:
                        continue
                    live_px, _src, _age = await _lq_price_with_fallback(symbol)
                    if live_px is None:
                        continue
                    sym_state['gann_levels'] = gann_calc_levels(symbol, live_px)
                    sym_state['gann_close_used'] = live_px
                    sym_state['gann_last_h1_time'] = now_utc
                    sym_state['gann_cycle_started_at'] = now_utc
                    sym_state['gann_level_status'] = {}
                    sym_state['gann_atr_cache'] = {}
                    for tf in sym_state['gann_monitor_tfs']:
                        if sym_state['gann_monitor_tfs'].get(tf):
                            tf_candles = await fetch_candles(symbol, tf, count=sym_state['gann_atr_period'] + 50)
                            if tf_candles:
                                sym_state['gann_atr_cache'][tf] = _gann_atr(tf_candles, sym_state['gann_atr_period'])
                    c_log(f'[{symbol}] Dynamic Gann recalculation at live_px={live_px}')
                    continue
                cycle_h = sym_state['gann_cycle_hours']
                last_h1 = await _gann_fetch_last_closed_anchor(symbol)
                if last_h1:
                    h1_time = last_h1['time']
                    if not sym_state['gann_last_h1_time'] or h1_time > sym_state['gann_last_h1_time']:
                        if not sym_state['gann_last_h1_time'] or (h1_time - sym_state['gann_last_h1_time']).total_seconds() / 3600.0 >= cycle_h:
                            h1_close = float(last_h1['close'])
                            sym_state['gann_levels'] = gann_calc_levels(symbol, h1_close)
                            sym_state['gann_close_used'] = h1_close
                            sym_state['gann_last_h1_time'] = h1_time
                            sym_state['gann_cycle_started_at'] = now_utc
                            sym_state['gann_level_status'] = {}
                            c_log(f'[{symbol}] New {cycle_h}h cycle started at {h1_close}')
                            from telegram_ui import send_tg_msg
                            await send_tg_msg(f"🔄 <b>تحديث دورة جان ({cycle_h}h)</b>\nالزوج: {symbol}\nإغلاق {_anchor_label()}: {h1_close:.5f}")
        except Exception as e:
            log_exception('gann_cycle_manager main loop', e)
        await asyncio.sleep(60)


# ── Global Ledger Reconciliation ──
async def global_ledger_reconciliation() -> None:
    global _recon_consecutive_mismatches
    c_log('Global ledger reconciliation started.')
    while True:
        try:
            await asyncio.sleep(RECONCILIATION_INTERVAL_SECONDS)
            _metaapi_conn = market_data._metaapi_conn
            if bot_state.get('connection_state', CONN_RUNNING) != CONN_RUNNING or not _metaapi_conn:
                continue
            try:
                broker_positions = _metaapi_conn.terminal_state.positions
                if not isinstance(broker_positions, list):
                    raise TypeError(f"get_positions() returned {type(broker_positions).__name__}")
            except Exception as e:
                log_exception('global_ledger_reconciliation get_positions', e)
                continue
            broker_ids = {str(p['id']) for p in broker_positions if p.get('id')}
            known_ids = set()
            for sym, ss in bot_state['symbol_state'].items():
                for tid, tr in ss.get('gann_open_trades', {}).items():
                    if tr.get('is_real'):
                        known_ids.add(str(tid))
            ghost_ids = broker_ids - known_ids; missing_ids = known_ids - broker_ids
            if ghost_ids:
                async with _fail_counter_lock:
                    _recon_consecutive_mismatches += 1
                    mismatches = _recon_consecutive_mismatches
                from telegram_ui import send_tg_msg
                await send_tg_msg(
                    f"🚨 <b>تحذير مطابقة الحساب المستقل:</b>\n"
                    f"يوجد {len(ghost_ids)} صفقة مفتوحة على الوسيط لا يعرفها البوت.\n"
                    f"IDs: {ghost_ids}"
                )
                if mismatches >= _RECON_MISMATCH_HALT_THRESHOLD:
                    await set_connection_state(CONN_HALTED,
                        f"{mismatches} consecutive independent reconciliation checks found unmanaged broker positions.")
            else:
                if _recon_consecutive_mismatches > 0:
                    c_log("Reconciliation recovered.")
                async with _fail_counter_lock:
                    _recon_consecutive_mismatches = 0
            if missing_ids:
                c_log(f"Reconciliation note: {len(missing_ids)} bot-tracked trade(s) not currently on broker.")
        except Exception as e:
            log_exception('global_ledger_reconciliation main loop', e)


# ── Diagnostics ──
async def gann_run_diagnostics() -> str:
    lines = ["<b>🩺 تشخيص أسباب عدم فتح الصفقات</b>\n"]
    conn_state = bot_state.get('connection_state', CONN_RUNNING)
    conn_ok = conn_state == CONN_RUNNING
    lines.append(f"1️⃣ حالة الاتصال: {'✅ RUNNING' if conn_ok else f'🛑 {conn_state}'}")
    if not conn_ok:
        lines.append(f"   السبب: {bot_state.get('connection_state_reason', '-')}")
    dam_blocked = _is_within_dam_restricted_window()
    dam_now = datetime.now(timezone.utc) + timedelta(hours=3)
    filter_on = bot_state.get('prot_dam_time_filter', True)
    lines.append(f"2️⃣ فلتر أوقات دمشق: {'مفعّل' if filter_on else '🔴 معطّل'} | الوقت الآن (DAM): {dam_now:%H:%M}")
    if filter_on and dam_blocked:
        lines.append("   🛑 داخل نافذة محظورة الآن.")
    overall_allowed = await is_trading_allowed()
    lines.append(f"3️⃣ الخلاصة العامة is_trading_allowed(): {'✅ مسموح' if overall_allowed else '🛑 ممنوع'}\n")
    active_symbols = [s for s, on in bot_state['active_symbols'].items() if on]
    if not active_symbols:
        lines.append("⚠️ لا يوجد أي رمز مفعّل.")
        return "\n".join(lines)
    for symbol in active_symbols:
        sym_state = bot_state['symbol_state'][symbol]
        lines.append(f"━━━━━━━━━━━━━━\n<b>{symbol}</b>")
        q = live_quotes.get(symbol)
        ws_age = (time.monotonic() - q['ts']) if q else None
        if q is None:
            lines.append("📡 تغذية MetaApi (WS): 🛑 <b>لم تصل ولا تيك واحد</b>")
        elif ws_age > _QUOTE_STALE_SECONDS:
            lines.append(f"📡 تغذية MetaApi (WS): 🛑 <b>متوقفة منذ {ws_age:.0f}s</b>")
        else:
            lines.append(f"📡 تغذية MetaApi (WS): ✅ حية ({ws_age:.1f}s)")
        cycle_active = sym_state.get('gann_cycle_active', False)
        n_levels = len(sym_state.get('gann_levels', []))
        lines.append(f"دورة جان نشطة: {'✅' if cycle_active else '🛑'}  |  عدد المستويات: {n_levels}")
        if not cycle_active or n_levels == 0:
            lines.append("↳ 🛑 السكانر بيتخطى هذا الرمز بالكامل.")
            continue
        flt_type = sym_state['trend_filter_type']; ttf = sym_state['trend_timeframe']
        entry_mode = sym_state['gann_entry_mode']
        lines.append(f"وضع الدخول: {entry_mode}  |  فلتر الاتجاه: {flt_type} ({ttf})")
        macro_trend_up = None
        if entry_mode == 'touch_trend':
            p_vwap = sym_state['trend_vwap_period'] if flt_type == 'vwap' else 0
            p_ema = sym_state['trend_ema_period'] if flt_type == 'ema' else 0
            max_period = max(p_vwap, p_ema, 100)
            try:
                trend_candles = await fetch_candles(symbol, ttf, count=max(max_period + 10, 120))
            except Exception as e:
                trend_candles = []; lines.append(f"🛑 فشل جلب بيانات الاتجاه: {e}")
            if not trend_candles:
                lines.append(f"🛑 لا توجد بيانات اتجاه ({ttf})")
            else:
                df_trend = pd.DataFrame(trend_candles)
                current_trend_close = float(trend_candles[-1]['close'])
                if flt_type == 'vwap':
                    df_trend['Typical_Price'] = (df_trend['high'] + df_trend['low'] + df_trend['close']) / 3
                    df_trend['VWAP'] = (df_trend['Typical_Price'] * df_trend['volume']).rolling(window=p_vwap).sum() / df_trend['volume'].rolling(window=p_vwap).sum()
                    current_vwap = df_trend.iloc[-1]['VWAP']
                    if pd.isna(current_vwap): current_vwap = current_trend_close
                    macro_trend_up = current_trend_close > current_vwap
                    lines.append(f"الاتجاه (VWAP{p_vwap}): {'صاعد' if macro_trend_up else 'هابط'}")
                elif flt_type == 'ema':
                    df_trend['EMA'] = df_trend['close'].ewm(span=p_ema, adjust=False).mean()
                    current_ema = df_trend.iloc[-1]['EMA']
                    macro_trend_up = current_trend_close > current_ema
                    lines.append(f"الاتجاه (EMA{p_ema}): {'صاعد' if macro_trend_up else 'هابط'}")
        levels = gann_active_levels(symbol)
        margin = sym_state['gann_touch_margin_pts'] * SYMBOL_INFO[symbol]['pip_value']
        enabled_tfs = [tf for tf, on in sym_state['gann_monitor_tfs'].items() if on]
        if not enabled_tfs:
            lines.append("🛑 لا يوجد أي فريم مفعّل.")
            continue
        master_px = await fetch_master_price(symbol)
        if master_px is None:
            lines.append("🛑 بيانات غير كافية من OANDA.")
            continue
        for tf in enabled_tfs:
            already_open = any(isinstance(v, dict) and v.get('tf') == tf for v in sym_state['gann_open_trades'].values())
            if already_open:
                lines.append(f"[{tf}] 🛑 يوجد صفقة مفتوحة بالفعل.")
                continue
            try:
                candles = await fetch_candles(symbol, tf, count=sym_state['gann_atr_period'] + 50)
            except Exception as e:
                lines.append(f"[{tf}] 🛑 فشل جلب الشموع: {e}"); continue
            if not candles or len(candles) < 3:
                lines.append(f"[{tf}] 🛑 بيانات غير كافية."); continue
            live_px = master_px; trend_up = True
            if entry_mode == 'touch_trend':
                if macro_trend_up is None:
                    lines.append(f"[{tf}] 🛑 لا يمكن التحقق من الاتجاه."); continue
                trend_up = macro_trend_up
            if entry_mode == 'touch_trend' and macro_trend_up is not None:
                directional_levels = [lv for lv in levels if (lv['dir'] == 'dn') == trend_up]
            else:
                directional_levels = levels
            nearest = None
            for lv in directional_levels:
                combo_key = f"{lv['key']}_{tf}" if bot_state['prot_allow_multi_tf'] else lv['key']
                status = sym_state['gann_level_status'].get(combo_key)
                dist = abs(live_px - lv['price']); is_buy = (lv['dir'] == 'dn')
                if nearest is None or dist < nearest['dist']:
                    nearest = {'dist': dist, 'price': lv['price'], 'status': status, 'is_buy': is_buy}
            if nearest is None:
                lines.append(f"[{tf}] لا توجد مستويات متوافقة."); continue
            within_margin = nearest['dist'] <= margin
            reason_blocked = []
            if nearest['status'] == 'used': reason_blocked.append('المستوى مستخدم بالفعل')
            if not within_margin: reason_blocked.append(f"بعيد عن الهامش ({nearest['dist']:.3f} > {margin:.3f})")
            exec_mode = bot_state.get('gann_execution_mode', 'instant')
            closed_close = float(candles[-1]['close'])
            spike_limit = bot_state.get('gann_spike_limit_pts', 20) * SYMBOL_INFO[symbol]['pip_value']
            if within_margin and exec_mode == 'close':
                if abs(closed_close - nearest['price']) > margin:
                    reason_blocked.append(f"وضع Close: الإغلاق بعيد")
            elif within_margin and exec_mode == 'hybrid':
                if abs(live_px - closed_close) > spike_limit:
                    reason_blocked.append(f"وضع Hybrid: قفزة سعرية")
            status_icon = '✅ جاهز للدخول' if (within_margin and not reason_blocked) else ('🛑 ' + ' | '.join(reason_blocked))
            dir_lbl = 'دعم/شراء 🟢' if nearest['is_buy'] else 'مقاومة/بيع 🔴'
            lines.append(f"[{tf}] السعر: {live_px:.2f} | أقرب مستوى [{dir_lbl}]: {nearest['price']:.2f} (فرق {nearest['dist']:.3f}) | {status_icon}")
    return "\n".join(lines)


async def export_diag_log_excel() -> None:
    from telegram_ui import send_tg_msg, send_tg_document
    log = list(bot_state.get('diag_log', []))
    if not log:
        await send_tg_msg("لا يوجد سجل تشخيص محفوظ بعد.")
        return
    df = pd.DataFrame(log)
    if 'ts' in df.columns:
        df['الوقت (DAM)'] = df['ts'].apply(lambda t: _utc_to_dam(t).strftime('%Y-%m-%d %H:%M:%S') if pd.notna(t) else '')
        df = df.drop(columns=['ts'])
    preferred_order = ['الوقت (DAM)', 'symbol', 'tf', 'master_px', 'trend_up', 'margin',
                       'nearest_compatible_level', 'nearest_dist', 'within_margin',
                       'touch_attempted', 'skip_reason']
    cols = [c for c in preferred_order if c in df.columns] + [c for c in df.columns if c not in preferred_order]
    df = df[cols]
    fname = f"DiagLog_{datetime.now(timezone.utc).strftime('%Y%m%d_%H%M%S')}.xlsx"
    try:
        with pd.ExcelWriter(fname, engine='openpyxl') as writer:
            if 'symbol' in df.columns:
                for sym in sorted(df['symbol'].dropna().unique()):
                    df[df['symbol'] == sym].to_excel(writer, sheet_name=str(sym)[:31], index=False)
            else:
                df.to_excel(writer, sheet_name='diag_log', index=False)
        await send_tg_document(fname, f"📊 <b>سجل تشخيص تفصيلي كامل</b>\n{len(log)} سطر")
    finally:
        if os.path.exists(fname): os.remove(fname)


async def export_live_trades_excel() -> None:
    from telegram_ui import send_tg_msg, send_tg_document
    hist = list(bot_state.get('live_trade_history', []))
    if not hist:
        await send_tg_msg("لا يوجد سجل صفقات حية مغلقة بعد.")
        return
    wb = openpyxl.Workbook(); ws = wb.active; ws.title = "الصفقات الحية"
    headers = ["الزوج", "TF", "حقيقية/وهمية", "اتجاه", "وقت الفتح (DAM)", "وقت الإغلاق (DAM)",
               "المدة (د)", "مستوى الدخول", "الدخول الفعلي", "انزلاق الدخول", "TP", "SL",
               "سعر الإغلاق", "النتيجة", "ربح ($)", "مؤكد من الوسيط؟", "سبب الإغلاق",
               "BE مفعّل؟", "مصدر التغذية", "عمر التغذية (ms)", "نوع التنفيذ"]
    ws.append(headers)
    gray_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
    for cell in ws[1]: cell.fill = gray_fill; cell.font = Font(bold=True)
    _OUTCOME_DISPLAY = {'WIN': 'WIN ✅', 'LOSS': 'LOSS ❌', 'BREAK_EVEN': 'BREAK_EVEN ⚖️'}
    _TRIGGER_DISPLAY = {'touch': 'لمس مباشر ⚡', 'close': 'إغلاق شمعة ⏳', 'hybrid': 'تنفيذ هجين 🛡️'}
    green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    be_fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")
    running_bal = 0.0; n_win = n_loss = n_be = 0
    for tr in hist:
        pnl = tr.get('pnl') or 0.0; running_bal += pnl
        outcome = tr.get('outcome')
        if outcome == 'WIN': n_win += 1
        elif outcome == 'LOSS': n_loss += 1
        elif outcome == 'BREAK_EVEN': n_be += 1
        def _dam(iso):
            if not iso: return ''
            try:
                dt = datetime.fromisoformat(iso); return _utc_to_dam(dt).strftime('%Y-%m-%d %H:%M:%S')
            except Exception: return str(iso)
        row = [tr.get('symbol'), tr.get('tf'), 'حقيقية' if tr.get('is_real') else 'وهمية',
               'BUY 📈' if tr.get('is_buy') else 'SELL 📉', _dam(tr.get('opened_at')), _dam(tr.get('closed_at')),
               tr.get('duration_min'), tr.get('level_price'), tr.get('entry'), tr.get('entry_slippage'),
               tr.get('tp'), tr.get('sl'), tr.get('exit_price'), _OUTCOME_DISPLAY.get(outcome, outcome), pnl,
               '✅' if tr.get('pnl_confirmed_from_broker') else '⚠️ تقديري',
               tr.get('close_reason'), '✅' if tr.get('be_activated') else '—',
               tr.get('feed_source') or '—', tr.get('feed_age_ms'),
               _TRIGGER_DISPLAY.get(tr.get('trigger_type'), 'غير مسجَّل')]
        ws.append(row)
        fill = green_fill if outcome == 'WIN' else red_fill if outcome == 'LOSS' else be_fill if outcome == 'BREAK_EVEN' else None
        if fill:
            for col in range(1, len(headers) + 1): ws.cell(row=ws.max_row, column=col).fill = fill
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                          top=Side(style='thin'), bottom=Side(style='thin'))
    for row in ws.iter_rows():
        for cell in row: cell.border = thin_border; cell.alignment = Alignment(horizontal='center', vertical='center')
    for i in range(1, len(headers) + 1): ws.column_dimensions[get_column_letter(i)].width = 20.0
    fname = f"LiveTrades_{datetime.now(timezone.utc).strftime('%Y%m%d_%H%M%S')}.xlsx"
    wb.save(fname)
    try:
        total = len(hist); wr = round(100 * n_win / max(1, n_win + n_loss), 1)
        await send_tg_document(fname, f"📒 <b>سجل الصفقات الحية</b>\n{total} صفقة | WR: {wr}% | صافي: {running_bal:+.2f}$")
    finally:
        if os.path.exists(fname): os.remove(fname)


async def export_execution_details_report() -> str | None:
    hist = list(bot_state.get('live_trade_history', []))
    if not hist: return None
    rows = []
    for tr in hist:
        symbol = tr.get('symbol', ''); tf = tr.get('tf', '')
        is_buy = tr.get('is_buy'); pnl = tr.get('pnl', 0.0)
        outcome = tr.get('outcome', '')
        pv = SYMBOL_INFO.get(symbol, {}).get('pip_value', 0.01)
        exec_slip = tr.get('exec_slippage')
        slip_pips = round(exec_slip / pv, 1) if exec_slip is not None and pv else None
        method_raw = tr.get('exec_method')
        method_label = {'limit': 'Phase 1 (Limit IOC)', 'market_fallback': 'Phase 2 (Market FOK)'}.get(method_raw, method_raw or '')
        try:
            opened_dam = _utc_to_dam(datetime.fromisoformat(tr.get('opened_at'))).strftime('%Y-%m-%d %H:%M:%S') if tr.get('opened_at') else ''
            closed_dam = _utc_to_dam(datetime.fromisoformat(tr.get('closed_at'))).strftime('%Y-%m-%d %H:%M:%S') if tr.get('closed_at') else ''
        except Exception:
            opened_dam = ''; closed_dam = ''
        rows.append({'Pair': symbol, 'TF': tf, 'Direction': 'BUY' if is_buy else 'SELL',
                     'Open Time (DAM)': opened_dam, 'Close Time (DAM)': closed_dam,
                     'Level Price': tr.get('level_price'), 'Entry Price': tr.get('entry'),
                     'Exit Price': tr.get('exit_price'), 'PnL ($)': round(pnl, 2),
                     'Result': outcome, 'Execution Latency (ms)': tr.get('exec_latency_ms'),
                     'Execution Method': method_label, 'Final Slippage (Pips)': slip_pips,
                     'IOC Failure Reason': tr.get('exec_ioc_fail_reason') or ''})
    df = pd.DataFrame(rows)
    fname = f"Execution_Report_{datetime.now(timezone.utc).strftime('%Y%m%d_%H%M%S')}.xlsx"
    try:
        with pd.ExcelWriter(fname, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='Execution Details', index=False)
            for col_idx in range(1, len(df.columns) + 1):
                writer.sheets['Execution Details'].column_dimensions[
                    writer.sheets['Execution Details'].cell(row=1, column=col_idx).column_letter].width = 22
    except Exception as e:
        log_exception('export_execution_details_report', e); return None
    return fname
