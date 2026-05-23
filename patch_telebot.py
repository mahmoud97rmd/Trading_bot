#!/usr/bin/env python3
"""
patch_telebot.py
شغّله في Termux هكذا:
    pip install openpyxl --break-system-packages
    python3 patch_telebot.py
"""
import re, shutil, os

SRC = 'telebot.py'
shutil.copy(SRC, SRC + '.bak')   # نسخة احتياطية

with open(SRC, 'r', encoding='utf-8') as f:
    c = f.read()

errors = []

# ══════════════════════════════════════════════════════════════
# 1. bot_state ── إضافة حقول جديدة
# ══════════════════════════════════════════════════════════════
OLD = "    'use_time_filter': False, 'use_f_cons': False, 'cons_count': 3,"
NEW = (
    "    'use_time_filter': False, 'use_danger_filter': False,\n"
    "    'use_stoch_deep': True, 'use_stoch_mid': True, 'use_stoch_shal': True,\n"
    "    'use_f_cons': False, 'cons_count': 3,"
)
if OLD in c: c = c.replace(OLD, NEW); print("✅ 1. bot_state updated")
else: errors.append("❌ 1. bot_state not found")

# ══════════════════════════════════════════════════════════════
# 2. get_filters_keyboard ── تحديث شامل
# ══════════════════════════════════════════════════════════════
OLD = '''def get_filters_keyboard():
    t_icon = "🟢" if bot_state['use_trend_filter'] else "🔴"
    s_icon = "🟢" if bot_state['use_smart_filter'] else "🔴"
    return {"inline_keyboard": [
        [{"text": f"1. فلتر الترند الأساسي (50/150): {t_icon}", "callback_data": "toggle_trend"}],
        [{"text": f"2. الفلتر الذكي (تجاهل الترند 10/90): {s_icon}", "callback_data": "toggle_smart"}],
        [{"text": f"3. ثبات الترند ({bot_state['cons_count']} شموع): {'🟢' if bot_state['use_f_cons'] else '🔴'}", "callback_data": "toggle_f_cons"}],
        [{"text": f"4. Time Filter (08-18): {'🟢' if bot_state['use_time_filter'] else '🔴'}", "callback_data": "toggle_time"}],
        [{"text": "🔙 القائمة الرئيسية", "callback_data": "menu_main"}]
    ]}'''
NEW = '''def get_filters_keyboard():
    t_icon  = "🟢" if bot_state['use_trend_filter']  else "🔴"
    s_icon  = "🟢" if bot_state['use_smart_filter']  else "🔴"
    d_icon  = "🟢" if bot_state['use_danger_filter'] else "🔴"
    dp_icon = "🟢" if bot_state['use_stoch_deep']    else "🔴"
    md_icon = "🟢" if bot_state['use_stoch_mid']     else "🔴"
    sh_icon = "🟢" if bot_state['use_stoch_shal']    else "🔴"
    return {"inline_keyboard": [
        [{"text": f"1. فلتر الترند الأساسي (50/150): {t_icon}", "callback_data": "toggle_trend"}],
        [{"text": f"2. الفلتر الذكي (تجاهل الترند 10/90): {s_icon}", "callback_data": "toggle_smart"}],
        [{"text": f"3. ثبات الترند ({bot_state['cons_count']} شموع): {'🟢' if bot_state['use_f_cons'] else '🔴'}", "callback_data": "toggle_f_cons"}],
        [{"text": f"4. Time Filter (08-17 UTC): {'🟢' if bot_state['use_time_filter'] else '🔴'}", "callback_data": "toggle_time"}],
        [{"text": f"5. 🚫 حظر 19:00-21:30 دمشق: {d_icon}", "callback_data": "toggle_danger"}],
        [{"text": "━━━ مستويات الستوكاستيك ━━━", "callback_data": "noop"}],
        [{"text": f"DEEP 10/90: {dp_icon}", "callback_data": "toggle_stoch_deep"},
         {"text": f"MID  15/85: {md_icon}", "callback_data": "toggle_stoch_mid"},
         {"text": f"SHAL 20/80: {sh_icon}", "callback_data": "toggle_stoch_shal"}],
        [{"text": "🔙 القائمة الرئيسية", "callback_data": "menu_main"}]
    ]}'''
if OLD in c: c = c.replace(OLD, NEW); print("✅ 2. get_filters_keyboard updated")
else: errors.append("❌ 2. get_filters_keyboard not found")

# ══════════════════════════════════════════════════════════════
# 3. Stoch signals ── BACKTEST (indent=16 spaces)
# ══════════════════════════════════════════════════════════════
OLD = (
    "                # --- Stochastic Triple-Net Logic ---\n"
    "                buy_deep_stoch = (prev['K'] <= 10) and (curr['K'] > 10)\n"
    "                buy_mid_stoch  = (10 < prev['K'] <= 15) and (curr['K'] > 15)\n"
    "                buy_shal_stoch = (15 < prev['K'] <= 20) and (curr['K'] > 20)\n"
    "\n"
    "                sell_deep_stoch = (prev['K'] >= 90) and (curr['K'] < 90)\n"
    "                sell_mid_stoch  = (85 <= prev['K'] < 90) and (curr['K'] < 85)\n"
    "                sell_shal_stoch = (80 <= prev['K'] < 85) and (curr['K'] < 80)"
)
NEW = (
    "                # --- Stochastic Triple-Net Logic ---\n"
    "                buy_deep_stoch = (prev['K'] <= 10) and (curr['K'] > 10) and bot_state['use_stoch_deep']\n"
    "                buy_mid_stoch  = (10 < prev['K'] <= 15) and (curr['K'] > 15) and bot_state['use_stoch_mid']\n"
    "                buy_shal_stoch = (15 < prev['K'] <= 20) and (curr['K'] > 20) and bot_state['use_stoch_shal']\n"
    "\n"
    "                sell_deep_stoch = (prev['K'] >= 90) and (curr['K'] < 90) and bot_state['use_stoch_deep']\n"
    "                sell_mid_stoch  = (85 <= prev['K'] < 90) and (curr['K'] < 85) and bot_state['use_stoch_mid']\n"
    "                sell_shal_stoch = (80 <= prev['K'] < 85) and (curr['K'] < 80) and bot_state['use_stoch_shal']"
)
if OLD in c: c = c.replace(OLD, NEW); print("✅ 3. Backtest stoch toggles updated")
else: errors.append("❌ 3. Backtest stoch signals not found")

# ══════════════════════════════════════════════════════════════
# 4. Danger filter ── BACKTEST
# ══════════════════════════════════════════════════════════════
OLD = "                if bot_state['use_time_filter'] and not (8 <= curr['time'].hour <= 17): continue"
NEW = (
    "                if bot_state['use_time_filter'] and not (8 <= curr['time'].hour <= 17): continue\n"
    "                if bot_state['use_danger_filter']:\n"
    "                    _dh = (curr['time'].hour + 3) % 24\n"
    "                    _dm = curr['time'].minute\n"
    "                    if (_dh == 16) or (_dh == 17) or (_dh == 18 and _dm <= 30): continue"
)
if OLD in c: c = c.replace(OLD, NEW); print("✅ 4. Backtest danger filter added")
else: errors.append("❌ 4. Backtest time filter not found")

# ══════════════════════════════════════════════════════════════
# 5. Stoch signals ── LIVE SCANNER (indent=24 spaces)
# ══════════════════════════════════════════════════════════════
OLD = (
    "                        buy_deep_stoch = (prev['K'] <= 10) and (curr['K'] > 10)\n"
    "                        buy_mid_stoch  = (10 < prev['K'] <= 15) and (curr['K'] > 15)\n"
    "                        buy_shal_stoch = (15 < prev['K'] <= 20) and (curr['K'] > 20)\n"
    "\n"
    "                        sell_deep_stoch = (prev['K'] >= 90) and (curr['K'] < 90)\n"
    "                        sell_mid_stoch  = (85 <= prev['K'] < 90) and (curr['K'] < 85)\n"
    "                        sell_shal_stoch = (80 <= prev['K'] < 85) and (curr['K'] < 80)"
)
NEW = (
    "                        buy_deep_stoch = (prev['K'] <= 10) and (curr['K'] > 10) and bot_state['use_stoch_deep']\n"
    "                        buy_mid_stoch  = (10 < prev['K'] <= 15) and (curr['K'] > 15) and bot_state['use_stoch_mid']\n"
    "                        buy_shal_stoch = (15 < prev['K'] <= 20) and (curr['K'] > 20) and bot_state['use_stoch_shal']\n"
    "\n"
    "                        sell_deep_stoch = (prev['K'] >= 90) and (curr['K'] < 90) and bot_state['use_stoch_deep']\n"
    "                        sell_mid_stoch  = (85 <= prev['K'] < 90) and (curr['K'] < 85) and bot_state['use_stoch_mid']\n"
    "                        sell_shal_stoch = (80 <= prev['K'] < 85) and (curr['K'] < 80) and bot_state['use_stoch_shal']"
)
if OLD in c: c = c.replace(OLD, NEW); print("✅ 5. Live scanner stoch toggles updated")
else: errors.append("❌ 5. Live scanner stoch signals not found")

# ══════════════════════════════════════════════════════════════
# 6. Danger filter ── LIVE SCANNER
# ══════════════════════════════════════════════════════════════
OLD = (
    "                if bot_state['use_time_filter'] and not (8 <= h <= 17):\n"
    "                    bot_state['market_data'][tf] = f\"⏸ خمول | {c[-1]['close']}\""
)
NEW = (
    "                _danger_now = False\n"
    "                if bot_state['use_danger_filter']:\n"
    "                    _lh = (datetime.now(timezone.utc).hour + 3) % 24\n"
    "                    _lm = datetime.now(timezone.utc).minute\n"
    "                    _danger_now = (_lh == 16) or (_lh == 17) or (_lh == 18 and _lm <= 30)\n"
    "                if (bot_state['use_time_filter'] and not (8 <= h <= 17)) or _danger_now:\n"
    "                    bot_state['market_data'][tf] = f\"⏸ خمول | {c[-1]['close']}\""
)
if OLD in c: c = c.replace(OLD, NEW); print("✅ 6. Live scanner danger filter added")
else: errors.append("❌ 6. Live scanner time filter not found")

# ══════════════════════════════════════════════════════════════
# 7. Excel output ── 2 sheets + تنسيق ألوان
# ══════════════════════════════════════════════════════════════
OLD = '''        if trade_logs:
            import openpyxl
            from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
            
            df_logs = pd.DataFrame(trade_logs)
            total_trades = win_count + loss_count
            win_rate = round(win_count / total_trades * 100, 1) if total_trades > 0 else 0
            dd_pct = round(max_dd / peak_equity * 100, 1) if peak_equity > 0 else 0
            
            # فصل الصفقات المنفذة عن المرفوضة
            executed = df_logs[~df_logs['Outcome'].astype(str).str.contains('REJECTED')]
            rejected = df_logs[df_logs['Outcome'].astype(str).str.contains('REJECTED')]
            
            xlsx_filename = csv_filename.replace('.csv', '.xlsx')
            
            summary_data = {
                'البند': ['✅ الربح الكلي', '❌ الخسارة الكلية', '💰 المحصلة النهائية', '🎯 نسبة الفوز', '📉 أقصى سحب (DD)', '🔄 بريك إيفن'],
                'القيمة': [
                    f'{win_count} صفقة | +${round(total_win,2)}',
                    f'{loss_count} صفقة | -${abs(round(total_loss,2))}',
                    f'${round(total_prof,2)}',
                    f'{win_rate}% ({total_trades} صفقة)',
                    f'${round(max_dd,2)} ({dd_pct}%)',
                    str(be_count)
                ]
            }
            df_summary = pd.DataFrame(summary_data)
            
            with pd.ExcelWriter(xlsx_filename, engine='openpyxl') as writer:
                # ورقة 1: الصفقات المنفذة + الملخص
                executed.to_excel(writer, sheet_name='الصفقات', index=False, startrow=0)
                df_summary.to_excel(writer, sheet_name='الصفقات', index=False, startrow=len(executed)+3)
                
                # ورقة 2: الصفقات المرفوضة
                if not rejected.empty:
                    rejected.to_excel(writer, sheet_name='المرفوضة', index=False)
                
                # تنسيق ورقة الصفقات
                ws = writer.sheets['الصفقات']
                green_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
                red_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
                header_fill = PatternFill(start_color='2E4057', end_color='2E4057', fill_type='solid')
                header_font = Font(color='FFFFFF', bold=True)
                
                for cell in ws[1]:
                    cell.fill = header_fill
                    cell.font = header_font
                
                for row in ws.iter_rows(min_row=2, max_row=len(executed)+1):
                    outcome_val = str(row[8].value) if len(row) > 8 else ''
                    if outcome_val == 'WIN':
                        for cell in row: cell.fill = green_fill
                    elif outcome_val == 'LOSS':
                        for cell in row: cell.fill = red_fill
                
                for col in ws.columns:
                    max_len = max((len(str(cell.value or '')) for cell in col), default=10)
                    ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 30)
            
            await send_tg_document(xlsx_filename, f"📊 التقرير التفصيلي\\nالربح الصافي: ${round(total_prof, 2)}\\nنسبة الفوز: {win_rate}%\\nأقصى DD: ${round(max_dd,2)}")
            os.remove(xlsx_filename)'''

# The existing CSV block
OLD_CSV = '''        if trade_logs:
            df_logs = pd.DataFrame(trade_logs)
            total_trades = win_count + loss_count
            win_rate = round(win_count / total_trades * 100, 1) if total_trades > 0 else 0
            dd_pct = round(max_dd / peak_equity * 100, 1) if peak_equity > 0 else 0
            summary_row = pd.DataFrame([
                {'Timeframe': '══════════════', 'Type': '', 'Entry Time': '', 'Exit Time': '', 'Entry Price': '✅ الربح الكلي:', 'TP': '', 'SL': '', 'Pips': '', 'Outcome': f'{win_count} صفقة | +${round(total_win,2)}', 'Profit ($)': ''},
                {'Timeframe': '', 'Type': '', 'Entry Time': '', 'Exit Time': '', 'Entry Price': '❌ الخسارة الكلية:', 'TP': '', 'SL': '', 'Pips': '', 'Outcome': f'{loss_count} صفقة | -${abs(round(total_loss,2))}', 'Profit ($)': ''},
                {'Timeframe': '', 'Type': '', 'Entry Time': '', 'Exit Time': '', 'Entry Price': '💰 المحصلة النهائية:', 'TP': '', 'SL': '', 'Pips': '', 'Outcome': f'${round(total_prof,2)}', 'Profit ($)': ''},
                {'Timeframe': '', 'Type': '', 'Entry Time': '', 'Exit Time': '', 'Entry Price': '🎯 نسبة الفوز:', 'TP': '', 'SL': '', 'Pips': '', 'Outcome': f'{win_rate}% ({total_trades} صفقة)', 'Profit ($)': ''},
                {'Timeframe': '', 'Type': '', 'Entry Time': '', 'Exit Time': '', 'Entry Price': '📉 أقصى سحب (DD):', 'TP': '', 'SL': '', 'Pips': '', 'Outcome': f'${round(max_dd,2)} ({dd_pct}%)', 'Profit ($)': ''},
                {'Timeframe': '', 'Type': '', 'Entry Time': '', 'Exit Time': '', 'Entry Price': '🔄 بريك إيفن:', 'TP': '', 'SL': '', 'Pips': '', 'Outcome': str(be_count), 'Profit ($)': ''},
                {'Timeframe': '══════════════', 'Type': '', 'Entry Time': '', 'Exit Time': '', 'Entry Price': '', 'TP': '', 'SL': '', 'Pips': '', 'Outcome': '', 'Profit ($)': ''}
            ])
            pd.concat([df_logs, summary_row]).to_csv(csv_filename, index=False)
            await send_tg_document(csv_filename, f"📊 التقرير التفصيلي متاح الآن.\nالربح الصافي: ${round(total_prof, 2)}")
            os.remove(csv_filename)'''

NEW_EXCEL = '''        if trade_logs:
            from openpyxl.styles import PatternFill, Font
            df_logs = pd.DataFrame(trade_logs)
            total_trades = win_count + loss_count
            win_rate = round(win_count / total_trades * 100, 1) if total_trades > 0 else 0
            dd_pct = round(max_dd / peak_equity * 100, 1) if peak_equity > 0 else 0

            executed = df_logs[~df_logs['Outcome'].astype(str).str.contains('REJECTED')].copy()
            rejected = df_logs[df_logs['Outcome'].astype(str).str.contains('REJECTED')].copy()

            xlsx_filename = csv_filename.replace('.csv', '.xlsx')
            summary_data = {
                'البند': ['✅ الربح الكلي','❌ الخسارة الكلية','💰 المحصلة النهائية','🎯 نسبة الفوز','📉 أقصى سحب (DD)','🔄 بريك إيفن'],
                'القيمة': [
                    f'{win_count} صفقة | +${round(total_win,2)}',
                    f'{loss_count} صفقة | -${abs(round(total_loss,2))}',
                    f'${round(total_prof,2)}',
                    f'{win_rate}% ({total_trades} صفقة)',
                    f'${round(max_dd,2)} ({dd_pct}%)',
                    str(be_count)
                ]
            }
            df_summary = pd.DataFrame(summary_data)
            with pd.ExcelWriter(xlsx_filename, engine='openpyxl') as writer:
                executed.to_excel(writer, sheet_name='الصفقات', index=False)
                df_summary.to_excel(writer, sheet_name='الملخص', index=False)
                if not rejected.empty:
                    rejected.to_excel(writer, sheet_name='المرفوضة', index=False)
                # تلوين
                ws = writer.sheets['الصفقات']
                gf = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
                rf = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
                hf = PatternFill(start_color='2E4057', end_color='2E4057', fill_type='solid')
                for cell in ws[1]:
                    cell.fill = hf; cell.font = Font(color='FFFFFF', bold=True)
                outcome_col = [cell.value for cell in ws[1]].index('Outcome') + 1 if 'Outcome' in [cell.value for cell in ws[1]] else 9
                for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
                    val = str(row[outcome_col-1].value) if len(row) >= outcome_col else ''
                    if val == 'WIN':
                        for cell in row: cell.fill = gf
                    elif val == 'LOSS':
                        for cell in row: cell.fill = rf
                for col in ws.columns:
                    ws.column_dimensions[col[0].column_letter].width = min(max((len(str(cell.value or '')) for cell in col), default=8)+3, 28)

            await send_tg_document(xlsx_filename, f"📊 <b>التقرير التفصيلي</b>\n✅ ربح: +${round(total_win,2)} ({win_count} صفقة)\n❌ خسارة: -${abs(round(total_loss,2))} ({loss_count} صفقات)\n💰 صافي: ${round(total_prof,2)}\n🎯 فوز: {win_rate}%\n📉 DD: ${round(max_dd,2)}")
            os.remove(xlsx_filename)'''

if OLD_CSV in c:
    c = c.replace(OLD_CSV, NEW_EXCEL)
    print("✅ 7. Excel 2-sheet output added")
else:
    errors.append("❌ 7. CSV output block not found")

# ══════════════════════════════════════════════════════════════
# 8. Advanced backtest function ── قبل run_oanda_backtest
# ══════════════════════════════════════════════════════════════
ADVANCED_FUNC = '''
async def run_advanced_backtest(days=7):
    """تقرير متقدم على غرار MT5 Strategy Tester"""
    if bot_state['is_backtesting']:
        await send_tg_msg("⚠️ يوجد باك تيست قيد المعالجة.")
        return
    bot_state['is_backtesting'] = True
    start_dt = datetime.now(timezone.utc) - timedelta(days=days)
    await send_tg_msg(f"⏳ <b>Advanced Backtest</b>\\nمن: {start_dt.strftime('%Y-%m-%d')} ({days} أيام)\\n⏳ قد يستغرق بضع دقائق...")

    trade_logs, blocked_logs = [], []
    total_prof, peak_equity, max_dd = 0.0, 0.0, 0.0
    total_win, total_loss, win_count, loss_count, be_count = 0.0, 0.0, 0, 0, 0
    long_win, long_loss, short_win, short_loss = 0, 0, 0, 0
    all_profits, consec_win, consec_loss = [], 0, 0
    max_consec_win, max_consec_loss = 0, 0
    max_consec_win_usd, max_consec_loss_usd = 0.0, 0.0
    cur_streak_usd = 0.0

    try:
        for tf in bot_state['timeframes']:
            if not bot_state['active_tfs'][tf]: continue
            c_data = await fetch_oanda_candles('XAU_USD', tf, 5000)
            if len(c_data) < 150: continue
            df = calculate_indicators(pd.DataFrame(c_data).sort_values('time').reset_index(drop=True))
            safe_start = max(3, bot_state['cons_count'])

            for i in df[df['time'] >= start_dt].index:
                if i < safe_start: continue
                curr, prev = df.loc[i], df.loc[i-1]
                if bot_state['use_time_filter'] and not (8 <= curr['time'].hour <= 17): continue
                if bot_state['use_danger_filter']:
                    _dh = (curr['time'].hour + 3) % 24
                    _dm = curr['time'].minute
                    if (_dh == 16) or (_dh == 17) or (_dh == 18 and _dm <= 30): continue

                b_ema, s_ema = True, True
                cons = bot_state['cons_count'] if bot_state['use_f_cons'] else 1
                for j in range(cons):
                    cc = df.loc[i-j]
                    if not (cc['ema50'] > cc['ema150']): b_ema = False
                    if not (cc['ema150'] > cc['ema50']): s_ema = False
                trend_buy  = b_ema if bot_state['use_trend_filter'] else True
                trend_sell = s_ema if bot_state['use_trend_filter'] else True

                buy_deep_stoch  = (prev['K'] <= 10) and (curr['K'] > 10)  and bot_state['use_stoch_deep']
                buy_mid_stoch   = (10 < prev['K'] <= 15) and (curr['K'] > 15) and bot_state['use_stoch_mid']
                buy_shal_stoch  = (15 < prev['K'] <= 20) and (curr['K'] > 20) and bot_state['use_stoch_shal']
                sell_deep_stoch = (prev['K'] >= 90) and (curr['K'] < 90)  and bot_state['use_stoch_deep']
                sell_mid_stoch  = (85 <= prev['K'] < 90) and (curr['K'] < 85) and bot_state['use_stoch_mid']
                sell_shal_stoch = (80 <= prev['K'] < 85) and (curr['K'] < 80) and bot_state['use_stoch_shal']

                stoch_buy_any  = buy_deep_stoch  or buy_mid_stoch  or buy_shal_stoch
                stoch_sell_any = sell_deep_stoch or sell_mid_stoch or sell_shal_stoch

                buy_deep_v  = buy_deep_stoch  and (True if bot_state['use_smart_filter'] else trend_buy)
                buy_mid_v   = buy_mid_stoch   and trend_buy
                buy_shal_v  = buy_shal_stoch  and trend_buy
                sell_deep_v = sell_deep_stoch and (True if bot_state['use_smart_filter'] else trend_sell)
                sell_mid_v  = sell_mid_stoch  and trend_sell
                sell_shal_v = sell_shal_stoch and trend_sell
                buy_sig  = buy_deep_v  or buy_mid_v  or buy_shal_v
                sell_sig = sell_deep_v or sell_mid_v or sell_shal_v

                if not (buy_sig or sell_sig):
                    if stoch_buy_any or stoch_sell_any:
                        blocked_logs.append({
                            'Timeframe': tf, 'Type': 'BUY (BLOCKED)' if stoch_buy_any else 'SELL (BLOCKED)',
                            'Entry Time': (curr['time'] + timedelta(hours=3)).strftime('%Y-%m-%d %H:%M'),
                            'Entry Price': curr['close'], 'Reason': 'REJECTED (Trend blocked)'
                        })
                    continue

                if i + 1 >= len(df): continue
                next_c = df.loc[i + 1]
                entry_p = next_c['open']
                entry_t = next_c['time']
                m = 1 if buy_sig else -1
                act_ent = entry_p + (m * bot_state['spread_pips'] * bot_state['pip_value'])
                tp_dist = bot_state['tp_pips'][tf] * bot_state['pip_value']
                sl_dist = bot_state['sl_pips'][tf] * bot_state['pip_value']
                if bot_state['use_atr']:
                    tp_dist = curr['atr'] * bot_state['atr_mult_tp']
                    sl_dist = curr['atr'] * bot_state['atr_mult_sl']
                tp_p = round(act_ent + (m * tp_dist), 2)
                sl_p = round(act_ent - (m * sl_dist), 2)
                tol_val = bot_state['tp_tolerance_pips'] * bot_state['pip_value']
                eff_tp_p = (tp_p - tol_val) if buy_sig else (tp_p + tol_val)
                max_exit_time = min(entry_t + timedelta(hours=72), datetime.now(timezone.utc))
                outcome, exit_t = "EXPIRED", max_exit_time
                val_candles = await fetch_oanda_candles('XAU_USD', '1m', 4320, max_exit_time)
                be_activated = False
                be_target = act_ent + (m * 20 * bot_state['pip_value'])
                for vc in [v for v in val_candles if v['time'] >= entry_t]:
                    if buy_sig:
                        if bot_state['use_be'] and not be_activated and vc['high'] >= be_target:
                            sl_p = act_ent; be_activated = True
                        if vc['low'] <= sl_p: outcome, exit_t = ("BREAK-EVEN" if be_activated and sl_p == act_ent else "LOSS"), vc['time']; break
                        if vc['high'] >= eff_tp_p: outcome, exit_t = "WIN", vc['time']; break
                    else:
                        if bot_state['use_be'] and not be_activated and vc['low'] <= be_target:
                            sl_p = act_ent; be_activated = True
                        if vc['high'] >= sl_p: outcome, exit_t = ("BREAK-EVEN" if be_activated and sl_p == act_ent else "LOSS"), vc['time']; break
                        if vc['low'] <= eff_tp_p: outcome, exit_t = "WIN", vc['time']; break

                if outcome == "BREAK-EVEN": p_usd = 0.0; be_count += 1
                elif outcome in ["WIN","LOSS"]:
                    p_usd = round(abs(act_ent-(tp_p if outcome=="WIN" else sl_p))*100*bot_state['lot_size'],2)*(1 if outcome=="WIN" else -1)
                else: p_usd = 0.0

                # Streak tracking
                if outcome == "WIN":
                    total_win += p_usd; win_count += 1
                    consec_win += 1; cur_streak_usd += p_usd
                    if consec_win > max_consec_win: max_consec_win = consec_win; max_consec_win_usd = cur_streak_usd
                    consec_loss = 0; cur_streak_usd = 0.0 if outcome != "WIN" else cur_streak_usd
                    if buy_sig: long_win += 1
                    else: short_win += 1
                elif outcome == "LOSS":
                    total_loss += p_usd; loss_count += 1
                    consec_loss += 1; cur_streak_usd += p_usd
                    if consec_loss > max_consec_loss: max_consec_loss = consec_loss; max_consec_loss_usd = cur_streak_usd
                    consec_win = 0; cur_streak_usd = 0.0 if outcome != "LOSS" else cur_streak_usd
                    if buy_sig: long_loss += 1
                    else: short_loss += 1

                total_prof += p_usd
                peak_equity = max(peak_equity, total_prof)
                max_dd = max(max_dd, peak_equity - total_prof)
                all_profits.append(p_usd)

                b_type = "DEEP(10)" if buy_deep_v else "MID(15)" if buy_mid_v else "SHAL(20)"
                s_type = "DEEP(90)" if sell_deep_v else "MID(85)" if sell_mid_v else "SHAL(80)"
                _dh = (curr['time'].hour + 3) % 24
                trade_logs.append({
                    'Timeframe': tf,
                    'Type': f"BUY {b_type}" if buy_sig else f"SELL {s_type}",
                    'Entry Time': (entry_t + timedelta(hours=3)).strftime('%Y-%m-%d %H:%M'),
                    'Exit Time':  (exit_t + timedelta(hours=3)).strftime('%Y-%m-%d %H:%M'),
                    'Entry Price': round(act_ent, 2), 'TP': tp_p, 'SL': sl_p,
                    'Pips': round(abs(act_ent-(tp_p if outcome=="WIN" else sl_p))/bot_state['pip_value'],1) if outcome in ["WIN","LOSS"] else 0,
                    'Outcome': outcome, 'Profit ($)': p_usd,
                    'Hour_Damascus': _dh, 'Weekday': curr['time'].strftime('%a')
                })

        if not trade_logs:
            await send_tg_msg("⚠️ لم يتم العثور على صفقات."); return

        # ─── حساب الإحصائيات المتقدمة ───
        total_trades = win_count + loss_count
        win_rate = round(win_count/total_trades*100,1) if total_trades > 0 else 0
        dd_pct = round(max_dd/peak_equity*100,1) if peak_equity > 0 else 0
        profit_factor = round(total_win/abs(total_loss),2) if total_loss != 0 else 999
        expected_payoff = round(total_prof/total_trades,2) if total_trades > 0 else 0
        recovery_factor = round(total_prof/max_dd,2) if max_dd > 0 else 999
        wins_only = [p for p in all_profits if p > 0]
        losses_only = [p for p in all_profits if p < 0]
        avg_win  = round(sum(wins_only)/len(wins_only),2) if wins_only else 0
        avg_loss = round(sum(losses_only)/len(losses_only),2) if losses_only else 0
        largest_win  = round(max(wins_only),2) if wins_only else 0
        largest_loss = round(min(losses_only),2) if losses_only else 0

        # Entries by hour (Damascus)
        df_t = pd.DataFrame(trade_logs)
        hour_counts = df_t[df_t['Outcome'].isin(['WIN','LOSS'])].groupby('Hour_Damascus').size()
        hour_pnl    = df_t[df_t['Outcome'].isin(['WIN','LOSS'])].groupby('Hour_Damascus')['Profit ($)'].sum()
        day_counts  = df_t[df_t['Outcome'].isin(['WIN','LOSS'])].groupby('Weekday').size()

        def bar_chart(data_dict, width=20):
            if not data_dict: return "(لا يوجد بيانات)"
            mx = max(data_dict.values()) if data_dict.values() else 1
            lines = []
            for k, v in sorted(data_dict.items()):
                bar = "█" * int(v / mx * width)
                lines.append(f"  {str(k):>4} |{bar:<{width}}| {v}")
            return "\\n".join(lines)

        hour_chart = bar_chart(hour_counts.to_dict())
        day_chart  = bar_chart(day_counts.to_dict())

        report = (
            f"📊 <b>Advanced Strategy Report — {days} يوم</b>\\n"
            f"━━━━━━━━━━━━━━━━━━━━━━━\\n"
            f"<b>💰 الأرباح</b>\\n"
            f"  صافي الربح:      ${round(total_prof,2)}\\n"
            f"  إجمالي الربح:    +${round(total_win,2)}\\n"
            f"  إجمالي الخسارة:  -${abs(round(total_loss,2))}\\n"
            f"  Profit Factor:   {profit_factor}\\n"
            f"  Expected Payoff: ${expected_payoff}\\n"
            f"  Recovery Factor: {recovery_factor}\\n"
            f"━━━━━━━━━━━━━━━━━━━━━━━\\n"
            f"<b>📉 السحب (Drawdown)</b>\\n"
            f"  أقصى DD:         ${round(max_dd,2)} ({dd_pct}%)\\n"
            f"━━━━━━━━━━━━━━━━━━━━━━━\\n"
            f"<b>📈 الصفقات</b>\\n"
            f"  الإجمالي:        {total_trades}\\n"
            f"  فوز:             {win_count} ({win_rate}%)\\n"
            f"  خسارة:           {loss_count} ({round(100-win_rate,1)}%)\\n"
            f"  Long  Win/Loss:  {long_win}/{long_loss}\\n"
            f"  Short Win/Loss:  {short_win}/{short_loss}\\n"
            f"  بريك إيفن:       {be_count}\\n"
            f"━━━━━━━━━━━━━━━━━━━━━━━\\n"
            f"<b>🔢 إحصاءات الصفقات</b>\\n"
            f"  أكبر ربح:        +${largest_win}\\n"
            f"  أكبر خسارة:      ${largest_loss}\\n"
            f"  متوسط الربح:     +${avg_win}\\n"
            f"  متوسط الخسارة:   ${avg_loss}\\n"
            f"  أكبر سلسلة فوز:  {max_consec_win} صفقة (+${round(max_consec_win_usd,2)})\\n"
            f"  أكبر سلسلة خسارة:{max_consec_loss} صفقة (-${abs(round(max_consec_loss_usd,2))})\\n"
            f"━━━━━━━━━━━━━━━━━━━━━━━\\n"
            f"<b>🕐 توزيع بالساعة (دمشق):</b>\\n<pre>{hour_chart}</pre>\\n"
            f"<b>📅 توزيع بالأيام:</b>\\n<pre>{day_chart}</pre>"
        )
        await send_tg_msg(report)

        # ─── Excel للتقرير المتقدم ───
        if trade_logs:
            from openpyxl.styles import PatternFill, Font
            xlsx_adv = f"ADV_Report_{datetime.now().strftime('%H%M%S')}.xlsx"
            df_exec = pd.DataFrame(trade_logs).drop(columns=['Hour_Damascus','Weekday'], errors='ignore')
            df_blocked = pd.DataFrame(blocked_logs) if blocked_logs else pd.DataFrame()
            stats_data = {
                'المقياس': ['صافي الربح','إجمالي الربح','إجمالي الخسارة','Profit Factor','Expected Payoff','Recovery Factor',
                             'أقصى Drawdown','DD%','إجمالي الصفقات','فوز','خسارة','نسبة الفوز','بريك إيفن',
                             'Long Win/Loss','Short Win/Loss','أكبر ربح','أكبر خسارة','متوسط ربح','متوسط خسارة',
                             'أكبر سلسلة فوز','أكبر سلسلة خسارة'],
                'القيمة': [f'${round(total_prof,2)}',f'+${round(total_win,2)}',f'-${abs(round(total_loss,2))}',
                            profit_factor, expected_payoff, recovery_factor,
                            f'${round(max_dd,2)}',f'{dd_pct}%', total_trades, win_count, loss_count,
                            f'{win_rate}%', be_count,
                            f'{long_win}/{long_loss}', f'{short_win}/{short_loss}',
                            f'+${largest_win}', f'${largest_loss}', f'+${avg_win}', f'${avg_loss}',
                            f'{max_consec_win} (+${round(max_consec_win_usd,2)})',
                            f'{max_consec_loss} (-${abs(round(max_consec_loss_usd,2))})']
            }
            df_stats = pd.DataFrame(stats_data)
            with pd.ExcelWriter(xlsx_adv, engine='openpyxl') as writer:
                df_exec.to_excel(writer, sheet_name='الصفقات', index=False)
                df_stats.to_excel(writer, sheet_name='الإحصاءات', index=False)
                if not df_blocked.empty:
                    df_blocked.to_excel(writer, sheet_name='المرفوضة', index=False)
                ws = writer.sheets['الصفقات']
                gf = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
                rf = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
                hf = PatternFill(start_color='1F3864', end_color='1F3864', fill_type='solid')
                for cell in ws[1]: cell.fill = hf; cell.font = Font(color='FFFFFF', bold=True)
                for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
                    if len(row) >= 9:
                        val = str(row[8].value)
                        if val == 'WIN':
                            for cell in row: cell.fill = gf
                        elif val == 'LOSS':
                            for cell in row: cell.fill = rf
                for col in ws.columns:
                    ws.column_dimensions[col[0].column_letter].width = min(max((len(str(c.value or '')) for c in col), default=8)+3, 28)
            await send_tg_document(xlsx_adv, f"📊 Advanced Report — {days} يوم")
            os.remove(xlsx_adv)

    except Exception as e:
        c_log(f"❌ خطأ Advanced BT: {e}")
        await send_tg_msg(f"❌ خطأ: {e}")
    finally:
        bot_state['is_backtesting'] = False

'''

INSERT_BEFORE = "async def run_oanda_backtest"
if INSERT_BEFORE in c:
    c = c.replace(INSERT_BEFORE, ADVANCED_FUNC + INSERT_BEFORE)
    print("✅ 8. run_advanced_backtest function added")
else:
    errors.append("❌ 8. run_oanda_backtest not found to insert before")

# ══════════════════════════════════════════════════════════════
# 9. Callback handlers ── أزرار جديدة
# ══════════════════════════════════════════════════════════════
OLD_CB = "        elif d == \"toggle_time\": bot_state['use_time_filter'] = not bot_state['use_time_filter']; await edit_tg_msg(chat_id, msg_id, \"🎛 <b>فلاتر وشروط التداول:</b>\", get_filters_keyboard())"
NEW_CB = (
    "        elif d == \"toggle_time\": bot_state['use_time_filter'] = not bot_state['use_time_filter']; await edit_tg_msg(chat_id, msg_id, \"🎛 <b>فلاتر وشروط التداول:</b>\", get_filters_keyboard())\n"
    "        elif d == \"toggle_danger\": bot_state['use_danger_filter'] = not bot_state['use_danger_filter']; await edit_tg_msg(chat_id, msg_id, \"🎛 <b>فلاتر وشروط التداول:</b>\", get_filters_keyboard())\n"
    "        elif d == \"toggle_stoch_deep\": bot_state['use_stoch_deep'] = not bot_state['use_stoch_deep']; await edit_tg_msg(chat_id, msg_id, \"🎛 <b>فلاتر وشروط التداول:</b>\", get_filters_keyboard())\n"
    "        elif d == \"toggle_stoch_mid\":  bot_state['use_stoch_mid']  = not bot_state['use_stoch_mid'];  await edit_tg_msg(chat_id, msg_id, \"🎛 <b>فلاتر وشروط التداول:</b>\", get_filters_keyboard())\n"
    "        elif d == \"toggle_stoch_shal\": bot_state['use_stoch_shal'] = not bot_state['use_stoch_shal']; await edit_tg_msg(chat_id, msg_id, \"🎛 <b>فلاتر وشروط التداول:</b>\", get_filters_keyboard())"
)
if OLD_CB in c: c = c.replace(OLD_CB, NEW_CB); print("✅ 9. New callback handlers added")
else: errors.append("❌ 9. toggle_time callback not found")

# ══════════════════════════════════════════════════════════════
# 10. Backtest menu ── إضافة 7 أيام + Advanced
# ══════════════════════════════════════════════════════════════
OLD_BT_MENU = (
    "        elif d == \"menu_backtest\":\n"
    "            kb = {\"inline_keyboard\": [[{\"text\": \"📊 1 يوم\", \"callback_data\": \"bto_1\"}, {\"text\": \"📊 3 أيام\", \"callback_data\": \"bto_3\"}], [{\"text\": \"🔙 رجوع\", \"callback_data\": \"menu_main\"}]]}\n"
    "            await edit_tg_msg(chat_id, msg_id, \"اختر المدة أو أرسل /backtest:\", kb)"
)
NEW_BT_MENU = (
    "        elif d == \"menu_backtest\":\n"
    "            kb = {\"inline_keyboard\": [\n"
    "                [{\"text\": \"📊 1 يوم\", \"callback_data\": \"bto_1\"}, {\"text\": \"📊 3 أيام\", \"callback_data\": \"bto_3\"}, {\"text\": \"📊 7 أيام\", \"callback_data\": \"bto_7\"}],\n"
    "                [{\"text\": \"🔬 Advanced Report (MT5 Style)\", \"callback_data\": \"bto_adv_7\"}],\n"
    "                [{\"text\": \"🔙 رجوع\", \"callback_data\": \"menu_main\"}]\n"
    "            ]}\n"
    "            await edit_tg_msg(chat_id, msg_id, \"اختر المدة أو أرسل /backtest YYYY-MM-DD:\", kb)"
)
if OLD_BT_MENU in c: c = c.replace(OLD_BT_MENU, NEW_BT_MENU); print("✅ 10. Backtest menu updated (7d + Advanced)")
else: errors.append("❌ 10. menu_backtest not found")

# ══════════════════════════════════════════════════════════════
# 11. bto_ handler ── إضافة 7 أيام + Advanced
# ══════════════════════════════════════════════════════════════
OLD_BTO = (
    "        elif d.startswith(\"bto_\"):\n"
    "            days = int(d.split('_')[1])\n"
    "            asyncio.create_task(run_oanda_backtest(datetime.now(timezone.utc) - timedelta(days=days), mode='candle'))"
)
NEW_BTO = (
    "        elif d.startswith(\"bto_adv_\"):\n"
    "            adv_days = int(d.split('_')[2])\n"
    "            asyncio.create_task(run_advanced_backtest(days=adv_days))\n"
    "        elif d.startswith(\"bto_\"):\n"
    "            days = int(d.split('_')[1])\n"
    "            asyncio.create_task(run_oanda_backtest(datetime.now(timezone.utc) - timedelta(days=days), mode='candle'))"
)
if OLD_BTO in c: c = c.replace(OLD_BTO, NEW_BTO); print("✅ 11. bto_adv handler added")
else: errors.append("❌ 11. bto_ handler not found")

# ══════════════════════════════════════════════════════════════
# كتابة الملف النهائي
# ══════════════════════════════════════════════════════════════
with open(SRC, 'w', encoding='utf-8') as f:
    f.write(c)

print("\n" + "="*40)
if errors:
    print("⚠️ بعض التغييرات لم تُطبَّق:")
    for e in errors: print(f"  {e}")
else:
    print("🎉 جميع التغييرات طُبِّقت بنجاح!")
print("="*40)
print("💡 تشغيل: python3 telebot.py")
